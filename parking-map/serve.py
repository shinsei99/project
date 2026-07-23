#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
駐車場 配置図ビューア（複数駐車場切り替え対応）
 - 車室レイアウトは template.html に固定
 - 空き状況（契約者・賃料等）は起動/再読込の度に Dropbox のレントロールxlsxから取得
"""
import http.server, socketserver, json, webbrowser, threading, datetime, os, re, sys, time

BASE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(BASE, "template.html")
DROPBOX_ROOT = ("/Users/apple/Library/CloudStorage/Dropbox-大京商事　株式会社/共有フォルダ/"
                "（★必読★）新共有フォルダ/物件・管理/")
XLSX_PARKING = DROPBOX_ROOT + "レントロール一覧（駐車場他）.xlsx"
XLSX_MANSION = DROPBOX_ROOT + "レントロール一覧（マンション）.xlsx"
PORT = 8522

# cols = (No, 現況, 契約者, 区分, 賃料, 保証金, 契約日) の列番号（1始まり）
LOTS = {
    "yokozutsumi": {
        "file": XLSX_PARKING, "sheet": "角屋（横堤）モータープール",
        "start": 3, "end": None, "cols": (1, 2, 3, 4, 5, 6, 7),
    },
    "daikyo": {
        "file": XLSX_PARKING, "sheet": "大京モータープール(横堤P）",
        "start": 3, "end": None, "cols": (1, 2, 3, 4, 5, 6, 7),
    },
    "belliere": {
        # 駐車場は1〜7番の7区画（コインパーキング行は当駐車場とは無関係のため除外）
        "file": XLSX_MANSION, "sheet": "コーポ・ラ・ベリエール",
        "start": 60, "end": 66, "cols": (1, 2, 3, 4, 5, 9, 10),
    },
    "honjonishi": {
        # 図面上は26台分だが、当社管理は19〜26番の8区画のみ（1〜18番は管理外）
        "file": XLSX_PARKING, "sheet": "本庄西駐車場",
        "start": 3, "end": None, "cols": (1, 2, 3, 4, 5, 6, 7),
    },
    "juso": {
        "file": XLSX_PARKING, "sheet": "十三駐車場",
        "start": 3, "end": None, "cols": (1, 2, 3, 4, 5, 6, 7),
    },
    "shigino2": {
        "file": XLSX_PARKING, "sheet": "A-3476 鴫野東2丁目第2駐車場",
        "start": 3, "end": None, "cols": (1, 2, 3, 4, 5, 6, 7),
    },
    "eiwa3": {
        # このシートには保証金・契約日の列がない（列8は存在せず常にNoneになる）
        "file": XLSX_PARKING, "sheet": "E-3482 永和3丁目駐車場",
        "start": 3, "end": None, "cols": (1, 2, 3, 4, 5, 8, 8),
    },
}


def _s(v):
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y/%m/%d")
    return str(v).replace("　", " ").strip()


def _no(raw):
    """区画Noを数値化できれば数値、できなければ文字列（S1/N1等）のまま返す。
    「81　軽」のように末尾に注記が付く場合は先頭の数字部分だけを採用する。"""
    try:
        return int(raw)
    except (ValueError, TypeError):
        s = _s(raw)
        m = re.match(r"^(\d+)", s)
        return int(m.group(1)) if m else s


def build_data(cfg, wb):
    """指定シートを区画No→占有状況の辞書にする"""
    ws = wb[cfg["sheet"]]
    c_no, c_genkyo, c_who, c_kubun, c_chin, c_ho, c_date = cfg["cols"]
    end = cfg["end"] or ws.max_row
    data = {}
    for r in range(cfg["start"], end + 1):
        raw_no = ws.cell(r, c_no).value
        who = _s(ws.cell(r, c_who).value)
        no = _no(raw_no)
        if no == "":
            if not who or "blank_no" not in cfg:
                continue
            no = cfg["blank_no"]
        genkyo = _s(ws.cell(r, c_genkyo).value)
        kubun = _s(ws.cell(r, c_kubun).value)
        chin = ws.cell(r, c_chin).value
        ho = ws.cell(r, c_ho).value
        keiyakubi = _s(ws.cell(r, c_date).value)
        if genkyo == "空室" or not who:
            data[no] = {"v": 1}
        elif "オーナー" in who:
            data[no] = {"o": 1}
        else:
            data[no] = {"n": who, "k": kubun, "chin": chin, "ho": ho, "d": keiyakubi}
    return data


def build_all_data():
    from openpyxl import load_workbook
    wbs = {}
    all_data = {}
    for lot_id, cfg in LOTS.items():
        f = cfg["file"]
        if f not in wbs:
            wbs[f] = load_workbook(f, data_only=True)
        all_data[lot_id] = build_data(cfg, wbs[f])
    return all_data


# --- 自己修復ロジック ---------------------------------------------------
# launchd経由で起動したPythonプロセスは、Dropbox(CloudStorage)のFull Disk
# Access権限がまれに無効化された状態で立ち上がり、PermissionErrorになる。
# この状態はプロセスを作り直す（unload/load相当）と直ることが分かっているため、
# 権限エラーが連続したら自プロセスを終了し、launchdのKeepAliveで自動再起動させる。
# ただし恒久的な権限剥奪の場合に無限再起動しないよう、直近の再起動回数を
# logs/_restart_state.json に記録して上限を設ける。
RESTART_STATE = os.path.join(BASE, "logs", "_restart_state.json")
MAX_RESTARTS_PER_WINDOW = 5
RESTART_WINDOW_SEC = 1800
MAX_PERM_FAILS = 2

_perm_fail_count = 0


def _should_self_restart():
    try:
        with open(RESTART_STATE, encoding="utf-8") as f:
            state = json.load(f)
    except Exception:
        state = {}
    now = datetime.datetime.now().timestamp()
    if now - state.get("window_start", 0) > RESTART_WINDOW_SEC:
        state = {"window_start": now, "count": 0}
    if state.get("count", 0) >= MAX_RESTARTS_PER_WINDOW:
        return False
    state["count"] = state.get("count", 0) + 1
    try:
        os.makedirs(os.path.dirname(RESTART_STATE), exist_ok=True)
        with open(RESTART_STATE, "w", encoding="utf-8") as f:
            json.dump(state, f)
    except Exception:
        pass
    return True


def _handle_permission_error(e):
    global _perm_fail_count
    _perm_fail_count += 1
    print("xlsx読込エラー(権限) %d回目: %s" % (_perm_fail_count, e), flush=True)
    if _perm_fail_count < MAX_PERM_FAILS:
        return
    if _should_self_restart():
        print("Full Disk Access再評価のためプロセスを再起動します", flush=True)
        os._exit(1)  # launchdのKeepAlive=trueで自動再起動される
    else:
        print("再起動回数の上限に達したため再起動を見合わせます（要手動確認）", flush=True)


def build_all_data_safe():
    global _perm_fail_count
    try:
        data = build_all_data()
        _perm_fail_count = 0
        return data, None
    except PermissionError as e:
        _handle_permission_error(e)
        return None, e
    except Exception as e:
        return None, e


def _startup_warmup():
    """起動直後に先読みしておき、ブラウザで開いた瞬間から反映済みにする"""
    for attempt in range(6):
        data, err = build_all_data_safe()
        if err is None:
            print("起動時プリロード成功", flush=True)
            return
        print("起動時プリロード失敗(%d/6): %s" % (attempt + 1, err), flush=True)
        time.sleep(5)


def render():
    tpl = open(TEMPLATE, encoding="utf-8").read()
    all_data, err = build_all_data_safe()
    if err is not None:
        all_data = {lot_id: {} for lot_id in LOTS}
        note = "⚠ xlsx読込エラー: %s" % err
    else:
        note = "自動反映"
    stamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    html = tpl.replace("__ALL_DATA_JSON__", json.dumps(all_data, ensure_ascii=False))
    html = html.replace("__UPDATED__", "%s（%s）" % (stamp, note))
    return html


class Handler(http.server.BaseHTTPRequestHandler):
    def do_GET(self):
        if self.path.split("?")[0] not in ("/", "/index.html"):
            self.send_response(404)
            self.end_headers()
            return
        body = render().encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, *a):
        pass


if __name__ == "__main__":
    url = "http://localhost:%d" % PORT
    if "--daemon" not in sys.argv:
        threading.Timer(0.8, lambda: webbrowser.open(url)).start()
    threading.Thread(target=_startup_warmup, daemon=True).start()
    socketserver.TCPServer.allow_reuse_address = True
    with socketserver.TCPServer(("", PORT), Handler) as httpd:
        print("配置図サーバー起動: %s  （Ctrl+Cで終了）" % url)
        print("レントロール: %s" % XLSX_PARKING)
        print("レントロール: %s" % XLSX_MANSION)
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("\n終了しました。")
