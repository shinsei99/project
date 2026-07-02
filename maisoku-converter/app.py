"""他社マイソク → 自社 不動産案内書 Excel 変換ツール

Claude Code CLI を subprocess で呼び出して画像解析。
APIキー不要・Claude Pro/Max サブスクリプションで動作。
テンプレート XLS をそのまま利用し、値セルのみ書き換えて出力。
"""
from __future__ import annotations

import base64
import io
import json
import re
import subprocess
import tempfile
from pathlib import Path

import streamlit as st
import streamlit.components.v1 as components
from PIL import Image, ImageDraw
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import (
    TwoCellAnchor, OneCellAnchor, AnchorMarker,
)
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

# ── streamlit-drawable-canvas 互換パッチ ──────────────────────────────────────
# Streamlit 1.50 で image_to_url が移動されたため最小限の互換スタブを追加。
# ※ 背景画像は background_image パラメータを使わず init_drawing に直接埋め込む方式に
#   変更したため、このパッチが呼ばれることは基本的にない。
try:
    import streamlit.elements.image as _st_img
    if not hasattr(_st_img, "image_to_url"):
        def _image_to_url(image, width=-1, clamp=False, channels="RGB",
                          output_format="auto", image_id=""):
            if isinstance(image, str):
                return image
            try:
                from streamlit.runtime import get_instance, exists
                if exists():
                    buf = io.BytesIO()
                    if hasattr(image, "convert"):
                        image.convert("RGB").save(buf, "PNG")
                    return get_instance().media_file_mgr.add(
                        buf.getvalue(), "image/png", image_id)
            except Exception:
                pass
            return ""
        _st_img.image_to_url = _image_to_url
except Exception:
    pass

# ─── パス定数 ────────────────────────────────────────────────────────────────
APP_DIR      = Path(__file__).parent
LOGO_PATH    = APP_DIR / "company_logo.png"

# ─── 間取り図トレーサー（Gemini image-to-image）─────────────────────────────
try:
    from google import genai as _genai
    from google.genai import types as _gtypes
    import sys as _sys
    _sys.path.insert(0, "/Users/apple/madori-tracer")
    from config import GEMINI_API_KEY as _GEMINI_KEY  # 環境変数 or .secret_key から取得
    _gemini_client = _genai.Client(api_key=_GEMINI_KEY)
    TRACER_AVAILABLE = True
except Exception:
    TRACER_AVAILABLE = False

_TRACER_TYPE_RULES = {
    "マンション": """
- Emphasize corner structural pillars (マンション柱): solid filled black squares at all corners and wall junctions
- Include PS (パイプスペース) and MB (メーターボックス) labels
- Balcony (バルコニー) has thin border and white fill
- 浴室 and 洗面室 are always separate adjacent rooms — label both
""",
    "戸建て": """
- Include staircase (階段) with step lines if present
- Include entrance (玄関) with step/platform indication
- Garden/parking area (庭・駐車場) if shown: thin border, white fill, label
- No structural pillars (柱) unless explicitly shown
- Include 和室 (Japanese-style room) with tatami lines if present
""",
    "1K・1R": """
- Very simple layout: one main room + bathroom/kitchen unit
- Kitchen counter along wall: sink rectangle + stove circles
- Unit bath (ユニットバス): single rectangle with bathtub oval + toilet oval inside
- Compact layout — preserve exact proportions carefully
""",
    "その他": "",
}

_TRACER_BASE = """## RULE 1 — TRACING OPERATION (NOT DESIGN)
You must trace the original spatial layout exactly.
Every room, wall, door, and equipment item must stay in its EXACT original position and size.
Do NOT rearrange, move, add, or omit ANY element. Do NOT apply your own design preferences.

## RULE 2 — STYLE CONVERSION ONLY
Change ONLY the visual style: colors → black & white, simplified icons, clean fonts.
Spatial structure must be 100% identical to the input.

## RULE 3 — ROOM LABELING (MANDATORY)
Label EVERY room with its exact Japanese name from the original.
NEVER merge two separate rooms into one label.
- 浴室 and 洗面室 are ALWAYS separate rooms with separate labels
- 物入 and クローゼット are separate rooms
- All storage spaces must be individually labeled

## RULE 4 — EQUIPMENT POSITION (MANDATORY)
Place ALL equipment in the EXACT same position as the original.
Kitchen equipment (sink + stove) belongs ONLY inside the LDK/kitchen area — NEVER in bedrooms.
Do NOT relocate any equipment based on "typical" floor plan assumptions.

## RULE 5 — VISUAL STYLE
- Background: pure white
- Outer walls: thick solid black lines
- Interior partition walls: medium black lines
- All room fills: white (remove ALL color — no orange, tan, blue, gray, green)
- Room labels: copy EXACTLY as shown in the original image
  * If original shows name only (e.g. "玄関", "トイレ"): write just the name
  * If original shows name+size as one combined label (e.g. "洋室6帖"): write as ONE line — do NOT add a separate size line below
  * If original shows name and size on separate lines (e.g. "LDK" above "11.9帖"): write them on two separate lines
  * NEVER duplicate: if "6帖" is already inside the label, do NOT add "6帖" again below it
- Equipment icons (ultra-simplified):
  * Toilet: rectangle + oval bowl
  * Bathtub: rectangle + oval tub
  * Sink/洗面台: rectangle + circle
  * Kitchen: counter rectangle + sink box + 3 stove circles
  * Washing machine: rectangle + large circle
- Doors: straight line + quarter-circle arc (dashed)
- Windows: three parallel lines across wall
- PS / MB / 棚: keep as small text labels
- North compass: circle with filled black arrow + "N", bottom-right corner
"""


def _encode_image(image: Image.Image) -> bytes:
    img = image.copy()
    if img.width > 1400:
        r = 1400 / img.width
        img = img.resize((1400, int(img.height * r)), Image.LANCZOS)
    buf = io.BytesIO()
    img.convert("RGB").save(buf, "JPEG", quality=92)
    return buf.getvalue()


def _trace_madori(
    image: Image.Image,
    floor_type: str = "マンション",
    correction: str = "",
    prev_result: bytes | None = None,
) -> bytes:
    """間取り図画像をGeminiで白黒図面に引き直す。correction+prev_resultで再修正も可能。"""
    type_rules = _TRACER_TYPE_RULES.get(floor_type, "")

    if correction and prev_result:
        # 再修正：前回の白黒結果をベースに指定箇所だけ修正
        prev_img = Image.open(io.BytesIO(prev_result)).convert("RGB")
        img_bytes = _encode_image(prev_img)
        prompt = f"""This is a black-and-white floor plan that was already generated.
Make ONLY the following correction and leave everything else EXACTLY as-is:

"{correction}"

Do NOT redraw, restyle, or move anything that is not mentioned above.
Keep all room labels, wall positions, icons, and proportions identical to this image.
Output the corrected floor plan in the same black-and-white style."""

    elif correction:
        img_bytes = _encode_image(image)
        prompt = f"""Redraw this floor plan as a clean black-and-white floor plan.
Apply this correction: "{correction}"

{_TRACER_BASE}
## FLOOR TYPE: {floor_type}
{type_rules}"""

    else:
        img_bytes = _encode_image(image)
        prompt = f"""Redraw this floor plan as a clean black-and-white floor plan.

{_TRACER_BASE}
## FLOOR TYPE: {floor_type}
{type_rules}"""

    resp = _gemini_client.models.generate_content(
        model="gemini-3.1-flash-image",
        contents=[
            _gtypes.Part(inline_data=_gtypes.Blob(mime_type="image/jpeg", data=img_bytes)),
            _gtypes.Part(text=prompt),
        ],
        config=_gtypes.GenerateContentConfig(response_modalities=["IMAGE", "TEXT"]),
    )
    for part in resp.candidates[0].content.parts:
        if part.inline_data is not None:
            return part.inline_data.data
    texts = [p.text for p in resp.candidates[0].content.parts if p.text]
    raise RuntimeError("画像が生成されませんでした。\n" + "\n".join(texts)[:300])

# ── カスタムコンポーネント（画像配置エディタ）────────────────────────────────
_placement_editor = components.declare_component(
    "placement_editor",
    path=str(APP_DIR / "placement_component"),
)

def _pil_to_b64(img: Image.Image, max_size: int = 320) -> str:
    thumb = img.copy()
    thumb.thumbnail((max_size, max_size), Image.LANCZOS)
    buf = io.BytesIO()
    thumb.convert("RGB").save(buf, "JPEG", quality=82)
    return "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode()
COMPANY_JSON = APP_DIR / "company_info.json"

# ─── 特徴アイコン（73種）─────────────────────────────────────────────────────
FEATURES_DIR = APP_DIR / "features"

def _load_feature_catalog() -> list[tuple[str, Path]]:
    """features/ 内の `NN_ラベル.png` を index 順に [(label, path), ...] で返す。"""
    out: list[tuple[int, str, Path]] = []
    if FEATURES_DIR.exists():
        for p in FEATURES_DIR.glob("*.png"):
            stem = p.stem
            if "_" in stem:
                num, _, label = stem.partition("_")
                try:
                    out.append((int(num), label, p))
                except ValueError:
                    out.append((9999, stem, p))
    out.sort(key=lambda t: t[0])
    return [(label, path) for _, label, path in out]

FEATURE_CATALOG = _load_feature_catalog()              # [(label, path), ...]
FEATURE_LABELS  = [lbl for lbl, _ in FEATURE_CATALOG]   # ラベルのみ

# テンプレート XLS（サンプルをそのまま使用）
TEMPLATE_XLS = {
    "賃貸":         APP_DIR / "templates" / "賃貸案内書_blank.xls",
    "売買マンション": APP_DIR / "templates" / "売却マンション案内書_blank.xls",
    "売買一般":     APP_DIR / "templates" / "売却一般案内書_blank.xls",
}

# ─── 会社情報デフォルト ──────────────────────────────────────────────────────
COMPANY_DEFAULTS: dict = {
    "商号":           "大京商事株式会社",
    "宅建免許番号":    "大阪府知事（9）第27334",
    "所在地":         "大阪市都島区東野田町２－３－１４",
    "TEL":            "０６－６３５３－０４１８",
    "FAX":            "０６－６３５３－０２８０",
    "担当者":         "鷲見",
    "MAIL":           "info@daikyocorp.co.jp",
    "URL":            "http://www.daikyocorp.co.jp",
    "建設業免許番号":  "",
    "取引態様":       "売主",
    "チラシ":         "可",
    "情報誌":         "要確認",
    "インターネット": "要確認",
    "報酬形態":       "",
}

def load_company_info() -> dict:
    if COMPANY_JSON.exists():
        try:
            return json.loads(COMPANY_JSON.read_text(encoding="utf-8"))
        except Exception:
            pass
    return COMPANY_DEFAULTS.copy()

def save_company_info(info: dict) -> None:
    COMPANY_JSON.write_text(json.dumps(info, ensure_ascii=False, indent=2), encoding="utf-8")

# ─── Claude CLI ───────────────────────────────────────────────────────────────
CLAUDE_BIN     = "/opt/homebrew/bin/claude"
CLAUDE_TIMEOUT = 600

CLAUDE_PROMPT = """\
このファイルは不動産マイソク（物件チラシ）の画像です（ファイル名: {filename}）。
Read ツールでこのファイルを開いて内容を確認してください。
以下の JSON のみを返答してください（説明文・前置き一切不要）。

{{
  "template_type": "賃貸 / 売買マンション / 売買一般 のいずれか（マンション=区分所有、戸建・土地=一般）",
  "specs": {{
    "物件名":     "物件名がなければ間取り＋所在地から作成",
    "種目":       "例: 分譲貸マンション / 中古戸建 / 売地",
    "賃料":       "賃貸のみ。例: 128000",
    "管理費":     "賃貸のみ。なければ空欄",
    "保証金":     "賃貸のみ",
    "敷金ヶ月":   "賃貸のみ。数字のみ: 1",
    "礼金ヶ月":   "賃貸のみ。数字のみ: 1",
    "更新料":     "賃貸のみ",
    "価格":       "売買のみ。例: 350",
    "所在地":     "住所",
    "交通":       "最寄り線と駅名。例: 地下鉄堺筋線・長堀鶴見緑地線 長堀橋駅",
    "交通2":      "徒歩分・バス等の補足。例: 地下鉄長堀鶴見緑地線 松屋町駅 徒歩5分",
    "物件名":     "建物名",
    "間取り":     "例: 1LDK",
    "間取り詳細": "例: 洋室10.5・LDK30.5",
    "構造":       "例: SRC / RC / 木",
    "階数_地上":  "数字のみ: 10",
    "部屋階":     "数字のみ: 9",
    "専有面積㎡": "数字のみ: 95.84",
    "専有面積坪": "数字のみ: 28.99",
    "バルコニー向き": "例: 南",
    "現況":       "例: 空家 / 居住中",
    "築年号":     "例: 昭和 / 平成",
    "築年":       "数字のみ: 57",
    "築月":       "数字のみ: 12",
    "引渡":       "例: 即時 / 相談",
    "設備":       "テキスト: 電気・都市ガス・上水道・公共下水道",
    "土地面積㎡": "売買のみ。数字のみ",
    "土地面積坪": "売買のみ。数字のみ",
    "延床面積㎡": "売買一般のみ。数字のみ",
    "延床面積坪": "売買一般のみ。数字のみ",
    "建ぺい率":   "売買一般のみ。数字のみ: 50",
    "容積率":     "売買一般のみ。数字のみ: 100",
    "坪単価":     "売買マンションのみ。例: 120万円",
    "借地料":     "売買マンションのみ。なければ空欄",
    "修繕積立金": "売買マンションのみ。例: 12000",
    "持分":       "売買マンションのみ。敷地権割合など",
    "総戸数":     "売買マンションのみ。数字のみ",
    "施主":       "売買マンションのみ。なければ空欄",
    "施工":       "売買マンションのみ。施工会社",
    "管理会社":   "売買マンションのみ。委託先",
    "管理形態":   "売買マンションのみ。例: 全部委託",
    "管理人":     "売買マンションのみ。例: 日勤 / 巡回",
    "権利":       "売買のみ。例: 所有権 / 借地権",
    "私道負担":   "売買一般のみ。なければ空欄",
    "地目":       "売買一般のみ。例: 宅地",
    "地勢":       "売買一般のみ。例: 平坦",
    "都市計画":   "売買一般のみ。例: 市街化区域",
    "その他規制": "売買一般のみ。なければ空欄",
    "その他":     ""
  }},
  "catchphrases": [
    "自社風キャッチコピー1（15文字以内）",
    "自社風キャッチコピー2（15文字以内）",
    "自社風キャッチコピー3（15文字以内）"
  ],
  "regions": [
    {{"種類": "外観写真", "x1": 0.0, "y1": 0.0, "x2": 0.0, "y2": 0.0}},
    {{"種類": "室内写真", "x1": 0.0, "y1": 0.0, "x2": 0.0, "y2": 0.0}},
    {{"種類": "間取り図", "x1": 0.0, "y1": 0.0, "x2": 0.0, "y2": 0.0}},
    {{"種類": "地図",     "x1": 0.0, "y1": 0.0, "x2": 0.0, "y2": 0.0}}
  ],
  "features": ["該当する特徴ラベルのみを配列で（下記リストの語句と完全一致）"]
}}

regions: x1,y1=左上、x2,y2=右下（画像全体に対する割合 0.0〜1.0）
catchphrases: 元のコピーを使わず自社独自にリライト
features: マイソク記載・図面・写真から判断し、確実に該当する特徴のみを次のリストから完全一致で選ぶ（推測で増やさない／該当なしは []）:
{feature_list}
"""

REGION_COLORS = {
    "外観写真": (220, 50,  50),
    "室内写真": (50,  120, 220),
    "間取り図": (50,  180, 80),
    "地図":    (160, 60,  200),
}

# ─── ユーティリティ ───────────────────────────────────────────────────────────

def convert_to_image(file_bytes: bytes, filename: str) -> Image.Image:
    ext = Path(filename).suffix.lower()
    if ext == ".pdf":
        try:
            import fitz
            doc = fitz.open(stream=file_bytes, filetype="pdf")
            pix = doc[0].get_pixmap(matrix=fitz.Matrix(2.0, 2.0))
            return Image.open(io.BytesIO(pix.tobytes("png"))).convert("RGB")
        except ImportError:
            raise RuntimeError("PDF変換には pymupdf が必要です: pip install pymupdf")
    return Image.open(io.BytesIO(file_bytes)).convert("RGB")


def analyze_with_claude(image: Image.Image) -> dict:
    analysis_img = image.copy()
    if analysis_img.width > 2000:
        r = 2000 / analysis_img.width
        analysis_img = analysis_img.resize((2000, int(analysis_img.height * r)), Image.LANCZOS)

    # ── 解析前に向き（縦横・回転）を自動補正 ────────────────────────────────
    # PDF・画像どちらもここに PIL 画像として集約されるので 1 箇所で両経路を補正。
    try:
        from pdf_orient import ensure_upright_image
        _buf = io.BytesIO()
        analysis_img.save(_buf, "PNG")
        png_bytes = ensure_upright_image(_buf.getvalue())
        analysis_img = Image.open(io.BytesIO(png_bytes)).convert("RGB")
    except Exception:
        pass  # 向き補正に失敗しても元画像で続行

    with tempfile.TemporaryDirectory(prefix="maisoku_") as tmp:
        p = Path(tmp) / "maisoku.png"
        analysis_img.save(p, "PNG")
        cmd = [CLAUDE_BIN, "-p", CLAUDE_PROMPT.format(
                   filename=p.name, feature_list="／".join(FEATURE_LABELS)),
               "--output-format", "json", "--tools", "Read",
               "--add-dir", tmp, "--dangerously-skip-permissions", "--model", "sonnet"]
        try:
            proc = subprocess.run(cmd, cwd=tmp, capture_output=True,
                                  text=True, timeout=CLAUDE_TIMEOUT)
        except FileNotFoundError:
            raise RuntimeError("`claude` コマンドが見つかりません。")
        except subprocess.TimeoutExpired:
            raise RuntimeError(f"解析タイムアウト（{CLAUDE_TIMEOUT}秒）。")

        if proc.returncode != 0:
            raise RuntimeError(f"Claude エラー (code={proc.returncode}):\n{proc.stderr[:400]}")

        try:
            result = json.loads(proc.stdout)
        except json.JSONDecodeError:
            raise RuntimeError("Claude レスポンスの JSON 解析失敗。")

        if result.get("is_error"):
            raise RuntimeError(f"Claude エラー: {result.get('result')}")

        raw = result.get("result", "")
        m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", raw, re.DOTALL)
        json_str = m.group(1) if m else raw.strip()
        if not json_str.startswith("{"):
            m2 = re.search(r"\{.*\}", json_str, re.DOTALL)
            if m2:
                json_str = m2.group(0)
        return json.loads(json_str)


def crop_region(image: Image.Image, region: dict) -> Image.Image | None:
    x1, y1 = region.get("x1", 0), region.get("y1", 0)
    x2, y2 = region.get("x2", 0), region.get("y2", 0)
    if x2 <= x1 or y2 <= y1:
        return None
    w, h = image.size
    return image.crop((int(x1*w), int(y1*h), int(x2*w), int(y2*h)))


def crop_overlay(image: Image.Image, region: dict,
                 color=(40, 180, 90)) -> Image.Image:
    """元画像に、切り取り範囲を示す半透明の目安枠を重ねて返す。
    範囲外は暗く落とし、範囲は明るいまま＋カラー枠線で強調する。
    """
    base = image.convert("RGB")
    w, h = base.size
    x1 = max(0, min(w, int(region.get("x1", 0) * w)))
    y1 = max(0, min(h, int(region.get("y1", 0) * h)))
    x2 = max(0, min(w, int(region.get("x2", 1) * w)))
    y2 = max(0, min(h, int(region.get("y2", 1) * h)))
    # 全体を45%暗く → 範囲だけ元に戻す
    out = Image.blend(base, Image.new("RGB", (w, h), (0, 0, 0)), 0.45)
    if x2 > x1 and y2 > y1:
        out.paste(base.crop((x1, y1, x2, y2)), (x1, y1))
    d  = ImageDraw.Draw(out)
    lw = max(3, w // 250)
    d.rectangle([x1, y1, x2 - 1, y2 - 1], outline=color, width=lw)
    return out


# ─── テンプレート XLS → openpyxl 変換 ───────────────────────────────────────

def _load_xls_as_openpyxl(xls_path: Path) -> tuple[openpyxl.Workbook,
                                                     openpyxl.worksheet.worksheet.Worksheet]:
    """XLS テンプレートを openpyxl Workbook として読み込む（構造・値を転写）。"""
    import xlrd

    xls  = xlrd.open_workbook(str(xls_path), formatting_info=True)
    xs   = xls.sheet_by_index(0)
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = "案内書"

    # 右側(AZ列〜)の種目／用途地域 選択肢欄は入力参考用でこのアプリでは不要なため、
    # 本体(写真＋スペック＋会社情報＝AX列まで=50列)のみ出力ワークブックへ転写する。
    MAX_COL = 50  # AX 列

    # ページ設定
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered   = True   # 余白を上下に均等分散
    ws.page_margins.left   = 0.16
    ws.page_margins.right  = 0.16
    ws.page_margins.top    = 0.16
    ws.page_margins.bottom = 0.16
    ws.page_margins.header = 0.0
    ws.page_margins.footer = 0.0

    # 行高を少し拡大して用紙の縦をより埋める（下余白対策）
    ROW_SCALE = 1.12

    # 列幅
    for ci in range(min(xs.ncols, MAX_COL)):
        cw      = xs.colinfo_map.get(ci)
        col_ltr = get_column_letter(ci + 1)
        if cw:
            ws.column_dimensions[col_ltr].width  = max(cw.width / 256, 0.1)
            ws.column_dimensions[col_ltr].hidden = bool(cw.hidden)

    # 行高さ（ROW_SCALE 倍して縦を少し拡大）
    for ri in range(xs.nrows):
        rh = xs.rowinfo_map.get(ri)
        if rh:
            ws.row_dimensions[ri + 1].height = (rh.height / 20) * ROW_SCALE

    # マージセル（AX列以降の選択肢欄のマージは転写しない）
    for r1, r2, c1, c2 in xs.merged_cells:
        if c1 >= MAX_COL:        # 開始列が AY 以降 → 右側参考欄なのでスキップ
            continue
        try:
            ws.merge_cells(
                start_row=r1 + 1, start_column=c1 + 1,
                end_row=r2,       end_column=min(c2, MAX_COL),
            )
        except Exception:
            pass

    # xlrd line_style → openpyxl border style 変換テーブル
    _BORDER_STYLE = {
        1: "thin", 2: "medium", 3: "dashed", 4: "dotted",
        5: "thick", 6: "double", 7: "hair", 8: "mediumDashed",
        9: "dashDot", 10: "mediumDashDot", 11: "dashDotDot",
        12: "mediumDashDotDot", 13: "slantDashDot",
    }

    def _xlrd_color(idx: int) -> str:
        # 64 = automatic (black); それ以外はパレットから取得
        if idx == 64:
            return "000000"
        try:
            r, g, b = xls.colour_map.get(idx, (0, 0, 0))
            return f"{r:02X}{g:02X}{b:02X}"
        except Exception:
            return "000000"

    def _side(line_style: int, colour_index: int) -> Side | None:
        bs = _BORDER_STYLE.get(line_style)
        if bs is None:
            return None
        return Side(border_style=bs, color=_xlrd_color(colour_index))

    # ── xlrd xf → openpyxl スタイル変換ヘルパー（フォント/配置/塗り/表示形式/罫線） ──
    _HOR = {1: "left", 2: "center", 3: "right", 4: "fill",
            5: "justify", 6: "centerContinuous", 7: "distributed"}
    _VER = {0: "top", 1: "center", 2: "bottom", 3: "justify", 4: "distributed"}

    def _font_of(xf) -> Font:
        try:
            f    = xls.font_list[xf.font_index]
            name = f.name or "ＭＳ Ｐゴシック"
            size = (f.height or 180) / 20.0
            color = None
            # 0x7FFF = automatic（色指定なし）→ 既定（黒）に任せる
            if f.colour_index not in (0x7FFF, None):
                color = "FF" + _xlrd_color(f.colour_index)
            return Font(
                name=name, size=size,
                bold=bool(f.bold), italic=bool(f.italic),
                underline="single" if getattr(f, "underline_type", 0) else None,
                color=color,
            )
        except Exception:
            return Font(name="ＭＳ Ｐゴシック", size=9)

    def _align_of(xf) -> Alignment | None:
        try:
            a    = xf.alignment
            h    = _HOR.get(a.hor_align)
            v    = _VER.get(a.vert_align)
            wrap = bool(a.text_wrapped)
            rot  = a.rotation if isinstance(a.rotation, int) and 0 <= a.rotation <= 180 else 0
            if h or v or wrap or rot:
                return Alignment(horizontal=h, vertical=v,
                                 wrap_text=wrap, text_rotation=rot)
        except Exception:
            pass
        return None

    def _fill_of(xf) -> PatternFill | None:
        try:
            bg = xf.background
            if bg.fill_pattern == 1:  # solid のみ忠実転写
                return PatternFill(
                    fill_type="solid",
                    fgColor="FF" + _xlrd_color(bg.pattern_colour_index),
                )
        except Exception:
            pass
        return None

    def _numfmt_of(xf) -> str | None:
        try:
            fmt = xls.format_map.get(xf.format_key)
            if fmt and fmt.format_str and fmt.format_str != "General":
                return fmt.format_str
        except Exception:
            pass
        return None

    def _border_of(xf) -> Border | None:
        b      = xf.border
        left   = _side(b.left_line_style,   b.left_colour_index)
        right  = _side(b.right_line_style,  b.right_colour_index)
        top    = _side(b.top_line_style,    b.top_colour_index)
        bottom = _side(b.bottom_line_style, b.bottom_colour_index)
        if any(s is not None for s in (left, right, top, bottom)):
            return Border(
                left=left   or Side(border_style=None),
                right=right or Side(border_style=None),
                top=top     or Side(border_style=None),
                bottom=bottom or Side(border_style=None),
            )
        return None

    # xf_index 単位でスタイルをキャッシュ（ユニークは数十個程度）
    _style_cache: dict = {}

    def _style(xf_idx):
        if xf_idx not in _style_cache:
            xf = xls.xf_list[xf_idx]
            _style_cache[xf_idx] = (
                _font_of(xf), _align_of(xf), _fill_of(xf),
                _numfmt_of(xf), _border_of(xf),
            )
        return _style_cache[xf_idx]

    # セル値・スタイル転写（フォント/配置/表示形式/罫線を忠実に。背景色は付けない）
    # ※ AX列(=50)までのみ。右側の選択肢欄は出力に含めない。
    last_row = 1   # 実際に内容/罫線がある最終行（印刷範囲の下端＝余白対策）
    for r in range(xs.nrows):
        for c in range(min(xs.ncols, MAX_COL)):
            cell = ws.cell(r + 1, c + 1)
            try:
                xf_idx = xs.cell_xf_index(r, c)
            except Exception:
                xf_idx = None

            font = align = numfmt = border = None
            if xf_idx is not None:
                try:
                    font, align, _fill, numfmt, border = _style(xf_idx)
                except Exception:
                    pass

            v = xs.cell_value(r, c)
            if v != "":
                try:
                    cell.value = v
                except Exception:
                    pass

            # 空セルにもフォントを残す（後段で値が書かれる値セルが正しい体裁を継ぐ）
            # 背景色(塗りつぶし)は意図的に転写しない＝出力は背景色なし。
            try:
                cell.font = font or Font(name="ＭＳ Ｐゴシック", size=9)
                if align  is not None: cell.alignment     = align
                if numfmt is not None: cell.number_format = numfmt
                if border is not None: cell.border        = border
            except Exception:
                pass

            if v != "" or border is not None:
                if r + 1 > last_row:
                    last_row = r + 1

    # 印刷範囲: 本体（A〜AX列）×内容のある最終行まで。
    # 末尾の空行を含めないことで印刷時の下部余白を抑える。
    last_vis = 50  # AX 列
    last_row = min(max(last_row, 1), xs.nrows)
    ws.print_area = f"A1:{get_column_letter(last_vis)}{last_row}"

    return wb, ws


# ─── テンプレート値セル書き込み ──────────────────────────────────────────────

def _w(ws, row: int, col: str | int, value, sz: int | None = None,
       bold: bool | None = None):
    """マージ済みでも master cell に値を書き込む。
    既定ではテンプレート由来の書式（フォント/配置）をそのまま保持し、
    sz / bold を明示した場合のみフォントサイズ・太字を上書きする。
    """
    col_idx = column_index_from_string(col) if isinstance(col, str) else col
    try:
        cell = ws.cell(row, col_idx)
        cell.value = value
        if sz is not None or bold is not None:
            base = cell.font
            cell.font = Font(
                name=base.name or "ＭＳ Ｐゴシック",
                size=sz if sz is not None else (base.size or 9),
                bold=bold if bold is not None else bool(base.bold),
                italic=bool(base.italic),
                color=base.color,
            )
        # それ以外はテンプレートの書式を維持（フォント上書きしない）
    except Exception:
        pass


def _fill_catchphrase(ws, catchphrases: list[str]):
    """キャッチコピーを画像エリア上部(L1:AJ2)に右寄せ・折返しで大きめに配置。
    3様式とも上段はタイトル(A1:K2)の右側が空きなので共通で使える。
    """
    cp = "　".join(f"◆ {c}" for c in catchphrases if c)
    if not cp:
        return
    try:
        ws.merge_cells("L1:AH2")   # AI:AJ を右余白として残す
    except Exception:
        pass
    cell = ws["L1"]
    cell.value     = cp
    cell.font      = Font(name="ＭＳ Ｐゴシック", size=12, bold=True)
    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)


def _reshape_company_band(ws):
    """会社情報帯のロゴ枠を A-D に縮め、商号/免許ブロックの左端を G→E に寄せて
    各フィールドを 2 列分広げる。3様式とも帯構造（rows53-58）は共通。
    """
    thin = Side(style="thin", color="000000")

    # 1) 既存マージを解除
    for rng in ("G53:G56", "H53:W56", "G57:J57", "G58:J58"):
        try:
            ws.unmerge_cells(rng)
        except Exception:
            pass

    # 2) ラベル（商号/宅建免許/建設業）を G 列 → E 列へ移動
    import copy as _copy

    def _move(r, src_c, dst_c):
        s = ws.cell(r, src_c)
        d = ws.cell(r, dst_c)
        d.value     = s.value
        d.font      = _copy.copy(s.font)
        d.alignment = _copy.copy(s.alignment)
        s.value = None

    _move(53, 7, 5)   # 商号
    _move(57, 7, 5)   # ●宅建免許番号
    _move(58, 7, 5)   # ●建設業番号

    # 3) 対象領域 A53:W58 の罫線をクリア
    for r in range(53, 59):
        for c in range(1, 24):      # A..W
            ws.cell(r, c).border = Border()

    # 4) 新レイアウトでマージし直す
    for rng in ("E53:E56", "F53:W56", "E57:J57", "E58:J58"):
        try:
            ws.merge_cells(rng)
        except Exception:
            pass

    # 5) 各ボックスの外枠を thin で描画（共有辺は集約）
    edges: dict = {}

    def _add_box(c1, r1, c2, r2):
        for c in range(c1, c2 + 1):
            edges.setdefault((r1, c), {})["top"]    = thin
            edges.setdefault((r2, c), {})["bottom"] = thin
        for r in range(r1, r2 + 1):
            edges.setdefault((r, c1), {})["left"]  = thin
            edges.setdefault((r, c2), {})["right"] = thin

    _add_box(1,  53, 4,  58)   # ロゴ枠 A-D（帯全高）
    _add_box(5,  53, 5,  56)   # 商号ラベル E
    _add_box(6,  53, 23, 56)   # 商号入力 F-W
    _add_box(5,  57, 10, 57)   # 宅建免許ラベル E-J
    _add_box(11, 57, 18, 57)   # 宅建免許入力 K-R
    _add_box(19, 57, 20, 58)   # 担当者ラベル S-T
    _add_box(21, 57, 23, 58)   # 担当者入力 U-W
    _add_box(5,  58, 10, 58)   # 建設業ラベル E-J
    _add_box(11, 58, 18, 58)   # 建設業入力 K-R

    for (r, c), sides in edges.items():
        b = ws.cell(r, c).border
        ws.cell(r, c).border = Border(
            top=sides.get("top", b.top),
            bottom=sides.get("bottom", b.bottom),
            left=sides.get("left", b.left),
            right=sides.get("right", b.right),
        )


def _fill_company(ws, company: dict):
    """会社情報欄を書き込む（3様式共通: rows 53-58）。
    商号 H53 / 所在地 Z53 / チラシ AM53 / 取引態様 AP53 / 報酬形態 AT53 /
    TEL Z55 / 情報誌 AM55 / FAX Z56 / 宅建免許 K57 / 担当者 U57 /
    E-MAIL Z57 / インターネット AM57 / 建設業番号 K58 / URL Z58
    """
    _w(ws, 53, "F",  company.get("商号", ""),           sz=28, bold=True)
    _w(ws, 53, "Z",  company.get("所在地", ""))
    _w(ws, 53, "AM", company.get("チラシ", ""))
    _w(ws, 53, "AP", company.get("取引態様", ""))
    _w(ws, 53, "AT", company.get("報酬形態", ""))
    _w(ws, 55, "Z",  company.get("TEL", ""))
    _w(ws, 55, "AM", company.get("情報誌", ""))
    _w(ws, 56, "Z",  company.get("FAX", ""))
    _w(ws, 57, "K",  company.get("宅建免許番号", ""))
    _w(ws, 57, "U",  company.get("担当者", ""))
    _w(ws, 57, "Z",  company.get("MAIL", ""))
    _w(ws, 57, "AM", company.get("インターネット", ""))
    _w(ws, 58, "K",  company.get("建設業免許番号", ""))
    _w(ws, 58, "Z",  company.get("URL", ""))


def _fill_chintai(ws, specs: dict, company: dict, catchphrases: list[str]):
    """賃貸不動産案内書（標準様式）の値セルを書き換える。
    テンプレート: 賃貸案内書_blank.xls
    """
    _fill_catchphrase(ws, catchphrases)

    _w(ws, 3, "AL", specs.get("種目", ""))               # 種目

    rent = specs.get("賃料", "")                          # 賃料 AN5
    try:
        _w(ws, 5, "AN", int(re.sub(r"[^0-9]", "", str(rent))) if rent else "")
    except ValueError:
        _w(ws, 5, "AN", rent)

    _w(ws, 9,  "AM", specs.get("管理費", ""))             # 共益費/管理費
    _w(ws, 10, "AM", specs.get("保証金", ""))             # 保証料
    _w(ws, 10, "AW", specs.get("敷金ヶ月", ""))           # 敷金ヶ月
    _w(ws, 11, "AW", specs.get("礼金ヶ月", ""))           # 礼金ヶ月
    _w(ws, 12, "AM", specs.get("更新料", ""))             # 更新料

    _w(ws, 13, "AL", specs.get("所在地", ""))             # 所在地
    _w(ws, 17, "AL", specs.get("交通", ""))               # 交通
    _w(ws, 20, "AO", specs.get("交通2", ""))              # 他交通手段
    _w(ws, 22, "AO", specs.get("物件名", ""))             # 建物名

    _w(ws, 25, "AO", specs.get("構造", ""))               # 構造
    _w(ws, 25, "AS", specs.get("階数_地上", ""))          # 地上N階
    _w(ws, 26, "AP", specs.get("部屋階", ""))             # （ ）階部分

    _w(ws, 27, "AO", specs.get("間取り", ""))             # 間取り
    _w(ws, 27, "AT", specs.get("間取り詳細", ""))         # 詳細

    _w(ws, 29, "AQ", specs.get("専有面積㎡", ""))         # 専有面積㎡
    _w(ws, 29, "AV", specs.get("専有面積坪", ""))         # 専有面積坪
    _w(ws, 31, "AO", specs.get("バルコニー向き", ""))     # バルコニー向き
    _w(ws, 33, "AV", specs.get("現況", ""))               # 現況

    _w(ws, 34, "AO", specs.get("築年号", ""))             # 建築 年号
    try:
        _w(ws, 34, "AQ", int(specs.get("築年", 0)) or "")  # 築年
    except Exception:
        _w(ws, 34, "AQ", specs.get("築年", ""))
    try:
        _w(ws, 34, "AT", int(specs.get("築月", 0)) or "")  # 築月
    except Exception:
        _w(ws, 34, "AT", specs.get("築月", ""))

    _w(ws, 44, "AQ", specs.get("引渡", ""))               # 引渡
    if specs.get("設備"):
        _w(ws, 46, "AL", specs.get("設備", ""))           # 設備（既定文を上書き）

    _fill_company(ws, company)


def _price_man(price) -> object:
    """価格文字列を 万円 数値に変換（失敗時は元文字列）。"""
    try:
        return float(re.sub(r"[^0-9.]", "", str(price))) if price else ""
    except ValueError:
        return price


def _fill_mansion(ws, specs: dict, company: dict, catchphrases: list[str]):
    """売却不動産案内書 マンション（区分所有建物用）標準様式の値セルを書き換える。
    テンプレート: 売却マンション案内書_blank.xls
    """
    _fill_catchphrase(ws, catchphrases)

    _w(ws, 3, "AL", specs.get("種目", ""))               # 種目
    _w(ws, 5, "AO", _price_man(specs.get("価格", "")))    # 価格（万円）

    _w(ws, 9,  "AN", specs.get("坪単価", ""))             # 坪単価
    _w(ws, 9,  "AU", specs.get("借地料", ""))             # 借地料
    _w(ws, 10, "AN", specs.get("管理費", ""))             # 管理費
    _w(ws, 10, "AU", specs.get("修繕積立金", ""))         # 修繕積立金

    _w(ws, 11, "AL", specs.get("所在地", ""))             # 所在地
    _w(ws, 15, "AL", specs.get("交通", ""))               # 交通
    _w(ws, 18, "AO", specs.get("交通2", ""))              # 他交通手段
    _w(ws, 20, "AO", specs.get("物件名", ""))             # 建物名

    _w(ws, 23, "AO", specs.get("構造", ""))               # 構造
    _w(ws, 23, "AS", specs.get("階数_地上", ""))          # 地上N階
    _w(ws, 24, "AP", specs.get("部屋階", ""))             # （ ）階部分

    _w(ws, 25, "AO", specs.get("間取り", ""))             # 間取り
    _w(ws, 25, "AT", specs.get("間取り詳細", ""))         # 詳細

    _w(ws, 27, "AQ", specs.get("専有面積㎡", ""))         # 専有面積㎡
    _w(ws, 27, "AV", specs.get("専有面積坪", ""))         # 専有面積坪
    _w(ws, 29, "AO", specs.get("バルコニー向き", ""))     # バルコニー向き

    _w(ws, 31, "AO", specs.get("総戸数", ""))             # 総戸数
    _w(ws, 31, "AV", specs.get("現況", ""))               # 現況

    _w(ws, 32, "AO", specs.get("築年号", ""))             # 建築 年号
    try:
        _w(ws, 32, "AQ", int(specs.get("築年", 0)) or "")
    except Exception:
        _w(ws, 32, "AQ", specs.get("築年", ""))
    try:
        _w(ws, 32, "AT", int(specs.get("築月", 0)) or "")
    except Exception:
        _w(ws, 32, "AT", specs.get("築月", ""))

    _w(ws, 34, "AO", specs.get("施主", ""))               # 施主
    _w(ws, 35, "AO", specs.get("施工", ""))               # 施工
    _w(ws, 36, "AO", specs.get("管理会社", ""))           # 管理（委託先）
    _w(ws, 37, "AQ", specs.get("管理形態", ""))           # 管理形態
    _w(ws, 38, "AS", specs.get("管理人", ""))             # 管理人
    _w(ws, 39, "AO", specs.get("その他", ""))             # その他

    _w(ws, 41, "AO", specs.get("土地面積㎡", ""))         # 土地（敷地）面積㎡
    _w(ws, 41, "AU", specs.get("土地面積坪", ""))         # 同 坪
    _w(ws, 42, "AO", specs.get("持分", ""))               # 持分
    _w(ws, 43, "AO", specs.get("用途地域", ""))           # 用途地域
    _w(ws, 44, "AO", specs.get("権利", ""))               # 権利
    _w(ws, 45, "AQ", specs.get("引渡", ""))               # 引渡
    if specs.get("設備"):
        _w(ws, 47, "AL", specs.get("設備", ""))           # 設備（既定文を上書き）

    _fill_company(ws, company)


def _fill_ippan(ws, specs: dict, company: dict, catchphrases: list[str]):
    """売却不動産案内書 一般用（売地・戸建）標準様式の値セルを書き換える。
    テンプレート: 売却一般案内書_blank.xls
    """
    _fill_catchphrase(ws, catchphrases)

    _w(ws, 3, "AL", specs.get("種目", ""))               # 種目
    _w(ws, 5, "AO", _price_man(specs.get("価格", "")))    # 価格（万円）

    _w(ws, 9,  "AL", specs.get("所在地", ""))             # 所在地
    _w(ws, 13, "AL", specs.get("交通", ""))               # 交通
    _w(ws, 16, "AO", specs.get("交通2", ""))              # 他交通手段

    _w(ws, 18, "AQ", specs.get("土地面積㎡", ""))         # 土地面積㎡
    _w(ws, 18, "AV", specs.get("土地面積坪", ""))         # 土地面積坪
    _w(ws, 20, "AT", specs.get("私道負担", ""))           # 私道負担

    _w(ws, 26, "AO", specs.get("権利", ""))               # 権利
    _w(ws, 26, "AW", specs.get("現況", ""))               # 現況
    _w(ws, 27, "AO", specs.get("地目", ""))               # 地目
    _w(ws, 27, "AW", specs.get("地勢", ""))               # 地勢
    _w(ws, 28, "AO", specs.get("都市計画", ""))           # 都市計画
    _w(ws, 29, "AO", specs.get("用途地域", ""))           # 用途地域
    _w(ws, 30, "AO", specs.get("建ぺい率", ""))           # 建ぺい率
    _w(ws, 31, "AO", specs.get("容積率", ""))             # 容積率
    _w(ws, 32, "AO", specs.get("その他規制", ""))         # その他の規制

    _w(ws, 35, "AO", specs.get("構造", ""))               # 構造
    _w(ws, 35, "AU", specs.get("階数_地上", ""))          # 地上N階
    _w(ws, 37, "AO", specs.get("間取り", ""))             # 間取り
    _w(ws, 37, "AT", specs.get("間取り詳細", ""))         # 詳細
    _w(ws, 39, "AO", specs.get("延床面積㎡", ""))         # 延床面積㎡
    _w(ws, 39, "AT", specs.get("延床面積坪", ""))         # 延床面積坪

    _w(ws, 41, "AO", specs.get("築年号", ""))             # 建築 年号
    try:
        _w(ws, 41, "AQ", int(specs.get("築年", 0)) or "")
    except Exception:
        _w(ws, 41, "AQ", specs.get("築年", ""))
    try:
        _w(ws, 41, "AT", int(specs.get("築月", 0)) or "")
    except Exception:
        _w(ws, 41, "AT", specs.get("築月", ""))

    _w(ws, 43, "AO", specs.get("現況", ""))               # 現況（建物）
    _w(ws, 44, "AQ", specs.get("引渡", ""))               # 引渡
    if specs.get("設備"):
        _w(ws, 46, "AL", specs.get("設備", ""))           # 設備（既定文を上書き）

    _fill_company(ws, company)


# ─── 画像挿入 ─────────────────────────────────────────────────────────────────

def _insert_image(ws, pil_img: Image.Image,
                  col_from: int, row_from: int,
                  col_to: int,   row_to: int,
                  col_w: float = 3.89, row_h: float = 12.0,
                  measure: bool = False):
    """TwoCellAnchor で画像を指定セル範囲に収まるよう挿入する。
    measure=True のときは実際の列幅・行高から枠の実寸を算出し、白背景に
    アスペクト比を保って自動縮小（thumbnail）するため、枠からはみ出さない。
    """
    px_per_char = 7.0
    if measure:
        # 実際の列幅(文字単位)・行高(pt)から枠のピクセル寸法を算出
        w_px = 0.0
        for c in range(col_from, col_to):
            ltr = get_column_letter(c)
            cd  = ws.column_dimensions[ltr] if ltr in ws.column_dimensions else None
            w_px += (cd.width if cd and cd.width else 8.43) * px_per_char
        h_px = 0.0
        for r in range(row_from, row_to):
            rd = ws.row_dimensions[r] if r in ws.row_dimensions else None
            h_px += (rd.height if rd and rd.height else 15.0) * (96 / 72)
        w_px = round(w_px)
        h_px = round(h_px)
    else:
        # 1pt = 96/72 px (96dpi)。int() の丸め誤差を防ぐため round() を使用
        w_px = round((col_to - col_from) * col_w * px_per_char)
        h_px = round((row_to - row_from) * row_h * (96 / 72))
    w_px = max(1, w_px)
    h_px = max(1, h_px)

    img = pil_img.convert("RGB").copy()
    img.thumbnail((w_px, h_px), Image.LANCZOS)
    # 白背景にセンタリング
    canvas = Image.new("RGB", (w_px, h_px), (255, 255, 255))
    canvas.paste(img, ((w_px - img.width) // 2, (h_px - img.height) // 2))

    buf = io.BytesIO()
    # 96dpi を明示して保存（指定なしだと LibreOffice/Excel が 72dpi と解釈し133%に拡大する）
    canvas.save(buf, "PNG", dpi=(96, 96))
    buf.seek(0)
    xl = XLImage(buf)
    xl.width  = w_px
    xl.height = h_px

    anchor = TwoCellAnchor()
    anchor._from = AnchorMarker(col=col_from - 1, colOff=0,
                                row=row_from - 1, rowOff=0)
    anchor.to    = AnchorMarker(col=col_to - 1,   colOff=0,
                                row=row_to - 1,   rowOff=0)
    anchor.editAs = "twoCell"
    xl.anchor = anchor
    ws.add_image(xl)


def _insert_logo(ws, pil_img: Image.Image,
                 col_from: int, row_from: int,
                 col_to: int,   row_to: int,
                 margin: float = 0.92):
    """ロゴを枠(col_from..col_to-1 × row_from..row_to-1)内に、縦横比を保ったまま
    枠最大サイズで中央配置する。OneCellAnchor で実寸固定するので引き伸ばし歪み無し。
    """
    px_per_char = 7.0

    def _col_px(c):
        ltr = get_column_letter(c)
        cd  = ws.column_dimensions[ltr] if ltr in ws.column_dimensions else None
        return (cd.width if cd and cd.width else 8.43) * px_per_char

    def _row_px(r):
        rd = ws.row_dimensions[r] if r in ws.row_dimensions else None
        return (rd.height if rd and rd.height else 15.0) * (96 / 72)

    frame_w = sum(_col_px(c) for c in range(col_from, col_to))
    frame_h = sum(_row_px(r) for r in range(row_from, row_to))
    box_w   = max(1, frame_w * margin)
    box_h   = max(1, frame_h * margin)

    # 白背景に統合してアスペクト比キープで縮小
    img = pil_img.convert("RGBA")
    flat = Image.new("RGB", img.size, (255, 255, 255))
    flat.paste(img, mask=img.split()[3])
    flat.thumbnail((int(box_w), int(box_h)), Image.LANCZOS)
    lw, lh = flat.size

    # 枠中央に来るようオフセット
    off_x = max(0.0, (frame_w - lw) / 2)
    off_y = max(0.0, (frame_h - lh) / 2)

    buf = io.BytesIO()
    flat.save(buf, "PNG", dpi=(96, 96))
    buf.seek(0)
    xl = XLImage(buf)
    xl.width  = lw
    xl.height = lh

    EMU = 9525  # 1px = 9525 EMU (96dpi)
    anchor = OneCellAnchor()
    anchor._from = AnchorMarker(col=col_from - 1, colOff=int(off_x * EMU),
                                row=row_from - 1, rowOff=int(off_y * EMU))
    anchor.ext   = XDRPositiveSize2D(cx=int(lw * EMU), cy=int(lh * EMU))
    xl.anchor = anchor
    ws.add_image(xl)


def _feature_max_per_row(col_from: int, col_to: int) -> int:
    """特徴帯(col_from..col_to-1)に1行で並べられる最大個数（1個=最小2列換算）。"""
    return max(1, (col_to - col_from) // 2)


def _insert_feature_strip(ws, icon_paths: list,
                          col_from: int, col_to: int,
                          row_from: int, row_to: int,
                          per_row: int | None = None):
    """特徴アイコンを矩形(col_from..col_to-1 × row_from..row_to-1)内にグリッド配置。
    各アイコンは写真と同じ TwoCellAnchor でセル範囲に収める（確実に並ぶ）。
    per_row 指定時は1行あたりの個数を固定、未指定なら自動で最適化。
    """
    n = len(icon_paths)
    if n == 0:
        return

    ncols_span = col_to - col_from
    nrows_span = row_to - row_from

    if per_row and per_row > 0:
        per = max(1, min(int(per_row), n, ncols_span))
    else:
        # 帯のピクセル寸から、アイコンが最大になる行数(1〜3)を選ぶ
        px_per_char = 7.0
        def _cpx(c):
            ltr = get_column_letter(c)
            cd  = ws.column_dimensions[ltr] if ltr in ws.column_dimensions else None
            return (cd.width if cd and cd.width else 8.43) * px_per_char
        def _rpx(r):
            rd = ws.row_dimensions[r] if r in ws.row_dimensions else None
            return (rd.height if rd and rd.height else 15.0) * (96 / 72)
        W = sum(_cpx(c) for c in range(col_from, col_to))
        H = sum(_rpx(r) for r in range(row_from, row_to))
        best = None
        for rr in range(1, 4):
            pp   = -(-n // rr)
            size = min(W / pp, H / rr)
            if best is None or size > best[0]:
                best = (size, rr, pp)
        per = best[2]

    rows     = -(-n // per)
    cols_per = ncols_span / per
    rows_per = nrows_span / rows

    for i, pth in enumerate(icon_paths):
        gr, gc = divmod(i, per)
        c1 = col_from + round(gc * cols_per)
        c2 = col_from + round((gc + 1) * cols_per)
        r1 = row_from + round(gr * rows_per)
        r2 = row_from + round((gr + 1) * rows_per)
        if c2 <= c1: c2 = c1 + 1
        if r2 <= r1: r2 = r1 + 1
        try:
            _insert_image(ws, Image.open(pth), c1, r1, c2, r2, measure=True)
        except Exception:
            pass


# ─── 画像配置プレビューの初期配置 ─────────────────────────────────────────────

# 初期配置レイアウト: レイアウト名 -> [(枠ラベル, (x率, y率, 幅率, 高率)), ...]
# 番号①②③…の順。率は画像エリア幅(IA_W)・高さ(PREV_H)に対する割合。
PLACE_LAYOUTS: dict[str, list] = {
    "縦2分割": [   # 左列を上下分割＋右列を全高
        ("①左上", (0.00, 0.00, 0.55, 0.50)),
        ("②左下", (0.00, 0.50, 0.55, 0.50)),
        ("③右",   (0.55, 0.00, 0.45, 1.00)),
    ],
    "横2分割": [   # 上段を全幅＋下段を左右分割
        ("①上",   (0.00, 0.00, 1.00, 0.50)),
        ("②下左", (0.00, 0.50, 0.50, 0.50)),
        ("③下右", (0.50, 0.50, 0.50, 0.50)),
    ],
    "4分割": [
        ("①左上", (0.00, 0.00, 0.50, 0.50)),
        ("②右上", (0.50, 0.00, 0.50, 0.50)),
        ("③左下", (0.00, 0.50, 0.50, 0.50)),
        ("④右下", (0.50, 0.50, 0.50, 0.50)),
    ],
}


def _slot_rect_px(rect, ia_w: int, prev_h: int) -> dict:
    """(x率,y率,幅率,高率) → プレビュー画像エリアのピクセル配置 dict。"""
    xf, yf, wf, hf = rect
    return {
        "enabled": True,
        "x": round(xf * ia_w),
        "y": round(yf * prev_h),
        "w": max(20, round(wf * ia_w) - 4),
        "h": max(20, round(hf * prev_h)),
    }


def _default_slot_index(label: str, n_slots: int) -> int:
    """画像の種類から既定の枠番号(0始まり)を推定。"""
    cat = {"外観": 0, "間取": n_slots - 1, "室内": 1, "地図": 1}
    for k, v in cat.items():
        if k in label:
            return min(max(v, 0), n_slots - 1)
    return 0


# ─── メイン Excel 生成 ────────────────────────────────────────────────────────

def _widen_setubi(ws, tpl_key: str):
    """設備テキスト欄を1行下(52行目)まで拡張して右側のバランスを取る。
    （写真エリア下部に特徴帯を設けた分、右側を広げる調整）。
    """
    setubi_top = 47 if tpl_key == "売買マンション" else 46   # 設備テキスト先頭行
    box  = f"AL{setubi_top}:AX51"
    box2 = f"AL{setubi_top}:AX52"
    # 設備値セル(AL{top})の内容を退避→再マージ→書き戻し
    val  = ws[f"AL{setubi_top}"].value
    try:
        ws.unmerge_cells(box)
    except Exception:
        pass
    try:
        ws.merge_cells(box2)
    except Exception:
        pass
    ws[f"AL{setubi_top}"].value = val
    # 設備ラベル列(AK)も52行目まで罫線を合わせる
    thin = Side(style="thin", color="000000")
    for c in ("AK", "AX"):
        cell = ws[f"{c}52"]
        b = cell.border
        side = thin
        cell.border = Border(top=b.top, bottom=b.bottom,
                             left=(side if c == "AK" else b.left),
                             right=(side if c == "AX" else b.right))


def create_fudosan_excel(
    template_type: str,
    specs: dict,
    company_info: dict,
    catchphrases: list[str],
    image_placements: list[dict],   # [{"img": PIL, "c1": int, "r1": int, "c2": int, "r2": int}, ...]
    features: list[str] | None = None,       # 選択された特徴ラベル
    feature_per_row: int | None = None,      # 特徴の1行あたり個数（None=自動）
) -> bytes:
    """テンプレート XLS を読み込み、値セル・画像を書き換えて XLSX を返す。"""
    tpl_key  = template_type if template_type in TEMPLATE_XLS else "賃貸"
    tpl_path = TEMPLATE_XLS.get(tpl_key)

    if tpl_path is None or not tpl_path.exists():
        raise RuntimeError(f"テンプレート XLS が見つかりません: {tpl_path}")

    wb, ws = _load_xls_as_openpyxl(tpl_path)
    _reshape_company_band(ws)   # ロゴ枠 A-D 化＋商号/免許フィールド拡幅

    if tpl_key == "賃貸":
        _fill_chintai(ws, specs, company_info, catchphrases)
        col_w, row_h = 3.89, 12.0
    elif tpl_key == "売買マンション":
        _fill_mansion(ws, specs, company_info, catchphrases)
        col_w, row_h = 3.66, 13.3
    else:
        _fill_ippan(ws, specs, company_info, catchphrases)
        col_w, row_h = 3.66, 13.3

    # 画像挿入
    for pl in image_placements:
        pil = pl.get("img")
        if pil is None:
            continue
        try:
            _insert_image(ws, pil,
                          pl["c1"], pl["r1"],
                          pl["c2"], pl["r2"],
                          col_w=col_w, row_h=row_h)
        except Exception as e:
            pass  # 画像配置失敗は無視して続行

    # 特徴アイコン（写真エリア下部 A-AJ × rows44-51 に帯状配置）
    if features:
        cat = {lbl: pth for lbl, pth in FEATURE_CATALOG}
        paths = [cat[f] for f in features if f in cat]
        try:
            _insert_feature_strip(ws, paths, 1, 37, 45, 52, per_row=feature_per_row)
        except Exception:
            pass

    # ロゴ挿入（会社情報帯の左枠）
    if LOGO_PATH.exists():
        try:
            logo_pil = Image.open(LOGO_PATH)
            # ロゴ枠 A53:D58 内に、縦横比を保ったまま最大サイズで中央配置
            _insert_logo(ws, logo_pil, 1, 53, 5, 59)
        except Exception:
            pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def create_band_swap_excel(maisoku_img: Image.Image,
                           company_info: dict,
                           tpl_key: str = "賃貸") -> bytes:
    """【帯変えモード（時短）】他社マイソク画像を上部にそのまま貼り付け、
    下部の自社情報帯（ロゴ＋会社情報）だけ現行レイアウトで差し替えて出力する。
    AI解析・項目入力なしで即作成できる。
    """
    tpl_path = TEMPLATE_XLS.get(tpl_key, TEMPLATE_XLS["賃貸"])
    wb, ws = _load_xls_as_openpyxl(tpl_path)

    # 上部(rows1-52)の様式（値・罫線・マージ）を白紙化 → 画像で覆う
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= 52:
            try:
                ws.unmerge_cells(str(rng))
            except Exception:
                pass
    for r in range(1, 53):
        for c in range(1, 51):
            cell = ws.cell(r, c)
            cell.value  = None
            cell.border = Border()

    # 自社情報帯（rows53-58）を現行レイアウトで構築
    _reshape_company_band(ws)
    _fill_company(ws, company_info)
    if LOGO_PATH.exists():
        try:
            _insert_logo(ws, Image.open(LOGO_PATH), 1, 53, 5, 59)
        except Exception:
            pass

    # 他社マイソク画像を上部(A1:AX52)に貼付（縦横比保持で最大配置）
    try:
        _insert_image(ws, maisoku_img, 1, 1, 51, 53, measure=True)
    except Exception:
        pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── Streamlit UI ─────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="不動産案内書 変換ツール",
    layout="wide",
    page_icon="🏠",
    initial_sidebar_state="expanded",
)

# ── サイドバー: 会社情報 ─────────────────────────────────────────────────────
with st.sidebar:
    st.header("🏢 自社情報設定")
    st.caption("入力内容は「保存」ボタンで記録されます")
    company = load_company_info()
    ci: dict = {}

    def _si(label, key, default=""):
        return st.text_input(label, value=company.get(key, default))

    ci["商号"]           = _si("商号（会社名）",   "商号")
    ci["宅建免許番号"]    = _si("宅建免許番号",     "宅建免許番号")
    ci["建設業免許番号"]  = _si("建設業免許番号",   "建設業免許番号")
    ci["所在地"]         = _si("所在地",           "所在地")
    ci["TEL"]            = _si("TEL",              "TEL")
    ci["FAX"]            = _si("FAX",              "FAX")
    ci["担当者"]         = _si("担当者",           "担当者")
    ci["MAIL"]           = _si("E-MAIL",           "MAIL")
    ci["URL"]            = _si("URL",              "URL")

    def _ss(label, key, opts):
        cur = company.get(key, opts[0])
        return st.selectbox(label, opts, index=opts.index(cur) if cur in opts else 0)

    ci["取引態様"]       = _ss("取引態様",       "取引態様",       ["売主","貸主","代理","仲介","専任","一般"])
    ci["チラシ"]         = _ss("チラシ",         "チラシ",         ["可","不可","要確認"])
    ci["情報誌"]         = _ss("情報誌",         "情報誌",         ["可","不可","要確認"])
    ci["インターネット"] = _ss("インターネット", "インターネット", ["可","不可","要確認"])
    ci["報酬形態"]       = st.text_input("報酬形態", value=company.get("報酬形態", ""))

    if st.button("💾 自社情報を保存", type="primary", use_container_width=True):
        save_company_info(ci)
        st.success("✅ 保存しました")
        company = ci

    st.divider()
    st.caption(f"保存先: `{COMPANY_JSON.name}`")

    st.divider()
    st.subheader("🖼️ 会社ロゴ")
    st.caption("案内書下帯の左枠に挿入されます")
    if LOGO_PATH.exists():
        st.image(str(LOGO_PATH), use_container_width=True)
        st.caption("現在のロゴ")
    logo_upload = st.file_uploader("ロゴ画像をアップロード", type=["png","jpg","jpeg"], key="logo_upload")
    if logo_upload:
        try:
            logo_pil = Image.open(logo_upload).convert("RGBA")
            logo_pil.save(LOGO_PATH, "PNG")
            st.success("✅ ロゴを保存しました")
            st.rerun()
        except Exception as e:
            st.error(f"ロゴ保存エラー: {e}")

# ── メインエリア ──────────────────────────────────────────────────────────────
st.title("🏠 他社マイソク → 不動産案内書 Excel 変換ツール")

_MODE_FULL = "🤖 AI解析で作り直す（通常）"
_MODE_BAND = "⚡ 帯変え（時短：他社マイソク画像＋自社帯）"
mode = st.radio("モード", [_MODE_FULL, _MODE_BAND], horizontal=True, key="app_mode")

# ── 帯変えモード（時短）─────────────────────────────────────────────────────────
if mode == _MODE_BAND:
    st.caption("他社マイソクを画像のまま上部に貼り付け、下部の自社情報帯（ロゴ＋会社情報）"
               "だけ現行レイアウトで差し替えます。AI解析・項目入力なしで即出力。")
    bs_file = st.file_uploader(
        "他社マイソクをアップロード（PDF・PNG・JPG対応）",
        type=["pdf", "png", "jpg", "jpeg"], key="bs_upload",
    )
    if not bs_file:
        st.info("↑ 他社マイソクをアップロードしてください")
        st.stop()

    try:
        bs_img = convert_to_image(bs_file.getvalue(), bs_file.name)
    except Exception as e:
        st.error(f"読み込みエラー: {e}")
        st.stop()

    st.markdown("**✂️ 貼り付け範囲を調整** — 他社の帯（会社情報）を除いて、上に載せる範囲を指定")
    bs_c1, bs_c2 = st.columns([1, 2])
    with bs_c1:
        cut_bottom = st.slider("下カット（他社帯を除く位置）", 0.50, 1.00, 0.88, 0.01,
                               help="他社マイソク下部の会社情報帯を切り落とす位置")
        cut_top   = st.slider("上カット",   0.00, 0.40, 0.00, 0.01)
        cut_left  = st.slider("左カット",   0.00, 0.40, 0.00, 0.01)
        cut_right = st.slider("右カット",   0.60, 1.00, 1.00, 0.01)
    region = {"x1": cut_left, "y1": cut_top, "x2": cut_right, "y2": cut_bottom}
    bs_crop = crop_region(bs_img, region) or bs_img
    with bs_c2:
        st.caption("元画像（緑枠が貼り付け範囲・外側は除外されます）")
        st.image(crop_overlay(bs_img, region), use_container_width=True)

    st.markdown("**🖼️ 切り取り後のイメージ（この画像が案内書の上部に入ります）**")
    st.image(bs_crop, use_container_width=True)

    company_now = ci if ci.get("商号") else load_company_info()

    st.divider()
    if st.button("⚡ 帯変え Excel を作成", type="primary", use_container_width=True):
        with st.spinner("Excel を生成中..."):
            try:
                xlsx = create_band_swap_excel(bs_crop, company_now)
                pname = Path(bs_file.name).stem.replace(" ", "_").replace("/", "_")
                st.download_button(
                    "⬇️ Excel をダウンロード",
                    data=xlsx,
                    file_name=f"帯変え案内書_{pname}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
                st.success("✅ 作成しました。ダウンロードしてください。")
            except Exception as e:
                st.error(f"❌ 生成エラー: {e}")
    st.stop()

# ── 通常モード（AI解析）─────────────────────────────────────────────────────────
st.caption("他社チラシ（PDF・画像）をアップロードすると Claude AI が解析し、"
           "自社テンプレートに値を書き込んだ Excel を出力します。")

uploaded = st.file_uploader(
    "他社マイソクをアップロード（PDF・PNG・JPG対応）",
    type=["pdf", "png", "jpg", "jpeg"],
)

if not uploaded:
    st.info("↑ まずファイルをアップロードしてください")
    st.stop()

col_left, col_right = st.columns([1, 2])
with col_left:
    st.subheader("元のマイソク")
    try:
        original_image = convert_to_image(uploaded.getvalue(), uploaded.name)
        st.image(original_image, use_container_width=True)
    except Exception as e:
        st.error(f"プレビューエラー: {e}")
        st.stop()

with col_right:
    st.subheader("AI 解析")
    if st.button("🤖 AI で解析開始", type="primary", use_container_width=True):
        with st.spinner("Claude が解析中です（最大 10 分）..."):
            try:
                result = analyze_with_claude(original_image)
                st.session_state["analysis"]       = result
                st.session_state["original_image"] = original_image
                st.success("✅ 解析完了！下のタブで内容を確認・編集してください。")
                st.rerun()
            except Exception as e:
                st.error(f"❌ 解析エラー: {e}")

if "analysis" not in st.session_state:
    st.info("「AI で解析開始」ボタンを押してください。")
    st.stop()

analysis      = st.session_state["analysis"]
orig_img      = st.session_state.get("original_image", original_image)
specs         = analysis.get("specs", {})
catchphrases  = analysis.get("catchphrases", ["", "", ""])
regions       = analysis.get("regions", [])
detected_type     = analysis.get("template_type", "賃貸")
detected_features = [f for f in analysis.get("features", []) if f in FEATURE_LABELS]

st.divider()
st.subheader("📝 内容確認・編集")
tab_spec, tab_cp, tab_img = st.tabs(["物件スペック", "キャッチコピー", "切り抜き・配置"])

# ── スペック編集 ──────────────────────────────────────────────────────────────
with tab_spec:
    col_type, _ = st.columns([1, 3])
    with col_type:
        _TMPL_OPTS = ["賃貸", "売買マンション", "売買一般"]
        _di = next((i for i, t in enumerate(_TMPL_OPTS) if t in detected_type), 0)
        tmpl = st.radio("書類種別", _TMPL_OPTS, index=_di, horizontal=True)

    edited_specs: dict = {}
    c1, c2 = st.columns(2)

    if tmpl == "賃貸":
        lf = ["物件名","種目","賃料","管理費","保証金","敷金ヶ月","礼金ヶ月","更新料"]
        rf = ["所在地","交通","交通2","間取り","間取り詳細","専有面積㎡","専有面積坪",
              "構造","階数_地上","部屋階","バルコニー向き","築年号","築年","築月","現況","引渡"]
    elif tmpl == "売買マンション":
        lf = ["物件名","種目","価格","坪単価","借地料","管理費","修繕積立金",
              "土地面積㎡","土地面積坪","持分"]
        rf = ["所在地","交通","交通2","間取り","間取り詳細","専有面積㎡","専有面積坪",
              "構造","階数_地上","部屋階","バルコニー向き","総戸数","築年号","築年","築月",
              "施主","施工","管理会社","管理形態","管理人","用途地域","権利","現況","引渡"]
    else:  # 売買一般（売地・戸建）
        lf = ["物件名","種目","価格","土地面積㎡","土地面積坪","私道負担",
              "建ぺい率","容積率","その他規制"]
        rf = ["所在地","交通","交通2","間取り","間取り詳細","延床面積㎡","延床面積坪",
              "構造","階数_地上","築年号","築年","築月","地目","地勢","都市計画",
              "用途地域","権利","現況","引渡"]

    with c1:
        for f in lf:
            edited_specs[f] = st.text_input(f, value=str(specs.get(f, "")), key=f"s_{f}")
    with c2:
        for f in rf:
            edited_specs[f] = st.text_input(f, value=str(specs.get(f, "")), key=f"s_{f}")

    edited_specs["設備"]  = st.text_area("設備（設備一覧テキスト）",
                                         value=str(specs.get("設備", "")), height=70, key="s_setubi")
    edited_specs["その他"] = st.text_area("その他",
                                          value=str(specs.get("その他", "")), height=50, key="s_other")

# ── キャッチコピー編集 ────────────────────────────────────────────────────────
with tab_cp:
    edited_cps: list[str] = []
    for i in range(3):
        default = catchphrases[i] if i < len(catchphrases) else ""
        edited_cps.append(st.text_input(f"キャッチコピー {i+1}", value=default, key=f"cp_{i}"))

# ── 切り抜き・配置 ────────────────────────────────────────────────────────────
with tab_img:
    all_regions    = [r for r in regions if r.get("種類")]
    adjusted_regions: list[dict] = []

    # ── インタラクティブ枠編集（streamlit-drawable-canvas）──────────────────
    st.markdown("**🎯 領域を調整** — 枠をクリック選択 → ドラッグで移動 / 角をドラッグでリサイズ")

    try:
        from streamlit_drawable_canvas import st_canvas

        img_w, img_h = orig_img.size
        canvas_w = min(720, img_w)
        canvas_h = int(img_h * canvas_w / img_w)

        canvas_objects: list[dict] = []
        for reg in all_regions:
            kind  = reg.get("種類", "不明")
            color = REGION_COLORS.get(kind, (150, 150, 150))
            bx    = reg.get("x1", 0) * canvas_w
            by    = reg.get("y1", 0) * canvas_h
            bw    = (reg.get("x2", 0) - reg.get("x1", 0)) * canvas_w
            bh    = (reg.get("y2", 0) - reg.get("y1", 0)) * canvas_h
            if bw > 2 and bh > 2:
                canvas_objects.append({
                    "type": "rect",
                    "name": kind,   # 種類を name に保存 → 並び替えに依存しない
                    "left": round(bx, 1), "top": round(by, 1),
                    "width": round(bw, 1), "height": round(bh, 1),
                    "fill":   "rgba({},{},{},0.20)".format(*color),
                    "stroke": "rgb({},{},{})".format(*color),
                    "strokeWidth": 2, "selectable": True,
                })

        # 下絵を base64 JPEG に変換して fabric.js Image オブジェクトとして埋め込む
        # ※ background_image パラメータは image_to_url 経由のためStreamlit 1.50 で
        #   動作しない。init_drawing に直接入れることで image_to_url を完全に迂回する。
        _buf = io.BytesIO()
        _bg = orig_img.resize((canvas_w, canvas_h), Image.LANCZOS).convert("RGB")
        _bg.save(_buf, "JPEG", quality=65)
        _bg_b64 = "data:image/jpeg;base64," + base64.b64encode(_buf.getvalue()).decode()
        bg_fabric_obj = {
            "type": "image", "src": _bg_b64,
            "left": 0, "top": 0, "scaleX": 1.0, "scaleY": 1.0,
            "selectable": False, "evented": False,
            "lockMovementX": True, "lockMovementY": True,
            "lockScalingX": True, "lockScalingY": True,
            "hasControls": False, "hasBorders": False,
            "name": "_background_",
        }
        init_drawing = {
            "version": "4.4.0",
            "objects": [bg_fabric_obj] + canvas_objects,
        }

        canvas_result = st_canvas(
            fill_color       = "rgba(255,255,255,0.0)",
            stroke_width     = 2,
            background_color = "",
            update_streamlit = True,
            height           = canvas_h,
            width            = canvas_w,
            drawing_mode     = "transform",
            initial_drawing  = init_drawing,
            display_toolbar  = False,
            key              = "region_canvas",
        )

        # stroke色 → 種類の逆引き辞書（fabric.jsはnameを保存しないがstroke色は保存する）
        _stroke_to_kind = {
            "rgb({},{},{})".format(*v): k for k, v in REGION_COLORS.items()
        }

        # キャンバス結果から座標を取得
        if canvas_result.json_data and canvas_result.json_data.get("objects"):
            _obj_list = [o for o in canvas_result.json_data["objects"]
                         if o.get("name") != "_background_" and o.get("type") == "rect"]
            for idx, obj in enumerate(_obj_list):
                # stroke色で種類を特定。色が取れない場合は元のall_regionsの順番で補完
                stroke = obj.get("stroke", "")
                kind = _stroke_to_kind.get(stroke)
                if kind is None:
                    kind = all_regions[idx]["種類"] if idx < len(all_regions) else "不明"
                sx  = obj.get("scaleX", 1.0)
                sy  = obj.get("scaleY", 1.0)
                lft = obj.get("left", 0)
                top = obj.get("top",  0)
                ow  = obj.get("width",  0) * sx
                oh  = obj.get("height", 0) * sy
                adjusted_regions.append({
                    "種類": kind,
                    "x1": max(0.0, min(1.0, lft / canvas_w)),
                    "y1": max(0.0, min(1.0, top / canvas_h)),
                    "x2": max(0.0, min(1.0, (lft + ow) / canvas_w)),
                    "y2": max(0.0, min(1.0, (top + oh) / canvas_h)),
                })
        else:
            adjusted_regions = [
                r.copy() for r in all_regions
                if r.get("x2", 0) > r.get("x1", 0) and r.get("y2", 0) > r.get("y1", 0)
            ]

        # 座標リードアウト
        if adjusted_regions:
            coord_cols = st.columns(min(len(adjusted_regions), 4))
            for i, reg in enumerate(adjusted_regions):
                kind = reg["種類"]
                c    = REGION_COLORS.get(kind, (150, 150, 150))
                hx   = "#{:02X}{:02X}{:02X}".format(*c)
                with coord_cols[i % len(coord_cols)]:
                    st.markdown(
                        f"<b style='color:{hx}'>■ {kind}</b><br>"
                        f"x: {reg['x1']:.2f}–{reg['x2']:.2f}<br>"
                        f"y: {reg['y1']:.2f}–{reg['y2']:.2f}",
                        unsafe_allow_html=True,
                    )

    except ImportError:
        st.warning("streamlit-drawable-canvas が未インストール。スライダーで代替中。")
        for i, reg in enumerate(all_regions):
            kind  = reg.get("種類", "不明")
            color = REGION_COLORS.get(kind, (150, 150, 150))
            with st.expander(f"**{kind}** を調整する", expanded=(i == 0)):
                c_sl, c_pv = st.columns(2)
                with c_sl:
                    x1 = st.slider("← 左端", 0.0, 1.0, float(reg.get("x1", 0.0)), 0.01, key=f"adj_{i}_x1")
                    x2 = st.slider("→ 右端", 0.0, 1.0, float(reg.get("x2", 1.0)), 0.01, key=f"adj_{i}_x2")
                    y1 = st.slider("↑ 上端", 0.0, 1.0, float(reg.get("y1", 0.0)), 0.01, key=f"adj_{i}_y1")
                    y2 = st.slider("↓ 下端", 0.0, 1.0, float(reg.get("y2", 1.0)), 0.01, key=f"adj_{i}_y2")
                adj = {"種類": kind, "x1": x1, "y1": y1, "x2": x2, "y2": y2}
                with c_pv:
                    if x2 > x1 and y2 > y1:
                        pv = crop_region(orig_img, adj)
                        if pv:
                            st.image(pv, use_container_width=True)
                adjusted_regions.append(adj)

    # ── 切り抜きプレビュー ────────────────────────────────────────────────────
    st.divider()
    cropped: dict[str, Image.Image] = {}
    valid_adj = [r for r in adjusted_regions
                 if r.get("x2", 0) > r.get("x1", 0) and r.get("y2", 0) > r.get("y1", 0)]

    for reg in valid_adj:
        kind = reg["種類"]
        img  = crop_region(orig_img, reg)
        if img:
            key = (kind if kind not in cropped
                   else f"{kind}_{sum(1 for k in cropped if k.startswith(kind))+1}")
            cropped[key] = img

    # 清書済み画像を適用（rerun後も維持）
    _traced_img: Image.Image | None = st.session_state.get("madori_traced_img")
    _traced_key: str | None = st.session_state.get("madori_traced_key")
    if _traced_img is not None and _traced_key is not None and _traced_key in cropped:
        cropped[_traced_key] = _traced_img

    # プレビュー表示
    if cropped:
        prev_cols = st.columns(min(len(cropped), 4))
        for i, (kind, img) in enumerate(cropped.items()):
            c  = REGION_COLORS.get(kind, (150, 150, 150))
            hx = "#{:02X}{:02X}{:02X}".format(*c)
            with prev_cols[i % len(prev_cols)]:
                label = f"<span style='color:{hx}'>■ **{kind}**</span>"
                if kind == _traced_key and _traced_img is not None:
                    label += " ✅ 清書済み"
                st.markdown(label, unsafe_allow_html=True)
                st.image(img, use_container_width=True)

    # ── 間取り図 AI清書（間取り図トレーサー連携）────────────────────────────
    if cropped and TRACER_AVAILABLE:
        st.divider()
        st.markdown("**🏠 間取り図をAI清書**")

        _traced_key: str | None = st.session_state.get("madori_traced_key")
        _is_traced = _traced_key is not None and _traced_key in cropped

        # 対象画像を選択
        _crop_labels = list(cropped.keys())
        _default_idx = (
            _crop_labels.index(_traced_key) if _is_traced and _traced_key in _crop_labels
            else next((i for i, k in enumerate(_crop_labels) if "間取" in k), 0)
        )
        _ta, _tb = st.columns([3, 1])
        with _ta:
            _target_label = st.radio(
                "清書する画像を選択",
                _crop_labels,
                index=_default_idx,
                horizontal=True,
                key="tracer_target",
            )
        with _tb:
            _floor_type = st.selectbox(
                "図面タイプ", ["マンション", "戸建て", "1K・1R", "その他"],
                key="tracer_floor_type",
            )

        _t1, _t2, _t3 = st.columns([1, 1, 1])
        with _t1:
            _do_trace = st.button(
                "✨ AI清書実行" if not _is_traced else "🔄 再清書",
                key="btn_trace", use_container_width=True, type="primary",
            )
        with _t2:
            if _is_traced and st.button("↩ 元の切り抜きに戻す", key="btn_reset_trace", use_container_width=True):
                for _k in ["madori_traced_img", "madori_traced_bytes", "madori_traced_key"]:
                    st.session_state.pop(_k, None)
                st.rerun()

        # 清書済みの場合はbefore/after表示
        if _is_traced:
            _raw_for_show = crop_region(orig_img, next(
                (r for r in valid_adj if r.get("種類") == _traced_key), valid_adj[0]
            )) if valid_adj else cropped[_traced_key]
            _bc1, _bc2 = st.columns(2)
            with _bc1:
                st.caption(f"元の切り抜き（{_traced_key}）")
                st.image(_raw_for_show, use_container_width=True)
            with _bc2:
                st.caption("AI清書後 ✅")
                st.image(cropped[_traced_key], use_container_width=True)

        # 再修正テキスト入力（清書済みの場合のみ表示）
        _correction_text = ""
        if _is_traced:
            _correction_text = st.text_input(
                "修正指示（任意）",
                placeholder="例: 浴室と洗面室のラベルを分けて / 北コンパスを追加して",
                key="tracer_correction",
            )

        if _do_trace:
            # 元の切り抜き（生画像）を取得
            _target_reg = next(
                (r for r in valid_adj if r.get("種類") == _target_label), None
            )
            _raw_crop = crop_region(orig_img, _target_reg) if _target_reg else cropped.get(_target_label)
            if _raw_crop is None:
                st.error("対象画像が見つかりません。")
            else:
                # 再修正の場合は前回結果バイトをprev_resultとして渡す
                _prev_bytes = st.session_state.get("madori_traced_bytes") if _correction_text else None
                _spin_msg = (
                    f"修正指示で再生成中...（30秒ほど）" if _correction_text
                    else f"Geminiが「{_target_label}」を引き直し中...（30秒ほど）"
                )
                with st.spinner(_spin_msg):
                    try:
                        _result_bytes = _trace_madori(
                            _raw_crop, _floor_type,
                            correction=_correction_text,
                            prev_result=_prev_bytes,
                        )
                        _result_img = Image.open(io.BytesIO(_result_bytes)).convert("RGB")
                        st.session_state["madori_traced_img"]   = _result_img
                        st.session_state["madori_traced_bytes"] = _result_bytes
                        st.session_state["madori_traced_key"]   = _target_label
                        st.rerun()
                    except Exception as _e:
                        st.error(f"❌ 清書エラー: {_e}")

    elif cropped and not TRACER_AVAILABLE:
        st.divider()
        st.info("💡 間取り図AI清書: `/Users/apple/madori-tracer/config.py` にGemini APIキーを設定すると利用できます。")

    # ── 手動差し替え ──────────────────────────────────────────────────────────
    st.divider()
    with st.expander("✋ 手動アップロードで差し替え（任意）"):
        mu1, mu2, mu3, mu4 = st.columns(4)
        with mu1: mf1 = st.file_uploader("外観写真", type=["png","jpg","jpeg"], key="mf1")
        with mu2: mf2 = st.file_uploader("室内写真", type=["png","jpg","jpeg"], key="mf2")
        with mu3: mf3 = st.file_uploader("間取り図", type=["png","jpg","jpeg"], key="mf3")
        with mu4: mf4 = st.file_uploader("地図",     type=["png","jpg","jpeg"], key="mf4")

    for lbl, mf in {"外観写真": mf1, "室内写真": mf2, "間取り図": mf3, "地図": mf4}.items():
        if mf:
            cropped[lbl] = Image.open(mf).convert("RGB")

    # ── 画像配置プレビュー（インタラクティブ）────────────────────────────────
    st.divider()
    st.markdown("**🖼️ Excel 画像配置プレビュー**")
    st.caption("チップをクリックで使用/非使用切り替え｜ドラッグで移動｜右下角をドラッグでリサイズ")

    # テンプレートの画像エリア（3様式とも 左 cols A-AJ = 1-36, rows 3-51。
    # スペック欄は AK 列 = 37 列目から始まる共通レイアウト）
    IMG_COL_START = 1
    IMG_COL_END   = 36
    IMG_ROW_START = 4   # タイトル(1-2行)と画像の間にスペースを空ける（3行目は余白）
    # 特徴アイコンを選んでいる場合は下部(45-51行)を特徴帯に空けるため写真は44行まで
    IMG_ROW_END   = 44 if st.session_state.get("features_sel") else 51

    # プレビューキャンバス（A4横比率）
    PREV_W = 600
    PREV_H = int(PREV_W * (210 / 297))  # A4横比率 ≈ 424

    total_cols     = 50
    img_area_ratio = IMG_COL_END / total_cols  # ~0.72

    placements: list[dict] = []

    if cropped:
        images_b64 = {label: _pil_to_b64(img) for label, img in cropped.items()}
        IA_W       = int(PREV_W * img_area_ratio)  # 画像エリアのピクセル幅

        # ── 初期配置の指定（レイアウト＋画像ごとの枠番号）─────────────────
        st.markdown("**📍 初期配置** — レイアウトを選び、各画像をどの枠に置くか番号で割り当て（後でドラッグ調整も可）")
        layout_name = st.selectbox("レイアウト", list(PLACE_LAYOUTS.keys()), key="place_layout")
        slots       = PLACE_LAYOUTS[layout_name]
        slot_labels = [s[0] for s in slots]
        n_slots     = len(slots)
        st.caption("枠: " + " ｜ ".join(slot_labels))

        OPT_NONE = "未使用"
        assign: dict[str, str] = {}
        _sel_cols = st.columns(min(len(cropped), 4))
        for i, label in enumerate(cropped.keys()):
            with _sel_cols[i % len(_sel_cols)]:
                di = _default_slot_index(label, n_slots)
                assign[label] = st.selectbox(
                    label, slot_labels + [OPT_NONE], index=di, key=f"slot_{label}")

        # 割り当て → 初期配置 dict
        init_pl: dict[str, dict] = {}
        for label, choice in assign.items():
            if choice == OPT_NONE:
                init_pl[label] = {"enabled": False, "x": 0, "y": 0, "w": 80, "h": 80}
            else:
                rect = slots[slot_labels.index(choice)][1]
                init_pl[label] = _slot_rect_px(rect, IA_W, PREV_H)

        # レイアウト/割り当てが変わったらコンポーネントを作り直して反映（key を変える）
        _sig = layout_name + "|" + "|".join(f"{k}={v}" for k, v in sorted(assign.items()))
        _editor_key = f"placement_editor_{abs(hash(_sig)) % (10 ** 8)}"

        placement_result = _placement_editor(
            images           = images_b64,
            container_w      = PREV_W,
            container_h      = PREV_H,
            img_area_ratio   = img_area_ratio,
            initial_placements = init_pl,
            key              = _editor_key,
            default          = None,
        )

        # placement_result（ドラッグ後）優先、無ければ指定した初期配置
        pd_now   = placement_result or init_pl
        img_cols = IMG_COL_END - IMG_COL_START + 1
        img_rows = IMG_ROW_END - IMG_ROW_START + 1

        for label, pl in pd_now.items():
            if not pl.get("enabled", True):
                continue
            img = cropped.get(label)
            if img is None:
                continue
            px = pl.get("x", 0)
            py = pl.get("y", 0)
            pw = max(1, pl.get("w", IA_W // 2))
            ph = max(1, pl.get("h", PREV_H // 2))

            # 列: コンポーネントは画像エリア(0→IA_W)内に配置するので IA_W で正規化
            c1 = IMG_COL_START + int(px / IA_W * img_cols)
            r1 = IMG_ROW_START + int(py / PREV_H * img_rows)
            c2 = min(IMG_COL_END, c1 + max(1, int(pw / IA_W * img_cols)))
            r2 = min(IMG_ROW_END, r1 + max(1, int(ph / PREV_H * img_rows)))
            placements.append({"img": img, "c1": c1, "r1": r1, "c2": c2, "r2": r2})

    st.session_state["placements"] = placements
    st.session_state["cropped"]    = cropped

    # ── 特徴アイコン選択（写真エリア下部に掲載）──────────────────────────────
    st.divider()
    st.markdown("**🏷️ 特徴アイコン** — マイソクから自動選択。手動で追加・削除できます（写真エリア下部に掲載）")
    _feat_default = detected_features if "features_sel" not in st.session_state else None
    selected_features = st.multiselect(
        "掲載する特徴を選択",
        FEATURE_LABELS,
        default=_feat_default,
        key="features_sel",
        help="選ぶと写真エリア下部にアイコンが並びます。多いほど小さく配置されます。",
    )
    _MAX_PER_ROW = 18   # 帯(A-AJ=36列)に1行で並ぶ最大個数（1個=2列換算）
    if selected_features:
        _n = len(selected_features)
        _cap1 = min(_n, _MAX_PER_ROW)
        st.caption(f"1行に並べられる最大は **{_MAX_PER_ROW}個**（多いほど小さくなります）。"
                   f"現在 {_n}個 選択中。")
        feature_per_row = st.slider(
            "1行あたりのアイコン数", 1, _MAX_PER_ROW,
            value=_cap1, key="features_per_row",
            help="この数で折り返します。1行に収めたい場合は選択数（最大18）に設定。",
        )
        _cat = {l: p for l, p in FEATURE_CATALOG}
        _pcols = st.columns(min(_n, 12))
        for _i, _f in enumerate(selected_features):
            with _pcols[_i % len(_pcols)]:
                if _f in _cat:
                    st.image(str(_cat[_f]), use_container_width=True)
                st.caption(_f)
    else:
        feature_per_row = None
    st.session_state["selected_features"] = selected_features
    st.session_state["features_per_row_val"] = feature_per_row

# ── Excel 生成 ────────────────────────────────────────────────────────────────
st.divider()
if st.button("📊 不動産案内書 Excel を作成", type="primary", use_container_width=True):
    placements = st.session_state.get("placements", [])

    company_now = load_company_info()
    if ci.get("商号"):
        company_now = ci

    with st.spinner("Excel を生成中..."):
        try:
            xlsx = create_fudosan_excel(
                template_type  = tmpl,
                specs          = edited_specs,
                company_info   = company_now,
                catchphrases   = edited_cps,
                image_placements = placements,
                features       = st.session_state.get("selected_features", []),
                feature_per_row = st.session_state.get("features_per_row_val"),
            )
            pname = edited_specs.get("物件名", "物件").replace(" ", "_").replace("/", "_")
            fname = f"{tmpl}案内書_{pname}.xlsx"

            st.download_button(
                label="⬇️ Excel をダウンロード",
                data=xlsx,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
            st.success("✅ 生成完了！ダウンロードしてください。")
        except Exception as e:
            st.error(f"❌ 生成エラー: {e}")
            st.exception(e)
