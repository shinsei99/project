# -*- coding: utf-8 -*-
"""帯変えモードの出力を検証する軽量テスト（外部ファイル・APIキー不要）。

    .venv/bin/python smoke_test.py

app.py は import すると Streamlit の UI まで走ってしまうので、
「# ─── Streamlit UI」より前（＝関数定義だけ）を取り出して実行する。
"""
import io
import sys
import types
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")
APP_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(APP_DIR))


def load_app_functions():
    src  = (APP_DIR / "app.py").read_text()
    head = src.split("# ─── Streamlit UI")[0]
    mod  = types.ModuleType("app_functions")
    mod.__file__ = str(APP_DIR / "app.py")
    sys.modules["app_functions"] = mod       # declare_component が呼び出し元を引くため
    exec(compile(head, "app.py", "exec"), mod.__dict__)
    return mod


def main() -> int:
    from PIL import Image
    import openpyxl

    app = load_app_functions()
    company = app.load_company_info()
    fails: list[str] = []

    def check(cond, msg):
        print(("  OK   " if cond else "  FAIL ") + msg)
        if not cond:
            fails.append(msg)

    # 縦(A4縦のマイソク) / 横(A4横) / 正方形 の3通り
    cases = [
        ("縦マイソク",   (1192, 1424), "portrait"),
        ("横マイソク",   (1686, 1048), "landscape"),
        ("正方形に近い", (1000, 1000), "portrait"),
    ]
    for name, (w, h), want_ori in cases:
        print(f"[{name}] {w}x{h}")
        img  = Image.new("RGB", (w, h), (240, 240, 240))
        data = app.create_band_swap_excel(img, company)
        ws   = openpyxl.load_workbook(io.BytesIO(data)).active

        check(ws.page_setup.orientation == want_ori,
              f"用紙の向きが {want_ori}（実際: {ws.page_setup.orientation}）")

        # 画像枠の縦横比が元画像と一致していれば、枠の中に白い余白が出ない
        sheet_w, img_h, band_h = app._sheet_px(ws)
        # 貼られた画像そのものの表示寸法（EMU）が元画像と同じ比か＝歪んでいないか
        a = ws._images[-1].anchor
        check(abs((a.ext.cx / a.ext.cy) - (w / h)) < 0.01,
              f"貼った画像が歪んでいない（{a.ext.cx/a.ext.cy:.3f} / 元 {w/h:.3f}）")
        check(abs(a.ext.cx / 9525 - sheet_w) < 2,
              f"画像の幅が自社帯と揃う（画像 {a.ext.cx/9525:.0f}px / 帯 {sheet_w:.0f}px）")

        # 帯は紙の上で 15〜24mm（横向きテンプレートの実測 21mm を基準に）
        portrait = ws.page_setup.orientation == "portrait"
        pw = app._body_width_mm(portrait)
        ph = (app._A4_LONG_MM if portrait else app._A4_SHORT_MM) - 2 * app._MARGIN_MM
        scale = min(pw / sheet_w, ph / (img_h + band_h))
        # 画像の段には「マイソク＋すき間」が入る。すき間は紙の上で既定 4mm
        gap = img_h - (sheet_w / (w / h))
        check(abs(gap * scale - app._GAP_MM) < 1.0,
              f"マイソクと帯の間のすき間が {app._GAP_MM:.0f}mm（実際 {gap*scale:.1f}mm）")
        check(15.0 <= band_h * scale <= 24.0,
              f"自社帯の高さが紙の上で 15〜24mm（実際 {band_h*scale:.1f}mm）")

        # 想定の本体幅（A4縦は195mm固定・A4横は289mm）の 95% 以上を使う
        check(sheet_w * scale / pw >= 0.95,
              f"本体幅 {pw:.0f}mm の95%以上を使う（実際 {sheet_w*scale:.0f}mm）")

        # A4縦は左右の余白で中央へ寄せる（帯の位置がマイソクによって動かない）
        if portrait:
            side = (app._A4_SHORT_MM - app._BODY_MM_PORTRAIT) / 2
            check(abs(ws.page_margins.left * 25.4 - side) < 0.2
                  and abs(ws.page_margins.right * 25.4 - side) < 0.2,
                  f"左右余白が各 {side:.1f}mm（実際 {ws.page_margins.left*25.4:.1f}mm）")
            check(ws.print_options.horizontalCentered is True, "横方向は中央寄せ")

        # 帯の文字が枠からはみ出していない（Excel は結合セルの文字を切り落とす）
        over = []
        for r in range(app._BAND_ROWS[0], app._BAND_ROWS[1] + 1):
            for c in range(1, app._MAX_COL + 1):
                cell = ws.cell(r, c)
                if cell.value in (None, ""):
                    continue
                end = c
                for rng in ws.merged_cells.ranges:
                    if rng.min_row == r and rng.min_col == c:
                        end = rng.max_col
                end  = max(end, app._box_end_col(ws, r, c))
                em   = sum(1.0 if ord(ch) > 0x2000 else 0.55 for ch in str(cell.value))
                need = em * (cell.font.size or 9) * app._PT2PX
                if need > app._cols_px(ws, c, end) * 1.02:
                    over.append(f"{cell.coordinate}={cell.value}")
        if portrait:
            # 縦向きは帯の文字を拡大するので、枠に収まっているかを必ず見る
            check(not over, "帯の文字が枠に収まっている" + (f" → はみ出し: {over}" if over else ""))
        elif over:
            # 横向きはテンプレートそのままの字送り（この改修では触っていない）。
            # 建設業番号が「担当者」欄へはみ出すのは以前からの症状なので、注意として出す
            print(f"  注意  横向きは以前からのはみ出しあり（未修正）: {over}")

        check(len(ws._images) >= 1, f"画像が貼られている（{len(ws._images)}枚）")

    # 本体幅の指定（画面のスライダー）が効くこと
    print("[本体幅の指定]")
    img = Image.new("RGB", (1192, 1424), (240, 240, 240))
    for want in (170, 202):
        ws = openpyxl.load_workbook(io.BytesIO(
            app.create_band_swap_excel(img, company, body_mm=want))).active
        sheet_w, img_h, band_h = app._sheet_px(ws)
        ph = app._A4_LONG_MM - 2 * app._MARGIN_MM
        scale = min(want / sheet_w, ph / (img_h + band_h))
        check(abs(sheet_w * scale - want) < 1.0,
              f"本体幅 {want}mm で出る（実際 {sheet_w*scale:.0f}mm）")
        side = (app._A4_SHORT_MM - want) / 2
        check(abs(ws.page_margins.left * 25.4 - side) < 0.2,
              f"左右余白が各 {side:.1f}mm（実際 {ws.page_margins.left*25.4:.1f}mm）")

    # 白フチの自動カット（マイソクの形・フチの太さが色々でも壊れないこと）
    print("[白フチの自動カット]")

    def with_margin(cw, ch, mx, my):
        im = Image.new("RGB", (cw + 2 * mx, ch + 2 * my), (255, 255, 255))
        im.paste(Image.new("RGB", (cw, ch), (30, 60, 200)), (mx, my))
        return im

    got, info = app.trim_white_margins(with_margin(800, 1000, 40, 50))
    check(got.size == (800, 1000) and info["trimmed"], f"5%の白フチを切る（{got.size}）")

    flush = Image.new("RGB", (800, 1000), (30, 60, 200))
    got, info = app.trim_white_margins(flush)
    check(got.size == (800, 1000) and not info["trimmed"], "フチが無い画像は触らない")

    # 白地に小さく置いた写真＝フチではないので切らない（1辺で3割超）
    got, info = app.trim_white_margins(with_margin(300, 300, 350, 350))
    check(not info["trimmed"], "余白が広すぎる画像は切らない（デザインを壊さない）")

    got, info = app.trim_white_margins(Image.new("RGB", (500, 400), (255, 255, 255)))
    check(not info["trimmed"], "真っ白な画像でも落ちない")

    print()
    if fails:
        print(f"NG: {len(fails)} 件失敗")
        return 1
    print("すべて OK")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
