"""law_check_template.xlsx を生成するスクリプト。

  .venv/bin/python templates/generate_template.py

ヘッダ・列幅・タイトル帯を作り込んだ空テンプレートを出力する。
実データ行は excel_export_service.build() が _DATA_START_ROW(6) から追記する。
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUT = os.path.join(os.path.dirname(__file__), "law_check_template.xlsx")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FILL = PatternFill("solid", fgColor="2E5496")
META_FILL = PatternFill("solid", fgColor="EAF0F8")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "リーガルチェック報告書"

    widths = {"A": 26, "B": 30, "C": 22, "D": 22, "E": 16, "F": 50}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # タイトル帯
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "書類リーガルチェック報告書（重説・契約書・謄本・行政 4点クロスチェック）"
    t.font = Font(bold=True, size=14, color="FFFFFF")
    t.fill = TITLE_FILL
    t.alignment = CENTER
    ws.row_dimensions[1].height = 30

    # メタ行
    for row, (label, val_cell) in enumerate(
        [("物件所在", "B2"), ("売主区分", "B3")], start=2
    ):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(bold=True)
        lc.fill = META_FILL
        lc.border = BORDER
        ws[val_cell].border = BORDER
        ws.cell(row=row, column=3).border = BORDER
    # 右側サマリ
    ws["D2"] = "齟齬・リスク"
    ws["D3"] = "一致"
    for r in (2, 3):
        ws.cell(row=r, column=4).font = Font(bold=True)
        ws.cell(row=r, column=4).fill = META_FILL
        ws.cell(row=r, column=4).border = BORDER
        ws.cell(row=r, column=5).border = BORDER

    # 空行
    ws.row_dimensions[4].height = 6

    # ヘッダ行（row 5）
    headers = [
        "チェック項目",
        "🌐行政正解 / 📄謄本ファクト",
        "📝重説記載値",
        "🛒契約書記載値",
        "判定",
        "修正指示・アドバイス",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[5].height = 28

    # データ雛形行（row 6）に罫線・折返しスタイルを仕込む
    for col in range(1, 7):
        c = ws.cell(row=6, column=col)
        c.border = BORDER
        c.alignment = WRAP

    ws.freeze_panes = "A6"
    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
