"""検閲結果を「書類リーガルチェック報告書（Excel）」として出力する。

templates/law_check_template.xlsx を読み込み、openpyxl で結果行を追記。
🔴（齟齬・リスクあり）の行は行全体を薄い赤でハイライトする。

列構成:
  A: チェック項目
  B: 🌐行政正解値 / 📄謄本ファクト値
  C: 📝重説記載値
  D: 🛒契約書記載値
  E: 判定ステータス（🟢一致 / 🔴齟齬・リスクあり / ⚪確認不可）
  F: 修正指示・アドバイス
"""

from __future__ import annotations

import copy
import io
import os

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from models.legal_check_data import STATUS_NG, LegalCrossCheckData

_TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)), "templates", "law_check_template.xlsx"
)

_DATA_START_ROW = 6  # テンプレートのヘッダ下の最初のデータ行

# 🔴行のハイライト（薄い赤）
_NG_FILL = PatternFill("solid", fgColor="FCE4E4")
_NG_FONT = Font(color="C00000")
_CAT_FILL = PatternFill("solid", fgColor="E7EEF7")
_CAT_FONT = Font(bold=True, color="1F3864")
_WRAP = Alignment(wrap_text=True, vertical="top")


def build(data: LegalCrossCheckData) -> bytes:
    wb = load_workbook(_TEMPLATE_PATH)
    ws = wb.active

    # メタ情報（物件所在・売主区分・サマリ）
    ws["B2"] = data.address or "（住所未入力）"
    ws["B3"] = "宅地建物取引業者" if data.seller_is_pro else "個人"
    ws["E2"] = f"🔴 {data.ng_count} 件"
    ws["E3"] = f"🟢 {data.ok_count} 件"

    # テンプレート行のスタイルを雛形として保持
    tmpl_cells = {c: ws.cell(row=_DATA_START_ROW, column=c) for c in range(1, 7)}

    row = _DATA_START_ROW
    last_cat = None
    for res in data.results:
        # カテゴリ見出し行
        if res.category != last_cat:
            _apply_style(ws, row, tmpl_cells)
            ws.cell(row=row, column=1, value=f"■ {res.category}")
            for c in range(1, 7):
                cell = ws.cell(row=row, column=c)
                cell.fill = _CAT_FILL
                cell.font = _CAT_FONT
            row += 1
            last_cat = res.category

        _apply_style(ws, row, tmpl_cells)
        ws.cell(row=row, column=1, value=res.item)
        ws.cell(row=row, column=2, value=res.admin_value)
        ws.cell(row=row, column=3, value=res.explanation_value)
        ws.cell(row=row, column=4, value=res.contract_value)
        ws.cell(row=row, column=5, value=f"{res.icon} {res.status}")
        ws.cell(row=row, column=6, value=res.advice)

        for c in range(1, 7):
            ws.cell(row=row, column=c).alignment = _WRAP

        if res.status == STATUS_NG:
            for c in range(1, 7):
                cell = ws.cell(row=row, column=c)
                cell.fill = _NG_FILL
                cell.font = _NG_FONT
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _apply_style(ws, row: int, tmpl_cells: dict) -> None:
    """テンプレート行の罫線・フォントをコピー。"""
    for c, tmpl in tmpl_cells.items():
        cell = ws.cell(row=row, column=c)
        if tmpl.has_style:
            cell._style = copy.copy(tmpl._style)
