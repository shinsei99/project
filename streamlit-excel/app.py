import streamlit as st
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
import pandas as pd
import io
import re

st.set_page_config(
    page_title="Excel セル編集ツール",
    page_icon="📊",
    layout="wide",
)

# ── ユーティリティ ──────────────────────────────────────────────────────────

def parse_cell(ref: str):
    m = re.match(r"^([A-Z]+)(\d+)$", ref.strip().upper())
    if not m:
        return None, None
    try:
        return int(m.group(2)), column_index_from_string(m.group(1))
    except Exception:
        return None, None


def coerce_value(s: str):
    s = s.strip()
    if not s:
        return None
    if re.match(r"^0\d", s):
        return s
    try:
        if "." not in s and "e" not in s.lower():
            return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    return s


def validate(cell_str: str, value_str: str):
    """(正規化セル番地, 変換後の値, エラー文字列 or None) を返す"""
    ref = cell_str.strip().upper()
    if not ref:
        return None, None, "セル番地を入力してください"
    row, _ = parse_cell(ref)
    if not row:
        return None, None, f"セル番地が無効です: {cell_str}（例: A1、B3）"
    if value_str.strip() == "":
        return None, None, "値を入力してください"
    return ref, coerce_value(value_str), None


def ws_to_df(ws) -> pd.DataFrame:
    max_r, max_c = ws.max_row, ws.max_column
    if not max_r or not max_c:
        return pd.DataFrame()
    cols  = [get_column_letter(c) for c in range(1, max_c + 1)]
    index = list(range(1, max_r + 1))
    data  = [
        [ws.cell(row=r, column=c).value for c in range(1, max_c + 1)]
        for r in range(1, max_r + 1)
    ]
    return pd.DataFrame(data, columns=cols, index=index)


def build_output(file_bytes: bytes, edits: list) -> bytes:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    for e in edits:
        if e["sheet"] in wb.sheetnames:
            wb[e["sheet"]][e["cell"]] = e["value"]
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── セッション初期化 ────────────────────────────────────────────────────────

for key, default in [
    ("edits", []),
    ("file_bytes", None),
    ("fname", None),
    ("editing_index", None),   # int: 再編集中のインデックス / None: 追加モード
]:
    if key not in st.session_state:
        st.session_state[key] = default

# ── ヘッダー ────────────────────────────────────────────────────────────────

st.title("📊 Excel セル編集ツール")
st.caption("Excelファイルをアップロードし、セルを指定して値を変更・ダウンロードできます")

# ── ファイルアップロード ─────────────────────────────────────────────────────

uploaded = st.file_uploader(
    "Excelファイルをアップロード（.xlsx / .xlsm）",
    type=["xlsx", "xlsm"],
)

if not uploaded:
    st.info("👆 Excelファイル（.xlsx / .xlsm）をアップロードしてください")
    st.stop()

if st.session_state.fname != uploaded.name:
    st.session_state.file_bytes    = uploaded.read()
    st.session_state.fname         = uploaded.name
    st.session_state.edits         = []
    st.session_state.editing_index = None
    # フォーム入力をクリア
    st.session_state.pop("form_cell",  None)
    st.session_state.pop("form_value", None)

# ── メインレイアウト ─────────────────────────────────────────────────────────

left, right = st.columns([3, 2], gap="large")
wb = openpyxl.load_workbook(io.BytesIO(st.session_state.file_bytes), data_only=True)

# ── 左ペイン: プレビュー ─────────────────────────────────────────────────────

with left:
    sheet = st.selectbox("シートを選択", wb.sheetnames)
    ws    = wb[sheet]
    df    = ws_to_df(ws)

    display_df   = df.copy() if not df.empty else pd.DataFrame()
    changed_keys: set = set()

    for e in st.session_state.edits:
        if e["sheet"] == sheet:
            r, c = parse_cell(e["cell"])
            if r and c:
                col_lbl = get_column_letter(c)
                if not display_df.empty and r in display_df.index and col_lbl in display_df.columns:
                    display_df.at[r, col_lbl] = e["value"]
                    changed_keys.add((r, col_lbl))

    n_edits = len([e for e in st.session_state.edits if e["sheet"] == sheet])
    st.subheader(f"📄 {sheet}" + (f"　（{n_edits} 件編集中）" if n_edits else ""))

    if display_df.empty:
        st.info("このシートにデータがありません")
    else:
        def highlight_edits(df_):
            s = pd.DataFrame("", index=df_.index, columns=df_.columns)
            for (r, c) in changed_keys:
                if r in s.index and c in s.columns:
                    s.at[r, c] = "background-color:#fff3cd;font-weight:bold;color:#856404"
            return s

        st.dataframe(
            display_df.style.apply(highlight_edits, axis=None),
            use_container_width=True,
            height=440,
        )

# ── 右ペイン: 編集フォーム ────────────────────────────────────────────────────

with right:
    editing_idx = st.session_state.editing_index
    is_editing  = (editing_idx is not None and editing_idx < len(st.session_state.edits))

    # ── フォームヘッダー ──────────────────────────────────────────────────
    if is_editing:
        current = st.session_state.edits[editing_idx]
        st.subheader(f"🔧 再編集中: `{current['sheet']}!{current['cell']}`")
    else:
        st.subheader("✏️ セルを追加")

    # ── フォーム ─────────────────────────────────────────────────────────
    # key= を使うことで st.session_state 経由の事前入力が可能
    with st.form("cell_form", clear_on_submit=True):
        cell_input = st.text_input(
            "セル番地",
            key="form_cell",
            placeholder="例: A1、B3、C10",
            help="Excelのセル番地（列A-Z＋行番号）を入力",
        )
        value_input = st.text_input(
            "新しい値",
            key="form_value",
            placeholder="例: 100、テキスト、2024-01-01",
        )

        if is_editing:
            fc1, fc2 = st.columns(2)
            with fc1:
                update_btn = st.form_submit_button("✓ 更新する", type="primary", use_container_width=True)
            with fc2:
                cancel_btn = st.form_submit_button("キャンセル", use_container_width=True)
            add_btn = False
        else:
            add_btn    = st.form_submit_button("＋ 追加", type="primary", use_container_width=True)
            update_btn = False
            cancel_btn = False

    # ── フォーム送信処理 ──────────────────────────────────────────────────

    if cancel_btn:
        st.session_state.editing_index = None
        st.rerun()

    if update_btn:
        new_ref, new_val, err = validate(cell_input, value_input)
        if err:
            st.error(err)
        else:
            original_sheet = st.session_state.edits[editing_idx]["sheet"]
            # 同シート・同セルへの別エントリが存在する場合は統合して重複排除
            conflict = next(
                (j for j, ex in enumerate(st.session_state.edits)
                 if ex["sheet"] == original_sheet and ex["cell"] == new_ref and j != editing_idx),
                None,
            )
            if conflict is not None:
                st.session_state.edits[conflict]["value"] = new_val
                st.session_state.edits.pop(editing_idx)
            else:
                st.session_state.edits[editing_idx] = {
                    "sheet": original_sheet,
                    "cell":  new_ref,
                    "value": new_val,
                }
            st.session_state.editing_index = None
            st.toast("更新しました ✓")
            st.rerun()

    if add_btn:
        new_ref, new_val, err = validate(cell_input, value_input)
        if err:
            st.error(err)
        else:
            existing = next(
                (e for e in st.session_state.edits
                 if e["sheet"] == sheet and e["cell"] == new_ref),
                None,
            )
            if existing:
                existing["value"] = new_val
                st.toast(f"{sheet}!{new_ref} を更新しました ✓")
            else:
                st.session_state.edits.append({"sheet": sheet, "cell": new_ref, "value": new_val})
                st.toast(f"{sheet}!{new_ref} を追加しました ✓")
            st.rerun()

    # ── 編集リスト ────────────────────────────────────────────────────────

    st.divider()

    if not st.session_state.edits:
        st.info("セルを追加すると、ここに編集リストが表示されます")
    else:
        st.subheader(f"📝 編集リスト（{len(st.session_state.edits)} 件）")

        for i, e in enumerate(st.session_state.edits):
            is_this = (st.session_state.editing_index == i)
            c_text, c_edit, c_del = st.columns([5, 1, 1])

            with c_text:
                if is_this:
                    st.markdown(
                        f'<div style="background:#dbeafe;border-left:3px solid #3b82f6;'
                        f'padding:4px 10px;border-radius:4px;font-size:0.9rem">'
                        f'<code>{e["sheet"]}!{e["cell"]}</code>　→　<b>{e["value"]}</b> 🔧</div>',
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(f"`{e['sheet']}!{e['cell']}`　→　**{e['value']}**")

            with c_edit:
                # 編集中は 📌、それ以外は ✏️（クリックでトグル）
                if st.button("📌" if is_this else "✏️", key=f"edit_{i}",
                             help="編集を閉じる" if is_this else "この値を再編集"):
                    if is_this:
                        # 編集モードを閉じる
                        st.session_state.editing_index = None
                    else:
                        # このアイテムを編集モードに切り替え
                        # key= 付き text_input は session_state 経由で事前入力できる
                        st.session_state.editing_index  = i
                        st.session_state["form_cell"]  = e["cell"]
                        st.session_state["form_value"] = str(e["value"])
                    st.rerun()

            with c_del:
                if st.button("✕", key=f"del_{i}", help="この編集を削除"):
                    st.session_state.edits.pop(i)
                    # インデックス補正
                    if st.session_state.editing_index == i:
                        st.session_state.editing_index = None
                    elif (st.session_state.editing_index is not None
                          and st.session_state.editing_index > i):
                        st.session_state.editing_index -= 1
                    st.rerun()

        if st.button("すべてクリア", use_container_width=True):
            st.session_state.edits         = []
            st.session_state.editing_index = None
            st.rerun()

        st.divider()

        out_bytes = build_output(st.session_state.file_bytes, st.session_state.edits)
        stem      = st.session_state.fname.rsplit(".", 1)[0]
        out_name  = f"{stem}_edited.xlsx"

        st.download_button(
            "⬇️ 編集済みファイルをダウンロード",
            data=out_bytes,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )
