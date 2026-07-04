# -*- coding: utf-8 -*-
"""
オーナー送金・月次締めマネージャー (owner-payout-tracker)
賃貸・ビル管理の月次サイクルを「集金日別」に進捗管理し、オーナー報告書を自動生成する。

月次5ステップ:
  ① 請求書発行  ② 郵送・投函  ③ 入金確認  ④ オーナー報告書  ⑤ 送金・送付

これまでファイル名の「☑」で手管理していた締め状況を、本物の進捗ボードに置き換える。
"""
import json
import re
import sqlite3
from datetime import date
from io import BytesIO
from pathlib import Path

import streamlit as st

BASE = Path(__file__).parent
DB = BASE / "tracker.db"
SEED = BASE / "seed_properties.json"

STEPS = [
    ("s1_invoice", "請求書発行", "📄"),
    ("s2_mail", "郵送・投函", "✉️"),
    ("s3_payment", "入金確認", "💰"),
    ("s4_report", "オーナー報告書", "📊"),
    ("s5_remit", "送金・送付", "🏦"),
]

st.set_page_config(page_title="オーナー送金・月次締めマネージャー", page_icon="🏢", layout="wide")


# ─────────────────────────── DB ───────────────────────────
def conn():
    c = sqlite3.connect(DB)
    c.row_factory = sqlite3.Row
    return c


def init_db():
    with conn() as c:
        c.execute(
            """CREATE TABLE IF NOT EXISTS properties(
                id INTEGER PRIMARY KEY, pay_day TEXT, sort_day INTEGER,
                owner TEXT, property TEXT, type TEXT, frequency TEXT, note TEXT,
                mail_kubun TEXT, mail_method TEXT, mail_timing TEXT, mail_note TEXT)"""
        )
        c.execute(
            """CREATE TABLE IF NOT EXISTS status(
                property_id INTEGER, ym TEXT,
                s1_invoice INT DEFAULT 0, s2_mail INT DEFAULT 0, s3_payment INT DEFAULT 0,
                s4_report INT DEFAULT 0, s5_remit INT DEFAULT 0,
                memo TEXT DEFAULT '', updated TEXT DEFAULT '',
                PRIMARY KEY(property_id, ym))"""
        )
        c.execute(
            """CREATE TABLE IF NOT EXISTS reports(
                property_id INTEGER, ym TEXT,
                collected REAL DEFAULT 0, mgmt_rate REAL DEFAULT 5,
                mgmt_fee REAL DEFAULT 0, expenses TEXT DEFAULT '[]',
                remit REAL DEFAULT 0, comment TEXT DEFAULT '', updated TEXT DEFAULT '',
                PRIMARY KEY(property_id, ym))"""
        )
        # 初回シード投入
        n = c.execute("SELECT COUNT(*) FROM properties").fetchone()[0]
        if n == 0 and SEED.exists():
            for p in json.loads(SEED.read_text(encoding="utf-8")):
                c.execute(
                    """INSERT INTO properties VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (p["id"], p["pay_day"], p["sort_day"], p["owner"], p["property"],
                     p["type"], p["frequency"], p["note"], p["mail_kubun"],
                     p["mail_method"], p["mail_timing"], p["mail_note"]),
                )


def get_props():
    with conn() as c:
        return [dict(r) for r in c.execute(
            "SELECT * FROM properties ORDER BY sort_day, id").fetchall()]


def get_status(ym, pid):
    with conn() as c:
        r = c.execute("SELECT * FROM status WHERE property_id=? AND ym=?", (pid, ym)).fetchone()
        if r:
            return dict(r)
    return {k: 0 for k, _, _ in STEPS} | {"memo": ""}


def set_status(ym, pid, field, val):
    with conn() as c:
        c.execute(
            "INSERT INTO status(property_id,ym) VALUES(?,?) ON CONFLICT DO NOTHING", (pid, ym))
        c.execute(f"UPDATE status SET {field}=?, updated=? WHERE property_id=? AND ym=?",
                  (val, date.today().isoformat(), pid, ym))


def get_report(ym, pid):
    with conn() as c:
        r = c.execute("SELECT * FROM reports WHERE property_id=? AND ym=?", (pid, ym)).fetchone()
        if r:
            d = dict(r)
            d["expenses"] = json.loads(d["expenses"] or "[]")
            return d
    return {"collected": 0, "mgmt_rate": 5.0, "mgmt_fee": 0, "expenses": [], "remit": 0, "comment": ""}


def save_report(ym, pid, d):
    with conn() as c:
        c.execute(
            """INSERT INTO reports(property_id,ym,collected,mgmt_rate,mgmt_fee,expenses,remit,comment,updated)
               VALUES(?,?,?,?,?,?,?,?,?)
               ON CONFLICT(property_id,ym) DO UPDATE SET
               collected=excluded.collected,mgmt_rate=excluded.mgmt_rate,mgmt_fee=excluded.mgmt_fee,
               expenses=excluded.expenses,remit=excluded.remit,comment=excluded.comment,updated=excluded.updated""",
            (pid, ym, d["collected"], d["mgmt_rate"], d["mgmt_fee"],
             json.dumps(d["expenses"], ensure_ascii=False), d["remit"], d["comment"],
             date.today().isoformat()),
        )


# ─────────────────────── ヘルパー ───────────────────────
def progress(ym, pid):
    s = get_status(ym, pid)
    done = sum(1 for k, _, _ in STEPS if s.get(k))
    return done, len(STEPS)


def deadline_passed(sort_day):
    if sort_day >= 99:
        return False
    return date.today().day > sort_day + 3  # 締め日+3日で警告


def yen(v):
    try:
        return f"¥{int(round(v)):,}"
    except Exception:
        return "¥0"


# ─────────────────────── オーナー報告書 Excel ───────────────────────
def build_report_xlsx(prop, ym, rep):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "オーナー報告書"
    thin = Side(style="thin", color="999999")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr = PatternFill("solid", fgColor="1F3864")
    sub = PatternFill("solid", fgColor="D9E1F2")
    wf = Font(color="FFFFFF", bold=True, size=12)

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 26

    y, m = ym.split("-")
    ws.merge_cells("A1:D1")
    ws["A1"] = "管 理 業 務 報 告 書"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:D2")
    ws["A2"] = f"{y}年{int(m)}月分"
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["B4"] = f"オーナー：{prop['owner'] or prop['property']} 様"
    ws["B4"].font = Font(bold=True, size=12)
    ws["B5"] = f"対象物件：{prop['property']}"
    ws["B6"] = "大京商事株式会社　TEL 06-6353-0418 / FAX 06-6353-0280"
    ws["B6"].font = Font(size=9, color="666666")

    r = 8
    ws.merge_cells(f"B{r}:D{r}")
    ws[f"B{r}"] = "■ 収支明細"
    ws[f"B{r}"].fill = hdr
    ws[f"B{r}"].font = wf
    r += 1

    def line(label, val, fill=None, bold=False):
        nonlocal r
        ws.merge_cells(f"B{r}:C{r}")
        ws[f"B{r}"] = label
        ws[f"D{r}"] = yen(val)
        ws[f"D{r}"].alignment = Alignment(horizontal="right")
        for col in ("B", "C", "D"):
            ws[f"{col}{r}"].border = box
            if fill:
                ws[f"{col}{r}"].fill = fill
            if bold:
                ws[f"{col}{r}"].font = Font(bold=True)
        r += 1

    line("① ご入金（集金）合計", rep["collected"], sub, True)
    line(f"② 管理手数料（{rep['mgmt_rate']:g}%）", -rep["mgmt_fee"])
    for e in rep["expenses"]:
        line(f"③ {e.get('name', '経費')}", -float(e.get("amount", 0)))
    ws.merge_cells(f"B{r}:C{r}")
    ws[f"B{r}"] = "オーナー様お振込額"
    ws[f"D{r}"] = yen(rep["remit"])
    ws[f"D{r}"].alignment = Alignment(horizontal="right")
    for col in ("B", "C", "D"):
        ws[f"{col}{r}"].fill = PatternFill("solid", fgColor="FFF2CC")
        ws[f"{col}{r}"].font = Font(bold=True, size=13)
        ws[f"{col}{r}"].border = box
    r += 2

    if rep["comment"]:
        ws.merge_cells(f"B{r}:D{r}")
        ws[f"B{r}"] = "■ 特記事項"
        ws[f"B{r}"].fill = hdr
        ws[f"B{r}"].font = wf
        r += 1
        ws.merge_cells(f"B{r}:D{r+2}")
        ws[f"B{r}"] = rep["comment"]
        ws[f"B{r}"].alignment = Alignment(vertical="top", wrap_text=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════ UI ═══════════════════════════
init_db()

with st.sidebar:
    st.title("🏢 月次締めマネージャー")
    today = date.today()
    ym = st.text_input("対象年月 (YYYY-MM)", value=f"{today.year}-{today.month:02d}")
    if not re.match(r"^\d{4}-\d{2}$", ym):
        st.error("YYYY-MM 形式で入力してください")
        st.stop()
    page = st.radio("メニュー", ["📋 月次ダッシュボード", "📊 オーナー報告書作成", "🏘 物件マスター"])
    st.caption("大京商事株式会社 / 賃貸・ビル管理")

props = get_props()

# ── 全体サマリ ──
total_steps = len(props) * len(STEPS)
done_steps = 0
alerts = []
for p in props:
    d, _ = progress(ym, p["id"])
    done_steps += d
    if d < len(STEPS) and deadline_passed(p["sort_day"]):
        alerts.append((p, d))


# ══════════════ ① ダッシュボード ══════════════
if page.startswith("📋"):
    st.header(f"📋 {ym} 月次締めダッシュボード")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("管理先", f"{len(props)} 件")
    c2.metric("全体進捗", f"{done_steps}/{total_steps}",
              f"{(done_steps/total_steps*100 if total_steps else 0):.0f}%")
    fully = sum(1 for p in props if progress(ym, p["id"])[0] == len(STEPS))
    c3.metric("締め完了", f"{fully} 件")
    c4.metric("⚠️ 締切超過・未完", f"{len(alerts)} 件")

    if alerts:
        with st.expander(f"⚠️ 締切超過アラート（{len(alerts)}件）", expanded=True):
            for p, d in alerts:
                st.warning(f"**{p['pay_day']}｜{p['property']}**（{p['owner'] or '—'}）… "
                           f"進捗 {d}/{len(STEPS)}　未完ステップあり")

    st.divider()

    # 集金日でグルーピング
    groups = {}
    for p in props:
        groups.setdefault(p["pay_day"], []).append(p)
    order = sorted(groups.keys(), key=lambda k: (groups[k][0]["sort_day"], k))

    for g in order:
        gp = groups[g]
        gdone = sum(progress(ym, p["id"])[0] for p in gp)
        gtot = len(gp) * len(STEPS)
        st.subheader(f"🗓 {g}　"
                     f"（{sum(1 for p in gp if progress(ym,p['id'])[0]==len(STEPS))}/{len(gp)} 完了）")
        st.progress(gdone / gtot if gtot else 0)

        for p in gp:
            s = get_status(ym, p["id"])
            d = sum(1 for k, _, _ in STEPS if s.get(k))
            warn = "🔴 " if (d < len(STEPS) and deadline_passed(p["sort_day"])) else ""
            badge = "✅ " if d == len(STEPS) else ""
            with st.expander(f"{badge}{warn}{p['property']}　—　{p['owner'] or '（自社/委託）'}"
                             f"　［{d}/{len(STEPS)}］"):
                meta = []
                if p["type"]:
                    meta.append(f"種別: {p['type']}")
                if p["frequency"]:
                    meta.append(f"頻度: {p['frequency']}")
                if p["mail_method"]:
                    meta.append(f"郵送: **{p['mail_method']}**"
                                f"（{p['mail_kubun']}／{p['mail_timing']}）")
                st.caption("　｜　".join(meta) if meta else "—")
                if p["mail_note"]:
                    st.caption(f"📮 {p['mail_note']}")
                if p["note"]:
                    st.caption(f"📝 {p['note']}")

                cols = st.columns(len(STEPS))
                for i, (k, label, icon) in enumerate(STEPS):
                    with cols[i]:
                        v = st.checkbox(f"{icon}{label}", value=bool(s.get(k)),
                                        key=f"{ym}-{p['id']}-{k}")
                        if v != bool(s.get(k)):
                            set_status(ym, p["id"], k, int(v))
                            st.rerun()
                memo = st.text_input("メモ", value=s.get("memo", ""),
                                     key=f"memo-{ym}-{p['id']}")
                if memo != s.get("memo", ""):
                    set_status(ym, p["id"], "memo", memo)


# ══════════════ ② オーナー報告書作成 ══════════════
elif page.startswith("📊"):
    st.header(f"📊 オーナー報告書 自動作成　（{ym}）")
    owner_props = [p for p in props if p["owner"]]
    labels = {f"{p['property']} — {p['owner']}": p for p in owner_props}
    sel = st.selectbox("対象物件を選択", list(labels.keys()))
    p = labels[sel]
    rep = get_report(ym, p["id"])

    st.subheader("① 集金・手数料")
    a, b = st.columns(2)
    collected = a.number_input("ご入金（集金）合計 ¥", min_value=0.0,
                               value=float(rep["collected"]), step=1000.0, format="%.0f")
    rate = b.number_input("管理手数料率 %", min_value=0.0, max_value=100.0,
                          value=float(rep["mgmt_rate"]), step=0.5)
    mgmt_fee = round(collected * rate / 100)
    b.caption(f"管理手数料 = {yen(mgmt_fee)}")

    st.subheader("② 経費・立替（修繕・広告料など）")
    if "exp_rows" not in st.session_state or st.session_state.get("exp_pid") != p["id"] or st.session_state.get("exp_ym") != ym:
        st.session_state.exp_rows = rep["expenses"] or [{"name": "", "amount": 0}]
        st.session_state.exp_pid = p["id"]
        st.session_state.exp_ym = ym
    expenses = []
    for idx, e in enumerate(st.session_state.exp_rows):
        c1, c2, c3 = st.columns([3, 2, 1])
        nm = c1.text_input("項目", value=e.get("name", ""), key=f"en-{idx}")
        am = c2.number_input("金額 ¥", min_value=0.0, value=float(e.get("amount", 0)),
                             step=1000.0, format="%.0f", key=f"ea-{idx}")
        if c3.button("🗑", key=f"ed-{idx}"):
            st.session_state.exp_rows.pop(idx)
            st.rerun()
        if nm or am:
            expenses.append({"name": nm, "amount": am})
    if st.button("＋ 経費行を追加"):
        st.session_state.exp_rows.append({"name": "", "amount": 0})
        st.rerun()

    exp_total = sum(float(e["amount"]) for e in expenses)
    remit = collected - mgmt_fee - exp_total

    st.subheader("③ お振込額")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("集金合計", yen(collected))
    m2.metric("管理手数料", "−" + yen(mgmt_fee))
    m3.metric("経費合計", "−" + yen(exp_total))
    m4.metric("オーナー振込額", yen(remit))

    comment = st.text_area("特記事項（任意）", value=rep["comment"])

    c1, c2 = st.columns(2)
    if c1.button("💾 保存", type="primary", use_container_width=True):
        save_report(ym, p["id"], {
            "collected": collected, "mgmt_rate": rate, "mgmt_fee": mgmt_fee,
            "expenses": expenses, "remit": remit, "comment": comment})
        set_status(ym, p["id"], "s4_report", 1)  # 報告書作成済みに連動
        st.success("保存しました（ダッシュボードの『オーナー報告書』を☑にしました）")

    data = {"collected": collected, "mgmt_rate": rate, "mgmt_fee": mgmt_fee,
            "expenses": expenses, "remit": remit, "comment": comment}
    c2.download_button("⬇️ Excelダウンロード",
                       data=build_report_xlsx(p, ym, data),
                       file_name=f"オーナー報告書_{p['property']}_{ym}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       use_container_width=True)


# ══════════════ ③ 物件マスター ══════════════
else:
    st.header("🏘 物件マスター（管理先一覧）")
    st.caption("『管理先・入居者一覧』②毎月請求(支払日別)＋④郵送投函方法 より自動生成")
    import pandas as pd
    df = pd.DataFrame(props)[
        ["pay_day", "owner", "property", "type", "frequency",
         "mail_kubun", "mail_method", "mail_timing", "mail_note", "note"]]
    df.columns = ["集金日", "オーナー/委託先", "物件", "種別", "頻度",
                  "郵送区分", "郵送方法", "郵送時期", "郵送メモ", "備考"]
    st.dataframe(df, use_container_width=True, height=640, hide_index=True)
    st.info(f"全 {len(props)} 管理先。データ更新は『管理先・入居者一覧』Excel を再取込してください。")
