"""UI を起動せずパイプライン各段を検証するスモークテスト。"""
import os

from models.property_data import create_property_data, merge
from services import (
    comment_service,
    excel_export_service,
    pdf_export_service,
    registry_service,
)
from utils import parser

BASE = os.path.dirname(os.path.abspath(__file__))

# 1) PropertyData
data = create_property_data()
assert "所在地" in data and data["所在地"] == ""

# 2) マージ（空/不正キーを無視）
merge(data, {"所在地": " 東京都千代田区丸の内1-1-1 ", "用途地域": "商業地域",
             "建ぺい率": "80%", "容積率": "800%", "不明キー": "x", "地番": ""})
assert data["所在地"] == "東京都千代田区丸の内1-1-1"
assert "不明キー" not in data
assert data["地番"] == ""  # 空文字は上書きしない

# 3) 登記簿パーサ（テキスト直接）
land_text = "所　在  千代田区丸の内一丁目\n地　番  1番1\n地　目  宅地\n地　積  123.45㎡\n所有者  山田太郎\n抵当権設定"
land = parser.parse_land(land_text)
assert land["地目"] == "宅地", land
assert land["地積"].startswith("123"), land
assert "抵当権" in land["抵当権"], land
print("land parse:", land)

# 4) registry_service 統合
reg = registry_service.parse_registry(None, None)
assert isinstance(reg, dict)

# 5) コメント生成
merge(data, {"最寄駅": "東京駅（JR山手線）", "駅距離": "約 350m"})
comment = comment_service.generate_comment(data)
assert "商業地域" in comment and len(comment) > 50
print("comment:", comment)

# 6) Excel 出力（テンプレ自動生成含む）
tpl = os.path.join(BASE, "templates", "jyuusetsu_template.xlsx")
out_xlsx = os.path.join(BASE, "reports", "jyuusetsu_draft.xlsx")
excel_export_service.export_excel(data, comment, tpl, out_xlsx)
assert os.path.exists(tpl) and os.path.exists(out_xlsx)
from openpyxl import load_workbook
ws = load_workbook(out_xlsx).active
assert ws["B2"].value == "東京都千代田区丸の内1-1-1", ws["B2"].value
assert ws["B10"].value == "商業地域", ws["B10"].value
print("excel B2/B10/B11:", ws["B2"].value, ws["B10"].value, ws["B11"].value)

# 7) PDF 出力
out_pdf = os.path.join(BASE, "reports", "jyuusetsu_draft.pdf")
pdf_export_service.export_pdf(data, comment, out_pdf)
assert os.path.exists(out_pdf) and os.path.getsize(out_pdf) > 1000
print("pdf bytes:", os.path.getsize(out_pdf))

print("\nALL SMOKE TESTS PASSED")


# 9) 公式書式カタログ（レジストリが無い環境では skip）
from services import format_catalog  # noqa: E402

msg = format_catalog.status_message()
if msg:
    print("[skip] 公式書式カタログ:", msg)
else:
    cats = format_catalog.categories()
    assert cats, "分類が空"
    total = sum(len(format_catalog.formats_in(c)) for c in cats)
    assert total > 0, "書式が0本"
    # 自動入力書式は「重説へ入れると他書式へ波及する」ものが必ずある
    fanout = [f for c in cats for f in format_catalog.formats_in(c) if f.get("fanout_count")]
    assert fanout, "波及する書式が見つからない（レジストリの作りが変わった可能性）"
    print("[ok] 公式書式カタログ: %d分類 / %d本 / 波及型 %d本" % (len(cats), total, len(fanout)))

# 10) クロスチェック（legal-crosscheck から吸収した検閲エンジン）
import io as _io  # noqa: E402
from reportlab.pdfgen import canvas as _canvas  # noqa: E402
from reportlab.pdfbase import pdfmetrics as _pm  # noqa: E402
from reportlab.pdfbase.cidfonts import UnicodeCIDFont as _CID  # noqa: E402
from services import crosscheck_service, crosscheck_report_service  # noqa: E402

_pm.registerFont(_CID("HeiseiKakuGo-W5"))


def _mkpdf(lines):
    buf = _io.BytesIO()
    c = _canvas.Canvas(buf)
    c.setFont("HeiseiKakuGo-W5", 11)
    y = 800
    for ln in lines:
        c.drawString(50, y, ln)
        y -= 18
    c.save()
    return buf.getvalue()


_base = {"所在地": "大阪市都島区東野田町二丁目", "地番": "123番4", "地積": "165.28",
         "家屋番号": "123番4", "床面積": "98.55", "所有者": "検証太郎",
         "用途地域": "商業地域", "建ぺい率": "80%", "容積率": "600%"}

# 書類が無ければ「一致」は1件も出ない（比較していないのに緑を出さない）
_none = crosscheck_service.run(_base, None, None)
assert _none.ok_count == 0, "書類が無いのに一致が出ている: %d" % _none.ok_count

# わざと食い違わせた重説・契約書を入れると検出する
_exp = _mkpdf(["重要事項説明書", "地番 123番4", "用途地域 第一種住居地域",
               "建ぺい率 60%", "容積率 600%", "売主 検証太郎"])
_con = _mkpdf(["不動産売買契約書", "地番 123番9", "売主 検証太郎",
               "契約不適合責任の通知期間 引渡しから1年", "違約金 売買代金の30%"])
_cc = crosscheck_service.run(_base, _exp, _con, seller_is_pro=True)
_ng = {r.item for r in _cc.results if r.is_ng}
for _must in ("地番", "用途地域", "指定建ぺい率"):
    assert _must in _ng, "検出できていない: %s（検出=%s）" % (_must, _ng)
assert crosscheck_service.build_admin({}).building_coverage == 0.0, "モックが混入している"
assert len(crosscheck_report_service.build(_cc)) > 3000, "報告書Excelが生成できない"
print("[ok] クロスチェック: %d項目 / 🔴%d件検出 / 報告書OK" % (len(_cc.results), _cc.ng_count))

# ---------------------------------------------------------------------------
# 災害3項目（2026-08-23 実装）。**通信せずに**判定ロジックだけを確かめる。
# 実APIを叩くテストにすると、キーが無いPCや外出先で落ちて意味が薄れる。
from services import checkbox_fill, hazard_service
from services import reinfolib_client as _rc


def _stub(features_by_layer):
    def fake(layer, lat, lon, zoom, api_key=""):
        return features_by_layer.get(layer)
    return fake


def _poly(coords, props):
    return {"type": "Feature", "properties": props,
            "geometry": {"type": "Polygon", "coordinates": [coords]}}


_square = [[135.0, 34.0], [135.1, 34.0], [135.1, 34.1], [135.0, 34.1], [135.0, 34.0]]
_far = [[130.0, 30.0], [130.1, 30.0], [130.1, 30.1], [130.0, 30.1], [130.0, 30.0]]

_orig_fetch, _orig_key = _rc.fetch_features, _rc.get_api_key
_rc.get_api_key = lambda: "dummy"
try:
    # 区域の中 … 浸水深ランク2 = 0.5m以上3.0m未満（公式コードリストで確認済み）
    _rc.fetch_features = _stub({
        "XKT026": [_poly(_square, {"A31a_205": 2, "A31a_202": "淀川"})],
        "XKT029": [_poly(_square, {"A33_001": 1, "A33_002": 2,
                                   "A33_003": "28", "A33_005": "住吉山手Ⅰ"})],
        "XKT028": [_poly(_square, {"A40_003": "1.0～2.0m"})],
        "XKT027": [_poly(_square, {"A49_003": "0.3～1.0m"})],
    })
    _h = hazard_service.get_hazard_detail(34.05, 135.05)
    assert "0.5m以上3.0m未満" in _h["洪水浸水想定"]["値"], _h
    assert "淀川" in _h["洪水浸水想定"]["値"], _h
    assert _h["土砂災害"]["値"].startswith("土砂災害特別警戒区域"), _h
    assert "急傾斜地の崩壊" in _h["土砂災害"]["値"], _h
    assert "兵庫県" in _h["土砂災害"]["注意"], "兵庫県の利用制限が出ていない"
    assert "1.0～2.0m" in _h["津波"]["値"], _h
    assert _h["高潮浸水想定"]["値"].startswith("高潮浸水想定区域内"), _h

    # 区域の外 … 地物はあるが地点は含まれない → 「区域外」と言い切ってよい
    _rc.fetch_features = _stub({
        "XKT026": [_poly(_far, {"A31a_205": 2})],
        "XKT029": [_poly(_far, {"A33_001": 1, "A33_002": 1, "A33_003": "27"})],
        "XKT028": [_poly(_far, {"A40_003": "0.3～1.0m"})],
        "XKT027": [_poly(_far, {"A49_003": "0.3～1.0m"})],
    })
    _o = hazard_service.get_hazard(34.05, 135.05)
    assert _o["洪水浸水想定"].endswith("区域外（想定最大規模）"), _o
    assert _o["土砂災害"] == "土砂災害警戒区域外", _o
    assert _o["津波"] == "津波浸水想定区域外", _o
    assert _o["高潮浸水想定"] == "高潮浸水想定区域外", _o

    # 取得できない／地物が1件も無い → 空欄（＝判定不可。「区域外」と言わない）
    _rc.fetch_features = _stub({"XKT026": None, "XKT029": [], "XKT028": None,
                                "XKT027": None})
    _u = hazard_service.get_hazard(34.05, 135.05)
    assert _u == {"洪水浸水想定": "", "土砂災害": "", "津波": "",
                  "高潮浸水想定": ""}, _u
finally:
    _rc.fetch_features, _rc.get_api_key = _orig_fetch, _orig_key

# チェック欄の検出（□ の右に「内」「外」がある実書式の並びを再現）
_rows = {
    276: [(10, "土砂災害警戒区域"), (32, "□"), (34, "外"), (36, "・"), (37, "□"), (39, "内")],
    277: [(10, "土砂災害特別警戒区域"), (32, "□"), (34, "外"), (37, "□"), (39, "内")],
    280: [(10, "津波災害警戒区域"), (32, "□"), (34, "外"), (37, "□"), (39, "内")],
}
_cb = checkbox_fill.detect_hazard(_rows)
assert _cb == {"土砂災害警戒区域_内": "AK276", "土砂災害特別警戒区域_内": "AK277"}, _cb
# 特別警戒区域は警戒区域の中にある（土砂災害防止法9条）ので両方に■
assert checkbox_fill.marks({"土砂災害": "土砂災害特別警戒区域（急傾斜地の崩壊）"}, _cb) == {
    "AK277": "■", "AK276": "■"}
assert checkbox_fill.marks({"土砂災害": "土砂災害警戒区域（土石流）"}, _cb) == {"AK276": "■"}
# 区域外・未取得のときは書式の□をそのまま残す
assert checkbox_fill.marks({"土砂災害": "土砂災害警戒区域外"}, _cb) == {}
assert checkbox_fill.marks({"土砂災害": ""}, _cb) == {}
print("[ok] 災害3項目: 区域内/区域外/判定不可の3状態 ＋ チェック欄の自動■")

# ---------------------------------------------------------------------------
# 抵当権を土地/建物で分ける（2026-08-23）。
# 重説の権利部(乙区)は土地と建物で行が違うのに、1項目にまとめていたため
# **土地の抵当権が建物の欄に入っていた**。その再発防止。
from services import registry_service as _rs

_split = _rs._from_shared({"物件種別": "土地建物",
                           "土地": {"抵当権": "抵当権: A銀行"},
                           "建物": {"抵当権": "根抵当権: B信金"},
                           "抵当権": "有"})
assert _split["土地抵当権"] == "抵当権: A銀行", _split
assert _split["建物抵当権"] == "根抵当権: B信金", _split
# 謄本が土地だけなら、トップレベルの値は土地のものと分かる
_land_only = _rs._from_shared({"物件種別": "土地", "土地": {}, "建物": {},
                               "抵当権": "抵当権: C銀行"})
assert _land_only["土地抵当権"] == "抵当権: C銀行" and _land_only["建物抵当権"] == ""
# 土地建物の謄本で側が分からないときは**どちらにも入れない**（推測で欄を埋めない）
_ambiguous = _rs._from_shared({"物件種別": "土地建物", "土地": {}, "建物": {},
                               "抵当権": "有: D銀行"})
assert _ambiguous["土地抵当権"] == "" and _ambiguous["建物抵当権"] == "", _ambiguous

# 実書式の並び（土地の乙区 → 建物の乙区）を再現して検出を確かめる
_rows_rights = {
    90:  [(2, "土　　地")],
    98:  [(6, "権利部(乙区)"), (9, "所有権以外の権利に関する事項"), (22, "□"), (24, "地上権")],
    99:  [(22, "□"), (24, "抵当権")],
    100: [(22, "□"), (24, "根抵当権")],
    104: [(2, "建　　　　物")],
    112: [(6, "権利部(乙区)"), (9, "所有権以外の権利に関する事項"), (22, "□"), (24, "抵当権")],
    113: [(22, "□"), (24, "根抵当権")],
}
_rb = checkbox_fill.detect_rights(_rows_rights, ["V98", "AH98", "V112", "AH112"])
assert _rb["土地_抵当権"] == "V99" and _rb["土地_根抵当権"] == "V100", _rb
assert _rb["建物_抵当権"] == "V112" and _rb["建物_根抵当権"] == "V113", _rb
assert _rb["土地_詳細"] == "AH98" and _rb["建物_詳細"] == "AH112", _rb

# 土地の抵当権は土地の欄へ。建物が根抵当権なら「抵当権」には■を付けない（別の権利）
_w = checkbox_fill.marks({"土地抵当権": "抵当権: A銀行",
                          "建物抵当権": "根抵当権: B信金"}, _rb)
assert _w == {"V99": "■", "AH98": "抵当権: A銀行",
              "V113": "■", "AH112": "根抵当権: B信金"}, _w
# 値が無い側には何も書かない（書式の□をそのまま残す）
assert checkbox_fill.marks({"土地抵当権": "", "建物抵当権": ""}, _rb) == {}
print("[ok] 抵当権: 土地/建物を別の欄へ ＋ 抵当権と根抵当権の区別")

# ---------------------------------------------------------------------------
# 自社（宅建業者・宅建士）情報 — 書式1枚目の13欄（2026-08-23）
import sys as _sys, os as _os
_sys.path.insert(0, _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))
import company_profile as _cp
from services import agent_fields as _af

# **コード側に会社情報の値を書かない**（このリポジトリは public）。
# 実データは config/company_profile.json（gitignore・Dropboxで運ぶ）にあるので、
# 中身そのものはテストしない（PCによって在る／無いが変わるため）。
assert all(v == "" for v in _cp.DEFAULTS.values()), \
    "コードに会社情報の既定値が書かれている（public リポジトリなので置かない）"
_prof = _cp.load()
assert set(_prof) == {k for k, _l, _n in _cp.FIELDS}, _prof
# 未設定の環境では必須項目がすべて「要入力」に出る
assert set(_cp.missing({k: "" for k in _prof})) == set(_cp.REQUIRED)
assert _cp.parse_license("大阪府知事(10)27334号") == {
    "免許_知事名": "大阪府知事", "免許_更新回数": "10", "免許_番号": "27334"}
assert _cp.parse_license("よくわからない文字列") == {}

# A欄・B欄の並びを再現（B列より左がA欄）
_rows_agent = {
    12: [(2, "宅地建物取引業者"), (6, "A"), (35, "B")],
    13: [(6, "主たる事務所所在地"), (35, "主たる事務所所在地")],
    14: [(6, "TEL"), (35, "TEL")],
    15: [(6, "商号又は名称"), (35, "商号又は名称")],
    16: [(6, "代表者の氏名"), (35, "代表者の氏名")],
    18: [(6, "免許証番号"), (22, "（"), (26, "）第"), (33, "号")],
    19: [(2, "説明をする宅地建物取引士"), (6, "氏名"), (35, "氏名")],
    20: [(6, "登録番号"), (13, "（"), (21, "）第"), (33, "号")],
    21: [(6, "業務に従事する事務所名")],
    22: [(6, "事務所所在地")],
    24: [(2, "取引態様")],
}
_inputs_agent = [("M13", ""), ("AP13", ""), ("M14", ""), ("M15", ""), ("M16", ""),
                 ("M18", ""), ("W18", ""), ("AA18", ""), ("M19", ""),
                 ("N20", ""), ("W20", ""), ("M21", ""), ("M22", "")]
_cells = _af.detect(_rows_agent, _inputs_agent)   # 媒介（既定の立場）
assert _cells["商号"] == "M15" and _cells["代表者"] == "M16", _cells
assert _cells["免許_知事名"] == "M18" and _cells["免許_更新回数"] == "W18" \
    and _cells["免許_番号"] == "AA18", _cells
assert _cells["宅建士_氏名"] == "M19", _cells
assert _cells["宅建士_登録先"] == "N20" and _cells["宅建士_登録番号"] == "W20", _cells
# B欄（AP13）は自社では埋めない
assert "AP13" not in _cells.values()
_vals = _af.values({"商号": "テスト商事", "免許_番号": "12345", "代表者": ""}, _cells)
assert _vals["M15"] == "テスト商事" and _vals["AA18"] == "12345"
assert "M16" not in _vals, "空の代表者を書き込もうとしている"
# 宅建業者売主版は **A＝売主 / B・C＝媒介**。媒介なのにA欄へ入れると
# 「媒介なのに売主として署名した書面」になるので、立場ごとに欄を分ける
_rows_seller = {
    12: [(2, "A")],
    13: [(2, "取引態様"), (13, "売主（宅地建物取引業者）"), (38, "宅地建物取引士")],
    14: [(2, "免許証番号"), (23, "（"), (26, "）第"), (36, "号"), (38, "登録番号"),
         (45, "（"), (50, "）第")],
    15: [(2, "主たる事務所所在地"), (38, "氏名")],
    17: [(2, "商号又は名称")],
    18: [(2, "代表者の氏名"), (38, "TEL")],
    25: [(2, "宅地建物取引業者"), (6, "B"), (35, "C")],
    26: [(6, "主たる事務所所在地"), (35, "主たる事務所所在地")],
    28: [(6, "商号又は名称"), (35, "商号又は名称")],
}
_inputs_seller = [("M14", ""), ("X14", ""), ("AB14", ""), ("AT14", ""), ("BA14", ""),
                  ("M15", ""), ("AS15", ""), ("M17", ""), ("M18", ""),
                  ("M26", ""), ("M28", ""), ("AP26", ""), ("AP28", "")]
_both = _af.detect_all(_rows_seller, _inputs_seller)
assert _both["売主"]["商号"] == "M17", _both["売主"]
assert _both["売主"]["免許_知事名"] == "M14", _both["売主"]
assert _both["売主"]["宅建士_氏名"] == "AS15", _both["売主"]
# 媒介は下の B ブロック（A＝売主 を拾わない）
assert _both["媒介"]["商号"] == "M28", _both["媒介"]
assert _both["媒介"]["所在地"] == "M26", _both["媒介"]
# 一般売主版（売主が業者ではない）には売主ブロックが無い
assert _af.detect_all(_rows_agent, _inputs_agent)["売主"] == {}
# 宅建士は複数登録して案件ごとに選ぶ。選んだ1人が 宅建士_* に写ること、
# **登録番号が空なら書式に書き込まない**こと（他人の番号を使わない）を見る
_saved_list, _saved_sel = _cp.takken_list(), _cp.takken_selected()
try:
    _cp.save_takken([
        {"氏名": "テスト　一郎", "登録先": "大阪", "登録番号": "111111"},
        {"氏名": "テスト　二郎", "登録先": "大阪", "登録番号": ""},
    ], selected="テスト　二郎")
    _p2 = _cp.load()
    assert _p2["宅建士_氏名"] == "テスト　二郎", _p2
    assert _p2["宅建士_登録番号"] == "", _p2
    assert "宅建士_登録番号" in _cp.missing(_p2)
    # 登録番号が空の人を選んでいるときは、その欄に何も書かない
    _v2 = _af.values(_p2, _cells)
    assert _v2.get(_cells["宅建士_氏名"]) == "テスト　二郎"
    assert _cells["宅建士_登録番号"] not in _v2, "空の登録番号を書き込もうとしている"
    # 選び直すとその人の番号が入る
    _cp.save_takken(_cp.takken_list(), selected="テスト　一郎")
    assert _cp.load()["宅建士_登録番号"] == "111111"
finally:
    if _saved_list:
        _cp.save_takken(_saved_list, selected=_saved_sel)
print("[ok] 自社情報: 12欄を自動、立場（媒介／売主）で分ける、宅建士は選択制、空は書かない")

# ---------------------------------------------------------------------------
# 区域指定と法令チェック（2026-08-23）
from services import legal_area_service as _las

assert _las.applies("区域内（本山北）") is True
assert _las.applies("区域外") is False
assert _las.applies("") is False
_law_rows = {
    247: [(8, "□"), (10, "海岸法"), (22, "□"), (24, "津波防災地域づくりに関する法律"),
          (36, "□"), (38, "砂防法"), (50, "□"), (52, "地すべり等防止法")],
    248: [(8, "□"), (10, "急傾斜地法"), (22, "□"), (24, "森林法")],
    244: [(50, "□"), (52, "自然公園法")],
}
_lb = checkbox_fill.detect_laws(_law_rows)
assert _lb == {"法令_地すべり等防止法": "AX247", "法令_急傾斜地法": "H248",
               "法令_自然公園法": "AX244"}, _lb
# 区域内のものだけ■。区域外・判定不可は書式の□を残す
assert checkbox_fill.marks(
    {"急傾斜地崩壊危険区域": "区域内（本山北）", "地すべり防止区域": "区域外",
     "自然公園": ""}, _lb) == {"H248": "■"}
print("[ok] 区域指定: 区域内の法令だけ■（立地適正化・地区計画は自動では触らない）")

# ---------------------------------------------------------------------------
# 追加資料（任意アップロード）
from services import document_intake as _di

assert len(_di.DOCS) == 5
assert "管理費月額" in _di.EXTRA_FIELDS and "確認済証番号" in _di.EXTRA_FIELDS
# 何も上げなければ何も足さない／エラー印は取り込まない
assert _di.flatten({}) == {}
assert _di.flatten({"kanri": {"_error": "読めない", "管理費月額": ""}}) == {}
assert _di.flatten({"kanri": {"管理費月額": " 12,300円 "}}) == {"管理費月額": "12,300円"}
assert _di.parse("kanri", None) == {}
print("[ok] 追加資料: 5種類・読めない項目は取り込まない")

# ---------------------------------------------------------------------------
# 書式への割り当て（2026-08-27 の紙面確認で見つけた事故の再発防止）
from services import field_map as _fm
from services import intake_fill as _if

# ★□ のセルと、すでに文字が入っているセルには書かない。
#   □に書くとチェック欄が消え、見出しに書くと見出しが値に置き換わる。
_inputs = [
    {"cell": "L38", "label": "登記名義人と", "section": "", "checkbox": True},
    {"cell": "AX62", "label": "の表示／地　積", "section": "土地の表示", "has_text": True},
    {"cell": "V64", "label": "地　番／所　在", "section": "土地の表示"},
    {"cell": "AD64", "label": "地　目／地　番", "section": "土地の表示"},
    {"cell": "AO64", "label": "地　積／地　目", "section": "土地の表示"},
]
assert _fm.resolve(_inputs).get("地積") == "AO64", _fm.resolve(_inputs)
assert "所有者" not in _fm.resolve(_inputs)

# ★登記の所在（謄本）と住居表示は別の欄。住居表示の欄が無い書式には**書かない**
_addr = [
    {"cell": "H42", "label": "所　　在／①", "section": "（１）土地"},
    {"cell": "O77", "label": "□／住居表示", "section": "（2）建物"},
]
_m = _fm.resolve(_addr)
assert _m.get("登記所在") == "H42" and _m.get("所在地") == "O77", _m
_land_only = [{"cell": "F41", "label": "所　　在／①", "section": "（１）土地"}]
_m2 = _fm.resolve(_land_only)
assert _m2.get("登記所在") == "F41" and "所在地" not in _m2, _m2

# ★同じ登記の所在を「土地」と「建物」の2箇所へ書く
_rep = {}
_fm.resolve([
    {"cell": "H42", "label": "所　　在／①", "section": "（１）土地"},
    {"cell": "H44", "label": "所　　在／②", "section": "（１）土地"},     # 2筆目には書かない
    {"cell": "O76", "label": "□／所在", "section": "（2）建物"},
    {"cell": "O288", "label": "高潮／水害ハザードマップにおける宅地建物の所在",
     "section": "水害ハザードマップ"},                                    # ハザード欄にも書かない
], extra=_rep)
assert _rep == {"登記所在": ["O76"]}, _rep

# ★追加資料の値は、単位（円）を落として枠へ入れる
assert _if.values({"管理費月額": "12,300円", "管理形態": "全部委託", "管理組合名": ""},
                  {"管理費月額": "J457", "管理形態": "L467", "管理組合名": "AG467"}) == {
    "J457": "12,300", "L467": "全部委託"}

# ★実物の書式（レジストリがある環境だけ）で、当たり所を1本確かめる
from services import format_catalog as _fc
if _fc.available():
    _e = [f for f in _fc.load()["formats"]
          if f["name"].startswith("【ファイル４】") and "一般売主" in f["path"]]
    if _e:
        _mp = _e[0]["mapping"]
        assert _mp["登記所在"] == "H42" and _mp["所在地"] == "O77", _mp
        assert _mp["地積"] == "AW42" and _mp["床面積"] == "BC82", _mp
        assert "所有者" not in _mp, _mp
        assert _e[0].get("repeat") == {"登記所在": ["O76"]}, _e[0].get("repeat")
    _k = [f for f in _fc.load()["formats"]
          if f["name"].startswith("【ファイル５】") and "一般売主" in f["path"]]
    if _k:
        # 区分所有は「一棟」ではなく「専有部分」の構造・床面積を採る
        assert _k[0]["mapping"]["構造"] == "M50", _k[0]["mapping"]
        assert _k[0]["mapping"]["床面積"] == "AE52", _k[0]["mapping"]
        assert _k[0]["intake_cells"].get("管理費月額") == "J457", _k[0]["intake_cells"]
print("[ok] 書式への割り当て: □と見出しには書かない／登記所在と住居表示を分ける／"
      "土地と建物の2箇所／追加資料の欄")

# ---------------------------------------------------------------------------
# 謄本の読み取り（区分建物で家屋番号・種類が落ちる件の受け皿）
import registry_parser as _rp

_TOUKI = ("家屋番号　中野町一丁目1番20の1の307\n"
          "種　類　居宅\n構　造　鉄筋コンクリート造1階建\n床　面　積　72.59㎡\n")
_r = {"建物": {"家屋番号": "", "種類": "", "構造": "", "床面積": ""}}
_rp._fill_missing_building(_r, _TOUKI)
assert _r["建物"]["家屋番号"] == "中野町一丁目1番20の1の307", _r["建物"]
assert _r["建物"]["種類"] == "居宅", _r["建物"]
# AI が読めているものは上書きしない
_r2 = {"建物": {"家屋番号": "AIが読んだ値", "種類": "共同住宅", "構造": "", "床面積": ""}}
_rp._fill_missing_building(_r2, _TOUKI)
assert _r2["建物"]["家屋番号"] == "AIが読んだ値" and _r2["建物"]["種類"] == "共同住宅"
# 罫線の直後に値がある形でも拾う（車庫の謄本で実際にあった形）
assert _rp._cell("┃中野町一丁目1番20の1の61 │ ② 構 造 │") == "中野町一丁目1番20の1の61"
print("[ok] 謄本: 区分建物の家屋番号・種類を取りこぼしたら埋め戻す（既存値は残す）")

print("smoke test: all assertions passed")
