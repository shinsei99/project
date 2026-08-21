#!/usr/bin/env python3
"""公式書式への流し込みを実データで検証する。

確かめること:
  1. PropertyData の値が起点シート（重要事項説明書）の正しいセルに入るか
  2. **他の同梱書式（売買契約書など）の転記数式が壊れていないか**
     ← xlsx_patcher は無損失のはずだが、ここが壊れると自動入力の意味が消える
  3. 図形・画像・他シートが保持されているか（zip エントリ数で確認）

使い方:
    .venv/bin/python test_official_fill.py [書式のパス]
"""

import json
import os
import sys
import zipfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook  # noqa: E402

from services import field_map, official_format_service as ofs  # noqa: E402

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(HERE, "reports", "test")

# 検証用のダミー物件（実在の物件は使わない）
SAMPLE = {
    "所在地": "大阪市都島区東野田町二丁目",
    "地番": "123番4",
    "地目": "宅地",
    "地積": "165.28",
    "家屋番号": "123番4",
    "種類": "居宅",
    "構造": "木造スレート葺2階建",
    "床面積": "98.55",
    "所有者": "検証用 太郎",
    "用途地域": "第一種住居地域",
    "建ぺい率": "60",
    "容積率": "200",
    "高度地区": "第2種高度地区",
}


def main():
    reg_path = os.path.join(HERE, "data", "format_registry.json")
    if len(sys.argv) > 1:
        src = sys.argv[1]
    else:
        with open(reg_path, encoding="utf-8") as fh:
            reg = json.load(fh)
        cand = [f for f in reg["formats"] if "土地建物公簿用" in f["name"] and "excel版" in f["category"]]
        if not cand:
            print("対象書式が見つからない"); return 1
        src = os.path.join(reg["root"], cand[0]["path"])

    print("書式:", os.path.basename(src))
    scanned = ofs.scan(src)
    mapping = field_map.resolve(scanned["inputs"])
    print("起点シート:", scanned["driver"])
    print(field_map.coverage(mapping))

    cells = {cell: SAMPLE[field] for field, cell in mapping.items() if field in SAMPLE}
    print("書き込むセル:", len(cells), "→", ", ".join("%s=%s" % (c, v[:10]) for c, v in list(cells.items())[:6]))

    os.makedirs(OUT_DIR, exist_ok=True)
    dst = os.path.join(OUT_DIR, "filled_" + os.path.basename(src))
    ofs.fill(src, dst, scanned["driver"], cells)

    # --- 検証 ---
    ok = True
    with zipfile.ZipFile(src) as a, zipfile.ZipFile(dst) as b:
        na, nb = len(a.namelist()), len(b.namelist())
        print("zipエントリ: 元 %d → 出力 %d %s" % (na, nb, "OK" if na == nb else "★欠落"))
        ok &= (na == nb)

    wb = load_workbook(dst)
    ws = wb[scanned["driver"]]
    bad = []
    for field, cell in mapping.items():
        if field not in SAMPLE:
            continue
        got = ws[cell].value
        if str(got or "").strip() != SAMPLE[field]:
            bad.append("%s(%s): 期待=%s 実際=%s" % (field, cell, SAMPLE[field], got))
    print("値の一致: %d/%d %s" % (len(cells) - len(bad), len(cells), "OK" if not bad else "★不一致"))
    for b_ in bad:
        print("   ", b_)
    ok &= not bad

    # 転記数式が生きているか
    formulas = 0
    for sh in wb.sheetnames:
        if sh == scanned["driver"]:
            continue
        for row in wb[sh].iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("=") and scanned["driver"] in c.value:
                    formulas += 1
    print("他書式の転記数式: %d 本 %s" % (formulas, "OK" if formulas else "★消えている"))
    ok &= bool(formulas)

    print("\n出力:", dst)
    print("結果:", "合格" if ok else "不合格")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
