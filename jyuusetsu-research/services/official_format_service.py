"""公式書式（全宅連・ハトサポ）を読み、入力欄を自動抽出して流し込むサービス。

肝は「excel版自動入力書式」の作りにある。1ファイルに重要事項説明書・売買契約書・
実測清算確認書・引渡し書・媒介手数料支払承諾書が同梱されており、契約書側のセルは
**重説シートを参照する数式**になっている。したがって重説シートの共通入力欄だけ
埋めれば、残りの書式は Excel 側が自動で埋める。

  実測（【ファイル１】土地実測清算用・2026年4月版）:
    他書式から参照されている重説セル = 90種類 → それだけで5書類が完成する

入力欄は色で区別されている（同梱の「入力に際しての注意事項」シートに明記）:

  FFFFFF99 (黄)  共通入力欄。ここに入れると他書式の該当欄へ反映される
  FFCCFFCC (緑)  選択欄（ドロップダウンから選ぶ）
  FFFFCC99 (橙)  その書式だけに反映される入力欄

セル座標を人が1枚ずつ拾うのは200書式では現実的でないため、
**数式の参照先と色からプログラムで入力欄を割り出す**方針を採る。

書き込みは openpyxl ではなく services.xlsx_patcher を使う（openpyxl の再保存は
図形・画像・書式を落とすため。公式書式は罫線と図形が命）。
"""

from __future__ import annotations

import collections
import os
import re
from typing import Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from services import checkbox_fill, xlsx_patcher

# 入力欄の色（注意事項シートの凡例と実測が一致）
COLOR_SHARED = "FFFFFF99"   # 共通入力欄（他書式へ反映）
COLOR_SELECT = "FFCCFFCC"   # 選択欄
COLOR_LOCAL = "FFFFCC99"    # その書式のみ
INPUT_COLORS = (COLOR_SHARED, COLOR_SELECT, COLOR_LOCAL)

_CELL_RE = re.compile(r"([A-Z]{1,3})(\d+)")


def _split(cell: str):
    m = _CELL_RE.match(cell)
    return column_index_from_string(m.group(1)), int(m.group(2))


def _fill_rgb(cell) -> Optional[str]:
    f = cell.fill
    if f is None or f.fill_type != "solid":
        return None
    rgb = getattr(f.start_color, "rgb", None)
    return rgb if isinstance(rgb, str) else None


def _merged_index(ws) -> Dict[tuple, tuple]:
    """結合セル内の全座標 → 左上（値を持つ）座標 の対応表。

    **見出しが結合セルだと、左上以外は値が None になる。** これを解決しないと
    「地　積」のような列見出しが拾えず、対応表に穴があく（2026-08-21 実測で判明）。
    """
    idx = {}
    for rng in ws.merged_cells.ranges:
        anchor = (rng.min_row, rng.min_col)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                idx[(r, c)] = anchor
    return idx


def _text_at(ws, r: int, c: int, merged=None) -> str:
    if merged is not None:
        r, c = merged.get((r, c), (r, c))
    v = ws.cell(row=r, column=c).value
    if not isinstance(v, str):
        return ""
    t = v.strip().replace("\n", "")
    return "" if t.startswith("=") else t


def _label_for(ws, cell: str, max_left: int = 16, max_up: int = 14, merged=None) -> str:
    """入力欄の見出しを推定する。

    帳票には2つの型があるので両方を見る。
      (a) 「見出し ｜ 入力欄」  … 同じ行の左に見出しがある
      (b) 表形式               … 列見出しが数行〜十数行上にある（地番・地目・地積など）
    左と上の両方を拾い、「上見出し／左見出し」の形で返す。片方しか無ければそれだけ。
    左だけを見ていたときは 25項目中11項目しか当たらなかった（2026-08-21 実測）。
    """
    col, row = _split(cell)

    def scan_left(r: int) -> str:
        for c in range(col - 1, max(0, col - max_left), -1):
            t = _text_at(ws, r, c, merged)
            if t:
                return t
        return ""

    # 同じ行の左 → 無ければ1つ上の行の左（「指定建蔽率」の見出しが1行上にある型）
    left = scan_left(row) or scan_left(row - 1)

    up = ""
    for r in range(row - 1, max(0, row - max_up), -1):
        up = _text_at(ws, r, col, merged)
        if up:
            break

    parts = [p for p in (up[:20], left[:20]) if p]
    return "／".join(parts)


def driver_sheet(wb) -> Optional[str]:
    """他シートの数式が最も多く参照しているシート＝入力の起点を返す。

    自動入力書式では必ず重要事項説明書がこれになる（注意事項シートにも
    「入力は必ず重要事項説明書から始めてください」と書かれている）。
    """
    ref = collections.Counter()
    for sh in wb.sheetnames:
        for row in wb[sh].iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                for m in re.finditer(r"'([^']+)'!|([A-Za-z0-9_（）()一-龥ぁ-んァ-ヶー]+)!", v):
                    name = m.group(1) or m.group(2)
                    if name and name != sh and name in wb.sheetnames:
                        ref[name] += 1
    return ref.most_common(1)[0][0] if ref else None


def referenced_cells(wb, target: str) -> Dict[str, int]:
    """target シートのうち、他シートの数式から参照されているセルと参照回数。"""
    pat = re.compile(r"'?" + re.escape(target) + r"'?!\$?([A-Z]{1,3})\$?(\d+)")
    out = collections.Counter()
    for sh in wb.sheetnames:
        if sh == target:
            continue
        for row in wb[sh].iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("="):
                    for m in pat.finditer(v):
                        out[m.group(1) + m.group(2)] += 1
    return dict(out)


def scan(path: str) -> dict:
    """書式ファイル1本を解析し、入力欄の一覧を返す。

    戻り値:
      {
        "file": ..., "driver": "(1)重要事項説明書...", "sheets": [...],
        "inputs": [{"cell","color","label","fanout"}...],  # fanout=他書式への反映数
        "checkboxes": {"土砂災害警戒区域_内": "AK276", ...}
      }
    """
    wb = load_workbook(path)
    drv = driver_sheet(wb) or wb.sheetnames[0]
    ws = wb[drv]
    merged = _merged_index(ws)
    refs = referenced_cells(wb, drv)

    inputs: List[dict] = []
    seen = set()
    # 行ごとの文字列（チェックボックス欄の検出に使う）。
    # 起点シートは一度しか走査しないので、ここで一緒に集めておく
    row_strings: Dict[int, List[tuple]] = {}
    # まず「他書式へ効くセル」を優先して拾う
    for cell, n in refs.items():
        inputs.append({
            "cell": cell,
            "color": _fill_rgb(ws[cell]),
            "label": _label_for(ws, cell, merged=merged),
            "fanout": n,
        })
        seen.add(cell)
    # 起点シート内だけで完結する入力欄も拾う（重説そのものを出すときに要る）
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if isinstance(v, str):
                t = v.strip().replace("\n", "")
                if t and not t.startswith("="):
                    row_strings.setdefault(c.row, []).append((c.column, t))
            if c.coordinate in seen:
                continue
            rgb = _fill_rgb(c)
            if rgb in INPUT_COLORS:
                inputs.append({
                    "cell": c.coordinate,
                    "color": rgb,
                    "label": _label_for(ws, c.coordinate, merged=merged),
                    "fanout": 0,
                })
    inputs.sort(key=lambda d: (_split(d["cell"])[1], _split(d["cell"])[0]))
    return {
        "file": path,
        "name": os.path.basename(path),
        "driver": drv,
        "sheets": list(wb.sheetnames),
        "inputs": inputs,
        # 「□」のチェック欄（災害欄と権利部(乙区)の抵当権）。
        # テキストではなく■を入れる欄なので mapping と分けて持つ（checkbox_fill 参照）
        "checkboxes": _checkboxes(row_strings, [d["cell"] for d in inputs], drv),
    }


def _checkboxes(row_strings, input_cells, sheet_name: str = "") -> Dict[str, str]:
    """チェック欄（災害＋権利部）をまとめて検出する。"""
    out = checkbox_fill.detect_hazard(row_strings)
    out.update(checkbox_fill.detect_rights(row_strings, input_cells, sheet_name))
    return out


def fill(src_path: str, dst_path: str, driver: str, cells: Dict[str, str]) -> str:
    """起点シートの指定セルへ値を書き込む（無損失）。

    他書式は数式で追随するため、ここで書くのは起点シートだけでよい。
    空文字は書かない（書式の既定値・チェックボックスを保持するため）。
    """
    payload = {k: v for k, v in cells.items() if str(v or "").strip()}
    return xlsx_patcher.set_cells(src_path, dst_path, driver, payload)
