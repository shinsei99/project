# -*- coding: utf-8 -*-
"""追加資料（管理会社の重要事項調査報告書）の値を、書式の欄へ入れる。

## なぜ別立てなのか

`field_map` は **書式が「入力欄」として色を付けた／数式で参照しているセル**の中から
見出しで選ぶ。ところが区分所有の重説で肝心な

    通常の管理費（月額・滞納額）／計画修繕積立金（月額・滞納額・積立総額）

の欄は、**全宅連の書式では入力欄として色が付いていない**（2026-08-27 実測。
【ファイル５】区分所有建物用（敷地権）で確認）。したがって `field_map` からは
そもそも見えず、「画面には出るが書面には入らない」状態だった。

## どう当てるか — 見出しの右にある空欄をたどる

これらの欄は作りが規則正しい。

    B444:S446「当該住戸の計画修繕積立金等」 ｜ T445「月額」 ｜ W445:AB445（空欄） ｜ AC445「円」

つまり **「まとまりの見出し」と「行内の見出し」で場所を決め、その右にある
最初の空の結合セルへ書く**。行や列がずれても効くので、版が変わっても追随する。

**当てにいく項目は下の表に書いたものだけ**（白紙のセルを手当たり次第に埋めない）。
見つからなければ何もしない＝今までどおり空欄で出る。
"""

from __future__ import annotations

import re
from typing import Dict, List, Optional, Tuple

# (PropertyData のキー, まとまりの見出し, 行内の見出し, 単位)
# 行内の見出しが None のときは、まとまりの見出しのすぐ右を使う。
RULES: List[Tuple[str, str, Optional[str], str]] = [
    ("管理費月額",       r"通常の管理費",              r"^月\s*額$",            "円"),
    ("管理費等滞納額",   r"通常の管理費",              r"当該住戸の滞納額",      "円"),
    ("修繕積立金月額",   r"当該住戸の計画修繕積立金",   r"^月\s*額$",            "円"),
    ("修繕積立金総額",   r"既に積み立て",              None,                    "円"),
    ("管理形態",         r"管理の形態",                r"管理の形態",            ""),
    ("管理組合名",       r"管理の委託先|管理の形態",    r"管理組合の名称",        ""),
    ("管理会社名",       r"管理の委託先",              r"氏名（商号又は名称）",   ""),
]

FIELDS = [r[0] for r in RULES]

_SPACE = re.compile(r"[\s　]")
_UNIT_TAIL = re.compile(r"[円㎡%％]\s*$")


def _norm(text: str) -> str:
    return _SPACE.sub("", text or "")


def _coord(anchor: tuple) -> str:
    from openpyxl.utils import get_column_letter
    return "{}{}".format(get_column_letter(anchor[1]), anchor[0])


def _is_blank(ws, anchor: tuple) -> bool:
    v = ws.cell(row=anchor[0], column=anchor[1]).value
    return v is None or (isinstance(v, str) and not v.strip())


def _blank_anchor_right(ws, merged: Dict[tuple, tuple], row: int, col: int,
                        max_right: int = 40) -> Optional[str]:
    """(row, col) の右にある最初の空欄。**結合セルがあればそちらを優先**する。

    見出しと本来の記入枠のあいだに、幅1の空セルが挟まっていることがある
    （「月額」と `J457:Q457` のあいだの `H457` など）。そこに書くと枠から
    ずれて見えるので、**文字に突き当たるまでの間にある結合セル**を選ぶ。
    """
    first: Optional[tuple] = None
    seen = set()
    for c in range(col + 1, col + 1 + max_right):
        anchor = merged.get((row, c), (row, c))
        if anchor in seen:
            continue
        seen.add(anchor)
        if not _is_blank(ws, anchor):
            break                     # 文字に突き当たったら打ち切る
        if merged.get((row, c)) is not None:
            return _coord(anchor)     # 結合セル＝記入枠
        if first is None:
            first = anchor
    return _coord(first) if first else None


def _cell_before_unit(ws, merged: Dict[tuple, tuple],
                      row_strings: Dict[int, List[tuple]],
                      row: int, col: int, unit: str,
                      rows_down: int = 2, max_right: int = 60) -> Optional[str]:
    """見出しの右にある単位（「円」）を探し、**その1つ左の枠**を返す。

    金額欄は「見出し ｜ 記入枠 ｜ 円」の形で作られているので、単位を目印に
    すると枠を正確に取れる。単位は**次の行に回っていることがある**
    （「当該管理組合に既に積み立てられている…」は見出しが3行分の結合セルで、
    円は2行目にある）ので、数行下まで見る。
    """
    for r in range(row, row + rows_down + 1):
        for c, text in sorted(row_strings.get(r, [])):
            if c <= col or c > col + max_right:
                continue
            if _norm(text) != unit:
                continue
            anchor = merged.get((r, c - 1), (r, c - 1))
            if _is_blank(ws, anchor):
                return _coord(anchor)
            return None
    return None


def detect(ws, merged: Dict[tuple, tuple], row_strings: Dict[int, List[tuple]],
           block_span: int = 6) -> Dict[str, str]:
    """{PropertyData のキー: セル} を返す。見つからない項目は入れない。"""
    if not row_strings:
        return {}
    out: Dict[str, str] = {}
    used = set()
    rows = sorted(row_strings)

    def block_at(row: int) -> str:
        """その行が属するまとまりの見出し（左端 A〜F 列を上へ辿る）。"""
        for r in range(row, max(0, row - block_span) - 1, -1):
            for col, text in row_strings.get(r, []):
                if col <= 6 and _norm(text):
                    return _norm(text)
        return ""

    for field, block_pat, label_pat, unit in RULES:
        block_re = re.compile(block_pat)
        label_re = re.compile(label_pat) if label_pat else None
        for row in rows:
            if not block_re.search(block_at(row)):
                continue
            if label_re is None:
                # 行内の見出しが無い型。まとまりの見出し（左端）の右を使う
                cands = [(col, t) for col, t in row_strings[row] if col <= 6]
            else:
                cands = [(col, t) for col, t in row_strings[row]
                         if label_re.search(_norm(t))]
            for col, _t in cands:
                # 見出しが結合セルなら、その右端の外側から探す
                end = _merged_end(merged, row, col)
                cell = (_cell_before_unit(ws, merged, row_strings, row, end, unit)
                        if unit else None)
                if cell is None:
                    cell = _blank_anchor_right(ws, merged, row, end)
                if cell and cell not in used:
                    out[field] = cell
                    used.add(cell)
                    break
            if field in out:
                break
    return out


def _merged_end(merged: Dict[tuple, tuple], row: int, col: int) -> int:
    """(row, col) が結合セルなら、その右端の列番号を返す（違えば col）。"""
    anchor = merged.get((row, col))
    if anchor is None:
        return col
    end = col
    while merged.get((row, end + 1)) == anchor:
        end += 1
    return end


def values(data: Dict[str, str], cells: Dict[str, str]) -> Dict[str, str]:
    """{セル: 書き込む値}。**単位は書式に印刷済み**なので落とす（"12,300円" → "12,300"）。"""
    out: Dict[str, str] = {}
    for field, cell in (cells or {}).items():
        v = str(data.get(field, "") or "").strip()
        if not v:
            continue
        out[cell] = _UNIT_TAIL.sub("", v).strip()
    return out
