"""公式書式の「□」チェック欄を見つけて、調査結果から■を入れる。

扱うのは2種類。どちらも**テキストを流し込む欄ではない**。

1. 災害欄（土砂災害警戒区域の「□外・□内」）
2. 権利部(乙区)の「□抵当権・□根抵当権」（土地と建物で行が分かれている）

**なぜ要るのか（2026-08-23 実測で判明した不具合）**

`field_map` は災害3項目（洪水浸水想定・土砂災害・津波）を書式のセルへ割り当てて
いたが、割り当て先は **中身が `□` のチェックボックス**で、しかも3項目とも
**「外」側の□**を指していた。25本すべて同じ構造であることを実測で確認した。
`hazard_service` がスタブ（常に空文字）だった間は「空文字は書かない」規則に
救われて表面化していなかったが、災害の値が入るようになった瞬間、
**「外」の□が長い説明文に置き換わる**ところだった。

そこで災害3項目はテキストとして流し込まず、この仕組みでチェックだけを入れる。

**自動で触る欄／触らない欄（2026-08-23 オーナー判断）**

- **土砂災害警戒区域・特別警戒区域**: 該当したときだけ **「内」を■にする**。
  「外」は人が押す。土砂災害警戒区域データは兵庫県が
  「重要事項の説明等の根拠としないで下さい」としており、**区域外だと機械に
  言い切らせない**ため（該当時は画面に警告を出す）
- **津波災害警戒区域**: 自動化しない。取得できるのは **津波浸水想定**
  （津波防災地域づくり法8条）で、書式が問うている **津波災害警戒区域**（同53条）は
  知事が別に指定するもの。**別の制度なので機械では判定できない**
- **水害ハザードマップ欄**: 自動化しない。あの□は「該当する図面における
  当該宅地建物の所在（＝地図を添付して示す）」のチェックであって、
  浸水想定区域の内外を書く欄ではない

**特別警戒区域に当たるときは警戒区域にも■を入れる。**
土砂災害防止法9条により、特別警戒区域は警戒区域の中に指定されるため。
"""

import re
from typing import Dict, List, Tuple

from openpyxl.utils import column_index_from_string, get_column_letter

_CELL = re.compile(r"([A-Z]{1,3})(\d+)$")

MARK = "■"

KEY_ALERT = "土砂災害警戒区域_内"
KEY_SPECIAL = "土砂災害特別警戒区域_内"


def detect_hazard(row_strings: Dict[int, List[Tuple[int, str]]]) -> Dict[str, str]:
    """行ごとの文字列から、土砂災害欄の「内」チェックボックスのセルを見つける。

    `row_strings` は {行番号: [(列番号, 文字列), ...]}（数式は除いたもの）。

    書式の作り（実測・25本共通）:
        ［見出し 土砂災害警戒区域］ □ 外 ・ □ 内 → …
    見出しは行の左側にあり、□ の右2〜3セル以内に「外」「内」の字がある。
    **座標を書式ごとに手で持たない**（200本では持ちきれないし、改訂でずれる）。
    """
    out: Dict[str, str] = {}
    for row, items in row_strings.items():
        items = sorted(items)
        text = "".join(t for _, t in items)
        if "土砂災害特別警戒区域" in text:
            key = KEY_SPECIAL
        elif "土砂災害警戒区域" in text:
            key = KEY_ALERT
        else:
            continue
        for i, (col, t) in enumerate(items):
            if t != "□":
                continue
            for _, t2 in items[i + 1:i + 4]:
                if t2 in ("内", "外"):
                    if t2 == "内":
                        out[key] = "{}{}".format(get_column_letter(col), row)
                    break
    return out


def _hazard_marks(data: Dict[str, str], checkboxes: Dict[str, str]) -> Dict[str, str]:
    """PropertyData の土砂災害の値から、書き込むセル → ■ を作る。

    「区域外」「空欄（判定不可）」のときは**何も書かない**（書式の□を残す）。
    """
    value = str(data.get("土砂災害", "") or "").strip()
    if not value or "区域外" in value:
        return {}
    if not checkboxes:
        return {}

    out: Dict[str, str] = {}
    if "特別警戒区域" in value:
        if checkboxes.get(KEY_SPECIAL):
            out[checkboxes[KEY_SPECIAL]] = MARK
        # 特別警戒区域は警戒区域の中に指定される（土砂災害防止法9条）
        if checkboxes.get(KEY_ALERT):
            out[checkboxes[KEY_ALERT]] = MARK
    elif "警戒区域" in value:
        if checkboxes.get(KEY_ALERT):
            out[checkboxes[KEY_ALERT]] = MARK
    return out


# ---------------------------------------------------------------------------
# 権利部(乙区) — 抵当権・根抵当権のチェック欄（2026-08-23）
#
# 重説の「登記記録に記録された事項」は **土地と建物で行が分かれている**。
#   土地  権利部(乙区)  □地上権 □抵当権 □根抵当権 □賃借権   ［詳細欄］
#   建物  権利部(乙区)  □抵当権 □根抵当権 □賃借権            ［詳細欄］
# ところが PropertyData は `抵当権` を1つしか持っておらず、書式のマッピングも
# **建物側の詳細欄1つ**を指していた。そのため**土地の抵当権が建物の欄に入る**
# （2026-08-21 に発見、この日まで未修正）。土地/建物で分けて持ち、
# **その側のチェックと詳細欄だけ**に書く。
#
# 抵当権と根抵当権は別の権利なので、文字列を見て分ける。
# どちらの不動産のものか分からないとき（`抵当権` にしか値が無いとき）は**何も書かない**。

_RIGHT_WORDS = ("抵当権", "根抵当権")
_SIDES = ("土地", "建物")


def _infer_side(sheet_name: str) -> str:
    """書式そのものが土地専用／建物専用なら、その側とみなす。

    土地だけの重説（「重要事項説明書(土地の売買・交換用)」）には
    **「土　地」「建　物」の目印が無い**。1つしかないので書く必要が無いため。
    シート名で判断できるときだけ補う（両方入る書式は目印があるので補わない）。
    """
    name = str(sheet_name or "")
    has_land = "土地" in name
    has_building = ("建物" in name) or ("区分所有" in name)
    if has_land and not has_building:
        return "土地"
    if has_building and not has_land:
        return "建物"
    return ""


def detect_rights(row_strings: Dict[int, List[Tuple[int, str]]],
                  input_cells: List[str],
                  sheet_name: str = "") -> Dict[str, str]:
    """権利部(乙区)の抵当権・根抵当権の□と、その行の詳細入力欄を見つける。

    `input_cells` は書式の入力欄（色つきセル）の座標一覧。詳細欄は
    「乙区の見出し行にある入力欄のうち、□ではないもの」で拾う。
    `sheet_name` は土地専用／建物専用の書式を見分けるために使う。
    """
    out: Dict[str, str] = {}
    inputs = set(input_cells or [])
    side = ""
    otsu_row = None
    for row in sorted(row_strings):
        items = sorted(row_strings[row])
        texts = [t for _, t in items]
        flat = "".join(texts).replace(" ", "").replace("　", "")

        # ブロックの切り替わり（行の左端が「土　地」「建　物」）
        head = texts[0].replace(" ", "").replace("　", "") if texts else ""
        if head in _SIDES:
            side = head
        if not side:
            side = _infer_side(sheet_name)
        if not side:
            continue

        if "権利部(乙区)" in flat or "権利部（乙区）" in flat:
            otsu_row = row
            box_cols = {c for c, t in items if t == "□"}
            # 詳細欄＝この行の入力欄のうち□でないもの（値が空なので row_strings に出ない）
            for cell in inputs:
                m = _CELL.match(cell)
                if not m or int(m.group(2)) != row:
                    continue
                if column_index_from_string(m.group(1)) in box_cols:
                    continue
                out["{}_詳細".format(side)] = cell
                break
        if otsu_row is None or row < otsu_row:
            continue
        # 「□ 抵当権」「□ 根抵当権」の並び
        for i, (col, t) in enumerate(items):
            if t != "□":
                continue
            for _, t2 in items[i + 1:i + 3]:
                if t2 in _RIGHT_WORDS:
                    out["{}_{}".format(side, t2)] = "{}{}".format(
                        get_column_letter(col), row)
                    break
    return out


def _rights_marks(data: Dict[str, str], boxes: Dict[str, str]) -> Dict[str, str]:
    """土地抵当権 / 建物抵当権 の値から、チェックと詳細欄の書き込みを作る。"""
    out: Dict[str, str] = {}
    for side in _SIDES:
        text = str(data.get("{}抵当権".format(side), "") or "").strip()
        if not text or text in ("無", "なし", "無し"):
            continue
        if re.search(r"根抵当権", text) and boxes.get("{}_根抵当権".format(side)):
            out[boxes["{}_根抵当権".format(side)]] = MARK
        # 「根抵当権」だけのときに「抵当権」を立てない（別の権利なので）
        if re.search(r"(?<!根)抵当権", text) and boxes.get("{}_抵当権".format(side)):
            out[boxes["{}_抵当権".format(side)]] = MARK
        detail = boxes.get("{}_詳細".format(side))
        if detail:
            out[detail] = re.sub(r"^有\s*[:：]\s*", "", text)
    return out


def marks(data: Dict[str, str], checkboxes: Dict[str, str]) -> Dict[str, str]:
    """書式へ書き込む {セル: 値} をまとめて作る（災害＋権利部＋法令）。"""
    out = _hazard_marks(data, checkboxes)
    out.update(_rights_marks(data, checkboxes))
    out.update(_law_marks(data, checkboxes))
    return out


# ---------------------------------------------------------------------------
# 「都市計画法・建築基準法以外の法令に基づく制限」の法令チェック（2026-08-23）
#
# 書式には64の法律が「□ 法令名」の形で並んでいて、宅建士が1つずつ見ている。
# 全国データで区域が公開されていて、かつ **「区域内＝その法律の制限を受ける」と
# 言い切れるものだけ**を自動でチェックする。
#
# 立地適正化計画（都市再生特別措置法）は入れない。**区域内であることが制限を
# 意味しない**（届出義務は居住誘導区域"外"の行為で生じる）ため、解釈が要る。
# 都市計画道路・地区計画は 64法令ではなく「都市計画法」の欄の話なので触らない。

# PropertyData のキー → 書式の法令名
LAW_BY_FIELD = {
    "急傾斜地崩壊危険区域": "急傾斜地法",
    "地すべり防止区域": "地すべり等防止法",
    "自然公園": "自然公園法",
}
_LAW_NAMES = set(LAW_BY_FIELD.values())


def detect_laws(row_strings: Dict[int, List[Tuple[int, str]]]) -> Dict[str, str]:
    """「□ 法令名」の並びから、対象の法令の□セルを見つける。

    64件すべてを持つとレジストリが太るので、**自動で入れる3件だけ**を拾う。
    """
    out: Dict[str, str] = {}
    for row in sorted(row_strings):
        items = sorted(row_strings[row])
        for i, (col, t) in enumerate(items):
            if t != "□":
                continue
            for _, t2 in items[i + 1:i + 3]:
                name = t2.replace(" ", "").replace("　", "")
                if name in _LAW_NAMES:
                    out.setdefault("法令_" + name,
                                   "{}{}".format(get_column_letter(col), row))
                    break
    return out


def _law_marks(data: Dict[str, str], boxes: Dict[str, str]) -> Dict[str, str]:
    """区域内と判定できた法律にだけ■を入れる。区域外・判定不可は触らない。"""
    out: Dict[str, str] = {}
    for field, law in LAW_BY_FIELD.items():
        value = str(data.get(field, "") or "").strip()
        if not value or "区域外" in value:
            continue
        cell = boxes.get("法令_" + law)
        if cell:
            out[cell] = MARK
    return out
