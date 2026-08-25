from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "社内アプリ一覧"

headers = ["URL", "アプリ名", "進捗", "機能概要", "使い方"]

apps = [
    {
        "url": "http://usernoMac-mini.local:3000",
        "name": "マンション・ビル管理",
        "status": "開発中",
        "features": "・物件・部屋・入居者情報の一元管理\n・AIが書類（PDF等）から入居者情報を自動入力\n・修繕履歴の登録・管理\n・修繕請求書PDFをAIで読み込みExcel保存\n・修繕履歴のExcelエクスポート",
        "usage": "① ブラウザでURLを開く\n② ダッシュボードからマンション/ビルを選択\n③ 物件詳細→部屋を選択\n④「✨ AIで自動入力」でPDF等を投入→解析→反映\n⑤「📄 AI修繕読込」で請求書PDFを登録",
    },
    {
        "url": "http://usernoMac-mini.local:8503",
        "name": "見積書自動生成ツール",
        "status": "完成",
        "features": "・PDF/Excelの請求書・見積書をAIで品目自動抽出\n・上乗せ率を設定して自社見積書・請求書をExcel出力\n・品目の備考欄対応\n・複数ページ対応（長い請求書も自動改ページ）",
        "usage": "① ブラウザでURLを開く\n② PDF/Excelファイルをアップロード\n③ 上乗せ率を入力\n④「解析・生成」ボタンを押す\n⑤ プレビューを確認してExcelをダウンロード",
    },
    {
        "url": "http://usernoMac-mini.local:8504",
        "name": "物件管理 案内文ジェネレーター",
        "status": "完成",
        "features": "・入居者向け案内文・注意喚起チラシをAIで自動生成\n・目的別テンプレート（騒音/ゴミ/駐車/共有部/設備点検/その他）\n・トーン選択（丁寧/通常/強め）\n・A4印刷用Wordファイル（.docx）でダウンロード",
        "usage": "① ブラウザでURLを開く\n② 目的・対象・状況・トーンを選択\n③「生成」ボタンを押す\n④ 本文を確認・編集\n⑤ Wordファイルをダウンロードして印刷",
    },
    {
        "url": "http://usernoMac-mini.local:8505",
        "name": "マイソクコンバーター",
        "status": "開発中",
        "features": "・他社マイソク（チラシ）PDF/画像をAIで解析\n・自社テンプレートExcel（.xls）に自動転記\n・間取り図のドラッグ配置\n・ロゴ・会社情報の自動挿入\n・賃貸・売買テンプレート対応",
        "usage": "① ブラウザでURLを開く\n② 他社マイソクPDF/画像をアップロード\n③ AIが自動解析して自社テンプレートに転記\n④ 間取り図を切り抜いて配置\n⑤ Excelをダウンロード",
    },
    {
        "url": "http://usernoMac-mini.local:8511",
        "name": "間取り図トレーサー",
        "status": "完成",
        "features": "・カラー間取り図をAIが白黒図面に引き直し\n・PDF対応（ページ指定・手動クロップUI）\n・再生成で前回結果ベースに差分修正可能\n・マンション/戸建て/1K・1Rなど種別設定",
        "usage": "① ブラウザでURLを開く\n② 画像またはPDFをアップロード\n③ 間取り図の範囲をドラッグして切り抜く\n④ 種別を選択して「生成」\n⑤ 結果を確認、気になる点を入力して再生成も可能",
    },
    {
        "url": "http://usernoMac-mini.local:8512/project/theta-space/",
        "name": "THETAビューワー",
        "status": "完成",
        "features": "・RICOH THETAのパノラマ写真をブラウザで3D空間化\n・AIが奥行きを自動推定してリアルな立体表示\n・ブラウザ完結（インストール不要・無料）\n・物件内覧・案内用途に対応",
        "usage": "① ブラウザでURLを開く\n② THETAで撮影したエクイレクタングラー画像を選択\n③ AIが自動で奥行き解析（初回は数十秒かかる場合あり）\n④ マウス/タッチで3D空間を自由に見回す",
    },
]

# --- スタイル定義 ---
header_fill = PatternFill("solid", fgColor="1F3864")
header_font = Font(bold=True, color="FFFFFF", size=11)
even_fill = PatternFill("solid", fgColor="EBF0FA")
odd_fill = PatternFill("solid", fgColor="FFFFFF")
status_done_fill = PatternFill("solid", fgColor="C6EFCE")
status_wip_fill = PatternFill("solid", fgColor="FFEB9C")
status_done_font = Font(bold=True, color="276221")
status_wip_font = Font(bold=True, color="9C5700")
thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

col_widths = [42, 22, 10, 48, 48]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ヘッダー行
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[1].height = 22

# データ行
for row_idx, app in enumerate(apps, 2):
    fill = even_fill if row_idx % 2 == 0 else odd_fill
    values = [app["url"], app["name"], app["status"], app["features"], app["usage"]]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if col == 3:
            if val == "完成":
                cell.fill = status_done_fill
                cell.font = status_done_font
            else:
                cell.fill = status_wip_fill
                cell.font = status_wip_font
            cell.alignment = Alignment(horizontal="center", vertical="top")
        elif col in (4, 5):
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        else:
            cell.fill = fill
    ws.row_dimensions[row_idx].height = 100

ws.freeze_panes = "A2"

out = "/Users/apple/Downloads/社内アプリ一覧.xlsx"
wb.save(out)
print(f"saved: {out}")
