#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""レントロールに載っていない物件を洗い出し、**分かる範囲の入居者**を集める。

  ★前提（2026-09-05 オーナー説明）
    レントロールに載るのは「うちが直接家賃を預かっている物件」。
    **管理していても家賃はオーナーが直接受け取る物件は、載っていなくて正しい。**
    だから「載っていない＝抜け」ではない。ここでは事実だけ並べ、判断は人がする。

  入居者の手がかりは次の2つ。どちらも機械で確かに読める。
    ① 物件フォルダの `入居者/` にある箱の名前（例 2-6F_㈱高等進学塾）
    ② `01_契約` の契約書の中身（区画ごとに、いちばん新しい契約の賃借人）

  使い方: rentroll_gap.py          … 表にする
          rentroll_gap.py --csv    … CSVで出す（貼り付け用）
          rentroll_gap.py --xlsx   … 共有フォルダの ★確認★レントロール未収録一覧.xlsx を作り直す

  ★「フォルダ無し」と出たら、まず名寄せを疑うこと。2026-09-05 に10件が誤って
    「フォルダ無し」になっていた（隆生の施設は入れ物フォルダの中、湯浅ビルは
    括弧の中が本当の呼び名、ザ・プラザ2はⅡとの表記違い、井高野は「駅前」の有無）。
"""
import os, re, sys, csv, unicodedata, collections
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import read_contracts as RC

SF   = "/Users/apple/Library/CloudStorage/Dropbox-大京商事　株式会社/共有フォルダ"
MP   = os.path.join(SF, "物件・管理/管理物件")
RR   = {"ビル": "★要更新★レントロール一覧（ビル）.xlsx",
        "マンション": "★要更新★レントロール一覧（マンション）.xlsx",
        "駐車場": "★要更新★レントロール一覧（駐車場他）.xlsx"}
N = lambda s: unicodedata.normalize("NFC", s)


# ★全角ローマ数字は台帳とフォルダで書き方が割れる（ザ・プラザ2 ↔ ザ・プラザⅡ）
ROMAN = str.maketrans({"Ⅰ": "1", "Ⅱ": "2", "Ⅲ": "3", "Ⅳ": "4", "Ⅴ": "5",
                       "Ⅵ": "6", "Ⅶ": "7", "Ⅷ": "8", "Ⅸ": "9", "Ⅹ": "10"})


def nz(s):
    """名寄せ用に潰す。★ここが弱いと「フォルダ無し」が大量に出る（2026-09-05）。"""
    t = unicodedata.normalize("NFKC", str(s)).translate(ROMAN).lower()
    t = t.replace("(株)", "").replace("(有)", "").replace("(同)", "")
    t = re.sub(r"[（(][^）)]*[）)]", "", t)          # 「（鷲見修）」のようなオーナー名を落とす
    return re.sub(r"[\s　・･\-ー－_/／,、.。]", "", t)


def loose(s):
    """住所の言い方の揺れまで落とした形。「枚方高野道」と「枚方市高野道2丁目23-20」を寄せる。"""
    t = nz(s)
    t = re.sub(r"[0-9]+", "", t)
    t = re.sub(r"駅前|駅", "", t)
    return re.sub(r"[都道府県市区町村丁目番地号大字字]", "", t)


# ---------- レントロールにあるシート ----------
sheets = set()
for f in RR.values():
    wb = load_workbook(os.path.join(SF, f), read_only=True)
    sheets |= {nz(s) for s in wb.sheetnames}
    wb.close()


def in_rentroll(name):
    n = nz(name)
    return any(len(s) >= 3 and len(n) >= 3 and (s in n or n in s) for s in sheets)


# ---------- 台帳 ----------
wb = load_workbook(os.path.join(SF, "★要更新★管理物件台帳.xlsx"), data_only=True, read_only=True)
data = [[("" if c is None else str(c).strip()) for c in r]
        for r in wb["管理物件台帳"].iter_rows(values_only=True)]
wb.close()
ledger = [{"種別": r[0], "物件名": r[1], "分類": r[2], "担当": r[3] if len(r) > 3 else ""}
          for r in data[2:] if len(r) > 3 and r[1]]

# ---------- 物件フォルダの索引 ----------
SHELF = ("物件基本", "入居者", "修繕・点検", "解約・精算", "通知・案内", "記録・写真", "その他")
# ★区画の箱（「1番_チョン」「テナント①_…」）は物件ではないので中へ降りない
BOXNAME = re.compile(r"^(?:\d{1,3}番|テナント[①-⑳0-9]|プレート|BOX\s*\d|\d{1,2}[FＦ][_＿]|"
                     r"[0-9A-Za-z]{1,5}[-_]?\d{1,4}[_＿])")

folders = {}


def _add(name, path):
    folders.setdefault(nz(name), path)
    # ★「YUASAビル（㈲湯浅ビル）」のように、括弧の中に本当の呼び名が入っていることがある。
    #   nz() は括弧を落とすので、中身も別名として登録しておく（2026-09-05）。
    for m in re.finditer(r"[（(]([^）)]{2,24})[）)]", unicodedata.normalize("NFC", name)):
        a = nz(m.group(1))
        if len(a) >= 2 and not re.fullmatch(r"[0-9]+", a): folders.setdefault(a, path)


for kind in sorted(os.listdir(MP)):
    kp = os.path.join(MP, kind)
    if not os.path.isdir(kp) or kind.startswith("."): continue
    # ★_旧管理物件 も索引する。「フォルダ無し」と言われると人は探しに行ってしまう
    for p in sorted(os.listdir(kp)):
        pp = os.path.join(kp, N(p))
        if not os.path.isdir(pp): continue
        _add(N(p), pp)
        # ★「隆生福祉会（ゆめ）」のように、中に施設が並ぶ入れ物がある。1階層だけ降りる
        for q in sorted(os.listdir(pp)):
            qp = os.path.join(pp, N(q))
            if not os.path.isdir(qp) or N(q) in SHELF or BOXNAME.match(unicodedata.normalize("NFKC", N(q))):
                continue
            _add(N(q), qp)


def folder_of(name):
    n = nz(name)
    if n in folders: return folders[n]
    c = [v for k, v in folders.items() if len(k) >= 3 and len(n) >= 3 and (k in n or n in k)]
    if c: return max(c, key=len)
    # ★住所の言い方の揺れ（枚方高野道 ↔ 枚方市高野道2丁目23-20）
    ln = loose(name)
    if len(ln) >= 3:
        c = [v for k, v in folders.items() if len(loose(k)) >= 3
             and (loose(k) in ln or ln in loose(k))]
        if c: return max(c, key=len)
    return None


# ---------- 契約書から読んだもの（区画ごとに最新） ----------
newest = collections.defaultdict(dict)          # 物件 → 区画 → (日付, 賃借人)
for r in RC.scan():
    if not r["賃借人"]: continue
    k = r["区画"] or "―"
    cur = newest[nz(r["物件"])].get(k)
    if cur is None or r["日付"] > cur[0]:
        newest[nz(r["物件"])][k] = (r["日付"], r["賃借人"])

# ---------- 集計 ----------
out = []
for L in ledger:
    if L["分類"] not in ("管理", "自社"): continue
    if in_rentroll(L["物件名"]): continue
    d = folder_of(L["物件名"])
    boxes, froms = [], []
    if d:
        sp = os.path.join(d, "入居者")
        if os.path.isdir(sp):
            for b in sorted(os.listdir(sp)):
                if os.path.isdir(os.path.join(sp, b)): boxes.append(N(b))
        nm = newest.get(nz(os.path.basename(d)), {})
        froms = [(k, v[1], v[0]) for k, v in sorted(nm.items())]
    out.append({**L, "フォルダ": d, "入居者の箱": boxes, "契約書から": froms})

print("=" * 84)
print("レントロールに無い『管理・自社』物件と、分かっている入居者")
print("=" * 84)
print("  ※ 家賃をオーナーが直接受け取る物件は、載っていなくて正しい（抜けとは限らない）\n")
a = [o for o in out if o["入居者の箱"] or o["契約書から"]]
b = [o for o in out if not (o["入居者の箱"] or o["契約書から"])]
print(f"  入居者が分かるもの   {len(a):>3}物件")
print(f"  手がかりが無いもの   {len(b):>3}物件\n")

for o in a:
    print(f"── {o['物件名']}（{o['種別']}・{o['分類']}・{o['担当']}）")
    for b2 in o["入居者の箱"]:
        print(f"     箱   {b2}")
    for k, who, dt in o["契約書から"]:
        print(f"     契約 {k:<5} {who[:24]:<26} {RC.ymd(dt)}")
    print()

print("── 手がかりが無いもの")
for o in b:
    print(f"     {o['物件名']}（{o['種別']}・{o['分類']}・{o['担当']}）"
          f"{'  ※フォルダも無い' if not o['フォルダ'] else ''}")

if "--csv" in sys.argv:
    w = csv.writer(open("local/レントロール追加候補.csv", "w", newline="", encoding="utf-8-sig"))
    w.writerow(["物件名", "種別", "分類", "担当", "区画", "契約者", "根拠", "日付"])
    for o in a:
        for b2 in o["入居者の箱"]:
            m = re.match(r"^([^_]*)_(.*)$", b2)
            w.writerow([o["物件名"], o["種別"], o["分類"], o["担当"],
                        m.group(1) if m else "", m.group(2) if m else b2, "入居者の箱", ""])
        for k, who, dt in o["契約書から"]:
            w.writerow([o["物件名"], o["種別"], o["分類"], o["担当"], k, who,
                        "契約書", RC.ymd(dt)])
    print("\n  local/レントロール追加候補.csv に出した")


if "--xlsx" in sys.argv:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import datetime
    TH = Side(style="thin", color="BFBFBF")
    BD = Border(left=TH, right=TH, top=TH, bottom=TH)
    wbx = Workbook(); w = wbx.active; w.title = "レントロール未収録"
    w["A1"] = f"レントロールに載っていない『管理・自社』物件（{datetime.date.today():%Y-%m-%d} 時点）"
    w["A1"].font = Font(bold=True, size=13, color="1F4E79")
    w["A2"] = ("※ 家賃をオーナーが直接受け取る物件は、載っていなくて正しい（抜けとは限らない）。"
               "「載せる？」と「現在の契約者」を手で埋めてください。")
    w["A2"].font = Font(size=9, color="C00000")
    head = ["種別", "物件名", "分類", "担当", "物件フォルダ",
            "分かっている入居者（機械で読めた分）", "根拠", "載せる？", "現在の契約者（手で記入）"]
    for j, h in enumerate(head, 1):
        c = w.cell(row=4, column=j, value=h)
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BD
    r = 5
    for o in out:
        who = " ／ ".join([b2 for b2 in o["入居者の箱"]]
                         + [f"{k} {v}" for k, v, _ in o["契約書から"]]) or "（手がかり無し）"
        why = " ／ ".join(f"契約書 {RC.ymd(dt)}" for _, _, dt in o["契約書から"])
        fol = os.path.relpath(o["フォルダ"], MP) if o["フォルダ"] else "★フォルダ無し"
        for j, v in enumerate([o["種別"], o["物件名"], o["分類"], o["担当"], fol, who, why, "", ""], 1):
            c = w.cell(row=r, column=j, value=v)
            c.border = BD; c.alignment = Alignment(vertical="top", wrap_text=(j in (5, 6, 7)))
            if j == 5 and not o["フォルダ"]:
                c.font = Font(bold=True, color="C00000")
        r += 1
    for col, wd in zip("ABCDEFGHI", [12, 26, 8, 8, 34, 46, 26, 10, 24]):
        w.column_dimensions[col].width = wd
    w.freeze_panes = "A5"; w.auto_filter.ref = f"A4:I{r-1}"
    out_p = os.path.join(SF, "★確認★レントロール未収録一覧.xlsx")
    wbx.save(out_p)
    nf = sum(1 for o in out if not o["フォルダ"])
    print(f"\n  作り直した: {os.path.basename(out_p)}（{len(out)}物件 / フォルダ無し {nf}件）")
