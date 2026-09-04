#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""整理計画（確定版）を、物件を指定して実行する。
   使い方: apply_plan.py "大京本社ビル"        … 下見
           apply_plan.py "大京本社ビル" --go   … 実行
   削除はしない。移動だけ。対応表を残すので戻せる。"""
import os, sys, json, shutil, datetime, unicodedata, collections
from openpyxl import load_workbook

ROOT = "/Users/apple/Library/CloudStorage/Dropbox-大京商事　株式会社/共有フォルダ/（★必読★）新共有フォルダ"
PLAN = os.path.expanduser(f"~/Desktop/整理計画_確定版_{datetime.date.today():%Y%m%d}.xlsx")
REC  = os.path.expanduser("~/chatwork-ai-manager/local/")
N = lambda s: unicodedata.normalize("NFC", s)
GO = "--go" in sys.argv
props = [a for a in sys.argv[1:] if not a.startswith("--")]
if not props:
    print("物件名を指定してください（--all で全物件）"); sys.exit(1)

ws = load_workbook(PLAN)["整理計画"]
cols = [c.value for c in ws[1]]
rows = [dict(zip(cols, [c.value for c in r])) for r in ws.iter_rows(min_row=2)]

sel = rows if props == ["--all"] else [r for r in rows if r["物件"] in props]
move = [r for r in sel if r["棚"] and r["現在のパス"] != r["移動後のパス"] and not r["注意"]]
stay = [r for r in sel if not r["棚"]]
skip = [r for r in sel if r["注意"] and r["棚"]]

print("=" * 74)
print(("【実行】" if GO else "【下見】") + f"  対象物件: {'、'.join(props)}")
print("=" * 74)
print(f"  この物件のファイル {len(sel)}件")
print(f"    移動する          {len(move)}件")
print(f"    現状維持（判定できず） {len(stay)}件")
print(f"    ★注意ありで見送り    {len(skip)}件")
if skip:
    for r in skip[:6]:
        print(f"       [{r['注意']}] {os.path.basename(r['現在のパス'])[:50]}")

# ---- 事前点検 ----
ng = []
seen = collections.Counter(N(r["移動後のパス"]) for r in move)
for r in move:
    s = os.path.join(ROOT, N(r["現在のパス"]))
    d = os.path.join(ROOT, N(r["移動後のパス"]))
    if not os.path.exists(s): ng.append(("元が無い", r["現在のパス"]))
    elif os.path.exists(d):   ng.append(("先に同名あり", r["移動後のパス"]))
    elif seen[N(r["移動後のパス"])] > 1: ng.append(("行き先が重なる", r["移動後のパス"]))
if ng:
    print(f"\n★点検で問題 {len(ng)}件。中止します。")
    for k, p in ng[:12]: print(f"   [{k}] {p}")
    sys.exit(1)
print("\n  事前点検: 問題なし（元が全部あり／行き先に同名なし／重なりなし）")

print("\n  行き先の内訳:")
for k, v in collections.Counter(
        "/".join(r["移動後のパス"].split("/")[4:-1]) for r in move).most_common():
    print(f"    {k:<44}{v:>4}件")

if not GO:
    print("\n  ※下見のみ。実行するには --go を付ける")
    sys.exit(0)

log = []
for r in move:
    s = os.path.join(ROOT, N(r["現在のパス"]))
    d = os.path.join(ROOT, N(r["移動後のパス"]))
    os.makedirs(os.path.dirname(d), exist_ok=True)
    shutil.move(s, d)
    log.append({"from": r["現在のパス"], "to": r["移動後のパス"]})

# 空になったフォルダを片付ける
removed = []
for _ in range(8):
    for prop in ({r["物件"] for r in move} if props != ["--all"] else set(props)):
        pass
    for r in move:
        base = os.path.join(ROOT, "/".join(N(r["現在のパス"]).split("/")[:4]))
        for cur, dn, fn in os.walk(base, topdown=False):
            if cur == base: continue
            real = [f for f in fn if f != ".DS_Store" and not f.startswith("._")]
            if real or dn: continue
            ds = os.path.join(cur, ".DS_Store")
            if os.path.exists(ds): os.remove(ds)
            try:
                os.rmdir(cur); removed.append(N(os.path.relpath(cur, ROOT)))
            except OSError: pass
        break

os.makedirs(REC, exist_ok=True)
tag = "_".join(props)[:40].replace("/", "_")
rec = os.path.join(REC, f"移動記録_整理_{tag}_{datetime.date.today():%Y%m%d}.json")
json.dump({"date": str(datetime.date.today()), "root": ROOT, "props": props,
           "moved": log, "removed_empty_dirs": removed},
          open(rec, "w"), ensure_ascii=False, indent=2)
print(f"\n  移動 {len(log)}件 / 空フォルダを片付け {len(removed)}個")
print(f"  対応表: {rec}")
print("  戻すときは to → from へ move すればよい")
