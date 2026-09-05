#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""年度計画表に「報告書があるか」を自動でチェックする。

  計画表の1行ごとに、業者の作業報告書・点検報告書を探して、
  **その年度のうちに実施されたものだけ** ✓ を付ける。

    報告書   … ✓（今年度に実施） / ―（もともと報告書が出ない作業）
                / 未（前回はあるが今年度はまだ） / 空（記録が1件も無い）

  ★行の色で「いま何をすべきか」が分かるようにする（2026-09-05 オーナー指示）。
      緑  … 実施済み
      黄  … そろそろ（予定月が今月・来月）
      赤  … 超過（予定月を過ぎたのに記録が無い）
      灰  … まだ先（予定月がこれから）
    実施の記録 … 見つかった報告書の年月（過去のぶんも並べる。例 2024.09 2025.03 2026.03）
    根拠     … いちばん新しい報告書のファイル名

  ★年度は4月はじまり。2026年度＝2026年4月1日〜2027年3月31日。
    去年の報告書で「今年やった」ことにしない（2026-09-05に間違えた）。

  ★人が手で書いた「実施日」「備考」は触らない。別の列に書く。
  ★何度でも回せる。月報が増えたらまた実行すればチェックが増える。
  ★消防訓練のように、そもそも業者の報告書が出ないものは「―」にする。

  探す場所
    業者・設備/報告書/          … 業者ごとの点検報告書
    物件・管理/管理物件/         … 物件フォルダの修繕・点検など
    （_旧管理物件 は見ない）

  使い方: check_yearplan.py            … 下見（何件チェックが付くか）
          check_yearplan.py --go       … 書き込む
          check_yearplan.py --go --fy 2027  … 年度を指定して回す

  ★週1回、棚卸し（run_inventory_weekly.sh）と一緒に自動で回している。
    月報や報告書が増えるたびに ✓ と色が最新になる。
"""
import os, re, sys, unicodedata, collections
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

SF = "/Users/apple/Library/CloudStorage/Dropbox-大京商事　株式会社/共有フォルダ"
KG = os.path.join(SF, "物件・管理/管理業務")
GO = "--go" in sys.argv
N = lambda s: unicodedata.normalize("NFC", s)
K = lambda s: unicodedata.normalize("NFKC", N(s))

THIN = Side(style="thin", color="B4C6E7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
OKF = Font(bold=True, color="2E7D32")     # ✓ は緑
NAF = Font(color="808080")                # ― は灰
YETF = Font(bold=True, color="C55A11")    # 未 は橙

# 行の塗り分け（実施済み／そろそろ／超過／まだ先）
FILL_DONE = PatternFill("solid", fgColor="E2EFDA")   # 緑
FILL_SOON = PatternFill("solid", fgColor="FFF2CC")   # 黄
FILL_OVER = PatternFill("solid", fgColor="FFC7CE")   # 赤
FILL_LATER = PatternFill("solid", fgColor="F2F2F2")  # 灰
FILL_NA = PatternFill("solid", fgColor="FFFFFF")     # 白（報告書が出ないもの）

# いまの年月（年度の進み具合を測る基準）
import datetime as _dt
NOW = (_dt.date.today().year, _dt.date.today().month)


def fy_pos(m):
    """月を年度のはじめ（4月）からの順番に直す。4月=0 … 3月=11。"""
    return (m - 4) % 12 if m else None


def status(mark, plan_month):
    """行の状態を決める。plan_month は計画表の「月」の列の値。"""
    if mark == "―": return "報告書なし", FILL_NA
    if mark == "✓": return "実施済み", FILL_DONE
    if not plan_month: return "予定月が無い", FILL_LATER
    p, n = fy_pos(int(plan_month)), fy_pos(NOW[1])
    if p is None or n is None: return "予定月が無い", FILL_LATER
    if p < n: return "★超過", FILL_OVER
    if p <= n + 1: return "そろそろ", FILL_SOON
    return "まだ先", FILL_LATER

# 物件名の書き方の揺れを吸収する
PROP = [(r"サトウ|ｻﾄｳ|SATO", "サトウビルⅡ"), (r"天王寺", "大京天王寺ビル"), (r"本社", "大京本社ビル"),
        (r"西ビル", "大京西ビル"), (r"河合", "河合京橋ビル"), (r"クリスタル京橋|ｸﾘ京", "クリスタル京橋"),
        (r"ベリエール|ﾍﾞﾘ", "コーポ・ラ・ベリエール"), (r"トーヨーコーポ", "トーヨーコーポ"),
        (r"エレガンス", "エレガンス放出"), (r"湯浅|YUASA", "YUASAビル"), (r"パールハイム", "パールハイム高殿"),
        (r"クリスタル.?66|CC66|ＣＣ66", "クリスタルコート66"),
        (r"クリスタル.?26|CC26|ＣＣ26", "クリスタルコート26"),
        (r"メゾンドール|ﾒｿﾞﾝ|メゾン", "メゾンドール都島"), (r"ソフィア南森", "ソフィア南森町"),
        (r"ソフィア東野田", "ソフィア東野田"), (r"ライラック|ﾗｲﾗｯｸ", "ワタヤライラック"),
        (r"KSK|ＫＳＫ", "KSKビル"), (r"H・K|ＨＫ|HKビル", "H・Kビル"), (r"囲碁", "大京ビル（囲碁）"),
        (r"鶴見公民館", "鶴見公民館"), (r"角屋", "角屋マンション"), (r"三好", "三好マンション"),
        (r"グレイスフル", "グレイスフルハイム大和"), (r"カナン", "カナンハウス"),
        (r"ゆめあまみ|あまみ", "ゆめあまみ"), (r"ゆめ中央|中央保育", "ゆめ中央保育園"),
        (r"ゆめ長居|長居公園", "ゆめ長居公園"), (r"パラティース", "ゆめパラティース"),
        (r"ゆめ都島", "ゆめ都島")]
# 作業名 → 種類
KIND = [(r"訓練", "訓練"),
        (r"消防|防火|自火報|総合点検|機器点検|耐圧|連結送水", "消防"),
        (r"受水槽|貯水槽|水質|簡易専用|ブースター|揚水", "受水槽"),
        (r"EV|ＥＶ|エレベータ|昇降", "EV"),
        (r"消毒|防除|ゴキブリ|害虫", "害虫防除"),
        (r"絶縁|排水|ピット|ポンプ|曝気|ﾋﾞﾙﾋﾟｯﾄ|ビルピット", "排水・ピット"),
        (r"清掃|床|ガラス|網戸|ﾌｨﾙﾀ|フィルタ|換気|空調|洗浄|ﾜｯｸｽ|ワックス|"
         r"ｶｰﾍﾟｯﾄ|カーペット|ｸﾞﾘｽﾄﾗｯﾌﾟ|グリストラップ|空室|ｸﾛｽ", "清掃・空調"),
        (r"非常|火災通報|全熱|ファン|ﾌｧﾝ|給湯|GHP|設備定期", "設備点検")]
# もともと業者の報告書が出ないもの
NO_REPORT = {"訓練"}

# ★年度（4月はじまり）。この間の報告書だけを「今年度に実施」とみなす。
# ★年度は「いま」から自動で決める（4〜12月は その年 / 1〜3月は 前の年）。
#   週次ジョブから毎週回すので、年度が変わっても手で直さなくてよい。
#   別の年度を見たいときだけ --fy 2027 のように渡す。
def _this_fy():
    y, m = NOW
    return y if m >= 4 else y - 1

FY = _this_fy()
for _i, _a in enumerate(sys.argv):
    if _a == "--fy" and _i + 1 < len(sys.argv): FY = int(sys.argv[_i + 1])
FY_FROM = (FY, 4)
FY_TO   = (FY + 1, 3)


def in_fy(y, m):
    """年度のうちか。月が読めない（0）ものは年だけで見る。"""
    if not m: return y in (FY, FY + 1)
    return FY_FROM <= (y, m) <= FY_TO


# 業者名から作業の種類を当てる（ファイル名に作業名が無いとき用）
VEND_KIND = [(r"西武消毒|西部消毒", "害虫防除"), (r"日本昇降機|昇降機", "EV"),
             (r"ゆたか商会|ユタカ", "消防"), (r"ビケンテクノ|ﾋﾞｹﾝ", "設備点検"),
             (r"ベルポ|ﾍﾞﾙﾎﾟ", "清掃・空調"), (r"アイエスケー|ｱｲｴｽｹｰ", "受水槽"),
             (r"東洋ビルメンテナンス", "排水・ピット")]


def vendor_kind(s):
    for p, k in VEND_KIND:
        if re.search(p, K(s)): return k
    return None


def match(pats, s):
    for p, k in pats:
        if re.search(p, K(s)): return k
    return None


def ym(s):
    s = K(s)
    for pat, base in [(r"(20\d{2})[.\-年_]?\s*(\d{1,2})", 0),
                      (r"[Rr](\d{1,2})[.\-](\d{1,2})", 2018),
                      (r"[Hh](\d{1,2})[.\-](\d{1,2})", 1988)]:
        m = re.search(pat, s)
        if m and 1 <= int(m.group(2)) <= 12:
            y = int(m.group(1)) if base == 0 else base + int(m.group(1))
            if 2000 <= y <= 2030: return y, int(m.group(2))
    m = re.search(r"(20\d{2})", s)
    if m and 2000 <= int(m.group(1)) <= 2030: return int(m.group(1)), 0
    return None


def collect():
    """報告書を集める → {(物件, 種類): [(年, 月, ファイル名)]}"""
    rep = collections.defaultdict(list)
    for base in [os.path.join(SF, "業者・設備/報告書"), os.path.join(SF, "物件・管理/管理物件")]:
        for c, d, f in os.walk(base):
            rel = N(c).split("共有フォルダ/")[-1]
            if "_旧管理物件" in rel or "_アーカイブ" in rel: continue
            for x in sorted(f):
                if x == ".DS_Store": continue
                x = N(x); full = rel + "/" + x
                # ★「報告書かどうか」はファイル名だけで決めない。
                #   ベルポの報告書は「2023.10.4､10.24､10.25（全館床洗浄）.pdf」のように
                #   作業日と作業名だけで、「報告書」の語は親フォルダ側にある（2026-09-05）。
                if not re.search(r"報告|点検|完了|伝票|検査|測定|作業", K(full)): continue
                # ★物件名も種類も、ファイル名だけでなく親フォルダの名前から拾う。
                #   隆生の報告書は「ゆめあまみ/大阪西武消毒㈱/作業完了報告書/2025.5.15…pdf」
                #   のように、施設名も業者名もフォルダ側にしかない（2026-09-05）。
                # ★フォルダ名は「㈱ベルポ（ゆめあまみ）」のように業者名が先に来る。
                #   施設名は括弧の中にあるので、パス全体から探す（2026-09-05）。
                p = match(PROP, x) or match(PROP, full)
                k = match(KIND, x) or match(KIND, rel) or vendor_kind(rel + "/" + x)
                t = ym(x) or ym(rel)
                if p and k and t: rep[(p, k)].append((t[0], t[1], x))
    return rep


def run(book, sheet, col_prop, col_kind, col_done=None):
    """計画表1枚にチェックを入れる。col_prop=物件名が載っている列、col_kind=種類の列。"""
    wb = load_workbook(book)
    ws = wb[sheet]
    # 追加する列（既にあれば使い回す）
    head = {ws.cell(row=4, column=c).value: c for c in range(1, ws.max_column + 2)}
    base = ws.max_column
    cols = {}
    # ★色の意味は書いておかないと伝わらない（見出しの1つ上に凡例を置く）
    leg = [("実施済み", FILL_DONE), ("そろそろ（今月・来月）", FILL_SOON),
           ("★超過（予定月を過ぎた）", FILL_OVER), ("まだ先", FILL_LATER)]
    for i, (t, f) in enumerate(leg):
        c = ws.cell(row=3, column=1 + i * 2, value=t)
        c.fill = f; c.font = Font(bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ws.cell(row=3, column=9, value=f"※色は {NOW[0]}年{NOW[1]}月 時点。--go を掛け直すと更新される").font = Font(size=9, color="808080")

    for i, name in enumerate(["報告書", "状態", "実施の記録", "根拠"]):
        cols[name] = head.get(name) or (base + 1 + i)
        c = ws.cell(row=4, column=cols[name], value=name)
        c.fill = PatternFill("solid", fgColor="375623"); c.font = Font(bold=True, color="FFFFFF")
        c.border = BORDER; c.alignment = Alignment(horizontal="center")
    ws.column_dimensions[chr(64 + cols["報告書"])].width = 8
    ws.column_dimensions[chr(64 + cols["状態"])].width = 12
    ws.column_dimensions[chr(64 + cols["実施の記録"])].width = 30
    ws.column_dimensions[chr(64 + cols["根拠"])].width = 46
    # ★前に自動で入れた実施日（緑の斜体）は毎回消してから入れ直す。
    #   年度の判定を直す前に書いた日付が残ると、去年の日付が実施日に居座る（2026-09-05）。
    if col_done:
        for r in range(5, ws.max_row + 1):
            c = ws.cell(row=r, column=col_done)
            if c.value and c.font and c.font.italic and c.font.color \
               and getattr(c.font.color, "rgb", "") in ("002E7D32", "FF2E7D32", "2E7D32"):
                c.value = None
    ok = na = ng = yet = 0
    cnt = {}
    ck = cp = None; cm = None
    for r in range(5, ws.max_row + 1):
        # ★1列目（施設名／点検の種類）は「変わったときだけ」書いてある。
        #   空欄の行は直前の値が続いていると見る（2026-09-05）。
        v1 = ws.cell(row=r, column=1).value
        if v1: cp = v1
        v2 = ws.cell(row=r, column=2).value
        if isinstance(v2, int): cm = v2
        k = ws.cell(row=r, column=col_kind).value or ck
        if ws.cell(row=r, column=col_kind).value: ck = k
        body = ws.cell(row=r, column=col_prop).value
        if not body: continue
        kind = match(KIND, str(body)) or match(KIND, str(k))
        prop = match(PROP, str(body)) or match(PROP, str(cp or ""))
        mark = dates = src = ""
        if kind in NO_REPORT:
            mark, src, na = "―", "訓練は業者の報告書が出ない", na + 1
        elif prop:
            v = sorted(set(REP.get((prop, kind), [])))
            done = [t for t in v if in_fy(t[0], t[1])]
            if done:
                mark = "✓"; ok += 1
                src = done[-1][2][:44]
            elif v:
                mark = "未"; yet += 1
                src = f"前回 {v[-1][0]}.{v[-1][1]:02d}" if v[-1][1] else f"前回 {v[-1][0]}"
            else: ng += 1
            dates = " ".join(f"{y}.{m:02d}" if m else f"{y}" for y, m, _ in v[-6:])
        else: ng += 1
        # ★人が手で書いた実施日が空なら、報告書から分かった日付を薄い字で入れる。
        #   今年度のものだけ。去年の日付を実施日に書かない。
        if mark == "✓" and col_done and not ws.cell(row=r, column=col_done).value:
            latest = [t for t in sorted(set(REP.get((prop, kind), []))) if in_fy(t[0], t[1])][-1]
            c = ws.cell(row=r, column=col_done,
                        value=f"{latest[0]}.{latest[1]:02d}" if latest[1] else str(latest[0]))
            c.font = Font(color="2E7D32", italic=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER
        # ★行ぜんぶを状態の色で塗る
        pm = ws.cell(row=r, column=2).value
        st, fill = status(mark, pm if isinstance(pm, int) else cm)
        for cc in range(1, cols["根拠"] + 1):
            ws.cell(row=r, column=cc).fill = fill
        cnt[st] = cnt.get(st, 0) + 1
        for name, val in [("報告書", mark), ("状態", st), ("実施の記録", dates), ("根拠", src)]:
            c = ws.cell(row=r, column=cols[name], value=val or None)
            c.border = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=(name != "報告書"),
                                    horizontal="center" if name == "報告書" else "left")
            if name == "状態" and st == "★超過": c.font = Font(bold=True, color="9C0006")
            if name == "報告書":
                if val == "✓": c.font = OKF
                elif val == "―": c.font = NAF
                elif val == "未": c.font = YETF
    if GO: wb.save(book)
    return ok, na, ng, yet, cnt


REP = collect()
print("=" * 70)
print(("【実行】" if GO else "【下見】") + f"  {FY}年度（{FY}.4〜{FY+1}.3）に実施したものへチェックを入れる")
print("=" * 70)
print(f"  集めた報告書 {sum(len(v) for v in REP.values())}件 / 物件×種類 {len(REP)}通り\n")
for book, sheet, cp, ck, cd in [
        (os.path.join(KG, f"★{FY}年度★計画表・点検一覧表（自社・管理物件）.xlsx"), f"{FY}年度", 3, 1, 5),
        (os.path.join(KG, f"★{FY}年度★計画表・点検一覧表（隆生福祉会）.xlsx"), f"{FY}年度", 3, 1, 6)]:
    if not os.path.exists(book): print("  ★無い:", os.path.basename(book)); continue
    ok, na, ng, yet, cnt = run(book, sheet, cp, ck, cd)
    print(f"  {os.path.basename(book)[:38]:<40}✓{ok:>3} / 未{yet:>3} / ―{na:>3} / 記録なし{ng:>3}")
    print("      " + " / ".join(f"{k}{v}" for k, v in sorted(cnt.items(), key=lambda x: -x[1])))
if not GO: print("\n  ※下見のみ。--go で書き込む")
