#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""レントロール（現在の契約者の正典）と、物件フォルダのテナント箱を突き合わせ、
   現契約 / 旧契約 を判定する。読むだけ。1件も動かさない。"""
import os, re, unicodedata, collections
from openpyxl import load_workbook

ROOT = "/Users/apple/Library/CloudStorage/Dropbox-大京商事　株式会社/共有フォルダ/（★必読★）新共有フォルダ"
MP   = os.path.join(ROOT, "物件・管理/管理物件")
RR   = ["★要更新★レントロール一覧（ビル）.xlsx",
        "★要更新★レントロール一覧（マンション）.xlsx",
        "★要更新★レントロール一覧（駐車場他）.xlsx"]
N = lambda s: unicodedata.normalize("NFC", s)

def norm(s):
    # ★NFKC で半角カナ・全角英数を吸収する（ｵﾌﾟﾃｨｶﾙ = オプティカル、ＵＦＪ = UFJ）
    s = unicodedata.normalize("NFKC", str(s)).lower()
    s = re.sub(r"[\s　・･,，.。\-ー－_/\\'\"()（）\[\]【】]", "", s)
    s = re.sub(r"株式会社|有限会社|合同会社|医療法人|社会福祉法人|一般社団法人|㈱|㈲", "", s)
    for a, b in (("ⅰ","1"),("ⅱ","2"),("ⅲ","3"),("ii","2")): s = s.replace(a, b)
    return s

def names_of(s):
    """括弧の外と中の両方を候補にする（医療法人栄新会（東野田町ｸﾘﾆｯｸ）→ 両方）"""
    s = unicodedata.normalize("NFKC", str(s))
    out = [s] + re.findall(r"[(（]([^)）]{2,30})[)）]", s)
    out += [re.sub(r"[(（][^)）]*[)）]", "", s)]
    return [norm(x) for x in out if norm(x)]
def pstem(s):
    n = norm(s)
    for suf in ("ビル","bldg","マンション","駐車場","モータープール","駐輪場"):
        if n.endswith(suf) and len(n) > len(suf)+2: n = n[:-len(suf)]
    return n

# ---------- レントロールを読む ----------
rr = {}          # 物件stem -> {契約者norm: 号室}
rr_raw = {}      # 物件stem -> [(号室, 契約者)]
for f in RR:
    p = os.path.join(ROOT, f)
    if not os.path.exists(p): continue
    wb = load_workbook(p, data_only=True, read_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        pairs = []
        for row in ws.iter_rows(values_only=True):
            v = [("" if c is None else str(c).strip()) for c in row[:3]]
            if len(v) < 2 or not v[0] or not v[1]: continue
            if v[0] in ("号室/フロア","号室","区画","合計","計") or v[1] in ("契約者","合計"): continue
            if v[1] in ("", "空室", "空き", "-"): continue
            pairs.append((v[0], v[1]))
        if pairs:
            k = pstem(sn)
            rr_raw[k] = pairs
            tbl = {}
            for r, t in pairs:
                for nm in names_of(t):
                    tbl.setdefault(nm, r)
            rr[k] = tbl
    wb.close()
print(f"レントロール: {len(rr)}物件 / 契約者 {sum(len(v) for v in rr_raw.values())}件\n")

# ---------- 物件フォルダのテナント箱 ----------
ENT = re.compile(r"^\d+番|^\d+号|^[A-Z]?\d{3,4}(号|室|_)|^\d+[FＦ]|^\d+階|^[0-9]{3}_|^\d+[FＦ]【")
OLD = re.compile(r"解約|退去|明渡|不成立|終了")
res = collections.Counter(); rows = []
for kind in sorted(os.listdir(MP)):
    kp = os.path.join(MP, kind)
    if not os.path.isdir(kp): continue
    for prop in sorted(os.listdir(kp)):
        pp = os.path.join(kp, prop)
        if not os.path.isdir(pp): continue
        k = pstem(prop)
        cand = [kk for kk in rr if kk and (kk in k or k in kk) and min(len(kk), len(k)) >= 3]
        table = rr[max(cand, key=len)] if cand else None
        subs = [N(d) for d in os.listdir(pp) if os.path.isdir(os.path.join(pp, d))]
        tenants = [s for s in subs if ENT.search(s) or OLD.search(s)]
        for s in tenants:
            if OLD.search(s):
                v, why = "旧契約", "名前に解約とある"
            elif table is None:
                v, why = "判定できず", "レントロールに物件が無い"
            else:
                cands = names_of(re.sub(r"^[0-9A-Za-zＦF【】\[\]_番号室階\-]+", "", s)) + names_of(s)
                hit = [t for t in table if t and len(t) >= 3
                       and any(t in c or c.endswith(t) or t.endswith(c) for c in cands)]
                if hit:
                    v, why = "現契約", f"レントロール {table[max(hit,key=len)]} と一致"
                else:
                    v, why = "旧契約", "レントロールに載っていない"
            res[v] += 1
            rows.append((N(kind), N(prop), s, v, why))

print("=== 判定 ===")
for k, n in res.most_common(): print(f"  {k:<10}{n:>5}箱")
print(f"  合計       {sum(res.values()):>5}箱\n")

print("=== 大京本社ビルで検算 ===")
for kind, prop, s, v, why in rows:
    if prop == "大京本社ビル": print(f"  {v:<6} {s[:44]:<44} {why}")
print("\n=== レントロールに載っていない＝旧契約 と判定したもの（抜粋25）===")
n = 0
for kind, prop, s, v, why in rows:
    if v == "旧契約" and why == "レントロールに載っていない":
        print(f"  {prop[:22]:<22} / {s[:46]}"); n += 1
        if n >= 25: break
print("\n=== 判定できなかった物件 ===")
ng = sorted({prop for kind, prop, s, v, why in rows if v == "判定できず"})
print("  " + "、".join(ng[:20]) + (f" …ほか{len(ng)-20}物件" if len(ng) > 20 else ""))
