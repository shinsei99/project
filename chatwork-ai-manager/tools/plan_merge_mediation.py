#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""整理案①: 仲介フォルダにある自社管理物件の書類を、物件フォルダへ寄せる計画表を作る。
   読むだけ。1件も動かさない。 出力: ~/Desktop/整理計画_仲介→物件_YYYYMMDD.xlsx"""
import os, re, unicodedata, collections, datetime, hashlib

ROOT = "/Users/apple/Library/CloudStorage/Dropbox-大京商事　株式会社/共有フォルダ"
N = lambda s: unicodedata.normalize("NFC", s)
MED = "契約・書類/★仲介（賃貸・売買）"
MP  = "物件・管理/管理物件"

def norm(s):
    s = N(s).lower()
    s = re.sub(r"[（(\[][^）)\]]{0,24}[）)\]]", "", s)
    s = re.sub(r"[\s　・･,，.。\-ー－_/\\'\"]", "", s)
    for a, b in (("ⅰ","1"),("ⅱ","2"),("ⅲ","3"),("ii","2"),("iii","3")): s = s.replace(a, b)
    return s
def stem(s):
    n = norm(s)
    for suf in ("ビル","bldg","b・l・d","マンション","駐車場","モータープール"):
        if n.endswith(suf) and len(n) > len(suf) + 2: n = n[:-len(suf)]
    return n

# ---------- 標準の棚 ----------
SHELVES = [
    ("01_契約",        r"契約書|重説|重要事項|覚書|合意書|更新|定期借家|保証委託|入居申込|申込書|審査|媒介|念書|承諾"),
    ("02_解約・精算",   r"解約|退去|精算|清算|原状回復|明渡|返還|敷金"),
    ("03_請求・入金",   r"請求|領収|入金|送金|月次|収支|家賃|振込|明細|仲介手数料|広告料"),
    ("04_工事・修繕",   r"工事|修繕|リフォーム|リホーム|見積|施工|点検|保守|清掃|設備|昇降|エレベータ|消防"),
    ("05_図面・写真",   r"図面|竣工|平面|間取|配置|測量|公図|写真|マイソク|パース|zumen|境界"),
    ("06_権利・登記",   r"登記|謄本|全部事項|評価証明|固定資産|税|保険|証券|権利|決済"),
    ("08_検針・メーター", r"検針|メーター|水道|電気|ガス|使用量"),
    ("09_賃借人資料",   r"賃借人|入居者|テナント|車庫証|入居時確認"),
]
def shelf(below, fname):
    for n_, pat in SHELVES:
        if re.search(pat, below): return n_, "フォルダ名"
    for n_, pat in SHELVES:
        if re.search(pat, fname): return n_, "ファイル名"
    return "07_通知・その他", "受け皿"

# ---------- 管理物件の一覧 ----------
props = {}
for kind in sorted(os.listdir(os.path.join(ROOT, MP))):
    kp = os.path.join(ROOT, MP, kind)
    if not os.path.isdir(kp): continue
    for nm in sorted(os.listdir(kp)):
        if os.path.isdir(os.path.join(kp, nm)):
            props[stem(nm)] = (N(nm), N(kind))

# ---------- 仲介の案件フォルダ ----------
SHIKI = re.compile(r"^(買入申込書|専任媒介契約書|一般媒介契約書|仲介手数料契約書|★?表紙|書式|雛形|ひな形)")
cases = []
for sub in ("売買", "賃貸/★事業用（店舗・事務所）", "賃貸/★居住用", "賃貸/一時使用賃貸借契約書", "賃貸/定期借家契約"):
    base = os.path.join(ROOT, MED, sub)
    if not os.path.isdir(base): continue
    for nm in sorted(os.listdir(base)):
        p = os.path.join(base, nm)
        if os.path.isdir(p): cases.append((N(sub), N(nm), p))

def files_in(p):
    out = []
    for cur, dn, fn in os.walk(p):
        for f in fn:
            if f == ".DS_Store" or f.startswith("._"): continue
            out.append(os.path.join(cur, f))
    return out

rows = []
summary = collections.Counter(); summary_mb = collections.Counter()
for sub, nm, path in cases:
    fs = files_in(path)
    if not fs: continue
    k = stem(nm)
    hit = [(kk, v) for kk, v in props.items() if kk and (kk in k or k in kk) and min(len(kk), len(k)) >= 3]
    if SHIKI.match(nm):
        kind_of_move = "③書式・雛形へ"
        conf = "高"
        dest_base = "書式・雛形/仲介"
        prop_name = ""
    elif hit:
        best = max(hit, key=lambda x: len(x[0]))
        prop_name, kind = best[1]
        kind_of_move = "①物件フォルダへ寄せる"
        # 名前が完全一致なら確度・高、部分一致は要確認
        conf = "高" if stem(prop_name) == k else "要確認"
        dest_base = f"{MP}/{kind}/{prop_name}"
    else:
        kind_of_move = "②仲介に残す"
        conf = "高"
        dest_base = None
        prop_name = ""
    for f in fs:
        rel = N(os.path.relpath(f, ROOT))
        below = N(os.path.relpath(os.path.dirname(f), path))
        below = "" if below == "." else below
        fname = os.path.basename(rel)
        sz = os.path.getsize(f)
        if dest_base is None:
            new = rel
            sh = ""; why = ""
        else:
            sh, why = shelf(below + " " + nm, fname)
            new = f"{dest_base}/{sh}/{fname}" if kind_of_move.startswith("①") else f"{dest_base}/{fname}"
        rows.append({
            "区分": kind_of_move, "確度": conf, "案件フォルダ": nm, "仲介の種別": sub,
            "寄せ先の物件": prop_name, "棚": sh, "棚の決め手": why,
            "容量(KB)": round(sz / 1024, 1),
            "現在のパス": rel, "移動後のパス": new,
        })
        summary[kind_of_move] += 1; summary_mb[kind_of_move] += sz

# ---------- 衝突 ----------
dest = collections.Counter(r["移動後のパス"] for r in rows if r["区分"].startswith("①"))
col = {k for k, v in dest.items() if v > 1}
# 既存の物件フォルダに同名があるか
exist = set()
for r in rows:
    if r["区分"].startswith("①") and os.path.exists(os.path.join(ROOT, r["移動後のパス"])):
        exist.add(r["移動後のパス"])
for r in rows:
    p = r["移動後のパス"]
    r["注意"] = ("★行き先が重なる" if p in col else "") + (" ★既に同名あり" if p in exist else "")

print("=" * 74); print("整理案① 仲介 → 物件フォルダ（計画のみ・1件も動かしていない）"); print("=" * 74)
for k in ["①物件フォルダへ寄せる", "②仲介に残す", "③書式・雛形へ"]:
    print(f"  {k:<22} {summary[k]:>5}件  {summary_mb[k]/1024**2:>7.0f}MB")
print(f"  {'合計':<22} {sum(summary.values()):>5}件  {sum(summary_mb.values())/1024**2:>7.0f}MB")
c1 = collections.Counter(r["確度"] for r in rows if r["区分"].startswith("①"))
print(f"\n  ①のうち 確度・高 {c1['高']}件 / 要確認 {c1['要確認']}件")
print(f"  ★注意が要る行: {sum(1 for r in rows if r['注意'])}件")
print("\n  寄せ先の物件 上位12:")
cc = collections.Counter(r["寄せ先の物件"] for r in rows if r["区分"].startswith("①"))
for nm, n in cc.most_common(12): print(f"    {nm:<32} {n:>4}件")
print("\n  棚の内訳（①のみ）:")
for sh, n in collections.Counter(r["棚"] for r in rows if r["区分"].startswith("①")).most_common():
    print(f"    {sh:<18} {n:>4}件")

# ---------- Excel ----------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
wb = Workbook(); ws = wb.active; ws.title = "移動計画"
cols = list(rows[0].keys())
ws.append(cols)
for c in range(1, len(cols) + 1):
    ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
    ws.cell(1, c).fill = PatternFill("solid", fgColor="4472C4")
order = {"①物件フォルダへ寄せる": 0, "③書式・雛形へ": 1, "②仲介に残す": 2}
for r in sorted(rows, key=lambda x: (order[x["区分"]], x["確度"] != "要確認", x["案件フォルダ"])):
    ws.append([r[c] for c in cols])
for i, w in enumerate([22, 8, 40, 24, 26, 16, 12, 10, 78, 78, 20], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
yel = PatternFill("solid", fgColor="FFF2CC"); red = PatternFill("solid", fgColor="FCE4EC")
for i in range(2, ws.max_row + 1):
    if ws.cell(i, 2).value == "要確認":
        for c in range(1, len(cols) + 1): ws.cell(i, c).fill = yel
    if ws.cell(i, len(cols)).value:
        for c in range(1, len(cols) + 1): ws.cell(i, c).fill = red
out = os.path.expanduser(f"~/Desktop/整理計画_仲介から物件へ_{datetime.date.today():%Y%m%d}.xlsx")
wb.save(out)
print(f"\n書き出し: {out}  （{len(rows)}行）")
