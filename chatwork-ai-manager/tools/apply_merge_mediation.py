#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""整理案①の実行。計画表のうち「確度=高」かつ「注意なし」の行だけを移動する。
   --dry で下見、--go で実行。対応表を残すので戻せる。"""
import os, sys, json, shutil, datetime, unicodedata
from openpyxl import load_workbook

ROOT = "/Users/apple/Library/CloudStorage/Dropbox-大京商事　株式会社/共有フォルダ"
PLAN = os.path.expanduser(f"~/Desktop/整理計画_仲介から物件へ_{datetime.date.today():%Y%m%d}.xlsx")
REC  = os.path.expanduser("~/chatwork-ai-manager/local/")
GO = "--go" in sys.argv
N = lambda s: unicodedata.normalize("NFC", s)

ws = load_workbook(PLAN)["移動計画"]
cols = [c.value for c in ws[1]]
rows = [dict(zip(cols, [c.value for c in r])) for r in ws.iter_rows(min_row=2)]

targets = [r for r in rows
           if r["区分"] == "①物件フォルダへ寄せる" and r["確度"] == "高" and not r["注意"]]
print("=" * 74)
print(("【実行】" if GO else "【下見】") + f" 対象 {len(targets)}件 / 計画表 全{len(rows)}行")
print("=" * 74)

# 事前点検
ng = []
for r in targets:
    s = os.path.join(ROOT, N(r["現在のパス"]))
    d = os.path.join(ROOT, N(r["移動後のパス"]))
    if not os.path.exists(s): ng.append(("元が無い", r["現在のパス"]))
    elif os.path.exists(d):   ng.append(("先に同名あり", r["移動後のパス"]))
if ng:
    print(f"★点検で問題 {len(ng)}件。中止します。")
    for k, p in ng[:10]: print(f"   [{k}] {p}")
    sys.exit(1)
print("事前点検: 問題なし（元が全部あり、行き先に同名なし）\n")

import collections
byprop = collections.Counter(r["寄せ先の物件"] for r in targets)
byshelf = collections.Counter(r["棚"] for r in targets)
print("寄せ先の物件:")
for k, v in byprop.most_common(): print(f"   {k:<34} {v:>4}件")
print("\n棚:")
for k, v in byshelf.most_common(): print(f"   {k:<20} {v:>4}件")
sz = sum(r["容量(KB)"] for r in targets)
print(f"\n合計 {len(targets)}件 / {sz/1024:.0f}MB")

if not GO:
    print("\n※下見のみ。実行するには --go")
    sys.exit(0)

log = []
for r in targets:
    s = os.path.join(ROOT, N(r["現在のパス"]))
    d = os.path.join(ROOT, N(r["移動後のパス"]))
    os.makedirs(os.path.dirname(d), exist_ok=True)
    shutil.move(s, d)
    log.append({"from": r["現在のパス"], "to": r["移動後のパス"],
                "物件": r["寄せ先の物件"], "棚": r["棚"]})

# 空になった仲介の案件フォルダを片付ける（ファイルが1つも無いものだけ）
removed = []
MED = os.path.join(ROOT, "契約・書類/★仲介（賃貸・売買）")
for _ in range(6):                       # 内側から順に何度か回す
    for cur, dn, fn in os.walk(MED, topdown=False):
        if cur == MED: continue
        real = [f for f in fn if f != ".DS_Store" and not f.startswith("._")]
        if real or [d for d in dn]: continue
        ds = os.path.join(cur, ".DS_Store")
        if os.path.exists(ds): os.remove(ds)
        try:
            os.rmdir(cur); removed.append(N(os.path.relpath(cur, ROOT)))
        except OSError: pass

os.makedirs(REC, exist_ok=True)
rec = os.path.join(REC, f"移動記録_仲介から物件へ_{datetime.date.today():%Y%m%d}.json")
json.dump({"date": str(datetime.date.today()), "root": ROOT,
           "moved": log, "removed_empty_dirs": removed},
          open(rec, "w"), ensure_ascii=False, indent=2)
print(f"\n移動 {len(log)}件")
print(f"空になったフォルダを片付けた: {len(removed)}個")
for p in removed[:12]: print(f"   {p}")
print(f"\n対応表: {rec}")
print("戻すときは to → from へ move すればよい")
