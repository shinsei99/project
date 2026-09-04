#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""管理物件の最終整理計画。読むだけ。1件も動かさない。
   出力: ~/Desktop/整理計画_確定版_YYYYMMDD.xlsx

   ■ 物件フォルダの形
       {物件名}/契約/           賃貸借契約書・重説・覚書・登記・保険
       {物件名}/修繕・点検/      工事の見積・請求・報告・点検・清掃・設備
       {物件名}/図面・写真/
       {物件名}/お知らせ/        騒音・ごみ・停電・断水・検針などの通知
       {物件名}/その他/
       {物件名}/入居者/{部屋_名前}/   ← レントロールに載っている人
       {物件名}/解約済/{部屋_名前}/   ← 載っていない／名前に解約とある人
   ■ 棚が大きいときの分け方
       50件以下      そのまま
       50件超        「2021年以降」「2020年以前」に分ける
       2021年以降が100件超  その中だけ年ごと
"""
import os, re, unicodedata, collections, datetime, hashlib
from openpyxl import load_workbook

ROOT = "/Users/apple/Library/CloudStorage/Dropbox-大京商事　株式会社/共有フォルダ"
MP   = "物件・管理/管理物件"
RR   = ["★要更新★レントロール一覧（ビル）.xlsx",
        "★要更新★レントロール一覧（マンション）.xlsx",
        "★要更新★レントロール一覧（駐車場他）.xlsx"]
N = lambda s: unicodedata.normalize("NFC", s)

def norm(s):
    s = unicodedata.normalize("NFKC", str(s)).lower()
    s = re.sub(r"[\s　・･,，.。\-ー－_/\\'\"()（）\[\]【】]", "", s)
    s = re.sub(r"株式会社|有限会社|合同会社|医療法人|社会福祉法人|一般社団法人|㈱|㈲", "", s)
    for a, b in (("ⅰ","1"),("ⅱ","2"),("ⅲ","3"),("ii","2")): s = s.replace(a, b)
    return s
def names_of(s):
    s = unicodedata.normalize("NFKC", str(s))
    out = [s] + re.findall(r"[(（]([^)）]{2,30})[)）]", s) + [re.sub(r"[(（][^)）]*[)）]", "", s)]
    return [norm(x) for x in out if norm(x)]
def pstem(s):
    n = norm(s)
    for suf in ("ビル","bldg","マンション","駐車場","モータープール","駐輪場"):
        if n.endswith(suf) and len(n) > len(suf)+2: n = n[:-len(suf)]
    return n

# ---------- レントロール（現契約の正典） ----------
# ★シートごとに列の並びが違う。見出し行から「契約者」「現況」の位置を探す。
#   ビル      : 号室 | 契約者 | 家賃 …
#   マンション : 号室 | 現況 | 契約者 | 区分 …
#   駐車場    : 区画No | 現況 | 契約者 | 区分 …
rr = {}
for f in RR:
    p = os.path.join(ROOT, f)
    if not os.path.exists(p): continue
    wb = load_workbook(p, data_only=True, read_only=True)
    for sn in wb.sheetnames:
        data = [[("" if c is None else str(c).strip()) for c in row] for row in wb[sn].iter_rows(values_only=True)]
        ci = si = ri = None
        for row in data[:6]:
            for j, c in enumerate(row):
                if c == "契約者": ci = j
                elif c == "現況": si = j
                elif c in ("号室","号室/フロア","区画No","区画"): ri = j
            if ci is not None: break
        if ci is None: continue
        tbl = {}
        for row in data:
            if len(row) <= ci: continue
            t = row[ci]; rm = row[ri] if (ri is not None and len(row) > ri) else ""
            if not t or t in ("契約者","合計","計","空室","空き","-"): continue
            if si is not None and len(row) > si and row[si] in ("空室","空き","解約","退去"): continue
            for nm in names_of(t): tbl.setdefault(nm, rm or t)
        if tbl: rr[pstem(sn)] = tbl
    wb.close()
print(f"レントロール: {len(rr)}物件 / 契約者 {sum(len(v) for v in rr.values())}件（別名込み）\n")

# ---------- 棚 ----------
SHELVES = [
    ("解約・精算", r"解約|退去|明渡|原状回復|敷金返還|精算書|清算書|精算明細|清算明細"),
    ("契約",      r"契約書|契約証書|重説|重要事項|覚書|合意書|更新|定期借家|保証委託|入居申込|申込書|審査|媒介|念書|承諾|民法改正|特殊詐欺|鍵受領|鍵預|登記|謄本|全部事項|評価証明|保険|証券|抵当"),
    ("修繕・点検", r"工事|修繕|リフォーム|リホーム|見積|施工|点検|保守|清掃|設備|昇降|エレベータ|ＥＶ|EV|消防|防火|避難|廃棄物|貯水槽|受水槽|給水|排水|空調|請求|領収|明細|報告書|作業"),
    ("図面・写真", r"図面|竣工|平面|間取|配置|測量|公図|写真|マイソク|パース|看板|広告物|サイン|地図|矩計|DSCN|DSC\d|IMG[_\-]?\d|P\d{7}"),
    ("お知らせ",   r"お知らせ|おしらせ|通知|案内|貼紙|掲示|連絡|検針|メーター|水道|電気|ガス|使用量|騒音|ごみ|ゴミ|停電|断水|アンケート"),
]
OLD_SHELF = re.compile(r"^(0[1-9])_|^(契約|修繕・点検|図面・写真|お知らせ|その他|入居者|解約済)$")
ENTITY = re.compile(r"^\d+番|^\d+号|^[A-Z]?\d{3,4}(号|室|_)|^\d+[FＦ]|^\d+階|^[0-9]{3}_|^\d+[FＦ]【|"
                    r"（\d{4}[.\-/]\d{1,2}|～）|^\d{4}[.\-]\d{2}[.\-]\d{2}_")
KAIYAKU = re.compile(r"解約|退去|明渡|不成立|終了")

def shelf(below, fname):
    for n_, pat in SHELVES:
        if re.search(pat, below): return n_, "いまのフォルダ名"
    for n_, pat in SHELVES:
        if re.search(pat, fname): return n_, "ファイル名"
    return "", "★手がかりなし＝動かさない"

# ---------- 収集 ----------
files = []
for cur, dn, fn in os.walk(os.path.join(ROOT, MP)):
    rel = N(os.path.relpath(cur, ROOT))
    for f in fn:
        f = N(f)
        if f == ".DS_Store" or f.startswith("._"): continue
        files.append(rel + "/" + f)

# 実体（人の箱）の判定
sib = collections.defaultdict(lambda: collections.defaultdict(set))
for p in files:
    parts = p.split("/")
    if len(parts) < 6: continue
    for i in range(4, len(parts)-1):
        sib["/".join(parts[:i])][parts[i]].add(parts[-1])
entity = set()
for parent, kids in sib.items():
    ks = list(kids)
    for i, a in enumerate(ks):
        if ENTITY.search(a): entity.add((parent, a)); continue
        for b in ks[i+1:]:
            if kids[a] & kids[b]: entity.add((parent, a)); entity.add((parent, b))

def mtime(p):
    try: return os.path.getmtime(os.path.join(ROOT, p))
    except OSError: return 0
def year(p):
    t = mtime(p)
    return datetime.datetime.fromtimestamp(t).year if t else 0

# ---------- 1回目: 棚と人の箱を決める ----------
tmp = []
for p in files:
    parts = p.split("/")
    if len(parts) < 5:
        tmp.append({"物件":"", "種別":"", "棚":"", "人の箱":"", "決め手":"", "根拠":"",
                    "現在のパス":p, "注意":"★物件フォルダの外にある"}); continue
    kind, prop = parts[2], parts[3]
    fname = parts[-1]; below = "/".join(parts[4:-1])
    keep = None
    for i in range(4, len(parts)-1):
        if ("/".join(parts[:i]), parts[i]) in entity: keep = parts[i]
    if keep:
        table = None
        k = pstem(prop)
        cand = [kk for kk in rr if kk and (kk in k or k in kk) and min(len(kk), len(k)) >= 3]
        if cand: table = rr[max(cand, key=len)]
        if KAIYAKU.search(keep) or KAIYAKU.search(below):
            box, why = "解約済", "名前に解約とある"
        elif table is None:
            box, why = "", "★レントロールに物件が無い＝動かさない"
        else:
            cands = names_of(re.sub(r"^[0-9A-Za-zＦF【】\[\]_番号室階\-]+", "", keep)) + names_of(keep)
            hit = [t for t in table if t and len(t) >= 3
                   and any(t in c or c.endswith(t) or t.endswith(c) for c in cands)]
            box, why = ("入居者", f"レントロール {table[max(hit,key=len)]} と一致") if hit \
                       else ("解約済", "レントロールに載っていない")
        tmp.append({"物件":prop, "種別":kind, "棚":box, "人の箱":keep, "決め手":"実体フォルダ",
                    "根拠":why, "現在のパス":p, "注意":""})
    else:
        sh, why = shelf(below, fname)
        tmp.append({"物件":prop, "種別":kind, "棚":sh, "人の箱":"", "決め手":why,
                    "根拠":"", "現在のパス":p, "注意":""})

# ---------- 2回目: 大きい棚を5年区切りで割る ----------
def bucket(y):
    if y == 0 or y <= 2020: return "2020年以前"
    b = 2021 + ((y - 2021) // 5) * 5
    return f"{b}〜{b+4}年"

grp = collections.defaultdict(list)
for r in tmp:
    if r["物件"] and r["棚"] and not r["人の箱"]: grp[(r["物件"], r["種別"], r["棚"])].append(r)
for (prop, kind, sh), v in grp.items():
    if len(v) <= 50:
        for r in v: r["時期"] = ""
    else:
        for r in v: r["時期"] = bucket(year(r["現在のパス"]))

rows = []
for r in tmp:
    if not r["物件"] or not r["棚"]:
        r["移動後のパス"] = r["現在のパス"]; r["時期"] = r.get("時期","")
        if not r["注意"]: r["注意"] = "現状維持（判定できず）"
    elif r["人の箱"]:
        r["時期"] = ""
        r["移動後のパス"] = f"{MP}/{r['種別']}/{r['物件']}/{r['棚']}/{r['人の箱']}/{os.path.basename(r['現在のパス'])}"
    else:
        seg = f"/{r['時期']}" if r.get("時期") else ""
        r["移動後のパス"] = f"{MP}/{r['種別']}/{r['物件']}/{r['棚']}{seg}/{os.path.basename(r['現在のパス'])}"
    r["容量(KB)"] = round(os.path.getsize(os.path.join(ROOT, r["現在のパス"]))/1024, 1)
    rows.append(r)

# ---------- 衝突 ----------
dest = collections.Counter(r["移動後のパス"] for r in rows)
col = {k for k, v in dest.items() if v > 1}
def h(p):
    try:
        with open(os.path.join(ROOT, p), "rb") as f: return hashlib.sha256(f.read()).hexdigest()
    except OSError: return None
by = collections.defaultdict(list)
for r in rows:
    if r["移動後のパス"] in col: by[r["移動後のパス"]].append(r)
for d, rs in by.items():
    tag = "★中身が違う＝要個別対応" if len({h(r["現在のパス"]) for r in rs}) > 1 else "重複（中身は同じ）"
    for r in rs: r["注意"] = tag

# ---------- 集計 ----------
tgt = [r for r in rows if r["物件"]]
print("=" * 76); print("整理計画（確定版）── 計画のみ。1件も動かしていない"); print("=" * 76)
print(f"  対象 {len(tgt):,}件 / {len({r['物件'] for r in tgt})}物件\n")
stay=[r for r in tgt if not r["棚"]]
print(f"  ★現状維持（判定できず・動かさない） {len(stay):,}件\n")
print("  行き先:")
for k, v in collections.Counter(r["棚"] for r in tgt if r["棚"]).most_common():
    print(f"    {k:<12}{v:>6,}件")
print("\n  棚の決め手:")
for k, v in collections.Counter(r["決め手"] for r in tgt).most_common():
    print(f"    {k:<16}{v:>6,}件")
print("\n  人の箱:")
pb = [r for r in tgt if r["人の箱"]]
print(f"    入居者 {sum(1 for r in pb if r['棚']=='入居者'):>5,}件 / "
      f"{len({(r['物件'],r['人の箱']) for r in pb if r['棚']=='入居者'})}箱")
print(f"    解約済 {sum(1 for r in pb if r['棚']=='解約済'):>5,}件 / "
      f"{len({(r['物件'],r['人の箱']) for r in pb if r['棚']=='解約済'})}箱")
dd = collections.Counter(len(r["移動後のパス"].split("/"))-1 for r in tgt if r["棚"])
print("\n  移動後の深さ（新共有フォルダから）:")
for k in sorted(dd): print(f"    {k}階層 {dd[k]:>6,}件")
print(f"\n  時期で分けた棚: {len({(r['物件'],r['棚']) for r in tgt if r.get('時期')})}個")
print(f"  ★注意が要る行 {sum(1 for r in rows if r['注意']):,}件")

# ---------- Excel ----------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
COLS = ["物件","種別","棚","人の箱","時期","決め手","根拠","容量(KB)","現在のパス","移動後のパス","注意"]
wb = Workbook(); ws = wb.active; ws.title = "整理計画"
ws.append(COLS)
for c in range(1, len(COLS)+1):
    ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
    ws.cell(1, c).fill = PatternFill("solid", fgColor="4472C4")
for r in sorted(rows, key=lambda x: (not x["注意"], x["決め手"] != "★手がかりなし", x["物件"], x["棚"])):
    ws.append([r.get(c, "") for c in COLS])
for i, w in enumerate([26,12,12,30,18,16,30,10,80,80,24], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
red = PatternFill("solid", fgColor="FCE4EC"); yel = PatternFill("solid", fgColor="FFF2CC")
for i in range(2, ws.max_row+1):
    if ws.cell(i, 11).value:
        for c in range(1, len(COLS)+1): ws.cell(i, c).fill = red
    elif ws.cell(i, 6).value == "★手がかりなし":
        for c in range(1, len(COLS)+1): ws.cell(i, c).fill = yel

ws2 = wb.create_sheet("物件ごと")
ws2.append(["物件","種別","件数","契約","修繕・点検","図面・写真","お知らせ","その他","入居者","解約済","最大階層"])
for c in range(1, 12):
    ws2.cell(1, c).font = Font(bold=True, color="FFFFFF")
    ws2.cell(1, c).fill = PatternFill("solid", fgColor="4472C4")
g2 = collections.defaultdict(list)
for r in tgt: g2[(r["物件"], r["種別"])].append(r)
for (nm, kind), rs in sorted(g2.items(), key=lambda x: -len(x[1])):
    c = collections.Counter(r["棚"] for r in rs)
    ws2.append([nm, kind, len(rs), c["契約"], c["修繕・点検"], c["図面・写真"], c["お知らせ"],
                c["その他"], c["入居者"], c["解約済"],
                max(len(r["移動後のパス"].split("/"))-1 for r in rs)])
for i, w in enumerate([30,12,8,8,12,12,10,8,10,10,10], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"; ws2.auto_filter.ref = ws2.dimensions

out = os.path.expanduser(f"~/Desktop/整理計画_確定版_{datetime.date.today():%Y%m%d}.xlsx")
wb.save(out)
print(f"\n書き出し: {out}  （{len(rows):,}行 ＋ 物件ごと{len(g2)}行）")
