#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""整理計画 v2 ── 「物件フォルダの直下にあるもの」だけを仕分ける。
   ★1ファイルしか入っていないフォルダを畳むときの例外:
     入居者/ 解約・精算/ の下のテナントの箱は、中が1件でも畳まない。
     「誰の書類か」を表す箱なので、畳むと持ち主が分からなくなる。
   ★フォルダはバラさない。まるごと棚の下へ移す（中身は一切触らない）。
   読むだけ。1件も動かさない。

   {物件名}/契約/            {物件名}/修繕・点検/   {物件名}/記録・写真/
   {物件名}/お知らせ/
   {物件名}/入居者/{テナント}/ {物件名}/解約・精算/{テナント}/
"""
import os, re, sys, unicodedata, collections, datetime
from openpyxl import load_workbook

ROOT = "/Users/apple/Library/CloudStorage/Dropbox-大京商事　株式会社/共有フォルダ"
MP   = "物件・管理/管理物件"
RR   = ["★要更新★レントロール一覧（ビル）.xlsx",
        "★要更新★レントロール一覧（マンション）.xlsx",
        "★要更新★レントロール一覧（駐車場他）.xlsx"]
N = lambda s: unicodedata.normalize("NFC", s)

def norm(s):
    s = unicodedata.normalize("NFKC", str(s)).lower()
    s = re.sub(r"[\s　・･,，.。\-ー－_/\\'\"()（）\[\]【】]", "", s)
    s = re.sub(r"株式会社|有限会社|合同会社|医療法人|社会福祉法人|一般社団法人|㈱|㈲", "", s)
    for a, b in (("ⅰ","1"),("ⅱ","2"),("ⅲ","3"),("ii","2")): s = s.replace(a, b)
    return s
def names_of(s):
    s = unicodedata.normalize("NFKC", str(s))
    return [norm(x) for x in ([s] + re.findall(r"[(（]([^)）]{2,30})[)）]", s)
                              + [re.sub(r"[(（][^)）]*[)）]", "", s)]) if norm(x)]
def pstem(s):
    n = norm(s)
    for suf in ("ビル","bldg","マンション","駐車場","モータープール","駐輪場"):
        if n.endswith(suf) and len(n) > len(suf)+2: n = n[:-len(suf)]
    return n

# ---------- レントロール ----------
rr = {}
for f in RR:
    p = os.path.join(ROOT, f)
    if not os.path.exists(p): continue
    wb = load_workbook(p, data_only=True, read_only=True)
    for sn in wb.sheetnames:
        data = [[("" if c is None else str(c).strip()) for c in row] for row in wb[sn].iter_rows(values_only=True)]
        ci = si = ri = None
        for row in data[:6]:
            for j, c in enumerate(row):
                if c == "契約者": ci = j
                elif c == "現況": si = j
                elif c in ("号室","号室/フロア","区画No","区画"): ri = j
            if ci is not None: break
        if ci is None: continue
        tbl = {}
        for row in data:
            if len(row) <= ci: continue
            t = row[ci]; rm = row[ri] if (ri is not None and len(row) > ri) else ""
            if not t or t in ("契約者","合計","計","空室","空き","-"): continue
            if si is not None and len(row) > si and row[si] in ("空室","空き","解約","退去"): continue
            for nm in names_of(t): tbl.setdefault(nm, rm or t)
        if tbl: rr[pstem(sn)] = tbl
    wb.close()

# ---------- 物件マスタ（サブ物件の判定に使う） ----------
# ★物件フォルダの中に、マスタに独立登録されている別物件が入っていることがある
#   （例: その他物件/隆生福祉会（ゆめ）/ の下に ゆめあまみ・ゆめ中央保育園 など4施設）
#   法人でまとめる形は正しいので、そのまま残して「サブ物件」として扱う。
#   棚は各施設の中に作る（＝サブ物件フォルダ自体は動かさない）。
import sqlite3
try:
    _con = sqlite3.connect("/Users/apple/chatwork-ai-manager/data/app.db")
    MASTER = {norm(r[0]) for r in _con.execute("SELECT name FROM properties WHERE active=1")}
    MASTER |= {norm(r[0]) for r in _con.execute("SELECT name FROM shinsei_properties WHERE active=1")}
    _con.close()
except Exception:
    MASTER = set()

# ---------- 棚（上から順に当てる＝優先順位） ----------
SHELVES = [
    # ★「入居者各位」「〜のお願い」は宛名であって入居者の書類ではない。通知を先に当てる
    # ★「解約通知書」は解約、「ご使用水量等のお知らせ」は検針。先に当たる棚があるので
    #   通知・案内は「〜のお知らせ/通知/案内」で終わる形と、掲示物に限る（2026-09-05）
    ("通知・案内", r"各位|のお願い|お願いします|しております|警告|注意喚起|"
                   r"のお知らせ|のおしらせ|のご案内|貼紙|掲示|回覧|チラシ|ﾁﾗｼ|ビラ|ﾋﾞﾗ|配布|周知|"
                   r"^お知らせ|^通知|^案内"),
    ("解約・精算",  r"解約|退去|明渡|原状回復|現状回復|敷金返還|精算|清算"),
    ("入居者",     r"賃借人資料|入居者資料|テナント資料"),
    # ★掲示物（立入禁止の貼り紙）は通知・案内。鍵の受渡しは その他。引き継ぎ資料は 物件基本
    ("通知・案内",  r"立入禁止|立ち入り禁止|関係者以外"),
    ("その他",     r"鍵受領|鍵預り|鍵預かり|鍵の受領|鍵一覧|鍵番号|"
                   r"入金一覧|収納明細|家賃収納|連絡先|検針|メーター|水道|ガス|使用量|騒音|停電|断水|アンケート|郵便受|ポスト|暗証|解錠|メールボックス|メールBOX|ﾒｰﾙBOX|宅配BOX|宅配ボックス|駐輪|シール|ｼｰﾙ|オートロック|緊急|南京錠|ダイヤル錠|錠番号|鍵貸与|貸与|ATBB|支援サイト|レインズ|ﾚｲﾝｽﾞ|預り証|預かり証|受領書|受領証|連絡網"),
    # ★(?i) を付ける。「溝蓋.JPEG」のように拡張子が大文字だと拾えていなかった（2026-09-05）
    ("記録・写真", r"(?i)写真|魚眼|\.jpe?g$|\.png$|\.heic$|\.tiff?$|DSCN|DSC\d|IMG[_\-]?\d|P\d{7}"),
    ("修繕・点検", r"工事|修繕|リフォーム|リホーム|見積|施工|点検|検査|保守|清掃|設備|昇降|エレベータ|ＥＶ|消防|防火|避難|廃棄物|ゴミ|ごみ|貯水槽|受水槽|給水|排水|空調|請求|領収|報告|作業|統括|訓練|自火報|産業廃棄物|電気|保証書|エアコン|給湯|漏水|雨漏|塗装|防水|管理月報|月報|ビルメンテナンス|メンテナンス|ビルシステム|昇降機|消毒|防除|ベルポ|ビケンテクノ|日立|東洋|関電|大阪ガス|消火器|誘導灯|自動火災|感知器|受信機|設置届|天井|トイレ|溝蓋|蛇口|網戸|建具|不具合|故障"),
    # 物件基本 = その物件の土台。管理委託・登記・媒介・保険 ＋ その物件用のひな形・書式
    ("物件基本",   r"引継ぎ|引き継ぎ|引継書|引継内容|管理委託|業務委託|サブリース|管理契約|媒介|登記|謄本|全部事項|評価証明|地位承継|保険|証券|抵当|オーナー|所有者|雛型|雛形|ひな形|ひな型|書式|様式|テンプレート|催告書|チェックリスト|契約書の中身|表紙|竣工|図面|平面|矩計|配置図|測量|公図|求積|立面|断面|間取|マイソク|パース|地図|所在図|車庫証|保管場所|使用承諾|台帳|契約車両|駐車場契約書|駐車場申込書|月極駐車場申込書|承諾証明|委任状|封筒|送付状|送付書|入退館|手順|マニュアル|看板|募集|ポップ|ﾎﾟｯﾌﾟ|配置図面|一括契約|J-?COM|ＪＣＯＭ|ジェイコム|管理規約|管理組合|管理集会|議案書|定期利用契約書|契約書フォーム|契約書ﾌｫｰﾑ|駐車場賃貸借契約書|覚書|念書|合意書|確認書|承諾書|賃料表|価格明細|契約NO|契約No|契約ＮＯ|連絡網|契約者リスト|契約者ﾘｽﾄ|入居者リスト|入居者ﾘｽﾄ|テナントリスト|ﾃﾅﾝﾄﾘｽﾄ|契約者一覧"),
]
ENTITY = re.compile(r"^\d+番|^\d+号|^[A-Z]?\d{3,4}(号|室|_)|^\d+[FＦ]|^\d+階|^[0-9]{3}_|^\d+[FＦ]【|"
                    r"^[BＢ]?\d*[FＦ]?\d*[-ー_]?[ABＡＢ]?_|"          # 7F_ 6F_ B1F_ 4-A_ など
                    r"^\d+\s*[-ー~〜]\s*\d+\s*[FＦ階]|"              # ★2-6F（階の範囲）2026-09-05
                    r"（\d{4}[.\-/]\d{1,2}|～）|^\d{4}[.\-]\d{2}[.\-]\d{2}_")
KAIYAKU = re.compile(r"解約|退去|明渡|不成立|終了")
SHELFNAME = re.compile(r"^(0[1-9]_)?(物件基本|契約|修繕・点検|記録・写真|通知・案内|その他|お知らせ|入居者|解約・精算)$")

# ★「作成中」でも本当に作業中とは限らない（2026-09-05に間違えた）
#   「天王寺書類作成中」の中身は **2009年（平成21年）の引き継ぎ届出書類**だった。
#   16年前のものを「作業中だから触らない」と避けていた。
#   → 名前だけで決めず、**中の更新日が新しいものだけ**を作業中とみなす。
WIP = re.compile(r"作成中|作業中|進行中|途中|下書き|未整理|新規作成")

def _recently_touched(path, months=6):
    """中に半年以内に更新されたファイルがあるか（本当に作業中かの判定）。"""
    import time
    cut = time.time() - months * 30 * 86400
    try:
        for c, d, f in os.walk(path):
            for x in f:
                if x == ".DS_Store": continue
                try:
                    if os.path.getmtime(os.path.join(c, x)) >= cut: return True
                except OSError: pass
    except OSError: pass
    return False

def shelf_of(name, inner, path=None):
    if WIP.search(name) and path and os.path.isdir(path) and _recently_touched(path):
        return "", "★作業中の箱（最近も更新あり）＝動かさない"
    """name=直下の名前、inner=その中のファイル名を連ねたもの"""
    for sh, pat in SHELVES:
        if re.search(pat, name): return sh, "名前"
    for sh, pat in SHELVES:
        if inner and re.search(pat, inner): return sh, "中身のファイル名"
    return "", "★判定できず"

def box_for_name(pp, fname):
    """物件フォルダ pp の既にあるテナントの箱のうち、名前が fname に出てくるものを返す。

    ★「ベネッセ【天王寺】出店確認事項一覧シート.xlsx」が物件フォルダの直下に転がっていた。
      解約・精算に「9F_ベネッセ」の箱があるのだから、そこへ入るのが素直（2026-09-05）。
      箱の名前から区画の見出し（7-8F_ など）を落として、残った借主名で照合する。
    """
    t = norm(fname)
    if not t: return None
    best = None
    for sh in ("入居者", "解約・精算"):
        sp = os.path.join(pp, sh)
        if not os.path.isdir(sp): continue
        for b in os.listdir(sp):
            if not os.path.isdir(os.path.join(sp, b)): continue
            b = N(b)
            nb = norm(re.sub(r"^[^_]*_", "", b))
            if len(nb) >= 3 and nb in t and (best is None or len(nb) > len(best[2])):
                best = (sh, b, nb)
    return (best[0], best[1]) if best else None


def count(p):
    n = 0
    for c, d, f in os.walk(p): n += sum(1 for x in f if x != ".DS_Store" and not x.startswith("._"))
    return n
def inner_names(p, limit=40):
    out = []
    for c, d, f in os.walk(p):
        for x in f:
            if x == ".DS_Store" or x.startswith("._"): continue
            out.append(N(x))
            if len(out) >= limit: return " ".join(out)
    return " ".join(out)

only = [a for a in sys.argv[1:] if not a.startswith("--")]
JSON = "--json" in sys.argv
rows = []
for kind in sorted(os.listdir(os.path.join(ROOT, MP))):
    kp = os.path.join(ROOT, MP, kind)
    # ★「_旧管理物件」は、いま管理も仲介もしていない物件の置き場（2026-09-05）。
    #   現役の物件と混ざらないよう、整理の対象から外す。
    if not os.path.isdir(kp) or kind.startswith("_"): continue
    for prop in sorted(os.listdir(kp)):
        pp = os.path.join(kp, prop)
        if not os.path.isdir(pp): continue
        if only and N(prop) not in only: continue
        # ★20件以下の物件は棚を作らない（1〜2画面で全部見えるので、
        #   棚を7つ作ると1棚1件になってかえって探しにくい）。
        #   ただし3階層以上あるものは中身が埋もれているので対象に残す。
        if not only:
            _n=0; _mx=0
            for _c,_d,_f in os.walk(pp):
                _rel=os.path.relpath(_c,pp)
                _lv=0 if _rel=="." else len(_rel.split(os.sep))
                for _x in _f:
                    if _x==".DS_Store" or _x.startswith("._"): continue
                    _n+=1; _mx=max(_mx,_lv)
            if _n<=20 and _mx<3:
                continue
        k = pstem(prop)
        cand = [kk for kk in rr if kk and (kk in k or k in kk) and min(len(kk), len(k)) >= 3]
        table = rr[max(cand, key=len)] if cand else None
        for item in sorted(os.listdir(pp)):
            if item in (".DS_Store",) or item.startswith("._"): continue
            ip = os.path.join(pp, item); item = N(item)
            isdir = os.path.isdir(ip)
            if isdir and SHELFNAME.match(item):
                continue                                    # すでに棚
            # ★サブ物件の判定
            #   フォルダ名「そのもの」が物件名のときだけ。括弧つきの補足は許すが、
            #   「賃借人資料（エレガンス放出）」のように括弧の中だけ一致するものは除く
            #   （前置きが付いている＝その物件の書類であって、物件そのものではない）
            _bare = norm(re.sub(r"[（(][^）)]*[）)]\s*$", "", item))   # 末尾の括弧だけ落とす
            _cands = {norm(item), _bare} - {""}
            if isdir and (_cands & MASTER) and norm(item) != norm(prop) \
               and not re.match(r"^(賃借人資料|入居者|テナント資料|解約済|解約・精算)", item):
                rows.append({"物件": N(prop), "種別": N(kind), "種類": "フォルダ",
                             "名前": item, "中の件数": count(ip),
                             "行き先": "（動かさない）",
                             "決め手": "★サブ物件（マスタに独立登録あり）。中で棚分けする",
                             "現在のパス": f"{MP}/{kind}/{prop}/{item}",
                             "移動後のパス": f"{MP}/{kind}/{prop}/{item}"})
                continue
            n = count(ip) if isdir else 1
            # テナントの箱か？
            if isdir and (ENTITY.search(item) or KAIYAKU.search(item)
                          or re.match(r"^(賃借人資料|入居者|テナント資料)", item)):
                if item.startswith("解約済"):      # 既存フォルダは「解約済（物件名）」の名前
                    sh, why, box = "解約・精算", "この箱ごと『解約・精算』にする（中身そのまま）", "＊中身をそのまま"
                elif re.match(r"^(賃借人資料|入居者|テナント資料)", item):
                    sh, why, box = "入居者", "この箱ごと『入居者』にする（中身そのまま）", "＊中身をそのまま"
                elif KAIYAKU.search(item):
                    sh, why, box = "解約・精算", "名前に解約とある", item
                elif table is None:
                    # ★レントロールが無くても、名前の形がテナントの箱なら「入居者」へ入れる。
                    #   現/旧の判別はできないので、そこは人が見る前提（動かさないより良い）。
                    sh, why, box = "入居者", "レントロールに物件が無い（現/旧は要確認）", item
                else:
                    c2 = names_of(re.sub(r"^[0-9A-Za-zＦF【】\[\]_番号室階\-]+", "", item)) + names_of(item)
                    hit = [t for t in table if t and len(t) >= 3
                           and any(t in c or c.endswith(t) or t.endswith(c) for c in c2)]
                    if hit: sh, why, box = "入居者", f"レントロール {table[max(hit,key=len)]} と一致", item
                    else:   sh, why, box = "解約・精算", "レントロールに載っていない", item
            else:
                # ★「誰のものか」で決めるのが本筋。既にテナントの箱があるなら、
                #   名前が一致する書類はその箱へ入れる（棚のルールより先に見る）。
                b = box_for_name(pp, item)
                if b:
                    sh, box = b
                    why = f"既にある「{box}」の書類"
                else:
                    sh, why = shelf_of(item, inner_names(ip) if isdir else "", ip)
                    box = ""
            new = f"{MP}/{kind}/{prop}/{sh}/" + (f"{box}/" if box else "") + (item if not box else "") \
                  if sh else ""
            if sh and box: new = f"{MP}/{kind}/{prop}/{sh}/{box}"
            elif sh:       new = f"{MP}/{kind}/{prop}/{sh}/{item}"
            rows.append({"物件": N(prop), "種別": N(kind), "種類": "フォルダ" if isdir else "ファイル",
                         "名前": item, "中の件数": n, "行き先": sh or "（動かさない）",
                         "決め手": why,
                         "現在のパス": f"{MP}/{kind}/{prop}/{item}",
                         "移動後のパス": new or f"{MP}/{kind}/{prop}/{item}"})

if JSON:
    import json
    print(json.dumps(rows, ensure_ascii=False))
    sys.exit(0)

tot = sum(r["中の件数"] for r in rows)
print("=" * 74); print("整理計画 v2 ── 直下だけ仕分ける／フォルダはバラさない"); print("=" * 74)
print(f"  直下の項目 {len(rows)}個（中身 合計{tot:,}件）/ {len({r['物件'] for r in rows})}物件\n")
print("  行き先（項目数 / 中の件数）:")
c1 = collections.Counter(r["行き先"] for r in rows)
c2 = collections.Counter()
for r in rows: c2[r["行き先"]] += r["中の件数"]
for k, v in c1.most_common(): print(f"    {k:<14}{v:>5}個{c2[k]:>8,}件")
print("\n  決め手:")
for k, v in collections.Counter(r["決め手"] for r in rows).most_common():
    print(f"    {k:<28}{v:>5}個")

if only:
    print(f"\n=== {only[0]} の内訳 ===")
    for r in sorted(rows, key=lambda x: (x["行き先"], x["名前"])):
        print(f"  {r['行き先']:<12} ← [{r['種類']}] {r['名前'][:44]:<44} {r['中の件数']:>3}件  ({r['決め手']})")
else:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "整理計画v2"
    COLS = list(rows[0].keys()); ws.append(COLS)
    for c in range(1, len(COLS)+1):
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, c).fill = PatternFill("solid", fgColor="4472C4")
    for r in sorted(rows, key=lambda x: (x["行き先"] != "（動かさない）", x["物件"], x["行き先"])):
        ws.append([r[c] for c in COLS])
    for i, w in enumerate([26,12,10,46,10,14,28,74,74], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    yel = PatternFill("solid", fgColor="FFF2CC")
    for i in range(2, ws.max_row+1):
        if str(ws.cell(i, 7).value).startswith("★"):
            for c in range(1, len(COLS)+1): ws.cell(i, c).fill = yel
    out = os.path.expanduser(f"~/Desktop/整理計画v2_{datetime.date.today():%Y%m%d}.xlsx")
    wb.save(out); print(f"\n書き出し: {out}  （{len(rows)}行）")
