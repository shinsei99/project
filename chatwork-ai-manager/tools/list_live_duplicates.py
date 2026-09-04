#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""現役（アーカイブ除く）の完全重複を一覧にする。読むだけ。1件も消さない。

出力: ~/Desktop/重複一覧_現役_YYYYMMDD.xlsx
  シート1「グループ」… 1組1行。残す候補・消す候補・容量・判定
  シート2「明細」    … 1ファイル1行。残す/消す の別つき
"""
import os, sys, collections, unicodedata, hashlib, json, datetime, re

ROOT = "/Users/apple/Library/CloudStorage/Dropbox-大京商事　株式会社/共有フォルダ/（★必読★）新共有フォルダ"
ARC = unicodedata.normalize("NFC", "_アーカイブ（2027年7月削除予定）")
N = lambda s: unicodedata.normalize("NFC", s)
CACHE = os.path.expanduser("~/chatwork-ai-manager/local/dup_hash_cache.json")

# ---------- 収集 ----------
files, size, mtime = [], {}, {}
for cur, dn, fn in os.walk(ROOT):
    rel = N(os.path.relpath(cur, ROOT))
    if rel.startswith(ARC):
        continue                                  # ★アーカイブは最初から見ない
    for f in fn:
        f = N(f)
        if f in {".DS_Store", "desktop.ini"} or f.startswith("._"):
            continue
        p = (rel + "/" + f) if rel != "." else f
        full = os.path.join(ROOT, p)
        try:
            st = os.stat(full)
        except OSError:
            continue
        files.append(p); size[p] = st.st_size; mtime[p] = st.st_mtime

print(f"現役ファイル {len(files):,}件を照合します…")

cache = {}
os.makedirs(os.path.dirname(CACHE), exist_ok=True)
if os.path.exists(CACHE):
    try: cache = json.load(open(CACHE))
    except Exception: cache = {}
def h(p):
    key = f"{p}|{size[p]}|{int(mtime[p])}"
    if key in cache: return cache[key]
    try:
        with open(os.path.join(ROOT, p), "rb") as f:
            d = hashlib.sha256(f.read()).hexdigest()
    except OSError:
        d = None
    cache[key] = d
    return d

bysize = collections.defaultdict(list)
for p in files:
    if size[p] > 0: bysize[size[p]].append(p)
groups = collections.defaultdict(list)
for sz, ps in bysize.items():
    if len(ps) < 2: continue
    for p in ps:
        d = h(p)
        if d: groups[(sz, d)].append(p)
json.dump(cache, open(CACHE, "w"))
dups = {k: v for k, v in groups.items() if len(v) > 1}

# ---------- 残す1本を決める ----------
# 規則: ①階層が浅いものを残す ②同じなら更新が新しい ③同じならパス名順
def keep_of(paths):
    return sorted(paths, key=lambda p: (len(p.split("/")), -mtime[p], p))[0]

# ---------- 判定 ----------
ENT = re.compile(r"\d+番|\d+号|^\d{3}_|（\d{4}[.\-/]\d{1,2}|～）")
def verdict(paths):
    props = {"/".join(p.split("/")[:4]) for p in paths}
    ents  = any(ENT.search(seg) for p in paths for seg in p.split("/")[:-1])
    waste = size[paths[0]] * (len(paths) - 1)
    if len(props) == 1 and ents:
        return "★消さない方がよい（区画・部屋ごとの控え）"
    if waste >= 10 * 1024**2:
        return "★消してよい（容量が大きい・置き場の二重化）"
    if len({os.path.dirname(p).split("/")[0] for p in paths}) == 1 and len(props) == 1:
        return "要確認（同じ物件の中の重複）"
    return "消してよい（別フォルダに同じもの）"

rows_g, rows_d = [], []
for k, paths in sorted(dups.items(), key=lambda x: -(x[0][0] * (len(x[1]) - 1))):
    sz = k[0]; keep = keep_of(paths); v = verdict(paths)
    waste = sz * (len(paths) - 1)
    rows_g.append({
        "本数": len(paths),
        "1本の容量(MB)": round(sz / 1024**2, 3),
        "消せる容量(MB)": round(waste / 1024**2, 3),
        "判定": v,
        "ファイル名": os.path.basename(keep),
        "残す候補": keep,
        "消す候補": " ｜ ".join(p for p in sorted(paths) if p != keep),
    })
    for p in sorted(paths):
        rows_d.append({
            "グループ": len(rows_g),
            "残す/消す": "残す" if p == keep else "消す",
            "判定": v,
            "容量(MB)": round(size[p] / 1024**2, 3),
            "更新日": datetime.datetime.fromtimestamp(mtime[p]).strftime("%Y-%m-%d"),
            "階層": len(p.split("/")) - 1,
            "パス": p,
        })

# ---------- 集計を画面へ ----------
tot_extra = sum(r["本数"] - 1 for r in rows_g)
tot_waste = sum(r["消せる容量(MB)"] for r in rows_g)
print("=" * 72)
print(f"現役の完全重複: {len(rows_g):,}組 / 余分 {tot_extra:,}件 / {tot_waste:,.0f}MB")
print("=" * 72)
byv = collections.Counter(); bywv = collections.Counter()
for r in rows_g:
    byv[r["判定"]] += r["本数"] - 1
    bywv[r["判定"]] += r["消せる容量(MB)"]
for v, n in byv.most_common():
    print(f"  {v:<40} {n:>5}件 {bywv[v]:>9,.0f}MB")

# ---------- Excel ----------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
def sheet(ws, rows, widths):
    if not rows: return
    cols = list(rows[0].keys())
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, c).fill = PatternFill("solid", fgColor="4472C4")
        ws.cell(1, c).alignment = Alignment(vertical="center")
    for r in rows:
        ws.append([r[c] for c in cols])
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

ws1 = wb.active; ws1.title = "グループ"
sheet(ws1, rows_g, [6, 13, 14, 42, 40, 70, 90])
red = PatternFill("solid", fgColor="FFF2CC")
for i, r in enumerate(rows_g, 2):
    if r["判定"].startswith("★消さない"):
        for c in range(1, 8):
            ws1.cell(i, c).fill = red
ws2 = wb.create_sheet("明細")
sheet(ws2, rows_d, [8, 9, 42, 10, 12, 6, 110])

out = os.path.expanduser(f"~/Desktop/重複一覧_現役_{datetime.date.today():%Y%m%d}.xlsx")
wb.save(out)
print(f"\n書き出し: {out}")
print(f"  シート「グループ」{len(rows_g):,}行 / シート「明細」{len(rows_d):,}行")
