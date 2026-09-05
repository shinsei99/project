#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""物件フォルダの直下に散らばったファイルを、**中身を読んで**棚と箱へ入れる。

  サトウビルⅡと河合京橋ビルで手作業でやったことを、そのまま道具にしたもの。

    1. ファイルの本文を開く（.doc/.docx/.xls/.xlsx/.wps、PDFはSpotlightの本文）
    2. 誰のものかを拾う
         ① 契約書の「賃 借 人 ○○」欄        （read_contracts.lessee_of）
         ② 書面の宛名「○○ 様 / 殿」          （重説・覚書・念書はこれで分かる）
         ③ ファイル名の括弧「（高等進学塾）」    （物件名と階は落とす）
    3. 区画（階・号室）をファイル名と本文から拾う
    4. 行き先を決める
         名前が既存の箱と一致                → その箱
         名前＋区画＋レントロールの現契約者と一致 → 入居者/{区画}_{名前}
         一致しない                        → 解約・精算/{区画}_{名前}
         名前が読めない                     → 語のルール（plan_v2 と同じ）で棚へ
         それでも決まらない                  → 動かさない（人が見る）

  ★「たぶんこの階だろう」で現契約者の箱には入れない。名前が取れたものだけ箱に入れる。
  ★動かしたものは local/移動記録_直下の掃き出し_YYYYMMDD.json に残す（元に戻せる）。

  使い方: sweep_loose.py            … 全物件の下見
          sweep_loose.py 物件名      … 1物件だけ
          sweep_loose.py --go       … 実行
          sweep_loose.py --all      … 決まらなかったものも全部並べる
"""
import os, re, sys, json, shutil, sqlite3, subprocess, unicodedata, datetime, collections
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import read_contracts as RC

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB   = os.path.join(BASE, "data", "app.db")
SF   = "/Users/apple/Library/CloudStorage/Dropbox-大京商事　株式会社/共有フォルダ"
MP   = os.path.join(SF, "物件・管理/管理物件")
RRF  = ["★要更新★レントロール一覧（ビル）.xlsx",
        "★要更新★レントロール一覧（マンション）.xlsx",
        "★要更新★レントロール一覧（駐車場他）.xlsx"]
GO   = "--go" in sys.argv
ALL  = "--all" in sys.argv
ONLY = [a for a in sys.argv[1:] if not a.startswith("--")]
N    = lambda s: unicodedata.normalize("NFC", s)
K    = lambda s: unicodedata.normalize("NFKC", N(s))
norm = RC.norm
SHELF = ("物件基本", "入居者", "解約・精算", "修繕・点検", "記録・写真", "通知・案内", "その他")

# 棚の語のルールは plan_v2 と同じものを借りる（判定の食い違いを作らない）
_src = open(os.path.join(BASE, "tools", "plan_v2.py"), encoding="utf-8").read()
_ns = {"__name__": "plan_v2_rules"}
exec(_src[:_src.index("ENTITY = re.compile")], _ns)
SHELVES = _ns["SHELVES"]

VAR = str.maketrans("濱髙﨑齋來眞澤邊瀨", "浜高崎斎来真沢辺瀬")
NOTNAME = re.compile(r"^(?:案|案文|修正|修正分|最終|最終版|変更|削除|追加|訂正|旧|新|控|写|押印|製本|"
                     r"原本|正本|コピー|更新|再|差替|確認|未|済|済み|抜粋|参考|下書き|部分|一部|全部|"
                     r"前|後|前回|今回|ひな形|雛形|様式|書式|株|有|同|服部新版|新版|旧版|"
                     r"特殊詐欺追加|民法改正版|各位|御中|関係者|入居者|賃借人|貸主|借主)$")
ROOMONLY = re.compile(r"^[0-9]{1,4}\s*[FＦ階号室ABC]?([・,、\-ー~〜/]*[0-9]{1,4}\s*[FＦ階号室ABC]?)*$")
LESSOR = re.compile(r"大京商事|新誠|賃\s*貸\s*人|貸\s*主|管理会社|御中$")


def pstem(s):
    n = norm(s)
    for suf in ("ビル", "bldg", "マンション", "駐車場", "モータープール", "駐輪場"):
        if n.endswith(suf) and len(n) > len(suf) + 2: n = n[:-len(suf)]
    return n


def same_name(a, b):
    a, b = norm(a), norm(b)
    return bool(a and b and (a in b or b in a))


def clean_name(c, prop):
    """括弧や宛名から拾った文字列を、借主名らしい形に整える。"""
    c = re.sub(r"[★☆\s　]", "", c)
    cv = c.translate(VAR)
    for w in (prop, re.sub(r"[（(].*", "", prop), re.sub(r"(ビル|ビルディング|マンション)$", "", prop)):
        if not w or len(w) < 3: continue
        wv = K(w).translate(VAR)
        i = cv.find(wv)
        while i >= 0:
            c = c[:i] + c[i + len(wv):]; cv = cv[:i] + cv[i + len(wv):]
            i = cv.find(wv)
    # ★物件名のローマ字部分（MDXBLDG→MDX）も大小を無視して落とす。
    #   落とさないと「MDX3階」がそのまま借主名になる（2026-09-05）。
    stem = re.sub(r"[^A-Za-z]", "", K(prop))
    if len(stem) >= 3:
        c = re.sub(re.escape(stem[:3]) + r"[A-Za-z]*", "", c, flags=re.I)
    c = re.sub(r"^[-ー_]*\d{1,2}\s*[FＦ階][-ー_]?", "", c)
    c = re.sub(r"^物件名[：:]?", "", c)
    c = re.sub(r"^[-ー_]*\d{2,4}\s*号[室]?[-ー_]?", "", c)      # 「303号室西村進也」の頭
    c = re.sub(r"^号\s*室?", "", c)                            # 「号室㈱ゼンショー…」の頭
    c = re.sub(r"^室(?=[^\s])", "", c)                          # 「室大本悦子」の頭
    c = re.sub(r"^[ぁ-ん]{1,3}[・･]", "", c)                    # 「ポ・ラ・…」のような欠けた物件名
    # ★「コーポ・ラ・ベリエール303号室西村進也」のように、号室が名前の途中に入ることがある。
    #   物件名の表記ゆれ（長音記号の違い）で物件名が落ちきらないので、号室より後ろを名前とする。
    m = re.search(r"\d{2,4}\s*号\s*室?", c)
    if m and c[m.end():].strip("-ー_・･ "): c = c[m.end():]
    c = re.sub(r"^[-ー_・･]+|[-ー_・･]+$", "", c)
    if len(c) < 2 or NOTNAME.match(c) or ROOMONLY.match(K(c)) or LESSOR.search(c): return ""
    # ★物件名の一部（「片町」など）を借主名にしない
    if len(c) <= 6 and norm(c) and norm(c) in norm(prop): return ""
    if re.fullmatch(r"[\d\s.\-/年月日]+", c): return ""
    return c


# ★宛名は「大路 喜一郎 様」のように姓と名の間が空く（2026-09-05に踏んだ）。
#   空白で切ると「喜一郎」だけになり、別人の箱（宮下純子↔松山純子）に入ってしまう。
ATENA = re.compile(r"((?:[^\s　\n、。（(]{1,14})(?:[\s　][^\s　\n、。（(]{1,14}){0,2})\s*(?:様|殿)(?![々])")
# ★人ではなく「書式そのもの」を指す語。これを名前として拾うと
#   「1F_新-A3表裏-ビルP契約書」のような箱が出来てしまう（2026-09-05に実際に出来ていた）。
NOTNAME = re.compile(r"契約書|重説|重要事項|雛形|雛型|ひな形|ひな型|様式|書式|テンプレ|"
                     r"表裏|Ａ?[A-Z]?[0-9]表裏|新-|旧-|民法改正|法人用|個人用|見本|サンプル|"
                     r"登録番号あり|付帯設備|中身|表紙|白紙|未記入")
# 条文の切れ端を名前として拾わない
CLAUSE = re.compile(r"[、。]|するもの|とする|した|により|場合|とき|以下|次の|下記|本契約|更新|"
                    r"自動|引き続き|暴力団|使用人|家族|お住まい|皆様|各位|退去|解約通知|"
                    r"以降も|以後も|その後も|本物件|改造|模様替|も同$|^も|^同$|^[0-9]{4}[.\-/]")


def who_of(text, fname, prop):
    """誰のものかを拾う。(名前, どこから) を返す。"""
    w = RC.lessee_of(text)
    if w and not CLAUSE.search(w) and not NOTNAME.search(w):
        return re.sub(r"[\s　]+", "", w), "契約書の賃借人欄"
    t = K(text or "").replace("(株)", "㈱").replace("(有)", "㈲")
    for m in ATENA.finditer(t[:2500]):                 # 書面の宛名
        raw = m.group(1).strip()
        if CLAUSE.search(raw): continue
        c = clean_name(raw, prop)
        if c and not CLAUSE.search(c): return c, "書面の宛名"
    f = K(fname).replace("(株)", "㈱").replace("(有)", "㈲")
    for m in re.finditer(r"[（(]([^）)]{2,24})[）)]", f):
        if CLAUSE.search(m.group(1)) or NOTNAME.search(m.group(1)): continue
        c = clean_name(m.group(1), prop)
        if c and not CLAUSE.search(c): return c, "ファイル名の括弧"
    return "", ""


def room_of(fname, text):
    r = RC.room_of(fname)
    if r: return r
    m = re.search(r"名\s*称[：:][^\n]{0,40}?([B地下]?\d{1,2})\s*階", K(text or "")[:1500])
    if m: return m.group(1).replace("地下", "B")
    return ""


# ---------- レントロール ----------
rr = {}
for f in RRF:
    p = os.path.join(SF, f)
    if not os.path.exists(p): continue
    wb = load_workbook(p, data_only=True, read_only=True)
    for sn in wb.sheetnames:
        data = [[("" if c is None else str(c).strip()) for c in row]
                for row in wb[sn].iter_rows(values_only=True)]
        ci = si = ri = None
        for row in data[:6]:
            for j, c in enumerate(row):
                if c == "契約者": ci = j
                elif c == "現況": si = j
                elif c in ("号室", "号室/フロア", "区画No", "区画"): ri = j
            if ci is not None: break
        if ci is None: continue
        tbl = {}
        for row in data:
            if len(row) <= ci: continue
            t = row[ci]; rm = row[ri] if (ri is not None and len(row) > ri) else ""
            if not t or t in ("契約者", "合計", "計", "空室", "空き", "-"): continue
            if si is not None and len(row) > si and row[si] in ("空室", "空き", "解約", "退去"): continue
            if rm: tbl[re.sub(r"[^0-9A-Za-zＡ-Ｚ]", "", K(rm)).upper()] = t
        if tbl: rr[pstem(sn)] = tbl
    wb.close()


def table_of(prop):
    keys = {pstem(prop)}
    try:
        c = sqlite3.connect(DB)
        row = c.execute("SELECT name, aliases FROM properties WHERE active=1 AND "
                        "(name=? OR aliases LIKE ?)", (prop, f"%{prop}%")).fetchone()
        if row:
            keys.add(pstem(row[0]))
            for a in re.split(r"[\n,、]", row[1] or ""):
                if a.strip(): keys.add(pstem(a.strip()))
        c.close()
    except Exception:
        pass
    for k in keys:
        cand = [kk for kk in rr if kk and (kk in k or k in kk) and min(len(kk), len(k)) >= 3]
        if cand: return rr[max(cand, key=len)]
    return None


def rooms_of_box(name):
    s = K(name)
    m = re.match(r"^(\d{1,2})\s*[-ー~〜]\s*(\d{1,2})\s*[FＦ階]", s)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        if a <= b and b - a <= 20: return {str(x) for x in range(a, b + 1)}
    r = RC.room_of(name)
    return {r} if r else set()


def boxes_of(pp):
    out = []
    for sh in ("入居者", "解約・精算"):
        sp = os.path.join(pp, sh)
        if not os.path.isdir(sp): continue
        for b in sorted(os.listdir(sp)):
            if not os.path.isdir(os.path.join(sp, b)): continue
            b = N(b)
            if b.startswith(("賃借人資料", "入居者", "テナント資料", "解約済", "★")): continue
            out.append((sh, b, rooms_of_box(b), norm(re.sub(r"^[^_]*_", "", b))))
    return out


def box_name(room, who, fname="", kind=""):
    """箱の見出し。★「26番」の駐車区画に F を付けない（2026-09-05）。"""
    if not room: return who
    if re.search(rf"{re.escape(room)}\s*番", K(fname)) or kind in ("駐車場",):
        return f"{room}番_{who}"
    return f"{room}F_{who}" if re.fullmatch(r"\d{1,2}", room) else f"{room}_{who}"


def text_of(p):
    t = RC.text_of(p)
    if t.strip(): return t
    if p.lower().endswith(".pdf"):
        o = subprocess.run(["mdls", "-name", "kMDItemTextContent", "-raw", p],
                           capture_output=True, text=True).stdout
        if o.strip() and o.strip() != "(null)": return o
    return ""


def shelf_by_words(name):
    """★「解　約　通　知　書」のように全角スペースで割られた語が読めていなかった
    （2026-09-05）。空白を潰した形でも当てる。"""
    flat = re.sub(r"[\s　]+", "", name)
    for sh, pat in SHELVES:
        if re.search(pat, name) or re.search(pat, flat): return sh
    return ""


# ---------- 掃き出し ----------
plans = []
for kind in sorted(os.listdir(MP)):
    kp = os.path.join(MP, kind)
    # ★_旧管理物件 は普段は触らないが、物件名を名指ししたときだけ入る
    if not os.path.isdir(kp): continue
    if kind.startswith("_") and not ONLY: continue
    for prop in sorted(os.listdir(kp)):
        pp = os.path.join(kp, prop); prop = N(prop)
        if not os.path.isdir(pp): continue
        if ONLY and prop not in ONLY: continue
        tot = sum(1 for c, d, f in os.walk(pp) for x in f
                  if x != ".DS_Store" and not x.startswith("._"))
        if not ONLY and tot <= 20: continue          # 20件以下は棚を作らない決まり
        table = table_of(prop)
        boxes = boxes_of(pp)

        # 対象＝物件フォルダの直下 ＋ 入居者/解約・精算 に箱へ入らず直に置かれたもの
        # ★「このフォルダについて.txt」は物件フォルダの入口の案内なので直下に残す
        KEEP = ("このフォルダについて.txt", "README.txt", "はじめに.txt")
        targets = [(os.path.join(pp, x), "") for x in sorted(os.listdir(pp))
                   if os.path.isfile(os.path.join(pp, x)) and x != ".DS_Store"
                   and not x.startswith("._") and N(x) not in KEEP]
        for sh in ("入居者", "解約・精算"):
            sp = os.path.join(pp, sh)
            if not os.path.isdir(sp): continue
            targets += [(os.path.join(sp, x), sh) for x in sorted(os.listdir(sp))
                        if os.path.isfile(os.path.join(sp, x)) and x != ".DS_Store"
                        and not x.startswith("._")]

        for p, cur_shelf in targets:
            f = N(os.path.basename(p))
            # ★「A203_西井雄紀（2024.6.1～）.pdf」のように、人が既に区画と名前を
            #   付けているものは触らない。機械が読み違えて別人の箱へ入れる危険が高い。
            if cur_shelf and re.match(r"^[0-9A-Za-zＡ-Ｚ①-⑳]{1,6}[_＿]", K(f)):
                plans.append({"物件": prop, "種別": N(kind), "名前": f, "パス": p,
                              "今の棚": cur_shelf, "行き先": "（動かさない）",
                              "理由": "既に「区画_名前」で名前が付いている", "_pp": pp})
                continue
            t = text_of(p)
            who, src = who_of(t, f, prop)
            if not who:
                # ★ファイル名の「新内様」を拾う。ATENA は空白をまたいで広く取るので
                #   「駐車場契約書 新内様」だと書式語まで巻き込んで捨てられる。
                #   様・殿の直前の1語だけを見る狭い形も試す（2026-09-05）。
                kf = K(f).replace("(株)", "㈱").replace("(有)", "㈲")
                for pat in (ATENA, re.compile(r"([^\s　\n、。（(＿_]{2,14})[\s　]*[様殿](?![々])")):
                    m = pat.search(kf)
                    if not m: continue
                    g = m.group(1).strip()
                    if CLAUSE.search(g) or NOTNAME.search(g): continue
                    c = clean_name(g, prop)
                    if c: who, src = c, "ファイル名の宛名"; break
            room = room_of(f, t)
            cur = (table or {}).get(room, "")
            dest = why = ""
            if who:
                # ★2文字だけの一致（「恭子」「比良」）は別人の箱に入る危険があるので、
                #   区画も合っているときだけ認める（2026-09-05）。
                strict = len(norm(who)) >= 3
                hit = [b for b in boxes if b[3] and same_name(b[3], who)
                       and (not room or not b[2] or room in b[2])]
                if not hit and strict:
                    hit = [b for b in boxes if b[3] and same_name(b[3], who)]
                if hit and not strict and not room:
                    # ★当てはまる箱が **1つだけ** なら、2文字でも取り違えようがない。
                    #   「駐車場契約書　新内様.docx」が 26番_新内光希 の箱へ入る（2026-09-05）
                    hit = hit if len(hit) == 1 else []
                if hit:
                    dest, why = f"{hit[0][0]}/{hit[0][1]}", f"{src}「{who}」＝既にある箱"
                elif room and cur and same_name(who, cur):
                    dest, why = f"入居者/{box_name(room, cur, f, kind)}", f"{src}「{who}」＝レントロールの現契約者"
                elif room and cur:
                    dest, why = f"解約・精算/{box_name(room, who, f, kind)}", f"区画{room}は今{cur}＝この{who}は旧契約"
                elif room and table is not None:
                    dest, why = f"解約・精算/{box_name(room, who, f, kind)}", f"{src}「{who}」／区画{room}はレントロールに無い"
            # ★名前が読めなくても、区画に当てはまる箱が1つだけなら入れてよい
            #   （「念　書家賃トーヨー404号.doc」→ 入居者/404_○○。2026-09-05）
            if not dest and room:
                hit = [b for b in boxes if b[2] and room in b[2]]
                if len(hit) == 1:
                    dest, why = f"{hit[0][0]}/{hit[0][1]}", f"区画{room}の箱が1つだけ"

            if not dest:
                sh = shelf_by_words(f) or (shelf_by_words(" ".join(K(t).split())[:400]) if t else "")
                if sh and sh not in ("入居者", "解約・精算"):
                    dest, why = sh, "語のルール"
                elif cur_shelf:
                    why = f"★{cur_shelf}の中だが誰のものか読めない（{who or '名前なし'}／区画{room or '―'}）"
                else:
                    why = f"★決め手なし（{who or '名前が読めない'}／区画{room or '―'}）"
            plans.append({"物件": prop, "種別": N(kind), "名前": f, "パス": p,
                          "今の棚": cur_shelf, "行き先": dest or "（動かさない）",
                          "理由": why, "_pp": pp})

print("=" * 84)
print(("【実行】" if GO else "【下見】")
      + f"  直下の掃き出し （{len({p['物件'] for p in plans})}物件 / {len(plans)}件）")
print("=" * 84)
c = collections.Counter(p["行き先"].split("/")[0] for p in plans)
for k, v in c.most_common(): print(f"  {k:<14}{v:>5}件")
move = [p for p in plans if p["行き先"] != "（動かさない）"]
stay = [p for p in plans if p["行き先"] == "（動かさない）"]
print(f"\n  動かす {len(move)}件 / 動かさない {len(stay)}件"
      f"（{len(move)*100//max(1,len(plans))}% を仕分けできる）")

for prop in sorted({p["物件"] for p in plans}):
    ps = [p for p in plans if p["物件"] == prop and (ALL or p["行き先"] != "（動かさない）")]
    if not ps: continue
    print(f"\n── {prop}")
    for p in sorted(ps, key=lambda x: x["行き先"]):
        print(f"   {p['名前'][:38]:<40} → {p['行き先'][:28]:<30} {p['理由'][:36]}")

if not GO:
    print("\n  ※下見のみ。--go で実行")
    sys.exit(0)

log, ng = [], 0
for p in move:
    d = os.path.join(p["_pp"], p["行き先"])
    os.makedirs(d, exist_ok=True)
    t = os.path.join(d, p["名前"])
    if os.path.exists(t):
        print(f"  ★同名なので動かさない: {p['物件']}/{p['名前'][:36]}"); ng += 1; continue
    shutil.move(p["パス"], t)
    log.append({"from": N(p["パス"]), "to": N(t), "なぜ": p["理由"]})
rec = os.path.join(BASE, "local", f"移動記録_直下の掃き出し_{datetime.date.today():%Y%m%d}.json")
json.dump({"date": str(datetime.date.today()), "moved": log}, open(rec, "w"),
          ensure_ascii=False, indent=2)
print(f"\n  移動 {len(log)}件 / 動かせず {ng}件")
print(f"  対応表: {rec}")
