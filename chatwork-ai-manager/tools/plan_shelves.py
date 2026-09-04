#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""整理案②: 管理物件の中を「標準の棚9つ」へ振り分ける計画表を作る。
   読むだけ。1件も動かさない。出力: ~/Desktop/整理計画_棚わけ_YYYYMMDD.xlsx

   規則
     基本  物件・管理/管理物件/{種別}/{物件名}/{棚}/file                 = 5階層
     例外  実体フォルダ（区画番号・部屋番号・入居者・サブ物件）は1つだけ残す = 6階層
"""
import os, re, unicodedata, collections, datetime, hashlib

ROOT = "/Users/apple/Library/CloudStorage/Dropbox-大京商事　株式会社/共有フォルダ/（★必読★）新共有フォルダ"
PRE  = "物件・管理/管理物件"
N = lambda s: unicodedata.normalize("NFC", s)

SHELVES = [
    ("01_契約",        r"契約書|契約証書|重説|重要事項|覚書|合意書|更新|定期借家|保証委託|入居申込|申込書|審査|媒介|念書|承諾|雛型|雛形|民法改正|特殊詐欺|鍵受領|鍵預|鍵引渡"),
    ("02_解約・精算",   r"解約|退去|精算|清算|原状回復|明渡|返還|敷金|立退"),
    ("03_請求・入金",   r"請求|領収|入金|送金|月次|収支|家賃|賃料|振込|明細|滞納|広告料|仲介手数料"),
    ("04_工事・修繕",   r"工事|修繕|リフォーム|リホーム|見積|施工|点検|保守|清掃|設備|昇降|エレベータ|ＥＶ|EV|消防|防火|避難|廃棄物|ゴミ|ごみ|塗装|防水|貯水槽|受水槽|給水|排水|空調"),
    ("05_図面・写真",   r"図面|竣工|平面|間取|配置|測量|公図|写真|マイソク|パース|看板|広告物|サイン|地図|DSCN|DSC\d|IMG[_\-]?\d|P\d{7}|\.jpe?g$|\.png$"),
    ("06_権利・登記",   r"登記|謄本|全部事項|評価証明|固定資産|税|保険|証券|権利|決済|抵当"),
    ("08_検針・メーター", r"検針|メーター|水道|電気|ガス|使用量|電力"),
    ("09_賃借人資料",   r"賃借人|入居者|テナント|車庫証|入居時確認|名簿|台帳"),
    ("07_通知・その他",  r"お知らせ|通知|案内|貼紙|掲示|連絡|報告|一覧|書式|様式|アンケート|議事録|送付|封筒|宛名|シール"),
]
# すでに標準の棚に入っているものは、その棚のまま動かさない
ALREADY = re.compile(r"^(0[1-9])_")
ENTITY_PAT = re.compile(
    r"^\d+番|^\d+号|^[A-Z]?\d{3,4}(号|室|_)|^\d+F|^\d+階|^[0-9]{3}_|"
    r"^\d+Ｆ|^[０-９]+Ｆ|^\d+[FＦ]【|"
    r"（\d{4}[.\-/]\d{1,2}|～）|^\d{4}[.\-]\d{2}[.\-]\d{2}_")

def shelf(below, fname):
    # ① すでに標準の棚に入っているならそのまま
    first = below.split("/")[0] if below else ""
    if ALREADY.match(first):
        return first, "すでに棚に入っている"
    for n_, pat in SHELVES:
        if re.search(pat, below): return n_, "いまのフォルダ名"
    for n_, pat in SHELVES:
        if re.search(pat, fname): return n_, "ファイル名"
    return "07_通知・その他", "★手がかりなし"

# ---------- 収集 ----------
files = []
for cur, dn, fn in os.walk(os.path.join(ROOT, PRE)):
    rel = N(os.path.relpath(cur, ROOT))
    for f in fn:
        f = N(f)
        if f == ".DS_Store" or f.startswith("._"): continue
        files.append(rel + "/" + f)

# ---------- 実体フォルダの判定（兄弟が同じファイル名を持つ / 名前が区画・部屋の形） ----------
sib = collections.defaultdict(lambda: collections.defaultdict(set))
for p in files:
    parts = p.split("/")
    if len(parts) < 6: continue
    for i in range(4, len(parts) - 1):
        sib["/".join(parts[:i])][parts[i]].add(parts[-1])
entity = set()
for parent, kids in sib.items():
    names = list(kids)
    for i, a in enumerate(names):
        if ENTITY_PAT.search(a): entity.add((parent, a)); continue
        for b in names[i+1:]:
            if kids[a] & kids[b]: entity.add((parent, a)); entity.add((parent, b))

# ---------- 計画 ----------
rows = []
for p in files:
    parts = p.split("/")
    if len(parts) < 5:
        rows.append({"区分": "★対象外（物件フォルダの外）", "物件": "", "種別": parts[2] if len(parts) > 2 else "",
                     "棚": "", "決め手": "", "残す実体": "", "いまの位置": "",
                     "容量(KB)": round(os.path.getsize(os.path.join(ROOT, p))/1024, 1),
                     "現在のパス": p, "移動後のパス": p, "注意": "★物件フォルダの外に置かれている"})
        continue
    kind, prop = parts[2], parts[3]
    base = "/".join(parts[:4]); fname = parts[-1]
    below = "/".join(parts[4:-1])
    keep = None
    for i in range(4, len(parts) - 1):
        if ("/".join(parts[:i]), parts[i]) in entity: keep = parts[i]
    if keep:
        sh, why = "09_賃借人資料", "実体フォルダ"
        new = f"{base}/09_賃借人資料/{keep}/{fname}"
    else:
        sh, why = shelf(below, fname)
        new = f"{base}/{sh}/{fname}"
    rows.append({"区分": "棚へ入れる", "物件": prop, "種別": kind, "棚": sh, "決め手": why,
                 "残す実体": keep or "", "いまの位置": below or "（物件フォルダに直置き）",
                 "容量(KB)": round(os.path.getsize(os.path.join(ROOT, p))/1024, 1),
                 "現在のパス": p, "移動後のパス": new, "注意": ""})

# ---------- 衝突 ----------
dest = collections.Counter(r["移動後のパス"] for r in rows)
col = {k for k, v in dest.items() if v > 1}
def h(p):
    try:
        with open(os.path.join(ROOT, p), "rb") as f: return hashlib.sha256(f.read()).hexdigest()
    except OSError: return None
bysrc = collections.defaultdict(list)
for r in rows:
    if r["移動後のパス"] in col: bysrc[r["移動後のパス"]].append(r)
same = diff = 0
for d, rs in bysrc.items():
    hs = {h(r["現在のパス"]) for r in rs}
    tag = "★中身が違う＝要個別対応" if len(hs) > 1 else "行き先が重なる（中身は同じ＝重複）"
    if len(hs) > 1: diff += len(rs)
    else: same += len(rs)
    for r in rs: r["注意"] = tag

# ---------- 集計 ----------
tgt = [r for r in rows if r["区分"] == "棚へ入れる"]
print("=" * 76); print("整理案② 管理物件を標準の棚へ（計画のみ・1件も動かしていない）"); print("=" * 76)
print(f"  対象 {len(tgt):,}件 / 物件 {len({r['物件'] for r in tgt})}件")
moved = sum(1 for r in tgt if r["現在のパス"] != r["移動後のパス"])
print(f"  位置が変わる {moved:,}件 / そのまま {len(tgt)-moved:,}件")
print(f"\n  棚の内訳:")
for k, v in sorted(collections.Counter(r["棚"] for r in tgt).items()):
    print(f"    {k:<20}{v:>6,}件")
print(f"\n  棚の決め手:")
for k, v in collections.Counter(r["決め手"] for r in tgt).most_common():
    print(f"    {k:<20}{v:>6,}件")
dd = collections.Counter(len(r["移動後のパス"].split("/")) - 1 for r in tgt)
print(f"\n  移動後の深さ（新共有フォルダから）:")
for k in sorted(dd): print(f"    {k}階層 {dd[k]:>6,}件")
print(f"\n  ★注意が要る行 {sum(1 for r in rows if r['注意']):,}件"
      f"（中身が違う衝突 {diff}件 / 重複 {same}件 / 物件フォルダの外 "
      f"{sum(1 for r in rows if r['区分'].startswith('★対象外'))}件）")
print(f"  ★手がかりなしで07へ入る {sum(1 for r in tgt if r['決め手']=='★手がかりなし'):,}件 ← 目で見た方がよい")

# ---------- Excel ----------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
wb = Workbook(); ws = wb.active; ws.title = "棚わけ計画"
cols = list(rows[0].keys()); ws.append(cols)
for c in range(1, len(cols)+1):
    ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
    ws.cell(1, c).fill = PatternFill("solid", fgColor="4472C4")
for r in sorted(rows, key=lambda x: (not x["注意"], x["決め手"] != "★手がかりなし", x["物件"], x["棚"])):
    ws.append([r[c] for c in cols])
for i, w in enumerate([24, 26, 12, 18, 16, 26, 34, 10, 82, 82, 30], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
red = PatternFill("solid", fgColor="FCE4EC"); yel = PatternFill("solid", fgColor="FFF2CC")
for i in range(2, ws.max_row+1):
    if ws.cell(i, len(cols)).value:
        for c in range(1, len(cols)+1): ws.cell(i, c).fill = red
    elif ws.cell(i, 5).value == "★手がかりなし":
        for c in range(1, len(cols)+1): ws.cell(i, c).fill = yel

ws2 = wb.create_sheet("物件ごと")
ws2.append(["物件", "種別", "件数", "棚の数", "実体フォルダ", "手がかりなし", "移動後の最大階層"])
for c in range(1, 8):
    ws2.cell(1, c).font = Font(bold=True, color="FFFFFF")
    ws2.cell(1, c).fill = PatternFill("solid", fgColor="4472C4")
g = collections.defaultdict(list)
for r in tgt: g[(r["物件"], r["種別"])].append(r)
for (nm, kind), rs in sorted(g.items(), key=lambda x: -len(x[1])):
    ws2.append([nm, kind, len(rs), len({r["棚"] for r in rs}),
                len({r["残す実体"] for r in rs if r["残す実体"]}),
                sum(1 for r in rs if r["決め手"] == "★手がかりなし"),
                max(len(r["移動後のパス"].split("/"))-1 for r in rs)])
for i, w in enumerate([34, 12, 8, 8, 14, 14, 18], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"; ws2.auto_filter.ref = ws2.dimensions

out = os.path.expanduser(f"~/Desktop/整理計画_棚わけ_{datetime.date.today():%Y%m%d}.xlsx")
wb.save(out)
print(f"\n書き出し: {out}  （{len(rows):,}行 ＋ 物件ごと{len(g)}行）")
