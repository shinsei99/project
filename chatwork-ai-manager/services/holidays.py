"""会社の休業日（年間休暇スケジュール）を読む。

オーナー管理の Excel「年間休暇スケジュール<年>.xlsx」が正典。
1シート＝1か月のカレンダーで、**オレンジ色に塗られたマスがその日の休み**。

読み方（実物を見て確かめた形・2026-08-21）:
  - 日付の数字は B/D/F/H/J/L/N 列（日〜土）にあり、行は 5,7,9,… と1行おき
  - **色はその1つ下の行の同じ列**に塗られている（日付マスが2行1組のため）
  - 色は「テーマ色5（アクセント2＝オレンジ）」。RGBでは入っていない
  - 前月・翌月の日も並ぶので、「1から始まる連続した並び」だけをその月とみなす

元ファイルは Google Drive（CloudStorage）上にあり、**launchd の常時起動プロセスからは
読めないことがある**（`/bin/bash` にフルディスクアクセスが要る）。そこで読めたときに
DB へ写しておき、読めないときは写しを使う。＝ 1年分は一度読めれば以後ずっと効く。
"""
import datetime
import os

from db.connection import get_conn, query
from services import settings

DAY_COLS = ("B", "D", "F", "H", "J", "L", "N")
ORANGE_THEMES = {5}          # テーマ5＝アクセント2（Office既定でオレンジ）
ORANGE_RGBS = {"FFED7D31", "FFFFC000", "FFF79646", "FFFFA500", "FFFF9900", "FFE36C0A"}


def schedule_path() -> str:
    return settings.get_setting("holiday_schedule_path", "") or ""


def _is_orange(cell) -> bool:
    f = cell.fill
    if not f or f.fill_type != "solid":
        return False
    c = f.start_color
    theme = getattr(c, "theme", None)
    if isinstance(theme, int) and theme in ORANGE_THEMES:
        return True
    rgb = getattr(c, "rgb", None)
    return isinstance(rgb, str) and rgb.upper() in ORANGE_RGBS


def _month_days(ws):
    """(日, 列, 行) を「その月の1日〜末日」だけ順に返す。前月・翌月の分は捨てる。"""
    # 曜日の見出し行（日 月 火 …）より上は見ない。
    # ★ここを見ないと、見出し上の「月」の数字（1月シートの H2=1）を「1日」と誤認して
    #   そこから月が始まったことにしてしまう。2026-08-21 に実際に1月が0件になった。
    head = 0
    for r in range(1, ws.max_row + 1):
        if str(ws[f"{DAY_COLS[0]}{r}"].value or "").strip() in ("日", "Sun", "SUN"):
            head = r
            break
    seq = []
    for r in range(head + 1, ws.max_row + 1):
        for col in DAY_COLS:
            v = ws[f"{col}{r}"].value
            if isinstance(v, int) and 1 <= v <= 31:
                seq.append((v, col, r))
    out, started, prev = [], False, None
    for v, col, r in seq:
        if not started:
            if v != 1:
                continue          # 前月の日
            started, prev = True, 0
        if v != prev + 1:
            break                 # 翌月の1に戻った＝ここで終わり
        out.append((v, col, r))
        prev = v
    return out


def parse(path: str = None) -> dict:
    """{'YYYY-MM-DD': '休み'} を返す。ファイルが読めなければ例外。"""
    from openpyxl import load_workbook
    path = path or schedule_path()
    if not path or not os.path.exists(path):
        raise FileNotFoundError(f"休暇スケジュールが見つかりません: {path}")
    wb = load_workbook(path)
    found = {}
    for ws in wb.worksheets:
        name = str(ws.title)
        if not name.endswith("月"):
            continue
        month = int(name[:-1])
        year = ws["B2"].value
        if not isinstance(year, int):
            year = _year_from_name(path)
        for day, col, r in _month_days(ws):
            # 色は「日付の1つ下の行」に付く。念のため日付セル自身も見る。
            if _is_orange(ws[f"{col}{r + 1}"]) or _is_orange(ws[f"{col}{r}"]):
                try:
                    found[datetime.date(year, month, day).isoformat()] = "休み"
                except ValueError:
                    pass
    return found


def _year_from_name(path: str) -> int:
    import re
    m = re.search(r"(20\d{2})", os.path.basename(path))
    return int(m.group(1)) if m else datetime.date.today().year


def refresh(path: str = None) -> dict:
    """スケジュールを読み直して DB に写す。"""
    try:
        found = parse(path)
    except Exception as e:
        return {"count": 0, "error": f"{type(e).__name__}: {e}"}
    years = sorted({d[:4] for d in found})
    src = os.path.basename(path or schedule_path())
    with get_conn() as conn:
        for y in years:               # その年ぶんを入れ替える（消えた休みも反映する）
            conn.execute("DELETE FROM holidays WHERE holiday_date LIKE ?", (f"{y}-%",))
        for d, note in sorted(found.items()):
            conn.execute("INSERT OR REPLACE INTO holidays (holiday_date, note, source) "
                         "VALUES (?, ?, ?)", (d, note, src))
    settings.set_state("holidays_synced_at", datetime.datetime.now().isoformat(timespec="seconds"))
    return {"count": len(found), "error": None, "years": years}


# ── 日曜・祝日は Excel を見ずに必ず休業日にする（2026-08-30 オーナー指示）────────────
#
# なぜ要るか:
#   2026-08-30（日）に carryover_1000 が動いて3人へ催促が飛んだ。
#   調べると、**Excel のその日のマスに色が塗られていなかった**。
#   8月シートの最終行（30・31）と9月シートの先頭行（30・31）は、前月/翌月と混ざる行で、
#   どちらも丸ごと無色だった。同じ理由で 2026-03-01（日）も落ちていた。
#   ＝ **人が塗り忘れると、その日は営業日として扱われて社員に催促が飛ぶ。**
#
# オーナー指示（2026-08-30）:
#   「基本的に全ての日曜日と祝日は除外してください。カレンダーに印がついてなかったら
#     それはミスです」
#   → Excel は「土曜のどれが休みか」「お盆・年末年始」を決めるためのものと位置づけ、
#     **日曜と祝日は Excel に関係なく休業日**とする。塗り忘れがあっても事故にならない。
#
# 土曜は隔週で出勤日があるため（実測: 2026年は1/17・2/7・2/21…が出勤）、
# ここでは扱わない。土曜は今までどおり Excel が正。

_FIXED = {                       # (月, 日): 名前  ※国民の祝日に関する法律（2026年時点）
    (1, 1): "元日", (2, 11): "建国記念の日", (2, 23): "天皇誕生日",
    (4, 29): "昭和の日", (5, 3): "憲法記念日", (5, 4): "みどりの日", (5, 5): "こどもの日",
    (8, 11): "山の日", (11, 3): "文化の日", (11, 23): "勤労感謝の日",
}
_HAPPY_MONDAY = {                # (月, 第n週の月曜): 名前
    (1, 2): "成人の日", (7, 3): "海の日", (9, 3): "敬老の日", (10, 2): "スポーツの日",
}


def _nth_monday(year: int, month: int, nth: int) -> int:
    d = datetime.date(year, month, 1)
    d += datetime.timedelta(days=(7 - d.weekday()) % 7)   # その月の最初の月曜
    return (d + datetime.timedelta(days=7 * (nth - 1))).day


def _equinox(year: int, spring: bool) -> int:
    """春分・秋分の日。1980〜2099年で使える近似式（天文学的な定義の実用近似）。"""
    base = 20.8431 if spring else 23.2488
    return int(base + 0.242194 * (year - 1980) - (year - 1980) // 4)


def _national_holidays(year: int) -> dict:
    """その年の祝日（振替休日・国民の休日を含む）。{date: 名前}"""
    h = {}
    for (m, d), name in _FIXED.items():
        h[datetime.date(year, m, d)] = name
    for (m, nth), name in _HAPPY_MONDAY.items():
        h[datetime.date(year, m, _nth_monday(year, m, nth))] = name
    h[datetime.date(year, 3, _equinox(year, True))] = "春分の日"
    h[datetime.date(year, 9, _equinox(year, False))] = "秋分の日"

    # 振替休日: 祝日が日曜なら、その後の最初の平日（祝日でない日）が休みになる
    for d in sorted(list(h)):
        if d.weekday() == 6:
            n = d + datetime.timedelta(days=1)
            while n in h:
                n += datetime.timedelta(days=1)
            h[n] = "振替休日"
    # 国民の休日: 祝日に挟まれた平日（例 9/22）
    for d in sorted(list(h)):
        n2 = d + datetime.timedelta(days=2)
        mid = d + datetime.timedelta(days=1)
        if n2 in h and mid not in h and mid.weekday() != 6:
            h[mid] = "国民の休日"
    return h


_HOLIDAY_CACHE: dict = {}


def national_holiday_name(date_str: str) -> str:
    """祝日ならその名前、違えば空文字。"""
    d = datetime.date.fromisoformat(date_str)
    if d.year not in _HOLIDAY_CACHE:
        _HOLIDAY_CACHE[d.year] = _national_holidays(d.year)
    return _HOLIDAY_CACHE[d.year].get(d, "")


def is_holiday(date_str: str) -> bool:
    """休業日か。

    順番:
      ① 日曜        … Excel を見ずに必ず休業日（塗り忘れで事故らないように）
      ② 祝日        … 同上（振替休日・国民の休日を含む）
      ③ Excel の写し … 土曜のどれが休みか、お盆・年末年始など会社固有の休み
    """
    d = datetime.date.fromisoformat(date_str)
    if d.weekday() == 6:                       # ① 日曜
        return True
    if national_holiday_name(date_str):        # ② 祝日
        return True
    if query("SELECT 1 FROM holidays WHERE holiday_date=?", (date_str,)):
        return True
    year = date_str[:4]
    if not query("SELECT 1 FROM holidays WHERE holiday_date LIKE ? LIMIT 1", (f"{year}-%",)):
        if refresh().get("count"):
            return bool(query("SELECT 1 FROM holidays WHERE holiday_date=?", (date_str,)))
    return False


def excel_gaps(year: int) -> list:
    """**Excel の塗り忘れ**を洗い出す（日曜・祝日なのに Excel に無い日）。

    2026-08-30 の事故の再発を見つけるため。is_holiday はもう日曜・祝日を自前で
    休業日にするので実害は無いが、Excel 側も直しておかないと人が見たときに食い違う。
    """
    have = {r["holiday_date"] for r in query(
        "SELECT holiday_date FROM holidays WHERE holiday_date LIKE ?", (f"{year}-%",))}
    out = []
    d = datetime.date(year, 1, 1)
    while d.year == year:
        s = d.isoformat()
        why = "日曜" if d.weekday() == 6 else national_holiday_name(s)
        if why and s not in have:
            out.append((s, why))
        d += datetime.timedelta(days=1)
    return out


def list_for_month(year: int, month: int):
    return [r["holiday_date"] for r in query(
        "SELECT holiday_date FROM holidays WHERE holiday_date LIKE ? ORDER BY holiday_date",
        (f"{year}-{month:02d}-%",))]


def count() -> int:
    r = query("SELECT COUNT(*) AS n FROM holidays")
    return r[0]["n"] if r else 0
