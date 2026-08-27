#!/usr/bin/env python3
"""管理物件台帳の空欄を、社内にある確かな出どころから埋める（2026-08-27・オーナー依頼）。

**空欄だけ埋める。既に入っている値は絶対に上書きしない。**（人が入れた値のほうが確かなため）

出どころ（すべて共有フォルダの中の既存ファイル。外から持ってこない）:

  1. `★要更新★管理物件等資料及び空室.xlsx` の「物件住所　空室」シート
     → **住所 / 築年数 / 構造 / 戸数**。台帳と同じ項目がそのまま並んでいる
  2. `★要更新★レントロール一覧（ビル/マンション/駐車場他）.xlsx`
     → **戸数**（物件ごとのシートの区画行を数える）。1 の戸数が空のときの控え
  3. 日本郵便の郵便番号API（住所 → 郵便番号）
     → **郵便番号**。住所が入っている行だけ。**7桁が一意に決まったときだけ**書く

★物件名の突き合わせは `fix_ledger_folders.py` と同じ考え方で、**正規化して一致したものだけ**。
  部分一致で寄せると別物件の住所を入れてしまう（2026-08-27に実際にやらかした）。

**埋められない項目と、その理由**（2026-08-27に調べた。同じ調査を繰り返さないこと）

  築年数86 / 構造85 … **重要事項説明書からは埋められない**。
    `_アーカイブ/★重要事項説明書/` に124件あるが、大半は **.xls の白紙雛形と部屋ごとの契約書**で、
    ナレッジに取り込めているのは **7件だけ**。うち台帳の物件名と結びつくのは1件（角屋マンション）で、
    それも**既に埋まっている**。構造を抽出できたのは7件中1件。
    さらに拾える「年月」は**築年月ではなく契約日**（令和6年5月・2024年5月など）で、
    角屋マンションでは重説の「昭和56年5月」と台帳の「1977年04月」が食い違う。
    → **機械で埋めると契約日を築年数として入れてしまう。やらない。**

  交通95 / 連絡先2:100 / 補足93 / TEL1:80 / TEL2:78 / 連絡先1:61
    … 社内に構造化された出どころが無い。交通は住所から計算できるが、それは
      「事実」ではなく「こちらの計算結果」なので、入れるなら補足欄にその旨を残すこと。

使い方:
  python3 fill_ledger_blanks.py --dry-run     # 何が埋まるか見る
  python3 fill_ledger_blanks.py --no-postal   # 郵便番号APIを使わない（通信なし）
  python3 fill_ledger_blanks.py               # 実行（実行前に自動でバックアップ）
"""
from __future__ import annotations

import argparse
import os
import re
import shutil
import sys
import unicodedata
from datetime import datetime, timedelta

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from fix_ledger_folders import _ledger_path, norm, strip_paren  # noqa: E402

ADDR_BOOK = "★要更新★管理物件等資料及び空室.xlsx"
ADDR_SHEET = "物件住所　空室"
RENTROLLS = ["★要更新★レントロール一覧（ビル）.xlsx",
             "★要更新★レントロール一覧（マンション）.xlsx",
             "★要更新★レントロール一覧（駐車場他）.xlsx"]

COL = {"住所": 5, "郵便番号": 6, "築年数": 7, "構造": 8, "戸数": 9}   # 1始まり
NAME_COL = 2


def excel_date(v):
    """Excelのシリアル値なら日付文字列にする（築年数の列が日付で入っている）。"""
    if isinstance(v, datetime):
        return v.strftime("%Y年%m月")
    if isinstance(v, (int, float)) and 20000 < float(v) < 60000:
        return (datetime(1899, 12, 30) + timedelta(days=float(v))).strftime("%Y年%m月")
    return v


def read_address_book(root: str) -> dict:
    """{正規化した物件名: {住所, 築年数, 構造, 戸数}}"""
    import openpyxl
    p = os.path.join(root, ADDR_BOOK)
    if not os.path.exists(p):
        return {}
    wb = openpyxl.load_workbook(p, read_only=True)
    if ADDR_SHEET not in wb.sheetnames:
        wb.close()
        return {}
    out = {}
    for row in wb[ADDR_SHEET].iter_rows(values_only=True):
        cells = list(row) + [None] * 6
        name, addr, age, struct, units = cells[1], cells[2], cells[3], cells[4], cells[5]
        if not name or not isinstance(name, str):
            continue
        if not addr or not isinstance(addr, str) or "大阪" not in addr and "府" not in addr and "県" not in addr:
            # 住所が無い行（見出し・地区名）は物件行ではない
            if not addr:
                continue
        out[norm(name)] = {"住所": addr, "築年数": excel_date(age),
                           "構造": struct, "戸数": units}
    wb.close()
    return out


def read_unit_counts(root: str) -> dict:
    """レントロールのシート名＝物件名。区画（号室）行を数えて戸数にする。"""
    import openpyxl
    out = {}
    for f in RENTROLLS:
        p = os.path.join(root, f)
        if not os.path.exists(p):
            continue
        wb = openpyxl.load_workbook(p, read_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            n = 0
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                if i <= 2:                      # 1行目=物件名 2行目=見出し
                    continue
                if row and row[0] not in (None, ""):
                    n += 1
            if n:
                out[norm(sn)] = n
        wb.close()
    return out


def postal_of(address: str):
    """住所から郵便番号（〒NNN-NNNN）。一意に決まらなければ None。

    日本郵便のAPIは**町域まで**しか持たないので、番地・丁目を落として問い合わせる。
    候補が2件以上あるときは書かない（別の町域を入れてしまうため）。
    """
    try:
        sys.path.insert(0, os.path.expanduser("~"))
        import japanpost_api as jp
    except Exception:
        return None
    q = address.split("丁目")[0] if "丁目" in address else address
    q = re.sub(r"[0-9０-９].*$", "", q).strip()
    if len(q) < 4:
        return None
    try:
        r = jp.address_zip(freeword=q, limit=5)
    except Exception:
        return None
    items = r.get("addresses") or r.get("data") or []
    zips = {re.sub(r"\D", "", str(i.get("zip_code") or i.get("zipCode") or "")) for i in items}
    zips = {z for z in zips if len(z) == 7}
    if len(zips) != 1:
        return None
    z = zips.pop()
    return "〒%s-%s" % (z[:3], z[3:])


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--no-postal", action="store_true")
    args = ap.parse_args()

    import openpyxl
    src = _ledger_path()
    root = os.path.dirname(src)
    book = read_address_book(root)
    units = read_unit_counts(root)
    print("出どころ: 物件住所シート %d件 / レントロールの物件 %d件" % (len(book), len(units)))

    wb = openpyxl.load_workbook(src)
    ws = wb["管理物件台帳"]
    filled = {k: 0 for k in COL}
    filled["郵便番号"] = 0
    unmatched = []
    for r in ws.iter_rows(min_row=3, max_row=ws.max_row):
        name = r[NAME_COL - 1].value
        if not name or str(name).startswith("■"):
            continue
        row = r[0].row
        keys = {norm(name), norm(strip_paren(str(name)))}
        rec = next((book[k] for k in keys if k in book), None)
        got = []
        if rec:
            for col in ("住所", "築年数", "構造", "戸数"):
                cell = ws.cell(row=row, column=COL[col])
                v = rec.get(col)
                if cell.value in (None, "") and v not in (None, ""):
                    if not args.dry_run:
                        cell.value = v
                    filled[col] += 1
                    got.append("%s=%s" % (col, str(v)[:20]))
        else:
            unmatched.append(str(name))
        # 戸数はレントロールからも数えられる（住所シートに無いとき）
        # 駐車場・駐輪場は「戸」ではなく「区画」
        kind = str(ws.cell(row=row, column=1).value or "")
        unit_word = "区画" if ("駐車" in kind or "駐輪" in kind or "トランク" in kind) else "戸"
        cell = ws.cell(row=row, column=COL["戸数"])
        if cell.value in (None, ""):
            u = next((units[k] for k in keys if k in units), None)
            if u:
                if not args.dry_run:
                    cell.value = "%d%s" % (u, unit_word)
                filled["戸数"] += 1
                got.append("戸数=%d%s(レントロール)" % (u, unit_word))
        # 郵便番号は住所から
        if not args.no_postal:
            pc = ws.cell(row=row, column=COL["郵便番号"])
            ad = ws.cell(row=row, column=COL["住所"]).value
            if pc.value in (None, "") and ad:
                z = postal_of(str(ad))
                if z:
                    if not args.dry_run:
                        pc.value = z
                    filled["郵便番号"] += 1
                    got.append("郵便番号=%s" % z)
        if got:
            print("  %3d %-22s %s" % (row, str(name)[:22], " / ".join(got)))

    print("\n%s 埋めた数: %s" % ("[試算]" if args.dry_run else "[実行]",
                                 " / ".join("%s %d" % (k, v) for k, v in filled.items() if v)))
    if unmatched:
        print("住所シートに無かった物件 %d件: %s" % (len(unmatched), " / ".join(unmatched[:12])))
    if not args.dry_run and any(filled.values()):
        bak = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs",
                           "ledger-backup-%s.xlsx" % datetime.now().strftime("%Y%m%d-%H%M%S"))
        shutil.copy2(src, bak)
        print("バックアップ: %s" % bak)
        wb.save(src)
        print("保存した")
    print("★空欄だけを埋めている。既に入っていた値は一切触っていない。")


if __name__ == "__main__":
    main()
