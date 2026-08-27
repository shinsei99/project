#!/usr/bin/env python3
"""管理物件台帳の「フォルダ」列を、現行のDropboxのパスへ直す（2026-08-27・オーナー依頼）。

背景:
  台帳のフォルダ列は共有フォルダの旧構成（`管理物件/` `管理業務/` `駐車場/`）のままで、
  現行の `物件・管理/管理物件/{ビル|マンション|駐車場|その他物件}/<物件名>/` を指していない。
  とくに `管理業務/` は書式・FAX送信票・貼紙などのフォルダで、**物件とは無関係**。

  108物件のうち フォルダ列あり75 / 空33。実フォルダは101。

方針（★推測で書かない）:
  物件名とフォルダ名は綴りが揃っていない（フォルダ側は「大京ビル（囲碁）」のように
  オーナー名が括弧で付く、「SBP」＝「シェローバイクパーク」の略、住所がそのまま名前、など）。
  総当たりの部分一致で埋めると **別物件のフォルダを指す**（実際この照合で
  「サトウビルⅡ → サトウビル」「シェローバイクパーク今福鶴見 → シェローバイクパーク」
  「菊本ビル → 第二菊本ビル」という誤りが出た。正しくは SATOビルⅡ / SBP今福鶴見(大京P) /
  （菊本ビルのフォルダは無い））。

  そこで確度を3段階に分け、**`確` だけを自動で書き込む**。`要` は人が決める。

    確 … 正規化して完全一致 ／ フォルダ名から括弧書きを外すと完全一致
    候 … 別名表（SBP等）や住所表記の揺れで一意に決まった
    要 … 候補が複数、あるいは片方向の部分一致しかない（＝取り違えの危険がある）

使い方:
  python3 fix_ledger_folders.py --dry-run          # 何がどうなるか見る
  python3 fix_ledger_folders.py --report           # 「要確認」だけ一覧で出す
  python3 fix_ledger_folders.py                    # 『確』だけ書き込む
  python3 fix_ledger_folders.py --include-candidate # 『候』も書き込む（人が目を通してから）
"""
from __future__ import annotations

import argparse
import os
import re
import shutil
import sys
import unicodedata
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from ingest_properties import _ledger_path  # noqa: E402

BASE = "物件・管理/管理物件"
FOLDER_COL = 17          # Q列（1始まり）
NAME_COL = 2             # B列
KIND_COL = 1             # A列

# 人が決めた対応（機械では決められなかったものを、確認のうえここに書く）。
# ★ここに書いたものは他のどの規則よりも優先する。流し直しても壊れない。
OVERRIDES = {
    "枚方招堤南町": "ビル/枚方市招堤南町3丁目23-8",
    "枚方高野道": "ビル/枚方市高野道2丁目23-20（鷲見修）",
    "クリスタル京橋": "ビル/クリスタル京橋ビル（高野博行）",
    "ザ・プラザ": "その他物件/ザ・プラザⅡ（The PlazaⅡ）",
    # SBP＝シェローバイクパークの略（2026-08-27 オーナー確認）。
    # 27行と28行は同じ住所（鶴見区横堤1-3-23）で、27＝バイク駐輪・28＝自動車駐車の対。
    # 旧台帳は親フォルダ「シェローバイクパーク」を指していたが、実体は「SBP今福鶴見(大京P)」。
    "シェローバイクパーク今福鶴見": "駐車場/SBP今福鶴見(大京P)",
}

# フォルダ名だけで使われる略称・別名（実フォルダを見て確認したもの）
ALIASES = [
    ("シェローバイクパーク", "SBP"),
    ("ビル", "BLDG"),
    ("パーキング", "P"),
]


def norm(s: str) -> str:
    s = unicodedata.normalize("NFKC", str(s or "")).lower()
    return re.sub(r"[\s　・（）()\[\]【】\-－ー_,、.。]+", "", s)


def strip_paren(s: str) -> str:
    """フォルダ名末尾の括弧書き（オーナー名・別称）を落とす。"""
    return re.sub(r"[（(][^（()）]*[)）]\s*$", "", str(s or "")).strip()


def variants(s: str) -> set:
    out = {norm(s), norm(strip_paren(s))}
    for full, abbr in ALIASES:
        n = norm(s)
        if norm(full) in n:
            out.add(n.replace(norm(full), norm(abbr)))
        if norm(abbr) in n:
            out.add(n.replace(norm(abbr), norm(full)))
    # 住所そのものがフォルダ名のもの（枚方市招堤南町3丁目23-8）に寄せる
    out.add(re.sub(r"\d+丁目.*$", "", norm(s)))
    out.add(norm(s).replace("市", ""))
    return {v for v in out if v}


def scan_folders(root: str) -> dict:
    cats = {}
    base = os.path.join(root, BASE)
    for cat in sorted(os.listdir(base)):
        d = os.path.join(base, cat)
        if cat.startswith(".") or not os.path.isdir(d):
            continue
        cats[cat] = [x for x in sorted(os.listdir(d))
                     if not x.startswith(".") and os.path.isdir(os.path.join(d, x))]
    return cats


def classify(name: str, cats: dict, current: str | None = None):
    """(確度, 'カテゴリ/フォルダ名' or None, 補足) を返す。

    `current`（台帳にいま入っている旧パス）は強い手がかりになる。旧パスは
    **入口（先頭）が変わっただけで、末尾のフォルダ名は今も実在する**ことが多い
    （例: `管理業務/ビル/SATOビルⅡ/` の `SATOビルⅡ` は現存する）。
    物件名とフォルダ名の綴りが違っていても、旧パスの末尾なら人が付けた対応なので信用できる。
    """
    ov = OVERRIDES.get(str(name).strip())
    if ov:
        cat, f = ov.split("/", 1)
        return "確", (cat, f), "人が決めた対応（OVERRIDES）"
    if current:
        leaf = str(current).rstrip("/").split("/")[-1]
        hits = [(cat, f) for cat, names in cats.items() for f in names
                if norm(f) == norm(leaf) or norm(strip_paren(f)) == norm(leaf)]
        if len(hits) == 1:
            return "確", hits[0], "旧パス末尾「%s」が現存" % leaf
    nv = variants(name)
    exact, cand = [], []
    for cat, names in cats.items():
        for f in names:
            fv = variants(f)
            if norm(name) == norm(f) or norm(name) == norm(strip_paren(f)):
                exact.append((cat, f))
            elif nv & fv:
                cand.append((cat, f))
    if len(exact) == 1:
        return "確", exact[0], ""
    if len(exact) > 1:
        return "要", None, "同名フォルダが複数: " + " / ".join(f for _, f in exact)
    if len(cand) == 1:
        return "候", cand[0], "別名・表記ゆれで一致"
    if len(cand) > 1:
        return "要", None, "候補が複数: " + " / ".join(f for _, f in cand[:5])
    return "要", None, "該当フォルダが見つからない"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--report", action="store_true", help="要確認だけ出す")
    ap.add_argument("--include-candidate", action="store_true", help="『候』も書き込む")
    args = ap.parse_args()

    import openpyxl
    src = _ledger_path()
    root = os.path.dirname(src)
    cats = scan_folders(root)
    print("実フォルダ: %s 合計%d" % ({k: len(v) for k, v in cats.items()},
                                     sum(len(v) for v in cats.values())))

    wb = openpyxl.load_workbook(src)
    ws = wb["管理物件台帳"]
    rows = []
    for r in ws.iter_rows(min_row=3, max_row=ws.max_row):
        name = r[NAME_COL - 1].value
        if not name or str(name).startswith("■"):
            continue
        conf, hit, note = classify(str(name), cats, r[FOLDER_COL - 1].value)
        newpath = "%s/%s/%s/" % (BASE, hit[0], hit[1]) if hit else None
        rows.append((r[0].row, str(name), r[KIND_COL - 1].value,
                     r[FOLDER_COL - 1].value, conf, newpath, note))

    n = {"確": 0, "候": 0, "要": 0}
    for _, _, _, _, conf, _, _ in rows:
        n[conf] += 1
    print("判定: 確%d / 候%d / 要%d  （全%d件）\n" % (n["確"], n["候"], n["要"], len(rows)))

    if args.report:
        print("■ 要確認（人が決める）")
        for row, name, kind, cur, conf, new, note in rows:
            if conf == "要":
                print("  %3d %-22s 種別=%-6s  %s" % (row, name[:22], kind, note))
                print("        いまの記載: %s" % cur)
        return

    write_conf = {"確"} | ({"候"} if args.include_candidate else set())
    changed = 0
    for row, name, kind, cur, conf, new, note in rows:
        if conf not in write_conf or not new or cur == new:
            continue
        print("  %3d %-22s [%s] %s" % (row, name[:22], conf, note))
        print("        旧: %s" % cur)
        print("        新: %s" % new)
        if not args.dry_run:
            ws.cell(row=row, column=FOLDER_COL).value = new
        changed += 1

    print("\n%s 書き換え %d件（確%s）" % ("[試算]" if args.dry_run else "[実行]", changed,
                                         "＋候" if args.include_candidate else "のみ"))
    if not args.dry_run and changed:
        bak = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs",
                           "ledger-backup-%s.xlsx" % datetime.now().strftime("%Y%m%d-%H%M%S"))
        shutil.copy2(src, bak)
        print("バックアップ: %s" % bak)
        wb.save(src)
        print("保存した: %s" % src)
    print("\n★『要』は書き換えていない。`--report` で一覧を出して人が決めること。")


if __name__ == "__main__":
    main()
