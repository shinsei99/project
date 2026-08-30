#!/usr/bin/env python3
"""管理物件台帳（Excel）→ properties テーブル → 住所を座標に変換（GIS の土台）。

  python3 ingest_properties.py                 # 取込＋未取得ぶんだけ座標を取る
  python3 ingest_properties.py --no-geocode    # 取込だけ（外部アクセスなし）
  python3 ingest_properties.py --dry-run       # 何件入るか見るだけ
  python3 ingest_properties.py --regeocode     # 座標を取り直す（キャッシュは使う）
  python3 ingest_properties.py --stats         # 現在の登録状況

⚠️ Dropbox(CloudStorage) を読むので **ターミナルから実行**する（launchd常時起動は
   TCCで読めない。/bin/bash にフルディスクアクセスを付けてあれば常駐からも可）。
⚠️ オーナーの連絡先・電話番号の列は**意図的に取り込まない**（地図やLINEに出す情報ではない）。
"""
import argparse
import os
import re
import sys
import unicodedata

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from db.connection import get_conn, query  # noqa: E402
from db.migrate import migrate  # noqa: E402
from services import config, gis  # noqa: E402

# 台帳の列（2行目のヘッダ）→ DBの列
COLUMNS = {
    "種別": "category", "物件名": "name", "分類": "classification", "担当": "assignee_name",
    "住所": "address", "郵便番号": "postal_code", "築年数": "built", "構造": "structure",
    "戸数": "units", "交通": "access", "オーナー": "owner", "フォルダ": "folder",
}
# 取り込まない列（個人情報）
SKIP_COLUMNS = {"連絡先1", "TEL1", "連絡先2", "TEL2", "補足"}

LEDGER_NAME = "★要更新★管理物件台帳.xlsx"


def _ledger_path() -> str:
    """台帳Excelの場所。ナレッジ取込済みならDBが知っているので、それを使う。"""
    row = query("SELECT filepath FROM knowledge_documents WHERE filename=? AND active=1 "
                "ORDER BY id DESC LIMIT 1", (LEDGER_NAME,))
    if row and os.path.exists(row[0]["filepath"]):
        return row[0]["filepath"]
    root = config.get("knowledge_source_dir", "") or ""
    cand = os.path.join(root, LEDGER_NAME)
    if os.path.exists(cand):
        return cand
    raise SystemExit(f"台帳が見つかりません: {LEDGER_NAME}\n"
                     f"（Dropbox共有フォルダ直下。ターミナルから実行しているか確認）")


def _clean(v) -> str:
    if v is None:
        return ""
    s = unicodedata.normalize("NFKC", str(v)).replace("　", " ").strip()
    return re.sub(r"\s+", " ", s)


def _property_id(name: str) -> str:
    """物件名から安定キーを作る（行順が変わっても同じ物件は同じIDになる）。"""
    s = unicodedata.normalize("NFKC", name).lower()
    s = re.sub(r"[\s　]+", "", s)
    s = re.sub(r"[^\w一-鿿ぁ-んァ-ヶー]", "", s)
    return s[:60] or name[:60]


def read_ledger(path: str):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["管理物件台帳"] if "管理物件台帳" in wb.sheetnames else wb[wb.sheetnames[0]]
    header, rows = None, []
    for raw in ws.iter_rows(values_only=True):
        cells = [_clean(c) for c in raw]
        if header is None:
            if "物件名" in cells and "住所" in cells:
                header = cells
            continue
        if not any(cells):
            continue
        rec = {}
        for i, col in enumerate(header):
            if col in SKIP_COLUMNS or col not in COLUMNS:
                continue
            rec[COLUMNS[col]] = cells[i] if i < len(cells) else ""
        name = rec.get("name", "")
        # 「■ 自社」のような区切り行・小計行を捨てる
        if not name or name.startswith("■") or not rec.get("category"):
            continue
        rec["property_id"] = _property_id(name)
        rows.append(rec)
    wb.close()
    return rows


def upsert(rows, dry_run=False) -> dict:
    """既存の座標は保持したまま台帳の内容を反映する（座標を取り直させない）。"""
    added = updated = 0
    if dry_run:
        existing = {r["property_id"] for r in query("SELECT property_id FROM properties")}
        for r in rows:
            if r["property_id"] in existing:
                updated += 1
            else:
                added += 1
        return {"added": added, "updated": updated, "total": len(rows)}
    cols = ["property_id", "name", "category", "classification", "assignee_name", "address",
            "postal_code", "built", "structure", "units", "access", "owner", "folder"]
    with get_conn() as conn:
        seen = []
        for r in rows:
            vals = [r.get(c, "") for c in cols]
            cur = conn.execute(
                "UPDATE properties SET " + ", ".join(f"{c}=?" for c in cols[1:]) +
                ", active=1, updated_at=datetime('now','localtime') WHERE property_id=?",
                (*vals[1:], r["property_id"]),
            )
            if cur.rowcount:
                updated += 1
            else:
                conn.execute(
                    f"INSERT INTO properties ({', '.join(cols)}) VALUES ({', '.join('?' * len(cols))})",
                    vals)
                added += 1
            seen.append(r["property_id"])
        # 台帳から消えた物件は消さずに active=0（履歴と座標を残す）
        if seen:
            ph = ",".join("?" * len(seen))
            conn.execute(f"UPDATE properties SET active=0 WHERE property_id NOT IN ({ph})", seen)
    return {"added": added, "updated": updated, "total": len(rows)}


def geocode_missing(regeocode=False, limit=None) -> dict:
    """座標が無い物件だけ変換する。**同じ住所は二度外部へ問い合わせない**（キャッシュ）。"""
    sql = "SELECT * FROM properties WHERE active=1"
    if not regeocode:
        sql += " AND lat IS NULL AND (geo_status IS NULL OR geo_status NOT IN ('no_address','not_found'))"
    rows = query(sql + " ORDER BY id")
    if limit:
        rows = rows[:limit]
    ok = ng = skipped = 0
    for p in rows:
        if not (p["address"] or "").strip():
            with get_conn() as conn:
                conn.execute("UPDATE properties SET geo_status='no_address', "
                             "updated_at=datetime('now','localtime') WHERE id=?", (p["id"],))
            skipped += 1
            print(f"  − {p['name']}: 住所が空（台帳に未入力）")
            continue
        res = gis.geocode(p["address"])
        with get_conn() as conn:
            if res.get("ok"):
                conn.execute(
                    "UPDATE properties SET lat=?, lon=?, geo_source='gsi', geo_query=?, "
                    "geo_status='ok', updated_at=datetime('now','localtime') WHERE id=?",
                    (res["lat"], res["lon"], res["query"], p["id"]))
                ok += 1
                mark = "（キャッシュ）" if res.get("cached") else ""
                print(f"  ✓ {p['name']}: {res['lat']:.6f}, {res['lon']:.6f} {mark}")
            else:
                conn.execute("UPDATE properties SET geo_status=?, geo_query=?, "
                             "updated_at=datetime('now','localtime') WHERE id=?",
                             (res.get("status"), res.get("query"), p["id"]))
                ng += 1
                print(f"  ✗ {p['name']}: {res.get('error')}")
    return {"ok": ok, "failed": ng, "skipped": skipped}


def stats():
    rows = query("SELECT classification, category, COUNT(*) c, "
                 "SUM(CASE WHEN lat IS NOT NULL THEN 1 ELSE 0 END) g "
                 "FROM properties WHERE active=1 GROUP BY classification, category "
                 "ORDER BY classification, category")
    total = sum(r["c"] for r in rows)
    geo = sum(r["g"] for r in rows)
    print(f"登録物件 {total}件 / 座標あり {geo}件")
    for r in rows:
        print(f"  {r['classification'] or '?':<6} {r['category'] or '?':<10} "
              f"{r['c']:>3}件（座標 {r['g']}）")
    miss = query("SELECT name, address, geo_status FROM properties "
                 "WHERE active=1 AND lat IS NULL")
    if miss:
        print(f"\n座標なし {len(miss)}件:")
        for m in miss:
            print(f"  - {m['name']}（{m['geo_status'] or '未処理'}）住所: {m['address'] or '（空）'}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--no-geocode", action="store_true", help="取込だけ（外部アクセスなし）")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--regeocode", action="store_true", help="座標を取り直す")
    ap.add_argument("--stats", action="store_true")
    ap.add_argument("--limit", type=int)
    args = ap.parse_args()

    migrate()
    if args.stats:
        return stats()

    path = _ledger_path()
    print(f"台帳: {path}")
    rows = read_ledger(path)
    print(f"読み取り {len(rows)}件")
    res = upsert(rows, dry_run=args.dry_run)
    print(f"{'（試算）' if args.dry_run else ''}新規 {res['added']} / 更新 {res['updated']}")
    if args.dry_run:
        return
    if not args.no_geocode:
        print("\n住所→座標（国土地理院・1件1秒。キャッシュ済みは即時）:")
        g = geocode_missing(regeocode=args.regeocode, limit=args.limit)
        print(f"\n取得 {g['ok']} / 失敗 {g['failed']} / 住所なし {g['skipped']}")
    print()
    stats()


if __name__ == "__main__":
    main()
