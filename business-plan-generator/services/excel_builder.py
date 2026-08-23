# -*- coding: utf-8 -*-
"""事業計画案を Excel（.xlsx）で出力する。

会長の既存様式のレイアウトをできるだけ再現しつつ、
「前提条件」シートに入力値・計算式を置き、「事業計画案」シートは
それを参照する数式で構成する。→ 前提を変えると利回り・CF が Excel 上で再計算される。
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .proforma import Inputs, to_tsubo

JP_FONT = "ＭＳ Ｐゴシック"
YEN = '#,##0.0"万"'
YEN0 = '#,##0"万"'
PCT = '0.0"％"'

_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
HEAD_FILL = PatternFill("solid", fgColor="1F3864")   # ネイビー見出し
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")     # 薄い青（小計行）


def _f(sz=11, bold=False, color="000000"):
    return Font(name=JP_FONT, size=sz, bold=bold, color=color)


# ---------------------------------------------------------------------------
# 前提条件シート（入力＋計算式）
# ---------------------------------------------------------------------------
# (name, ラベル, 数値, 表示書式)
def _input_rows(inp: Inputs):
    return [
        ("土地代", "土地代（万）", inp.土地代, YEN0),
        ("建物代", "建物代（万）", inp.建物代, YEN0),
        ("消費税入力", "消費税（万・0で自動）", inp.消費税, YEN0),
        ("設備代", "設備代（万）", inp.設備代, YEN0),
        ("保証金", "保証金・協力金（万）", inp.保証金, YEN0),
        ("借入総額入力", "借入総額（万）", inp.借入総額, YEN0),
        ("土地評価額", "土地評価額（万）", inp.土地評価額, YEN0),
        ("建物評価額", "建物評価額（万）", inp.建物評価額, YEN0),
        ("借入金利", "借入金利（％）", inp.借入金利, PCT),
        ("借入年数", "借入年数（年）", inp.借入年数, "0"),
        ("月額賃料", "月額賃料（万）", inp.月額賃料, YEN),
        ("協力金返済月", "協力金返済/月（万）", inp.協力金返済月, YEN),
        ("固都税土地", "固定資産税・土地（万/年）", inp.固都税土地, YEN),
        ("固都税建物", "固定資産税・建物（万/年）", inp.固都税建物, YEN),
        ("火災保険", "火災保険（万/年）", inp.火災保険, YEN),
        ("リフォーム代", "リフォーム代（万）", inp.リフォーム代, YEN0),
        ("管理費月", "管理費計/月（万）", inp.管理費月, YEN),
        ("法定耐用年数", "法定耐用年数（年）", inp.法定耐用年数, "0"),
        ("築年数", "築年数（年）", inp.築年数, "0"),
        ("法人税", "法人税（万/年）", inp.法人税, YEN0),
        ("登免税土地率", "登録免許税・土地率（％）", inp.登免税土地率, PCT),
        ("登免税建物率", "登録免許税・建物率（％）", inp.登免税建物率, PCT),
        ("仲介料率", "仲介料率（％）", inp.仲介料率, PCT),
        ("仲介料加算", "仲介料加算（万）", inp.仲介料加算, YEN0),
        ("消費税率", "消費税率（％）", inp.消費税率, PCT),
        ("抵当権率", "抵当権設定率（％）", inp.抵当権率, PCT),
        ("取得税土地率", "不動産取得税・土地率（％）", inp.取得税土地率, PCT),
        ("取得税建物率", "不動産取得税・建物率（％）", inp.取得税建物率, PCT),
        ("土地取得税減額", "土地取得税 減額係数", inp.土地取得税減額, "0.00"),
        ("印紙入力", "印紙（万・0で自動）", inp.印紙, YEN),
        ("司法書士その他", "司法書士・その他（万）", inp.司法書士その他, YEN),
        ("予備費", "予備費（万）", inp.予備費, YEN),
    ]


# 計算式行 (name, ラベル, 数式テンプレ, 書式)。{名前} は前提セル参照に置換
def _calc_rows():
    return [
        ("消費税", "消費税", "IF({消費税入力}>0,{消費税入力},{建物代}*{消費税率}/100)", YEN),
        ("売買価格", "売買価格（土地+建物）", "{土地代}+{建物代}", YEN),
        ("登免税土地", "登録免許税（土地）", "{土地評価額}*{登免税土地率}/100", YEN),
        ("登免税建物", "移転登記（建物）", "{建物評価額}*{登免税建物率}/100", YEN),
        # 仲介手数料 速算式（宅建業法上限, realestate-calc 準拠, 消費税込）
        ("仲介料", "仲介料",
         "IF({売買価格}<=0,0,(IF({売買価格}>400,{売買価格}*{仲介料率}/100+{仲介料加算},"
         "IF({売買価格}>200,{売買価格}*4/100+2,{売買価格}*5/100)))*(1+{消費税率}/100))", YEN),
        # 印紙（0入力なら 2024軽減税率の段階表で自動）
        ("印紙", "印紙税",
         "IF({印紙入力}>0,{印紙入力},"
         "IF({売買価格}>50000,48,IF({売買価格}>10000,16,IF({売買価格}>5000,6,"
         "IF({売買価格}>1000,3,IF({売買価格}>500,1,IF({売買価格}>100,0.5,"
         "IF({売買価格}>10,0.1,0.02))))))))", YEN),
        ("借入総額", "借入総額", "{借入総額入力}", YEN),
        ("抵当権設定", "抵当権設定", "{借入総額}*{抵当権率}/100", YEN),
        ("取得税土地", "不動産取得税（土地）", "{土地評価額}*{土地取得税減額}*{取得税土地率}/100", YEN),
        ("取得税建物", "不動産取得税（建物）", "{建物評価額}*{取得税建物率}/100", YEN),
        ("諸経費", "諸経費 合計",
         "{登免税土地}+{登免税建物}+{仲介料}+{抵当権設定}+{取得税土地}+{取得税建物}+{印紙}+{司法書士その他}+{予備費}", YEN),
        ("総事業費", "総事業費",
         "{土地代}+{建物代}+{消費税}+{設備代}+{諸経費}", YEN),
        ("自己資金", "自己資金", "{総事業費}-{保証金}-{借入総額}", YEN),
        ("年収", "年収", "({月額賃料}-{協力金返済月})*12", YEN),
        ("満室年収", "満室年収", "{月額賃料}*12", YEN),
        ("金利平均", "金利（平均）", "{借入総額}*{借入金利}/100*0.5", YEN),
        ("償却年数", "償却年数", "MAX({法定耐用年数}-{築年数},2)", "0"),
        ("償却", "償却（定額法）", "IF({建物代}>0,{建物代}*0.9/{償却年数},0)", YEN),
        ("管理費年", "管理費（年）", "{管理費月}*12", YEN),
        ("運営経費", "運営経費 計",
         "{固都税土地}+{固都税建物}+{火災保険}+{管理費年}+{リフォーム代}", YEN),
        ("支出計償却込", "支出計（償却込）", "{運営経費}+{金利平均}+{償却}", YEN),
        ("返済年", "返済（年）", "IF({借入年数}>0,{借入総額}/{借入年数},0)", YEN),
        ("返済月", "返済（月）", "{返済年}/12", YEN),
        ("実利回り", "実利回り",
         "IF({総事業費}>0,({年収}-{運営経費}-{金利平均}-{償却})/{総事業費}*100,0)", PCT),
        ("経費込利回り", "経費・金利込利回り",
         "IF({総事業費}>0,({年収}-{運営経費}-{金利平均})/{総事業費}*100,0)", PCT),
        ("単純利回り", "単純利回り", "IF({総事業費}>0,{年収}/{総事業費}*100,0)", PCT),
        ("CF借入あり", "CF（借入あり・税込）",
         "{年収}-{運営経費}-{金利平均}-{返済年}-{法人税}", YEN),
        ("CF借入なし", "CF（借入なし・税込）", "{年収}-{運営経費}-{法人税}", YEN),
    ]


def _build_assumptions(wb, inp: Inputs):
    ws = wb.create_sheet("前提条件")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    refs = {}
    r = 1
    ws.cell(r, 1, "■ 入力値（ここを変えると再計算されます）").font = _f(12, True, "1F3864")
    r += 1
    for name, label, val, fmt in _input_rows(inp):
        ws.cell(r, 1, label).font = _f(10)
        c = ws.cell(r, 2, float(val or 0))
        c.number_format = fmt
        c.font = _f(10, True)
        c.fill = PatternFill("solid", fgColor="FFF2CC")  # 入力=薄黄
        refs[name] = f"前提条件!$B${r}"
        r += 1

    r += 1
    ws.cell(r, 1, "■ 計算結果").font = _f(12, True, "1F3864")
    r += 1
    for name, label, tmpl, fmt in _calc_rows():
        ws.cell(r, 1, label).font = _f(10)
        formula = "=" + tmpl.format(**{k: v for k, v in refs.items()})
        c = ws.cell(r, 2, formula)
        c.number_format = fmt
        c.font = _f(10)
        refs[name] = f"前提条件!$B${r}"
        r += 1

    # 商圏データ（政府統計 e-Stat）。画面で取得したときだけ載せる。
    # 金融機関に出す計画で「この賃料で埋まるのか」を公的な数字で裏づけるため。
    if getattr(inp, "商圏データ", ""):
        r += 1
        ws.cell(r, 1, "■ 商圏データ（出典: 政府統計 e-Stat）").font = _f(12, True, "1F3864")
        r += 1
        for line in str(inp.商圏データ).splitlines():
            line = line.strip()
            if not line:
                continue
            ws.cell(r, 1, line).font = _f(9)
            r += 1
    return refs


# ---------------------------------------------------------------------------
# 事業計画案シート（プレゼン）
# ---------------------------------------------------------------------------
def _build_plan(wb, inp: Inputs, refs):
    ws = wb.active
    ws.title = "事業計画案"
    ws.sheet_view.showGridLines = False
    widths = [3, 22, 16, 3, 22, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    def title(text):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row, 1, text)
        c.font = _f(16, True, "FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = HEAD_FILL
        ws.row_dimensions[row].height = 26
        row += 1

    def section(text):
        nonlocal row
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row, 1, "  " + text)
        c.font = _f(12, True, "FFFFFF")
        c.fill = HEAD_FILL
        ws.row_dimensions[row].height = 20
        row += 1

    def kv(label, value, col=1, fmt=None, bold=False, ref=False):
        """label(col) / value(col+1)。ref=True なら value を数式参照扱い。"""
        lc = ws.cell(row, col, label)
        lc.font = _f(10, bold)
        vc = ws.cell(row, col + 1)
        if ref:
            vc.value = f"={value}"
        else:
            vc.value = value
        if fmt:
            vc.number_format = fmt
        vc.font = _f(10, bold)
        vc.alignment = Alignment(horizontal="right")

    def line(label_a, ref_a, label_b=None, ref_b=None, fmt=YEN, bold=False):
        nonlocal row
        kv(label_a, refs[ref_a], col=1, fmt=fmt, bold=bold, ref=True)
        if label_b:
            kv(label_b, refs[ref_b], col=4, fmt=fmt, bold=bold, ref=True)
        row += 1

    def text_row(text, bold=False, italic=False):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row, 1, text)
        c.font = Font(name=JP_FONT, size=10, bold=bold, italic=italic)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        row += 1

    # ---- タイトル ----
    title(f"{inp.物件名 or '（物件名）'}　事業計画案")

    # ---- 物件概要 ----
    section("物件概要")
    shiki = inp.敷地面積
    ws.cell(row, 1, "所 在 地").font = _f(10)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row, 2, inp.所在地).font = _f(10)
    row += 1
    ws.cell(row, 1, "敷地面積").font = _f(10)
    ws.cell(row, 2, f"{shiki:,.2f}㎡（約{to_tsubo(shiki):,.2f}坪）").font = _f(10)
    row += 1
    ws.cell(row, 1, "建　　物").font = _f(10)
    bldg = "　".join([x for x in [inp.建物構造, f"延床{inp.延床面積:,.2f}㎡（約{to_tsubo(inp.延床面積):,.2f}坪）" if inp.延床面積 else "", inp.戸数, inp.駐車場, (f"{inp.築年}築" if inp.築年 else "")] if x])
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row, 2, bldg).font = _f(10)
    row += 1
    ws.cell(row, 1, "都市計画").font = _f(10)
    toshi = "　".join([x for x in [inp.用途地域, (f"建ぺい率{inp.建ぺい率}％" if inp.建ぺい率 else ""), (f"容積率{inp.容積率}％" if inp.容積率 else "")] if x])
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row, 2, toshi).font = _f(10)
    row += 1
    ws.cell(row, 1, "交　　通").font = _f(10)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row, 2, inp.交通).font = _f(10)
    row += 1
    text_row(inp.スキーム文)

    # ---- 資金計画 ----
    section("資金計画（万円）")
    line("土地代", "土地代", "建物代", "建物代")
    line("消費税", "消費税", "設備代", "設備代")
    line("諸経費", "諸経費", "保証金", "保証金")
    line("総事業費", "総事業費", None, None, bold=True)
    line("借入総額", "借入総額", "自己資金", "自己資金", bold=True)

    # ---- 収入 ----
    section("収入（万円）")
    kv("月額賃料", refs["月額賃料"], col=1, fmt=YEN, ref=True)
    kv("×12ヶ月＝年収", refs["年収"], col=4, fmt=YEN, bold=True, ref=True)
    row += 1

    # ---- 支出 ----
    section("支出（万円／年）")
    line("固定資産税（土地）", "固都税土地", "固定資産税（建物）", "固都税建物")
    line("火災保険", "火災保険", "管理費（年）", "管理費年")
    line("リフォーム代", "リフォーム代", "金利（平均）", "金利平均")
    line("償却（定額法）", "償却", "支出計（償却込）", "支出計償却込", bold=True)

    # ---- 返済・利回り ----
    section("返済・利回り")
    line("返済（年）", "返済年", "返済（月）", "返済月")
    line("実利回り", "実利回り", "経費・金利込利回り", "経費込利回り", fmt=PCT, bold=True)
    line("単純利回り", "単純利回り", None, None, fmt=PCT, bold=True)

    # ---- キャッシュフロー ----
    section("キャッシュフロー（税込）")
    line("借入あり", "CF借入あり", "借入なし", "CF借入なし", bold=True)

    # ---- 諸費用内訳 ----
    section("諸費用内訳（万円）")
    line("登録免許税（土地）", "登免税土地", "登録免許税（建物）", "登免税建物")
    line("仲介料（3％＋6万＋税）", "仲介料", "抵当権設定", "抵当権設定")
    line("不動産取得税（土地）", "取得税土地", "不動産取得税（建物）", "取得税建物")
    line("印紙", "印紙", "司法書士・その他", "司法書士その他")
    line("予備費", "予備費", "諸経費 合計", "諸経費", bold=True)

    # ---- フッター ----
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row, 1, f"作成日: {inp.基準日}　／　大京商事株式会社")
    c.font = Font(name=JP_FONT, size=9, color="808080")
    c.alignment = Alignment(horizontal="right")

    ws.print_options.horizontalCentered = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_workbook(inp: Inputs) -> BytesIO:
    wb = Workbook()
    refs = _build_assumptions(wb, inp)
    _build_plan(wb, inp, refs)
    # 前提条件を後ろに
    wb.move_sheet("前提条件", offset=1)
    wb.active = wb["事業計画案"]
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
