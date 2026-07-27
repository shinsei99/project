#!/usr/bin/env python3
"""
入金突合（消込）システム

使い方:
    python match_payments.py \
        --bank CSV1.csv CSV2.csv \
        --rent 入金一覧A.xlsx 入金一覧B.xlsx [--dry-run]

銀行CSV: りそな形式・UFJ形式を自動判定（最大5ファイル）
入金一覧Excel: 複数マンション対応（最大20ファイル）
出力: {元ファイル名}_updated.xlsx  /  name_mapping.csv
"""

import os
import sys
import csv
import re
import argparse
import unicodedata
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Set, Tuple

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

try:
    import pykakasi
    _kks = pykakasi.kakasi()
    HAS_PYKAKASI = True
except ImportError:
    HAS_PYKAKASI = False

# name_mapping.csv はスクリプトと同階層に保存
SCRIPT_DIR = Path(__file__).parent
NAME_MAPPING_FILE = SCRIPT_DIR / 'name_mapping.csv'

# ハイライト色
FILL_EXACT    = PatternFill(fill_type='solid', fgColor='C6EFCE')  # 緑: 完全一致
FILL_SPLIT    = PatternFill(fill_type='solid', fgColor='FFEB9C')  # 黄: 分割入金
FILL_MISMATCH = PatternFill(fill_type='solid', fgColor='FFC7CE')  # 赤: 金額不一致
FILL_CLEAR    = PatternFill(fill_type=None)


# ═══════════════════════════════════════════════════════════════
# ユーティリティ
# ═══════════════════════════════════════════════════════════════

# 銀行振込名義は小文字カナ不使用のため、比較前に大文字へ統一
_SMALL_TO_LARGE = str.maketrans(
    'ァィゥェォャュョッヮヵヶ',
    'アイウエオヤユヨツワカケ'
)


def normalize_kana(text: str) -> str:
    """カナを正規化（半角→全角、小文字カナ→大文字、スペース除去）"""
    if not text:
        return ''
    s = unicodedata.normalize('NFKC', str(text))   # 半角カナ→全角
    s = s.translate(_SMALL_TO_LARGE)                # ジュ→ジユ など
    s = re.sub(r'[\s　]+', '', s)
    return s.upper()


def kanji_to_kana(text: str) -> str:
    """漢字→カタカナ変換（pykakasi）"""
    if not HAS_PYKAKASI or not text:
        return ''
    result = _kks.convert(text)
    kana = ''.join(item['kana'] for item in result)
    return normalize_kana(kana)


def canon_name(name: str) -> str:
    """契約者名の正規化キー（スペース・全角スペース除去）"""
    return name.replace(' ', '').replace('　', '').strip()


def parse_amount(val) -> int:
    if val is None:
        return 0
    s = str(val).replace(',', '').replace('¥', '').replace('￥', '').strip()
    try:
        return int(float(s))
    except ValueError:
        return 0


def parse_date(val) -> Optional[datetime]:
    if not val:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in ['%Y年%m月%d日', '%Y.%m.%d', '%Y/%m/%d', '%Y-%m-%d']:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


# ═══════════════════════════════════════════════════════════════
# 名義変換マスタ (name_mapping.csv)
# ═══════════════════════════════════════════════════════════════

def load_name_mapping(filepath: Path) -> Dict[str, Set[str]]:
    """
    name_mapping.csv → {正規化漢字キー: {正規化銀行カナ, ...}}
    CSV列: kanji_name, bank_kana, room, note
    """
    mapping: Dict[str, Set[str]] = {}
    if not filepath.exists():
        return mapping
    with open(filepath, encoding='utf-8-sig') as f:
        for row in csv.DictReader(f):
            kanji_key = canon_name(row.get('kanji_name', ''))
            bank_kana_norm = normalize_kana(row.get('bank_kana', ''))
            if kanji_key and bank_kana_norm:
                mapping.setdefault(kanji_key, set()).add(bank_kana_norm)
    return mapping


def save_name_mapping(filepath: Path, mapping: Dict[str, Set[str]],
                      rent_records: List[Dict]):
    """
    新テナントを追加してname_mapping.csvを保存。
    既存エントリは上書きしない（ユーザー修正を保持）。
    """
    # 既存CSV読み込み
    existing_rows: List[Dict] = []
    existing_keys: Set[str] = set()
    if filepath.exists():
        with open(filepath, encoding='utf-8-sig') as f:
            for row in csv.DictReader(f):
                existing_rows.append(row)
                existing_keys.add(canon_name(row.get('kanji_name', '')))

    # 新テナントを追加
    for rec in rent_records:
        key = canon_name(rec['tenant'])
        if key in existing_keys:
            continue
        # pykakasi で候補カナを生成
        candidate = kanji_to_kana(rec['tenant']) if HAS_PYKAKASI else ''
        existing_rows.append({
            'kanji_name': rec['tenant'],
            'bank_kana':  candidate,
            'room':       str(rec.get('room', '')),
            'note':       'auto' if candidate else 'manual required',
        })
        existing_keys.add(key)

    with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(
            f, fieldnames=['kanji_name', 'bank_kana', 'room', 'note'])
        writer.writeheader()
        writer.writerows(existing_rows)


# ═══════════════════════════════════════════════════════════════
# 銀行データ読み込み
# ═══════════════════════════════════════════════════════════════

def _detect_bank_format(filepath: str) -> str:
    with open(filepath, encoding='cp932') as f:
        first = f.readline()
    if '照会口座' in first or '勘定日' in first:
        return 'risona'
    return 'ufj'


def _load_risona(filepath: str) -> List[Dict]:
    """りそな形式CSV（ヘッダー行あり）"""
    records = []
    with open(filepath, encoding='cp932') as f:
        for row in csv.DictReader(f):
            if row.get('取引区分', '') != '振込':
                continue
            amount = parse_amount(row.get('入金金額（円）', ''))
            if amount <= 0:
                continue
            name_raw = row.get('摘要', '').strip()
            records.append({
                'source':     Path(filepath).name,
                'date':       parse_date(row.get('勘定日', '')),
                'amount':     amount,
                'name_raw':   name_raw,
                'name_norm':  normalize_kana(name_raw),
            })
    return records


def _load_ufj(filepath: str) -> List[Dict]:
    """UFJ形式CSV（先頭列=2が明細行）"""
    records = []
    with open(filepath, encoding='cp932') as f:
        for row in csv.reader(f):
            if not row or row[0] != '2' or len(row) < 6:
                continue
            amount = parse_amount(row[5])
            if amount <= 0:
                continue
            name_raw = row[3].strip() if len(row) > 3 else ''
            records.append({
                'source':     Path(filepath).name,
                'date':       parse_date(row[1]),
                'amount':     amount,
                'name_raw':   name_raw,
                'name_norm':  normalize_kana(name_raw),
            })
    return records


def load_bank_data(filepaths: List[str]) -> List[Dict]:
    """複数銀行CSVを統合して返す"""
    all_records: List[Dict] = []
    for fp in filepaths:
        fmt = _detect_bank_format(fp)
        loader = _load_risona if fmt == 'risona' else _load_ufj
        recs = loader(fp)
        print(f"  {Path(fp).name} [{fmt}] → {len(recs)}件")
        all_records.extend(recs)
    return all_records


# ═══════════════════════════════════════════════════════════════
# 入金一覧Excel読み込み
# ═══════════════════════════════════════════════════════════════

def _find_header_row(ws) -> int:
    """ヘッダー行を検出（最初の5行内で '部屋' と '契約者' が揃う行）"""
    for row_idx in range(1, 6):
        vals = set()
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row_idx, col).value
            if v:
                vals.add(str(v).strip())
        if '部屋' in vals and '契約者' in vals:
            return row_idx
    return 2


def _detect_cols(ws, header_row: int) -> Dict[str, int]:
    """列名キーワードから列番号を検出"""
    cols: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = str(ws.cell(header_row, col).value or '').strip()
        v = raw.replace('\n', '').replace('　', ' ').strip()
        if v == '部屋':
            cols['room'] = col
        elif '契約者' in v:
            cols['tenant'] = col
        elif '家賃' in v and 'room' not in cols:
            cols.setdefault('rent_start', col)
        elif v == '合計':
            cols['total'] = col
        elif 'UFJ自動引落' in v or 'ＵＦＪ自動引落' in v:
            cols['ufjdebit'] = col
        elif '個別振込' in v:
            cols['individual'] = col
        elif '入金日' in v:
            cols['payment_date'] = col
        elif '未収残' in v:
            cols['unpaid'] = col
        elif '承認' in v and 'payment_date' not in cols:
            cols['approval'] = col
        elif '備' in v and '考' in v:
            cols['notes'] = col
    return cols


def load_rent_list(filepath: str) -> Tuple[Dict[str, int], List[Dict]]:
    """入金一覧Excelを読み込んで (列マップ, テナントリスト) を返す"""
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    header_row = _find_header_row(ws)
    cols = _detect_cols(ws, header_row)

    if 'tenant' not in cols:
        raise ValueError(f"'契約者'列が見つかりません: {filepath}")

    rent_start = cols.get('rent_start', 3)

    records: List[Dict] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        tenant_val = ws.cell(row_idx, cols['tenant']).value
        if not tenant_val:
            continue
        tenant_str = str(tenant_val).strip()
        if not tenant_str:
            continue

        # ── 合計金額取得 ──────────────────────────────────────
        amount = 0
        if 'total' in cols:
            amount = parse_amount(ws.cell(row_idx, cols['total']).value)
        elif 'ufjdebit' in cols or 'individual' in cols:
            v_ufj = parse_amount(ws.cell(row_idx, cols['ufjdebit']).value) if 'ufjdebit' in cols else 0
            v_ind = parse_amount(ws.cell(row_idx, cols['individual']).value) if 'individual' in cols else 0
            amount = v_ufj or v_ind

        # data_only でも取れない場合は費目列を直接合算
        if amount == 0:
            end_col = min(
                cols.get('total', cols.get('ufjdebit', cols.get('individual', rent_start + 8))) - 1,
                rent_start + 7
            )
            for c in range(rent_start, end_col + 1):
                v = ws.cell(row_idx, c).value
                if isinstance(v, (int, float)):
                    amount += int(v)

        if amount == 0:
            continue  # 空行スキップ

        records.append({
            'row':         row_idx,
            'room':        ws.cell(row_idx, cols.get('room', 1)).value,
            'tenant':      tenant_str,
            'tenant_norm': normalize_kana(tenant_str),
            'tenant_key':  canon_name(tenant_str),
            'amount':      amount,
        })

    return cols, records


# ═══════════════════════════════════════════════════════════════
# 突合ロジック
# ═══════════════════════════════════════════════════════════════

def _kana_candidates(tenant: str, tenant_key: str,
                     name_mapping: Dict[str, Set[str]]) -> Set[str]:
    """テナント名に対応する正規化カナ候補セットを返す"""
    candidates: Set[str] = set()
    # 1. name_mapping からの候補
    candidates.update(name_mapping.get(tenant_key, set()))
    # 2. pykakasi 変換
    if HAS_PYKAKASI:
        kana = kanji_to_kana(tenant)
        if kana:
            candidates.add(kana)
    # 3. 契約者名自体をカナ正規化（外国人名・カナ表記テナント向け）
    candidates.add(normalize_kana(tenant))
    candidates.discard('')
    return candidates


def match_payments(bank_records: List[Dict],
                   rent_records: List[Dict],
                   name_mapping: Dict[str, Set[str]]) -> List[Dict]:
    """
    優先順位:
      ① 完全一致（名義一致 + 金額一致）
      ② 表記揺れ一致（name_mapping 経由 + 金額一致）  ← ①と同ロジックで自動吸収
      ③ 分割入金（同日・同名義の合算 = 請求合計）
      ④ 金額不一致アラート（名義一致だが金額ズレ）
      × 未突合
    """
    used: Set[int] = set()
    results: List[Dict] = []

    for rent in rent_records:
        candidates = _kana_candidates(
            rent['tenant'], rent['tenant_key'], name_mapping)
        expected = rent['amount']

        # 名義マッチする銀行レコードを抽出
        matched: List[Tuple[int, Dict]] = [
            (i, b) for i, b in enumerate(bank_records)
            if i not in used and b['name_norm'] in candidates
        ]

        if not matched:
            results.append({
                'rent': rent, 'status': 'unmatched',
                'bank_records': [], 'matched_amount': 0,
                'diff': 0, 'match_date': None,
            })
            continue

        # ① / ② 金額完全一致
        found = False
        for i, bank in matched:
            if bank['amount'] == expected:
                used.add(i)
                results.append({
                    'rent': rent, 'status': 'exact',
                    'bank_records': [bank], 'matched_amount': bank['amount'],
                    'diff': 0, 'match_date': bank['date'],
                })
                found = True
                break

        if found:
            continue

        # ③ 分割入金（同日・同名義の合算）
        from collections import defaultdict
        date_groups: Dict[Optional[str], List[Tuple[int, Dict]]] = defaultdict(list)
        for i, bank in matched:
            date_key = bank['date'].strftime('%Y%m%d') if bank['date'] else 'unknown'
            date_groups[date_key].append((i, bank))

        split_found = False
        for date_key, group in date_groups.items():
            total = sum(b['amount'] for _, b in group)
            if total == expected:
                for i, _ in group:
                    used.add(i)
                results.append({
                    'rent': rent, 'status': 'split',
                    'bank_records': [b for _, b in group],
                    'matched_amount': total,
                    'diff': 0,
                    'match_date': group[0][1]['date'],
                })
                split_found = True
                break

        if split_found:
            continue

        # ④ 金額不一致アラート（最近似額を選択）
        best_i, best_bank = min(matched, key=lambda x: abs(x[1]['amount'] - expected))
        results.append({
            'rent': rent, 'status': 'amount_mismatch',
            'bank_records': [best_bank], 'matched_amount': best_bank['amount'],
            'diff': best_bank['amount'] - expected,
            'match_date': best_bank['date'],
        })

    return results


# ═══════════════════════════════════════════════════════════════
# Excel 更新（フォーマット保持）
# ═══════════════════════════════════════════════════════════════

def update_excel(filepath: str, cols: Dict[str, int],
                 results: List[Dict], dry_run: bool = False) -> Tuple[str, int]:
    """
    突合結果をExcelに書き込む。
    - formulas保持のため data_only=False でロード
    - 入金日・未収残・備考のみ更新
    - 保存先: {元ファイル名}_updated.xlsx
    """
    if dry_run:
        return filepath, 0

    wb = load_workbook(filepath)  # formulas保持
    ws = wb.active
    updated = 0

    for result in results:
        rent    = result['rent']
        row     = rent['row']
        status  = result['status']
        diff    = result['diff']
        date    = result.get('match_date')
        banks   = result.get('bank_records', [])

        # ── 入金日 ────────────────────────────────────────────
        if 'payment_date' in cols and date and status != 'unmatched':
            cell = ws.cell(row, cols['payment_date'])
            cell.value = date
            cell.number_format = 'M/D'

        # ── 未収残 ────────────────────────────────────────────
        if 'unpaid' in cols and status != 'unmatched':
            ws.cell(row, cols['unpaid']).value = diff if diff != 0 else 0

        # ── 備考 / ステータス ─────────────────────────────────
        if 'notes' in cols:
            note_cell = ws.cell(row, cols['notes'])
            existing = str(note_cell.value or '').strip()
            if status == 'split':
                amounts_str = '+'.join(f"{b['amount']:,}" for b in banks)
                tag = f'[分割:{amounts_str}]'
                if tag not in existing:
                    note_cell.value = (tag + ' ' + existing).strip()
            elif status == 'amount_mismatch':
                tag = f'[差異{diff:+,}円]'
                if tag not in existing:
                    note_cell.value = (tag + ' ' + existing).strip()
            elif status == 'unmatched':
                tag = '【未突合】'
                if tag not in existing:
                    note_cell.value = (tag + ' ' + existing).strip()

        # ── 行ハイライト（入金日セルのみ）────────────────────
        if 'payment_date' in cols:
            fill_cell = ws.cell(row, cols['payment_date'])
            if status == 'exact':
                fill_cell.fill = FILL_EXACT
            elif status == 'split':
                fill_cell.fill = FILL_SPLIT
            elif status == 'amount_mismatch':
                fill_cell.fill = FILL_MISMATCH

        if status != 'unmatched':
            updated += 1

    out_path = str(filepath).replace('.xlsx', '_updated.xlsx')
    if out_path == filepath:
        out_path = filepath.replace('.xlsx', '_updated.xlsx')
    wb.save(out_path)
    return out_path, updated


# ═══════════════════════════════════════════════════════════════
# レポート表示
# ═══════════════════════════════════════════════════════════════

def print_report(filename: str, results: List[Dict]):
    exact     = [r for r in results if r['status'] == 'exact']
    split     = [r for r in results if r['status'] == 'split']
    mismatch  = [r for r in results if r['status'] == 'amount_mismatch']
    unmatched = [r for r in results if r['status'] == 'unmatched']

    print(f"\n{'='*60}")
    print(f"  {Path(filename).name}")
    print(f"{'='*60}")
    print(f"  ① 完全一致   : {len(exact):3}件")
    print(f"  ③ 分割入金   : {len(split):3}件")
    print(f"  ④ 金額不一致 : {len(mismatch):3}件  ← 要確認")
    print(f"  × 未突合     : {len(unmatched):3}件")

    if split:
        print(f"\n  【分割入金】")
        for r in split:
            t = r['rent']['tenant']
            amts = '+'.join(f"{b['amount']:,}" for b in r['bank_records'])
            total = r['matched_amount']
            dt = r['match_date'].strftime('%m/%d') if r['match_date'] else '?'
            print(f"    {t}  ({dt}) {amts} = {total:,}円")

    if mismatch:
        print(f"\n  【要確認：金額不一致】")
        for r in mismatch:
            t = r['rent']['tenant']
            exp = r['rent']['amount']
            act = r['matched_amount']
            diff = r['diff']
            dt = r['match_date'].strftime('%m/%d') if r['match_date'] else '?'
            print(f"    {t}  請求={exp:,}  入金={act:,}  差異={diff:+,}円  ({dt})")

    if unmatched:
        print(f"\n  【未突合テナント】")
        for r in unmatched:
            t = r['rent']['tenant']
            amt = r['rent']['amount']
            print(f"    {t}  {amt:,}円")


# ═══════════════════════════════════════════════════════════════
# エントリポイント
# ═══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='入金突合（消込）システム',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
例:
  python match_payments.py \\
      --bank ~/Downloads/①CSV_りそな.csv ~/Downloads/①CSV_UFJ.csv \\
      --rent ~/Downloads/入金一覧メゾン.xlsx ~/Downloads/入金一覧ベリエール.xlsx
        """)
    parser.add_argument('--bank', nargs='+', required=True,
                        help='銀行CSVファイル（最大5ファイル）')
    parser.add_argument('--rent', nargs='+', required=True,
                        help='入金一覧Excelファイル（最大20ファイル）')
    parser.add_argument('--mapping', default=str(NAME_MAPPING_FILE),
                        help=f'名義変換CSV（デフォルト: {NAME_MAPPING_FILE}）')
    parser.add_argument('--dry-run', action='store_true',
                        help='Excelを更新せず結果を表示のみ')
    args = parser.parse_args()

    # ── 銀行データ読み込み ─────────────────────────────────────
    print(f"\n▼ 銀行データ読み込み（{len(args.bank)}ファイル）")
    bank_records = load_bank_data(args.bank)
    print(f"  合計: {len(bank_records)}件の入金レコード")

    # ── 名義マスタ読み込み ─────────────────────────────────────
    mapping_path = Path(args.mapping)
    name_mapping = load_name_mapping(mapping_path)
    print(f"\n▼ 名義マスタ: {sum(len(v) for v in name_mapping.values())}エントリ ({mapping_path.name})")
    if not mapping_path.exists():
        print(f"  → 初回実行のため自動生成します")

    all_rent_records: List[Dict] = []

    # ── 各入金一覧Excelを処理 ──────────────────────────────────
    for excel_path in args.rent:
        print(f"\n▼ 処理中: {Path(excel_path).name}")
        try:
            cols, rent_records = load_rent_list(excel_path)
        except Exception as e:
            print(f"  エラー: {e}")
            continue

        print(f"  テナント数: {len(rent_records)}件")
        all_rent_records.extend(rent_records)

        # 突合実行
        results = match_payments(bank_records, rent_records, name_mapping)

        # レポート表示
        print_report(excel_path, results)

        # Excel更新
        if not args.dry_run:
            try:
                out_path, n = update_excel(excel_path, cols, results)
                print(f"\n  → 保存: {Path(out_path).name}  ({n}件更新)")
            except Exception as e:
                print(f"  Excel保存エラー: {e}")
        else:
            print(f"\n  [dry-run: Excel未更新]")

    # ── 名義マスタ保存 ─────────────────────────────────────────
    save_name_mapping(mapping_path, name_mapping, all_rent_records)
    print(f"\n▼ 名義マスタ保存: {mapping_path}")
    print(f"  ※ bank_kana列を確認・修正して再実行すると②表記揺れ突合が改善します\n")


if __name__ == '__main__':
    main()
