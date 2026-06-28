"""退去精算書テンプレート（seisan_template.xlsx）を生成するスクリプト。

一度実行してテンプレートを作成・コミットする。実行時に既存ファイルを上書きする。
    python templates/generate_template.py
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.properties import PageSetupProperties


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "退去精算書"

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")

    # 列幅（A4縦1ページ幅に収まる値）
    widths = {"A": 20, "B": 11, "C": 8, "D": 11, "E": 11, "F": 29}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # A4縦・1ページ幅フィットの印刷設定（出力側でも再設定するが基準値を持たせる）
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "portrait"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "9:9"

    # タイトル
    ws.merge_cells("A1:F1")
    ws["A1"] = "退去精算書 兼 原状回復見積書"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = center

    # 基本情報
    ws["A3"] = "物件名・部屋番号"
    ws["A4"] = "契約者名"
    ws["A5"] = "入居日 / 退去日"
    ws["A6"] = "総入居期間"
    ws["C5"] = "〜"
    for r in (3, 4, 5, 6):
        ws[f"A{r}"].font = Font(bold=True)

    # 明細ヘッダー（9行目）
    headers = [
        "工事・部材名", "業者見積総額(原価)", "入居者負担率",
        "入居者負担額", "オーナー負担額", "算出根拠メモ",
    ]
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=9, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = wrap_center   # 狭い列でも見出しが折り返して収まるように
        c.border = border
    ws.row_dimensions[9].height = 30   # 折り返し2行分の高さ

    # 明細スタイルの雛形（10行目に罫線だけ用意）
    for col in range(1, 7):
        c = ws.cell(row=10, column=col)
        c.border = border
        c.alignment = Alignment(vertical="center")

    out = os.path.join(os.path.dirname(__file__), "seisan_template.xlsx")
    wb.save(out)
    print(f"created: {out}")


if __name__ == "__main__":
    main()
