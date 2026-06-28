"""退去時確認書兼原状回復費用負担誓約書（Excel）出力サービス。

退去立会い時に損耗が発見された場合、入居者に署名させる誓約書を生成する。
参考フォーマット「taikyoji_kakuninsho_v1.pdf」の文言・レイアウトに基づき再構築。
ヘッダー表（物件名・部屋番号／契約者名／入居期間／鍵／喫煙／残置物）、
原状回復・修繕必要箇所の表、誓約文、署名欄（現住所／新住所／氏名 印／連絡先）で構成する。
"""

from __future__ import annotations

import io
import math

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.properties import PageSetupProperties

from models.restoration_data import RestorationData


THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOTTOM_BORDER = Border(bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# A4縦・6列レイアウト（修繕箇所=A:B / 仕様=C:D / 詳細等=E:F）
WIDTHS = {"A": 13, "B": 13, "C": 12, "D": 12, "E": 14, "F": 14}

# 誓約文（taikyoji_kakuninsho_v2.pdf の文言。喫煙・ペット条項を第1段落に標準で含む）
PLEDGE_LINES = [
    "　私は上記物件の契約解除に際し、貴社立会いのもと、賃貸借契約書に定められた特約制限"
    "（退去時クリーニング費用等）および、私が原状回復にあたって負担する上記修繕箇所の内容に"
    "ついて確認いたしました。なお、喫煙およびペットの飼育（無断飼育含む）に起因する汚損・破損・"
    "異臭等の修繕費用は、ガイドラインの定めに関わらず、通常損耗にあたらない入居者負担の"
    "損害賠償対象であることを承諾します。",
    "　修繕後に請求される工事代金の負担額については、別途提示される原状回復見積書（またはその"
    "概算額）に基づき、金額確定後速やかにお支払いする事を誓約いたします。また、後日送付される"
    "見積書の発行から7日以内に異議申し立てがない場合、当該金額に同意したものとみなします。",
    "　請求期日までに支払いが実行されなかった場合は、保証会社（または連帯保証人）への立替請求"
    "（代位弁済請求）がなされること、および保証会社が立替払いを行った後に私に対して求償権を"
    "行使することについて、一切の異議申し立てをいたしません。",
]

_PT_PER_LINE = 15   # 折り返し1行あたりの高さ(pt)
_PAD = 4            # セル内余白(pt)


def _weight(s: str) -> int:
    """文字列の表示幅（半角=1・全角=2）。"""
    return sum(2 if ord(ch) > 0x2E80 else 1 for ch in str(s))


def _wrap_lines(s: str, width_units: float) -> int:
    """Excel列幅 width_units（半角文字数相当）に対する折り返し行数。"""
    return max(1, math.ceil(_weight(s) / max(1.0, width_units)))


def _row_height(width_text_pairs) -> float:
    """(幅, テキスト) の組から、行に必要な高さ(pt)を返す。"""
    lines = max(_wrap_lines(t, w) for w, t in width_text_pairs)
    return lines * _PT_PER_LINE + _PAD


def _merge_set(ws, cell_range, value=None, font=None, align=None, fill=None,
               border=False, bottom_border=False):
    ws.merge_cells(cell_range)
    top_left = cell_range.split(":")[0]
    cell = ws[top_left]
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if fill:
        cell.fill = fill
    if border or bottom_border:
        from openpyxl.utils import range_boundaries
        c1, r1, c2, r2 = range_boundaries(cell_range)
        b = BOTTOM_BORDER if bottom_border else BORDER
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(row=r, column=c).border = b
    return cell


def build(data: RestorationData, issuer: dict, options: dict | None = None) -> bytes:
    """誓約書 xlsx をバイト列で返す。

    options: {keys_count, key_replacement_cost, smoking, leftover, witness_date}
    """
    options = options or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "誓約書"

    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w

    base = Font(size=11)
    bold = Font(size=11, bold=True)

    # ── タイトル ──
    _merge_set(ws, "A1:F2", "【退去時確認書兼原状回復費用負担誓約書】",
               font=Font(size=15, bold=True), align=CENTER)

    # ── ヘッダー表（◎項目）──
    def _d(d):
        return d.strftime("%Y年%m月%d日") if d else "―"

    key_cost = int(options.get("key_replacement_cost") or 0)
    keys_text = f"{options.get('keys_count', '')}　本"
    if key_cost > 0:
        keys_text += f"　（返却不足によるカギ交換代：¥{key_cost:,}）"
    y, m = data.residence_period
    period = f"{_d(data.move_in_date)}　〜　{_d(data.move_out_date)}　（入居期間：{y}年{m}ヶ月）"
    room = f"{data.property_name}　{data.room_number}".strip()

    rows = [
        ("◎物件名・部屋番号", room),
        ("◎契約者名", f"{data.tenant_name} 様" if data.tenant_name else ""),
        ("◎入居期間", period),
        ("◎鍵の返却", keys_text),
        ("◎喫煙の有無", f"{options.get('smoking', '有　　・　　無')}　（※電子タバコ・加熱式タバコ等を含む）"),
        ("◎ペット飼育の有無", f"{options.get('pet', '有　　・　　無')}　（※無断飼育、一時的な預かり、種類を問わずすべて含む）"),
        ("◎残置物の有無",
         f"{options.get('leftover', '有　　・　　無')}　（※「有」の場合、合意した残置物の所有権を放棄し、貴社の処分に同意します）"),
    ]
    r = 4
    for label, value in rows:
        _merge_set(ws, f"A{r}:B{r}", label, font=bold, align=LEFT, fill=HEADER_FILL, border=True)
        _merge_set(ws, f"C{r}:F{r}", value, font=base, align=LEFT, border=True)
        ws.row_dimensions[r].height = _row_height([(26, label), (52, value)])
        r += 1

    # ── 原状回復・修繕必要箇所 ──
    r += 1
    _merge_set(ws, f"A{r}:F{r}", "【原状回復・修繕必要箇所】", font=bold, align=LEFT)
    r += 1
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    detail_header = "詳細等（故意・過失、喫煙、ペット損耗の別を含む）"
    _merge_set(ws, f"A{r}:B{r}", "修繕箇所", font=bold, align=wrap_center, fill=HEADER_FILL, border=True)
    _merge_set(ws, f"C{r}:D{r}", "仕　様", font=bold, align=wrap_center, fill=HEADER_FILL, border=True)
    _merge_set(ws, f"E{r}:F{r}", detail_header, font=bold, align=wrap_center, fill=HEADER_FILL, border=True)
    ws.row_dimensions[r].height = _row_height([(26, "修繕箇所"), (28, detail_header)])
    r += 1

    # 入居者負担のある項目＋鍵交換代を転記
    item_rows: list[tuple[str, str, str, int]] = []
    for it in data.items:
        if it.tenant_amount > 0:
            spec = it.material_type
            if it.total_qty:
                spec += f"（{it.total_qty:g}{it.unit}）"
            item_rows.append((it.name, spec, f"入居者負担 ¥{it.tenant_amount:,}", it.tenant_amount))
    if key_cost > 0:
        item_rows.append(("鍵交換（返却不足）", "シリンダー交換", f"入居者負担 ¥{key_cost:,}", key_cost))

    rows_needed = max(len(item_rows), 5)
    for i in range(rows_needed):
        if i < len(item_rows):
            name, spec, detail, _ = item_rows[i]
            _merge_set(ws, f"A{r}:B{r}", name, font=base, align=LEFT, border=True)
            _merge_set(ws, f"C{r}:D{r}", spec, font=base, align=LEFT, border=True)
            _merge_set(ws, f"E{r}:F{r}", detail, font=base, align=LEFT, border=True)
        else:
            _merge_set(ws, f"A{r}:B{r}", None, border=True)
            _merge_set(ws, f"C{r}:D{r}", None, border=True)
            _merge_set(ws, f"E{r}:F{r}", None, border=True)
        ws.row_dimensions[r].height = 19
        r += 1

    if item_rows:
        total = sum(amt for *_, amt in item_rows)
        _merge_set(ws, f"A{r}:D{r}", "入居者負担合計（税込・概算）", font=bold,
                   align=Alignment(horizontal="right", vertical="center"), border=True)
        _merge_set(ws, f"E{r}:F{r}", f"¥{total:,}", font=bold,
                   align=Alignment(horizontal="right", vertical="center"), border=True)
        r += 1

    # ── 誓約文（喫煙・ペットは第1段落に標準で含む。残置物の所有権放棄は◎残置物欄に注記）──
    r += 1
    for line in PLEDGE_LINES:
        _merge_set(ws, f"A{r}:F{r}", line, font=base, align=LEFT_TOP)
        ws.row_dimensions[r].height = _row_height([(78, line)])
        r += 1

    # ── 署名日 ──
    r += 1
    _merge_set(ws, f"A{r}:F{r}", "令和　　　　年　　　　　月　　　　　日",
               font=base, align=Alignment(horizontal="right", vertical="center"))
    r += 1

    # ── 立会人（管理会社＝発行元情報）──
    contact = issuer.get("tel", "")
    if issuer.get("fax"):
        contact = f"{contact}　FAX {issuer['fax']}" if contact else f"FAX {issuer['fax']}"

    def _info_row(label, value):
        nonlocal r
        _merge_set(ws, f"A{r}:A{r}", label, font=bold, align=LEFT)
        _merge_set(ws, f"B{r}:F{r}", value, font=base, align=LEFT)
        ws.row_dimensions[r].height = _row_height([(13, label), (65, value)])
        r += 1

    _info_row("立会人", issuer.get("name", ""))
    _info_row("住　所", issuer.get("address", ""))
    _info_row("連絡先", contact)
    r += 1

    # ── 署名欄（入居者）※囲み罫線ではなく下罫線 ──
    def _sign_row(label, value_range, with_seal=False):
        nonlocal r
        _merge_set(ws, f"A{r}:A{r}", label, font=bold, align=LEFT)
        _merge_set(ws, value_range, None, align=LEFT, bottom_border=True)
        if with_seal:
            _merge_set(ws, f"F{r}:F{r}", "印", font=base, align=CENTER)
        ws.row_dimensions[r].height = 26
        r += 1

    _merge_set(ws, f"A{r}:F{r}", "＜入居者＞", font=bold, align=LEFT)
    ws.row_dimensions[r].height = 18
    r += 1
    _sign_row("現住所", f"B{r}:F{r}")
    _sign_row("新住所", f"B{r}:F{r}")
    _sign_row("氏　名", f"B{r}:E{r}", with_seal=True)
    _sign_row("連絡先", f"B{r}:F{r}")

    # ── A4縦・1ページ幅フィット ──
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "portrait"
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1   # 縦も1ページに収める（誓約書は必ずA4 1枚）
    for side in ("left", "right"):
        setattr(ws.page_margins, side, 0.55)
    for side in ("top", "bottom"):
        setattr(ws.page_margins, side, 0.5)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
