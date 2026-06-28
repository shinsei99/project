"""退去精算書（Excel）出力サービス。

`templates/seisan_template.xlsx` を読み込み、openpyxl で指定セルに
データを流し込む。テンプレートのレイアウト・罫線を崩さないよう、
明細テーブルは必要行数だけ既存スタイルを複製して書き込む。
"""

from __future__ import annotations

import copy
import io
import os

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.properties import PageSetupProperties

from models.restoration_data import RestorationData


TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)), "templates", "seisan_template.xlsx"
)

# 明細テーブルの開始行（テンプレートと一致させる）
ITEM_START_ROW = 10
COLUMN_HEADER_ROW = 9

# A4縦1ページ幅に収めるための列幅（文字数単位）
COLUMN_WIDTHS = {"A": 20, "B": 11, "C": 8, "D": 11, "E": 11, "F": 29}


def _setup_a4_print(ws, last_row: int) -> None:
    """A4縦・1ページ幅フィット・列見出し繰り返しの印刷設定を適用する。"""
    ws.page_setup.paperSize = 9          # A4
    ws.page_setup.orientation = "portrait"
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1         # 横は必ず1ページ幅に収める
    ws.page_setup.fitToHeight = 0        # 縦は複数ページ可
    ws.page_margins.left = 0.55
    ws.page_margins.right = 0.55
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3
    # 2ページ目以降にも列見出し行を繰り返す
    ws.print_title_rows = f"{COLUMN_HEADER_ROW}:{COLUMN_HEADER_ROW}"
    # 印刷範囲を明示
    ws.print_area = f"A1:F{last_row}"
    # 列幅を A4 に収まる値へ調整
    for col, w in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = w


def _fmt_date(d) -> str:
    return d.strftime("%Y年%m月%d日") if d else ""


def _yen(n: int) -> str:
    return f"¥{n:,}"


def build(data: RestorationData) -> bytes:
    """精算書 xlsx をバイト列で返す。"""
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # ---- ヘッダー・基本情報 ----
    ws["B3"] = f"{data.property_name}　{data.room_number}"
    ws["B4"] = f"{data.tenant_name} 様"
    ws["B5"] = _fmt_date(data.move_in_date)
    ws["D5"] = _fmt_date(data.move_out_date)
    ws["B6"] = data.residence_label

    # ---- 明細テーブル ----
    # テンプレートの ITEM_START_ROW 行のスタイルを雛形として複製
    template_cells = {
        col: ws.cell(row=ITEM_START_ROW, column=col)
        for col in range(1, 7)  # A〜F
    }

    row = ITEM_START_ROW
    for item in data.items:
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            tmpl = template_cells[col]
            if tmpl.has_style:
                cell._style = copy.copy(tmpl._style)
        ws.cell(row=row, column=1, value=item.name)                       # A 工事・部材名
        ws.cell(row=row, column=2, value=item.vendor_amount)             # B 業者見積総額（原価）
        ws.cell(row=row, column=3, value=f"{item.tenant_rate_pct}%")     # C 入居者負担率
        ws.cell(row=row, column=4, value=item.tenant_amount)            # D 入居者負担額
        ws.cell(row=row, column=5, value=item.owner_amount)            # E オーナー負担額
        memo = ws.cell(row=row, column=6, value=item.basis)            # F 算出根拠メモ
        memo.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        row += 1

    # ---- 最終サマリー ----
    summary_row = row + 1
    bold = Font(bold=True, size=12)

    ws.cell(row=summary_row, column=1, value="入居者負担総額").font = bold
    c = ws.cell(row=summary_row, column=4, value=data.total_tenant)
    c.font = bold

    ws.cell(row=summary_row + 1, column=1, value="預かり敷金").font = bold
    ws.cell(row=summary_row + 1, column=4, value=data.deposit).font = bold

    settlement = data.settlement
    if settlement >= 0:
        label = "差引精算額（敷金返還額）"
    else:
        label = "差引精算額（不足請求額）"
    ws.cell(row=summary_row + 2, column=1, value=label).font = Font(bold=True, size=13)
    sc = ws.cell(row=summary_row + 2, column=4, value=abs(settlement))
    sc.font = Font(bold=True, size=13, color="C00000")

    # 金額セルに通貨表示形式を付与
    for r in range(ITEM_START_ROW, summary_row + 3):
        for col in (2, 4, 5):
            cell = ws.cell(row=r, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '¥#,##0'

    # A4印刷設定（横・1ページ幅フィット）
    _setup_a4_print(ws, summary_row + 2)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
