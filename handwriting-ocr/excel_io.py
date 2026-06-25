"""Excelの読み込み・セル書き込みロジック。"""
from __future__ import annotations

import io
import re
from datetime import datetime, timedelta

from openpyxl import load_workbook

_SERIAL_MIN = 30000
_SERIAL_MAX = 60000
_HEADER_KEYWORDS = {"水道", "フロア"}


def _serial_to_label(serial: int) -> str:
    dt = datetime(1899, 12, 30) + timedelta(days=serial)
    return f"{dt.month}月{dt.day}日"


def _is_date_serial(val) -> bool:
    return isinstance(val, int) and _SERIAL_MIN <= val <= _SERIAL_MAX


def _cell_to_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return f"{val.month}月{val.day}日"
    return str(val).strip()


def extract_excel_text(file_bytes: bytes) -> tuple[str, list[str], dict[str, int]]:
    """ExcelをRow番号付きテキストに変換し、日付列マップも返す。

    Returns:
        text_repr   : Claude に渡す文字列
        date_list   : 検出した検針日リスト（列順）
        date_col_map: {「M月D日」: 列番号(1始まり)}
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # ── Step 1: ヘッダー行を見つけ、日付列マップを作る ──────────────
    # 「水道」「フロア」などのキーワードがある行がヘッダー行。
    # その行でシリアル値を持つ列を日付列として記録する。
    #
    # col_serial_map: {列番号: シリアル値}
    #   → テキスト生成時に「ヘッダー行の値のみ」日付ラベルへ置換するために使う
    col_serial_map: dict[int, int] = {}

    for r in range(1, min(21, ws.max_row + 1)):
        first_vals = [ws.cell(r, c).value for c in range(1, 5)]
        has_keyword = any(
            str(v).strip() in _HEADER_KEYWORDS for v in first_vals if v is not None
        )
        if not has_keyword:
            continue
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if _is_date_serial(val) and c not in col_serial_map:
                col_serial_map[c] = val

    # シリアル値が見つからない場合は文字列「M月D日」でフォールバック
    date_col_map: dict[str, int] = {}
    if col_serial_map:
        for c, serial in sorted(col_serial_map.items()):
            label = _serial_to_label(serial)
            if label not in date_col_map:
                date_col_map[label] = c
    else:
        for r in range(1, min(21, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                s = _cell_to_str(ws.cell(r, c).value)
                if re.fullmatch(r"\d{1,2}月\d{1,2}日", s) and s not in date_col_map:
                    date_col_map[s] = c

    # ── Step 2: Excelをテキスト化 ────────────────────────────────────
    # 日付列のセルが「ヘッダー行と同じシリアル値」の場合だけラベルに置換。
    # データ行のメーター値（たまたまシリアル範囲に入る）は置換しない。
    lines: list[str] = []
    for r in range(1, min(ws.max_row + 1, 150)):
        cells: list[str] = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if c in col_serial_map and val == col_serial_map[c]:
                # ヘッダー行と同じシリアル値 → 日付ラベルに置換
                s = _serial_to_label(val)
            else:
                s = _cell_to_str(val)
            cells.append(s)
        lines.append(f"Row{r}:\t" + "\t".join(cells))

    return "\n".join(lines), list(date_col_map.keys()), date_col_map


def extract_prev_values(file_bytes: bytes, compact_rows: str, prev_col: int) -> str:
    """前月の指示数をコンパクトテキストの各行に付加して返す。

    compact_rows の各行: "Row5: 子 / セブンイレブン"
    出力例:            "Row5: 子 / セブンイレブン  [前月:4,413]"

    日付シリアル値（ヘッダー行と同一値）は前月値とみなさず除外する。
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # ヘッダー行で prev_col の日付シリアル値を取得（除外用）
    header_serial_for_prev_col = None
    for r in range(1, min(21, ws.max_row + 1)):
        first_vals = [ws.cell(r, c).value for c in range(1, 5)]
        has_keyword = any(
            str(v).strip() in _HEADER_KEYWORDS for v in first_vals if v is not None
        )
        if has_keyword:
            val = ws.cell(r, prev_col).value
            if _is_date_serial(val):
                header_serial_for_prev_col = val
                break

    import re
    lines = []
    for line in compact_rows.splitlines():
        m = re.match(r"Row(\d+):", line)
        if not m:
            lines.append(line)
            continue
        r = int(m.group(1))
        val = ws.cell(row=r, column=prev_col).value
        # 日付シリアル（ヘッダーと同値）は除外
        if val == header_serial_for_prev_col:
            lines.append(line)
            continue
        if isinstance(val, (int, float)) and val > 0:
            lines.append(f"{line}  [前月:{val:,}]")
        else:
            lines.append(line)
    return "\n".join(lines)


def detect_target_date(file_bytes: bytes, date_col_map: dict[str, int]) -> str | None:
    """データが入っていない最初の日付列を自動検出して返す。

    ヘッダー行の日付シリアル値自体は「データ」とみなさず、
    実データ（メーター値）が全行空の列を「未入力列」として返す。
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    ordered = sorted(date_col_map.items(), key=lambda x: x[1])

    for label, col in ordered:
        has_data = False
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=col).value
            # 空・ゼロはスキップ
            if val is None or val == 0 or val == "":
                continue
            # 負の値は差引行の計算式結果（未入力時は -前月値 になる）のためスキップ
            if isinstance(val, (int, float)) and val < 0:
                continue
            # ヘッダー行の日付シリアル値はスキップ
            if _is_date_serial(val) and _serial_to_label(val) == label:
                continue
            has_data = True
            break
        if not has_data:
            return label

    return ordered[-1][0] if ordered else None


def extract_compact_text(file_bytes: bytes) -> str:
    """識別子のある行だけ抽出した最小限のExcel構造テキストを返す。

    差引行・空行・タイトル行を除き、Row番号 + 識別子（フロア/契約者名/種類）のみ出力する。
    これをClaudeプロンプトに使うことで処理を高速化する。
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    _NOISE = {"大京天王寺ビル", "R8年度", "R7年度", "R6年度"}

    lines: list[str] = []
    for r in range(1, ws.max_row + 1):
        id_cells = [_cell_to_str(ws.cell(r, c).value) for c in range(1, 5)]
        # 識別子列が全部空 → 差引行などのため除外
        if not any(id_cells[:3]):
            continue
        # ノイズ行（タイトルや注記）を除外
        first = id_cells[0]
        if any(kw in first for kw in _NOISE) or len(first) > 30:
            continue
        label = " / ".join(v for v in id_cells[:3] if v)
        lines.append(f"Row{r}: {label}")

    return "\n".join(lines)


def get_row_labels(file_bytes: bytes, row_indices: list[int]) -> dict[int, str]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    labels: dict[int, str] = {}
    for r in row_indices:
        parts = []
        for c in range(1, min(6, ws.max_column + 1)):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip():
                parts.append(str(v).strip())
        labels[r] = " / ".join(parts) if parts else f"Row {r}"
    return labels


def apply_cell_updates(template_bytes: bytes, updates: list[dict]) -> bytes:
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active
    for u in updates:
        try:
            r, c = int(u["row"]), int(u["col"])
            existing = ws.cell(row=r, column=c).value
            # 既にデータが入っているセルは絶対に上書きしない
            if existing is not None and existing != "" and existing != 0:
                continue
            ws.cell(row=r, column=c, value=u["value"])
        except (KeyError, ValueError, TypeError):
            continue
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
