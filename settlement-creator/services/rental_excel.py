"""賃貸 初回精算の帳票（Excel）出力サービス。

入居者用（精算明細書兼請求書）とオーナー用（初回送金精算明細書）の2シートを
openpyxl で生成する。発行企業情報をヘッダーに、振込先・支払期限・必要書類を
フッターに印字し、A4縦でそのまま配布できる体裁に仕上げる。
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.properties import PageSetupProperties

from models.rental_data import RentalData, RentLine
from services.issuer_store import ISSUER_FIELDS


_thin = Side(style="thin", color="999999")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_head_fill = PatternFill("solid", fgColor="D9E1F2")
_total_fill = PatternFill("solid", fgColor="FFF2CC")
_center = Alignment(horizontal="center", vertical="center")
_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
_right = Alignment(horizontal="right", vertical="center")

COLUMN_WIDTHS = {"A": 30, "B": 16, "C": 30}


def _setup_a4(ws, last_row: int) -> None:
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "portrait"
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.55
    ws.page_margins.right = 0.55
    ws.page_margins.top = 0.7
    ws.page_margins.bottom = 0.7
    ws.print_area = f"A1:C{last_row}"
    for col, w in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = w


def _issuer_lines(issuer: dict) -> list[str]:
    out = []
    if issuer.get("name"):
        out.append(issuer["name"])
    if issuer.get("representative"):
        out.append(f"代表者　{issuer['representative']}")
    if issuer.get("address"):
        out.append(issuer["address"])
    tel_fax = []
    if issuer.get("tel"):
        tel_fax.append(f"TEL {issuer['tel']}")
    if issuer.get("fax"):
        tel_fax.append(f"FAX {issuer['fax']}")
    if tel_fax:
        out.append("　".join(tel_fax))
    if issuer.get("registration_no"):
        out.append(f"登録番号 {issuer['registration_no']}")
    return out


def _write_sheet(ws, title: str, addressee: str, items: list[RentLine],
                 total_label: str, issuer: dict, footer: list[tuple[str, str]],
                 intro: str) -> None:
    for col, w in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = w

    # タイトル
    ws.merge_cells("A1:C1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=18)
    ws["A1"].alignment = _center
    ws.row_dimensions[1].height = 28

    # 宛名（左）／発行企業（右） を3行目以降に
    ws.merge_cells("A3:A3")
    ws["A3"] = f"{addressee} 様"
    ws["A3"].font = Font(bold=True, size=13)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")

    row = 3
    for line in _issuer_lines(issuer):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        c = ws.cell(row=row, column=2, value=line)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.font = Font(size=10)
        row += 1
    row = max(row, 6)

    # 導入文
    if intro:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row=row, column=1, value=intro).alignment = _left
        row += 1

    total = sum(i.amount for i in items)

    # 合計金額（上部に大きく）
    row += 1
    ws.cell(row=row, column=1, value=total_label).font = Font(bold=True, size=13)
    ws.cell(row=row, column=1).fill = _total_fill
    ws.cell(row=row, column=1).border = _border
    tc = ws.cell(row=row, column=2, value=total)
    tc.font = Font(bold=True, size=14, color="C00000")
    tc.number_format = '¥#,##0;¥-#,##0'
    tc.alignment = _right
    tc.fill = _total_fill
    tc.border = _border
    ws.cell(row=row, column=3, value="").fill = _total_fill
    ws.cell(row=row, column=3).border = _border
    ws.row_dimensions[row].height = 24
    row += 2

    # 明細ヘッダー
    for col, head in zip((1, 2, 3), ("項　目", "金額（円）", "備考")):
        c = ws.cell(row=row, column=col, value=head)
        c.fill = _head_fill
        c.font = Font(bold=True)
        c.alignment = _center
        c.border = _border
    row += 1

    for it in items:
        ws.cell(row=row, column=1, value=it.label).alignment = _left
        amt = ws.cell(row=row, column=2, value=it.amount)
        amt.number_format = '¥#,##0;¥-#,##0'
        amt.alignment = _right
        ws.cell(row=row, column=3, value=it.note).alignment = _left
        for col in (1, 2, 3):
            ws.cell(row=row, column=col).border = _border
        row += 1

    # 合計行
    ws.cell(row=row, column=1, value="合　計").font = Font(bold=True)
    ws.cell(row=row, column=1).alignment = _right
    tc2 = ws.cell(row=row, column=2, value=total)
    tc2.font = Font(bold=True, size=12)
    tc2.number_format = '¥#,##0;¥-#,##0'
    tc2.alignment = _right
    for col in (1, 2, 3):
        cell = ws.cell(row=row, column=col)
        cell.fill = _total_fill
        cell.border = _border
    row += 2

    # フッター（振込先・支払期限・必要書類など）
    for label, value in footer:
        if not value:
            continue
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = _left
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.cell(row=row, column=2, value=value).alignment = _left
        for col in (1, 2, 3):
            ws.cell(row=row, column=col).border = _border
        row += 1

    _setup_a4(ws, row - 1)


def build(data: RentalData, issuer: dict, issue_date: str = "") -> bytes:
    """入居者用・オーナー用の2シートを持つ賃貸精算帳票 xlsx を返す。"""
    issuer = {f: issuer.get(f, "") for f in ISSUER_FIELDS}
    bank = issuer.get("bank", "")

    wb = Workbook()

    # ---- 入居者用：精算明細書兼請求書 ----
    ws_t = wb.active
    ws_t.title = "入居者用請求書"
    tenant_footer = [
        ("【振込先】", bank),
        ("支払期限", data.payment_due),
        ("必要書類", data.required_docs),
        ("物件", f"{data.property_location}　{data.room_number}".strip()),
    ]
    _write_sheet(
        ws_t, "精算明細書 兼 請求書", data.tenant_name or "入居者",
        data.tenant_items, "ご請求金額", issuer, tenant_footer,
        "下記のとおりご請求申し上げます。",
    )

    # ---- オーナー用：初回送金精算明細書 ----
    ws_o = wb.create_sheet("オーナー用送金精算")
    owner_footer = [
        ("物件", f"{data.property_location}　{data.room_number}".strip()),
    ]
    _write_sheet(
        ws_o, "初回送金精算明細書", data.landlord_name or "オーナー",
        data.owner_items, "送金金額（オーナー手取り）", issuer, owner_footer,
        "下記のとおり初回送金を精算いたしました。",
    )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
