"""決済案内書（Excel）出力サービス。

買主用・売主用の2シートを持つワークブックを openpyxl で生成する。
実務でそのまま印刷・PDF化して配布できるA4縦レイアウトに仕上げる。
最終合計セルは太字＋塗りつぶしで強調する。
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.properties import PageSetupProperties

from models.settlement_data import SettlementData, SettlementDoc


# 当日の持ち物チェックリスト（役割別）
BELONGINGS = {
    "買主": [
        "ご実印", "ご印鑑証明書", "ご本人確認書類（運転免許証等）",
        "残代金・諸費用のご準備（お振込手配）", "住民票", "通帳・銀行届出印",
    ],
    "売主": [
        "ご実印", "ご印鑑証明書", "ご本人確認書類（運転免許証等）",
        "権利証（登記識別情報）", "通帳（着金確認用）", "鍵一式・関係書類",
    ],
}

_thin = Side(style="thin", color="999999")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_header_fill = PatternFill("solid", fgColor="305496")
_header_font = Font(bold=True, color="FFFFFF", size=11)
_total_fill = PatternFill("solid", fgColor="FFF2CC")  # 薄い黄色
_center = Alignment(horizontal="center", vertical="center")
_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
_right = Alignment(horizontal="right", vertical="center")

# A4縦1ページ幅に収める列幅
COLUMN_WIDTHS = {"A": 26, "B": 16, "C": 44}


def _setup_a4(ws, last_row: int) -> None:
    ws.page_setup.paperSize = 9          # A4
    ws.page_setup.orientation = "portrait"
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.55
    ws.page_margins.right = 0.55
    ws.page_margins.top = 0.7
    ws.page_margins.bottom = 0.7
    ws.print_area = f"A1:C{last_row}"
    for col, w in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = w


def _section_header(ws, row: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _header_fill
    c.font = _header_font
    c.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = _border


def _fmt_date(d) -> str:
    return d.strftime("%Y年%m月%d日") if d else "（決済日 未定）"


def _write_sheet(ws, data: SettlementData, doc: SettlementDoc) -> None:
    for col, w in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = w

    # タイトル
    ws.merge_cells("A1:C1")
    ws["A1"] = f"決済案内書（{doc.role}様用）"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = _center
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:C2")
    ws["A2"] = f"{doc.name or doc.role} 様"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    row = 4
    # 1. 決済日時・場所
    _section_header(ws, row, "1. 決済日時・場所")
    row += 1
    info = [
        ("物件所在", data.property_location or "—"),
        ("決済日（引き渡し日）", _fmt_date(data.settlement_date)),
        ("固都税起算", data.region_label),
    ]
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=1).border = _border
        ws.cell(row=row, column=1).alignment = _left
        m = ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        vc = ws.cell(row=row, column=2, value=value)
        vc.alignment = _left
        for col in (1, 2, 3):
            ws.cell(row=row, column=col).border = _border
        row += 1

    row += 1
    # 2. 決済時に必要な物
    _section_header(ws, row, "2. 決済時にお持ちいただく物")
    row += 1
    for item in BELONGINGS.get(doc.role, []):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row=row, column=1, value=f"☐  {item}")
        c.alignment = _left
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = _border
        row += 1

    row += 1
    # 3. 精算金内訳
    _section_header(ws, row, "3. 精算金内訳")
    row += 1
    # 列見出し
    for col, head in zip((1, 2, 3), ("項目", "金額", "計算内訳")):
        c = ws.cell(row=row, column=col, value=head)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
        c.font = Font(bold=True)
        c.alignment = _center
        c.border = _border
    row += 1

    for line in doc.lines:
        ws.cell(row=row, column=1, value=line.label).alignment = _left
        amt = ws.cell(row=row, column=2, value=line.amount)
        amt.number_format = '¥#,##0;¥-#,##0'
        amt.alignment = _right
        note = ws.cell(row=row, column=3, value=line.note)
        note.alignment = _left
        for col in (1, 2, 3):
            ws.cell(row=row, column=col).border = _border
        row += 1

    # 差引合計（強調）
    total_label = doc.total_label or "差引合計額"
    lc = ws.cell(row=row, column=1, value=total_label)
    lc.font = Font(bold=True, size=12)
    lc.alignment = _left
    tc = ws.cell(row=row, column=2, value=doc.total)
    tc.font = Font(bold=True, size=13, color="C00000")
    tc.number_format = '¥#,##0;¥-#,##0'
    tc.alignment = _right
    ws.cell(row=row, column=3, value="")
    for col in (1, 2, 3):
        cell = ws.cell(row=row, column=col)
        cell.fill = _total_fill
        cell.border = _border
    ws.row_dimensions[row].height = 22
    row += 1

    # 備考（自動文言）
    if doc.notes:
        row += 1
        _section_header(ws, row, "4. 備考・ご注意")
        row += 1
        for n in doc.notes:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            c = ws.cell(row=row, column=1, value=n)
            c.alignment = _left
            c.font = Font(color="C00000")
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = _border
            ws.row_dimensions[row].height = 30
            row += 1

    _setup_a4(ws, row - 1)


def build(data: SettlementData, buyer: SettlementDoc, seller: SettlementDoc) -> bytes:
    """買主用・売主用の2シートを持つ決済案内書 xlsx をバイト列で返す。"""
    wb = Workbook()
    ws_buyer = wb.active
    ws_buyer.title = "買主用決済案内書"
    _write_sheet(ws_buyer, data, buyer)

    ws_seller = wb.create_sheet("売主用決済案内書")
    _write_sheet(ws_seller, data, seller)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
